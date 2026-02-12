from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash, send_file, make_response, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import text, or_
from sqlalchemy.orm import joinedload
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta, date
import os
from dotenv import load_dotenv

# Carregar variáveis de ambiente do arquivo .env
load_dotenv()
import logging
from logging.handlers import RotatingFileHandler
import re
from shutil import copy2
import socket
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from io import BytesIO
import hashlib
import json
from decimal import Decimal
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
import atexit

# Configuração de logs simplificada
if not os.path.exists('logs'):
    os.makedirs('logs')

# Configurar logging apenas para console para evitar problemas de permissão
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]',
    handlers=[logging.StreamHandler()]
)

# SEMPRE usar templates HTML da pasta static
app = Flask(__name__, static_folder='static', static_url_path='/static')
print("Usando templates HTML do sistema")
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', os.urandom(32).hex())
# Configuração do banco de dados (Suporte a PostgreSQL para produção)
database_url = os.getenv('DATABASE_URL', 'sqlite:///saas_financeiro_v2.db')
if database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql://", 1)
app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Configurações de segurança e performance
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)  # Sessão expira em 24 horas
app.config['SESSION_COOKIE_SECURE'] = False  # Para desenvolvimento local
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# Configuração da pasta de upload para relatórios
app.config['UPLOAD_FOLDER'] = 'uploads'
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

# Configurar logging para Flask (simplificado)
if not app.debug:
    app.logger.setLevel(logging.INFO)
    app.logger.info('Sistema SAAS Financeiro iniciado')

db = SQLAlchemy(app)

# Função helper para verificar colunas (compatível com SQLite e PostgreSQL)
def verificar_coluna_existe(tabela, coluna):
    """Verifica se uma coluna existe em uma tabela (SQLite ou PostgreSQL)"""
    from sqlalchemy import text
    try:
        # Detectar tipo de banco
        dialect = db.engine.dialect.name

        if dialect == 'postgresql':
            result = db.session.execute(text("""
                SELECT column_name FROM information_schema.columns
                WHERE table_name = :tabela AND column_name = :coluna
            """), {'tabela': tabela, 'coluna': coluna})
            return result.fetchone() is not None
        else:  # sqlite
            result = db.session.execute(text(f"PRAGMA table_info({tabela})"))
            columns = [row[1] for row in result.fetchall()]
            return coluna in columns
    except Exception as e:
        print(f"⚠️ Aviso: Não foi possível verificar coluna {coluna}: {e}")
        db.session.rollback()  # Rollback em caso de erro
        return False

# Criar todas as tabelas automaticamente
with app.app_context():
    db.create_all()
    
    # Verificar e adicionar coluna tipo_produto_servico se não existir
    try:
        from sqlalchemy import text

        if not verificar_coluna_existe('lancamento', 'tipo_produto_servico'):
            print("Adicionando coluna 'tipo_produto_servico' ao banco de dados...")
            db.session.execute(text("ALTER TABLE lancamento ADD COLUMN tipo_produto_servico VARCHAR(20)"))
            db.session.commit()
            print("Coluna 'tipo_produto_servico' adicionada com sucesso!")
        else:
            print("Coluna 'tipo_produto_servico' já existe no banco de dados!")

        # Adicionar coluna itens_carrinho se não existir
        if not verificar_coluna_existe('lancamento', 'itens_carrinho'):
            print("Adicionando coluna 'itens_carrinho' ao banco de dados...")
            db.session.execute(text("ALTER TABLE lancamento ADD COLUMN itens_carrinho TEXT"))
            db.session.commit()
            print("Coluna 'itens_carrinho' adicionada com sucesso!")
        else:
            print("Coluna 'itens_carrinho' já existe no banco de dados!")

    except Exception as e:
        print(f"⚠️ Aviso: Não foi possível verificar/adicionar coluna: {str(e)}")
        print("Execute o script ATUALIZAR_BANCO_TIPO_PRODUTO.bat manualmente se necessário.")
        db.session.rollback()
    
    # Verificar e adicionar coluna 'usuario' na tabela sub_usuario_contador se não existir
    try:
        from sqlalchemy import text

        if not verificar_coluna_existe('sub_usuario_contador', 'usuario'):
            print("Adicionando coluna 'usuario' na tabela sub_usuario_contador...")
            db.session.execute(text("ALTER TABLE sub_usuario_contador ADD COLUMN usuario VARCHAR(50)"))
            db.session.commit()
            print("Coluna 'usuario' adicionada!")
            
            # Preencher registros existentes usando email ou nome como fallback
            print("Preenchendo campo 'usuario' nos registros existentes...")
            db.session.execute(text("""
                UPDATE sub_usuario_contador 
                SET usuario = COALESCE(email, nome, 'usuario_' || id)
                WHERE usuario IS NULL
            """))
            db.session.commit()
            print("Registros atualizados!")
        else:
            print("Coluna 'usuario' já existe na tabela sub_usuario_contador!")
            
    except Exception as e:
        print(f"⚠️ Aviso: Não foi possível verificar/adicionar coluna 'usuario': {str(e)}")
        db.session.rollback()
    
    # Verificar e adicionar coluna 'data_inicio_assinatura' na tabela empresa se não existir
    try:
        from sqlalchemy import text
        if not verificar_coluna_existe('empresa', 'data_inicio_assinatura'):
            print("Adicionando coluna 'data_inicio_assinatura' na tabela empresa...")
            db.session.execute(text("ALTER TABLE empresa ADD COLUMN data_inicio_assinatura DATETIME"))
            db.session.commit()
            print("Coluna 'data_inicio_assinatura' adicionada!")
            
            # Definir data_inicio_assinatura para empresas existentes baseado na data_criacao
            print("Definindo data_inicio_assinatura para empresas existentes...")
            db.session.execute(text("""
                UPDATE empresa 
                SET data_inicio_assinatura = data_criacao
                WHERE data_inicio_assinatura IS NULL
            """))
            db.session.commit()
            print("Datas de início definidas!")
        else:
            print("Coluna 'data_inicio_assinatura' já existe na tabela empresa!")
            
    except Exception as e:
        print(f"⚠️ Aviso: Não foi possível verificar/adicionar coluna 'data_inicio_assinatura': {str(e)}")
        db.session.rollback()
    
    # Verificar e adicionar coluna 'ativo' na tabela produto se não existir
    try:
        from sqlalchemy import text
        if not verificar_coluna_existe('produto', 'ativo'):
            print("Adicionando coluna 'ativo' na tabela produto...")
            db.session.execute(text("ALTER TABLE produto ADD COLUMN ativo BOOLEAN DEFAULT 1"))
            db.session.commit()
            print("Coluna 'ativo' adicionada na tabela produto!")
            
            # Definir ativo = 1 (True) para produtos existentes que tenham estoque > 0
            print("Ativando produtos com estoque > 0...")
            db.session.execute(text("""
                UPDATE produto 
                SET ativo = 1
                WHERE estoque > 0
            """))
            db.session.commit()
            print("Produtos com estoque ativados!")
        else:
            print("Coluna 'ativo' já existe na tabela produto!")
            
    except Exception as e:
        print(f"⚠️ Aviso: Não foi possível verificar/adicionar coluna 'ativo': {str(e)}")
        db.session.rollback()

    # Verificar e adicionar colunas de rastreamento de usuário na tabela lancamento
    try:
        from sqlalchemy import text

        colunas_adicionadas = []

        if not verificar_coluna_existe('lancamento', 'usuario_criacao_id'):
            print("Adicionando coluna 'usuario_criacao_id' na tabela lancamento...")
            db.session.execute(text("ALTER TABLE lancamento ADD COLUMN usuario_criacao_id INTEGER"))
            colunas_adicionadas.append('usuario_criacao_id')

        if not verificar_coluna_existe('lancamento', 'usuario_ultima_edicao_id'):
            print("Adicionando coluna 'usuario_ultima_edicao_id' na tabela lancamento...")
            db.session.execute(text("ALTER TABLE lancamento ADD COLUMN usuario_ultima_edicao_id INTEGER"))
            colunas_adicionadas.append('usuario_ultima_edicao_id')

        if not verificar_coluna_existe('lancamento', 'data_ultima_edicao'):
            print("Adicionando coluna 'data_ultima_edicao' na tabela lancamento...")
            db.session.execute(text("ALTER TABLE lancamento ADD COLUMN data_ultima_edicao DATETIME"))
            colunas_adicionadas.append('data_ultima_edicao')

        if colunas_adicionadas:
            db.session.commit()
            print(f"✅ Colunas de rastreamento adicionadas: {', '.join(colunas_adicionadas)}")

            # Preencher usuario_criacao_id com usuario_id para lançamentos existentes
            print("Preenchendo usuario_criacao_id nos lançamentos existentes...")
            db.session.execute(text("""
                UPDATE lancamento
                SET usuario_criacao_id = usuario_id
                WHERE usuario_criacao_id IS NULL
            """))
            db.session.commit()
            print("Lançamentos existentes atualizados!")
        else:
            print("Colunas de rastreamento de usuário já existem na tabela lancamento!")

    except Exception as e:
        print(f"⚠️ Aviso: Não foi possível verificar/adicionar colunas de rastreamento: {str(e)}")
        db.session.rollback()

    # ============================================================================
    # MIGRAÇÃO: Adicionar empresa_id a Cliente, Fornecedor, Venda, Compra
    # Para correção do bug BPO - isolamento multi-tenant
    # Suporta SQLite e PostgreSQL
    # ============================================================================
    try:
        from sqlalchemy import text, inspect

        # Detectar tipo de banco
        db_type = db.engine.name
        print(f"🔍 Banco de dados detectado: {db_type}")

        # Lista de tabelas que precisam do empresa_id
        tabelas = ['cliente', 'fornecedor', 'venda', 'compra']

        for tabela in tabelas:
            # Verificar se empresa_id já existe na tabela usando a função helper
            if not verificar_coluna_existe(tabela, 'empresa_id'):
                print(f"Adicionando coluna 'empresa_id' na tabela {tabela}...")

                # Adicionar coluna empresa_id
                db.session.execute(text(f"ALTER TABLE {tabela} ADD COLUMN empresa_id INTEGER"))
                db.session.commit()
                print(f"✅ Coluna 'empresa_id' adicionada na tabela {tabela}!")

                # Preencher empresa_id com base no usuario_id
                print(f"Preenchendo empresa_id na tabela {tabela} com base no usuario_id...")

                if db_type == 'sqlite':
                    db.session.execute(text(f"""
                        UPDATE {tabela}
                        SET empresa_id = (
                            SELECT empresa_id FROM usuario WHERE usuario.id = {tabela}.usuario_id
                        )
                        WHERE empresa_id IS NULL
                    """))
                else:  # PostgreSQL
                    db.session.execute(text(f"""
                        UPDATE {tabela}
                        SET empresa_id = u.empresa_id
                        FROM usuario u
                        WHERE {tabela}.usuario_id = u.id
                          AND {tabela}.empresa_id IS NULL
                    """))

                db.session.commit()
                print(f"✅ Registros da tabela {tabela} atualizados com empresa_id!")
            else:
                print(f"✓ Coluna 'empresa_id' já existe na tabela {tabela}!")

    except Exception as e:
        print(f"⚠️ Aviso: Não foi possível verificar/adicionar empresa_id: {str(e)}")
        db.session.rollback()

    # ============================================================================
    # MIGRAÇÃO: Adicionar nota_fiscal a Lancamento, Venda, Compra
    # ============================================================================
    try:
        from sqlalchemy import text

        # Lista de tabelas que precisam do nota_fiscal
        tabelas_nf = ['lancamento', 'venda', 'compra']

        for tabela in tabelas_nf:
            # Verificar se nota_fiscal já existe na tabela usando a função helper
            if not verificar_coluna_existe(tabela, 'nota_fiscal'):
                print(f"Adicionando coluna 'nota_fiscal' na tabela {tabela}...")
                db.session.execute(text(f"ALTER TABLE {tabela} ADD COLUMN nota_fiscal VARCHAR(50)"))
                db.session.commit()
                print(f"✅ Coluna 'nota_fiscal' adicionada na tabela {tabela}!")
            else:
                print(f"✓ Coluna 'nota_fiscal' já existe na tabela {tabela}!")

    except Exception as e:
        print(f"⚠️ Aviso: Não foi possível verificar/adicionar nota_fiscal: {str(e)}")
        db.session.rollback()

    # ============================================================================
    # MIGRAÇÃO: Colunas adicionais na tabela importacao
    # ============================================================================
    try:
        colunas_importacao = {
            'total_entradas': 'FLOAT DEFAULT 0.0',
            'total_saidas': 'FLOAT DEFAULT 0.0',
            'total_vendas': 'INTEGER DEFAULT 0',
            'total_compras': 'INTEGER DEFAULT 0',
            'lancamentos_ids': 'TEXT',
            'vendas_ids': 'TEXT',
            'compras_ids': 'TEXT'
        }

        for coluna, tipo in colunas_importacao.items():
            if not verificar_coluna_existe('importacao', coluna):
                print(f"Adicionando coluna '{coluna}' na tabela importacao...")
                db.session.execute(text(f"ALTER TABLE importacao ADD COLUMN {coluna} {tipo}"))
                db.session.commit()
                print(f"✅ Coluna '{coluna}' adicionada na tabela importacao!")
            else:
                print(f"✓ Coluna '{coluna}' já existe na tabela importacao!")
    except Exception as e:
        print(f"⚠️ Aviso: Não foi possível verificar/adicionar colunas na importacao: {str(e)}")
        db.session.rollback()

    # ============================================================================
    # MIGRAÇÃO: Hierarquia no Plano de Contas
    # ============================================================================
    try:
        # 1. Colunas para plano_conta
        novas_colunas_pc = {
            'codigo': 'VARCHAR(50)',
            'natureza': 'VARCHAR(20) DEFAULT "analitica"',
            'nivel': 'INTEGER DEFAULT 1',
            'pai_id': 'INTEGER',
            'empresa_id': 'INTEGER'
        }

        for col, tipo in novas_colunas_pc.items():
            if not verificar_coluna_existe('plano_conta', col):
                print(f"Adicionando coluna '{col}' na tabela plano_conta...")
                db.session.execute(text(f"ALTER TABLE plano_conta ADD COLUMN {col} {tipo}"))
                db.session.commit()
                print(f"✅ Coluna '{col}' adicionada!")

        # 2. Coluna para lancamento
        if not verificar_coluna_existe('lancamento', 'plano_conta_id'):
            print("Adicionando coluna 'plano_conta_id' na tabela lancamento...")
            db.session.execute(text("ALTER TABLE lancamento ADD COLUMN plano_conta_id INTEGER"))
            db.session.commit()
            print("✅ Coluna 'plano_conta_id' adicionada!")

    except Exception as e:
        print(f"⚠️ Aviso: Não foi possível verificar/adicionar colunas de hierarquia: {str(e)}")
        db.session.rollback()

    # ============================================================================
    # MIGRAÇÃO: Adicionar Índices para Performance
    # ============================================================================
    try:
        print("🔍 Verificando e criando índices para melhor performance...")

        # Lista de índices a serem criados
        indices = [
            # Índices compostos para Lancamento (queries mais comuns)
            ("idx_lancamento_empresa_data", "lancamento", ["empresa_id", "data_prevista"]),
            ("idx_lancamento_empresa_tipo_realizado", "lancamento", ["empresa_id", "tipo", "realizado"]),
            ("idx_lancamento_transferencia", "lancamento", ["transferencia_id"]),
            ("idx_lancamento_usuario_empresa", "lancamento", ["usuario_id", "empresa_id"]),

            # Índices para Cliente
            ("idx_cliente_empresa", "cliente", ["empresa_id"]),
            ("idx_cliente_usuario_empresa", "cliente", ["usuario_id", "empresa_id"]),

            # Índices para Fornecedor
            ("idx_fornecedor_empresa", "fornecedor", ["empresa_id"]),
            ("idx_fornecedor_usuario_empresa", "fornecedor", ["usuario_id", "empresa_id"]),

            # Índices para Venda
            ("idx_venda_empresa_data", "venda", ["empresa_id", "data_prevista"]),
            ("idx_venda_usuario_empresa", "venda", ["usuario_id", "empresa_id"]),

            # Índices para Compra
            ("idx_compra_empresa_data", "compra", ["empresa_id", "data_prevista"]),
            ("idx_compra_usuario_empresa", "compra", ["usuario_id", "empresa_id"]),

            # Índices para VinculoContador
            ("idx_vinculo_contador_status", "vinculo_contador", ["contador_id", "status"]),
            ("idx_vinculo_contador_empresa", "vinculo_contador", ["empresa_id", "status"]),
        ]

        for nome_indice, tabela, colunas in indices:
            try:
                # Usar CREATE INDEX IF NOT EXISTS (funciona em SQLite e PostgreSQL)
                colunas_str = ", ".join(colunas)
                sql = f"CREATE INDEX IF NOT EXISTS {nome_indice} ON {tabela}({colunas_str})"
                db.session.execute(text(sql))
                db.session.commit()
                print(f"  ✅ Índice '{nome_indice}' criado/verificado em {tabela}({colunas_str})")
            except Exception as e:
                print(f"  ⚠️  Erro ao criar índice '{nome_indice}': {str(e)}")
                db.session.rollback()

        print("✅ Verificação de índices concluída!")

    except Exception as e:
        print(f"⚠️ Aviso: Erro ao verificar/criar índices: {str(e)}")
        db.session.rollback()

    print("Banco de dados inicializado com todas as tabelas")

# Modelos do banco de dados
class Empresa(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    tipo_pessoa = db.Column(db.String(20), nullable=False, default='PJ')  # PF, PJ ou CONTADOR
    cpf = db.Column(db.String(20), unique=True, nullable=True)  # Para pessoa física
    cnpj = db.Column(db.String(20), unique=True, nullable=True)  # Para pessoa jurídica
    razao_social = db.Column(db.String(200), nullable=False)
    nome_fantasia = db.Column(db.String(200))
    email = db.Column(db.String(120))
    telefone = db.Column(db.String(20))
    endereco = db.Column(db.Text)
    tipo_empresa = db.Column(db.String(20), default='servicos')  # servicos, comercio, industria
    tipo_conta = db.Column(db.String(20), default='empresa')  # empresa, pessoa_fisica, contador_bpo, admin
    dias_assinatura = db.Column(db.Integer, default=30)  # Dias restantes de assinatura
    data_inicio_assinatura = db.Column(db.DateTime, nullable=True)  # Data de início da assinatura para cálculo automático
    plano_id = db.Column(db.Integer, db.ForeignKey('plano.id'), nullable=True)  # Plano de assinatura ativo
    ativo = db.Column(db.Boolean, default=True)
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)

    # Relacionamentos
    usuarios = db.relationship('Usuario', backref='empresa', lazy=True)
    plano_ativo = db.relationship('Plano', foreign_keys=[plano_id], backref='empresas_ativas')

class Usuario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    usuario = db.Column(db.String(50), nullable=False)  # Nome de usuário para login
    email = db.Column(db.String(120), nullable=False)  # Removido unique=True para permitir emails iguais em empresas diferentes
    senha = db.Column(db.String(200), nullable=False)
    telefone = db.Column(db.String(20))
    tipo = db.Column(db.String(20), default='usuario')  # admin, usuario_principal, usuario
    categoria_id = db.Column(db.Integer, db.ForeignKey('categoria_usuario.id'), nullable=True)  # Categoria personalizada
    ativo = db.Column(db.Boolean, default=True)
    pausado = db.Column(db.Boolean, default=False)
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)  # Relacionamento com empresa
    criado_por = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=True)  # Quem criou este usuário
    
    # Relacionamentos
    usuarios_criados = db.relationship('Usuario', backref=db.backref('criador', remote_side=[id]))
    categoria = db.relationship('CategoriaUsuario', foreign_keys=[categoria_id], backref='usuarios')

class Lancamento(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    descricao = db.Column(db.String(200), nullable=False)
    valor = db.Column(db.Float, nullable=False)
    tipo = db.Column(db.String(20), nullable=False)  # entrada ou saida
    categoria = db.Column(db.String(100), nullable=False)
    data_prevista = db.Column(db.Date, nullable=False)
    data_realizada = db.Column(db.Date)
    realizado = db.Column(db.Boolean, default=False)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)
    compra_id = db.Column(db.Integer, db.ForeignKey('compra.id'), nullable=True)  # Relacionamento com compra
    venda_id = db.Column(db.Integer, db.ForeignKey('venda.id'), nullable=True)  # Relacionamento com venda
    conta_caixa_id = db.Column(db.Integer, db.ForeignKey('conta_caixa.id'), nullable=True)  # Relacionamento com conta caixa
    cliente_id = db.Column(db.Integer, db.ForeignKey('cliente.id'), nullable=True)  # Relacionamento com cliente
    fornecedor_id = db.Column(db.Integer, db.ForeignKey('fornecedor.id'), nullable=True)  # Relacionamento com fornecedor
    nota_fiscal = db.Column(db.String(50), nullable=True)  # Campo para nota fiscal
    
    # Campos adicionais para mais detalhamento
    observacoes = db.Column(db.Text, nullable=True)  # Observações adicionais
    produto_servico = db.Column(db.String(200), nullable=True)  # Produto/Serviço relacionado
    tipo_produto_servico = db.Column(db.String(20), nullable=True)  # Tipo: produto ou serviço
    itens_carrinho = db.Column(db.Text, nullable=True)  # JSON com itens do carrinho (quando usar_carrinho=True)
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)  # Data de criação do registro

    # Rastreamento de usuários que criaram e editaram o lançamento
    usuario_criacao_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=True)  # Usuário que criou
    usuario_ultima_edicao_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=True)  # Último usuário que editou
    data_ultima_edicao = db.Column(db.DateTime, nullable=True)  # Data da última edição

    # Plano de contas (ID)
    plano_conta_id = db.Column(db.Integer, db.ForeignKey('plano_conta.id'), nullable=True)

    # Campos para transferências entre contas
    eh_transferencia = db.Column(db.Boolean, default=False)  # Indica se é uma transferência
    transferencia_id = db.Column(db.Integer, nullable=True)  # ID do lançamento par na transferência

    # Relacionamentos
    usuario = db.relationship('Usuario', backref='lancamentos', lazy=True, foreign_keys=[usuario_id])
    usuario_criacao = db.relationship('Usuario', foreign_keys=[usuario_criacao_id], lazy=True)
    usuario_ultima_edicao = db.relationship('Usuario', foreign_keys=[usuario_ultima_edicao_id], lazy=True)

class Cliente(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(200), nullable=False)
    email = db.Column(db.String(120))
    telefone = db.Column(db.String(20))
    cpf_cnpj = db.Column(db.String(20))
    endereco = db.Column(db.Text)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)  # Multi-tenancy
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)

    # Relacionamentos
    lancamentos = db.relationship('Lancamento', backref='cliente', lazy=True)

class Fornecedor(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(200), nullable=False)
    email = db.Column(db.String(120))
    telefone = db.Column(db.String(20))
    cpf_cnpj = db.Column(db.String(20))
    endereco = db.Column(db.Text)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)  # Multi-tenancy
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)

    # Relacionamentos
    lancamentos = db.relationship('Lancamento', backref='fornecedor', lazy=True)

class Produto(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(200), nullable=False)
    descricao = db.Column(db.Text)
    preco_custo = db.Column(db.Float, default=0)
    preco_venda = db.Column(db.Float, default=0)
    estoque = db.Column(db.Integer, default=0)
    ativo = db.Column(db.Boolean, default=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)

class Servico(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(200), nullable=False)
    descricao = db.Column(db.Text)
    preco = db.Column(db.Float, default=0)
    ativo = db.Column(db.Boolean, default=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)

class PlanoConta(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(200), nullable=False)
    codigo = db.Column(db.String(50))
    tipo = db.Column(db.String(20), nullable=False)  # entrada ou saida
    descricao = db.Column(db.Text)
    natureza = db.Column(db.String(20), default='analitica')  # sintetica ou analitica
    nivel = db.Column(db.Integer, default=1)
    pai_id = db.Column(db.Integer, db.ForeignKey('plano_conta.id'), nullable=True)
    ativo = db.Column(db.Boolean, default=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=True)
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relacionamentos
    filhos = db.relationship('PlanoConta', backref=db.backref('pai', remote_side=[id]), lazy='dynamic')
    lancamentos = db.relationship('Lancamento', backref='plano_conta', lazy=True)

class Importacao(db.Model):
    __tablename__ = 'importacao'
    
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    nome_arquivo = db.Column(db.String(255), nullable=False)
    data_importacao = db.Column(db.DateTime, default=datetime.now)
    total_lancamentos = db.Column(db.Integer, default=0)
    total_entradas = db.Column(db.Float, default=0.0)
    total_saidas = db.Column(db.Float, default=0.0)
    total_vendas = db.Column(db.Integer, default=0)
    total_compras = db.Column(db.Integer, default=0)
    sucessos = db.Column(db.Integer, default=0)
    erros = db.Column(db.Integer, default=0)
    status = db.Column(db.String(20), default='concluida')  # concluida, desfeita
    lancamentos_ids = db.Column(db.Text)  # JSON com IDs dos lançamentos criados
    vendas_ids = db.Column(db.Text)  # JSON com IDs das vendas criadas
    compras_ids = db.Column(db.Text)  # JSON com IDs das compras criadas
    
    # Relacionamentos
    usuario = db.relationship('Usuario', backref='importacoes')

class ContaCaixa(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(200), nullable=False)
    tipo = db.Column(db.String(50), nullable=False)  # conta_corrente, poupanca, caixa_fisico, cartao_credito, etc.

    # Informações bancárias
    banco = db.Column(db.String(200))
    agencia = db.Column(db.String(50))
    conta = db.Column(db.String(50))

    produto_servico = db.Column(db.String(200))
    tipo_produto_servico = db.Column(db.String(50))

    # Campo para nota fiscal
    nota_fiscal = db.Column(db.String(50))

    # Plano de contas (ID)
    plano_conta_id = db.Column(db.Integer, db.ForeignKey('plano_conta.id'), nullable=True)

    # Relacionamentos
    saldo_inicial = db.Column(db.Float, default=0.0)
    saldo_atual = db.Column(db.Float, default=0.0)
    ativo = db.Column(db.Boolean, default=True)
    descricao = db.Column(db.Text)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)

    # Relacionamentos
    lancamentos = db.relationship('Lancamento', backref='conta_caixa', lazy=True)

    def calcular_saldo_atual(self):
        """
        Calcula o saldo atual da conta baseado nos lançamentos realizados
        Fórmula: saldo_inicial + entradas_realizadas - saidas_realizadas
        """
        from sqlalchemy import and_

        # Buscar apenas lançamentos realizados desta conta
        lancamentos_realizados = Lancamento.query.filter(
            and_(
                Lancamento.conta_caixa_id == self.id,
                Lancamento.realizado == True
            )
        ).all()

        # Calcular totais
        entradas = sum([l.valor for l in lancamentos_realizados if l.tipo == 'entrada'])
        saidas = sum([l.valor for l in lancamentos_realizados if l.tipo == 'saida'])

        return self.saldo_inicial + entradas - saidas

class Venda(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    cliente_id = db.Column(db.Integer, db.ForeignKey('cliente.id'), nullable=False)
    produto = db.Column(db.String(200), nullable=False)
    valor = db.Column(db.Float, nullable=False)
    quantidade = db.Column(db.Integer, default=1)  # Adicionando campo quantidade
    tipo_venda = db.Column(db.String(20), default='produto')  # produto ou servico
    data_prevista = db.Column(db.Date, nullable=False)
    data_realizada = db.Column(db.Date)
    realizado = db.Column(db.Boolean, default=False)
    observacoes = db.Column(db.Text)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)  # Multi-tenancy
    nota_fiscal = db.Column(db.String(50), nullable=True)  # Campo para nota fiscal
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)

    # Novos campos para pagamento parcelado e desconto
    tipo_pagamento = db.Column(db.String(20), default='a_vista')  # a_vista ou parcelado
    numero_parcelas = db.Column(db.Integer, default=1)  # Número de parcelas
    valor_parcela = db.Column(db.Float)  # Valor de cada parcela
    desconto = db.Column(db.Float, default=0.0)  # Valor do desconto
    valor_final = db.Column(db.Float)  # Valor final após desconto

    # Relacionamentos
    cliente = db.relationship('Cliente', backref='vendas')
    usuario = db.relationship('Usuario', backref='vendas', lazy=True)
    lancamento_financeiro = db.relationship('Lancamento', backref='venda', uselist=False)
    parcelas = db.relationship('Parcela', backref='venda', lazy=True, cascade='all, delete-orphan')

class Compra(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    fornecedor_id = db.Column(db.Integer, db.ForeignKey('fornecedor.id'), nullable=False)
    produto = db.Column(db.String(200), nullable=False)
    valor = db.Column(db.Float, nullable=False)
    quantidade = db.Column(db.Integer, default=1)
    preco_custo = db.Column(db.Float, nullable=False)
    tipo_compra = db.Column(db.String(20), default='mercadoria')  # mercadoria ou servico
    data_prevista = db.Column(db.Date, nullable=False)
    data_realizada = db.Column(db.Date)
    realizado = db.Column(db.Boolean, default=False)
    observacoes = db.Column(db.Text)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)  # Multi-tenancy
    nota_fiscal = db.Column(db.String(50), nullable=True)  # Campo para nota fiscal
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)

    # Novos campos para pagamento parcelado
    tipo_pagamento = db.Column(db.String(20), default='a_vista')  # a_vista ou parcelado
    numero_parcelas = db.Column(db.Integer, default=1)  # Número de parcelas
    valor_parcela = db.Column(db.Float)  # Valor de cada parcela

    # Relacionamentos
    fornecedor = db.relationship('Fornecedor', backref='compras')
    usuario = db.relationship('Usuario', backref='compras', lazy=True)
    lancamento_financeiro = db.relationship('Lancamento', backref='compra', uselist=False)
    parcelas = db.relationship('Parcela', backref='compra', lazy=True, cascade='all, delete-orphan')

# Novo modelo para Parcelas
class Parcela(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    numero = db.Column(db.Integer, nullable=False)  # Número da parcela (1, 2, 3, etc.)
    valor = db.Column(db.Float, nullable=False)  # Valor da parcela
    data_vencimento = db.Column(db.Date, nullable=False)  # Data de vencimento
    data_pagamento = db.Column(db.Date)  # Data de pagamento (quando realizado)
    realizado = db.Column(db.Boolean, default=False)  # Se foi paga
    venda_id = db.Column(db.Integer, db.ForeignKey('venda.id'), nullable=True)  # Relacionamento com venda
    compra_id = db.Column(db.Integer, db.ForeignKey('compra.id'), nullable=True)  # Relacionamento com compra
    lancamento_id = db.Column(db.Integer, db.ForeignKey('lancamento.id'), nullable=True)  # Relacionamento com lançamento
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relacionamentos
    lancamento = db.relationship('Lancamento', backref='parcelas')

# Novas tabelas para vínculos e auditoria
class Vinculo(db.Model):
    """Tabela pivot para relacionar entidades do sistema"""
    id = db.Column(db.Integer, primary_key=True)
    lado_a_tipo = db.Column(db.String(50), nullable=False)  # 'venda', 'compra', 'lancamento', 'estoque'
    lado_a_id = db.Column(db.Integer, nullable=False)
    lado_b_tipo = db.Column(db.String(50), nullable=False)  # 'venda', 'compra', 'lancamento', 'estoque'
    lado_b_id = db.Column(db.Integer, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=True)
    
    # Relacionamentos
    usuario = db.relationship('Usuario', backref='vinculos')

class EventLog(db.Model):
    """Tabela para registrar eventos e evitar duplicação"""
    id = db.Column(db.Integer, primary_key=True)
    tipo_evento = db.Column(db.String(100), nullable=False)  # 'estoque_venda', 'estoque_compra', 'financeiro_venda', 'financeiro_compra'
    origem_tipo = db.Column(db.String(50), nullable=False)  # 'venda', 'compra'
    origem_id = db.Column(db.Integer, nullable=False)
    hash_evento = db.Column(db.String(64), unique=True, nullable=False)  # Hash único para idempotência
    dados_evento = db.Column(db.Text)  # JSON com dados do evento
    status = db.Column(db.String(20), default='processado')  # 'processado', 'erro', 'revertido'
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    
    # Relacionamentos
    usuario = db.relationship('Usuario', backref='event_logs')

# Novo modelo para Permissões
class Permissao(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    modulo = db.Column(db.String(50), nullable=False)  # lancamentos, clientes, fornecedores, estoque, vendas, compras, relatorios, configuracoes
    acao = db.Column(db.String(50), nullable=False)  # visualizar, criar, editar, deletar
    ativo = db.Column(db.Boolean, default=True)
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relacionamentos
    usuario = db.relationship('Usuario', backref='permissoes')

# Novo modelo para Categorias de Usuários Personalizadas
class CategoriaUsuario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    descricao = db.Column(db.Text)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=True)
    ativo = db.Column(db.Boolean, default=True)
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)
    criado_por = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=True)
    
    # Relacionamentos
    empresa = db.relationship('Empresa', backref='categorias_usuario')
    criador = db.relationship('Usuario', foreign_keys=[criado_por], backref='categorias_criadas')

# Novo modelo para Permissões de Categoria
class PermissaoCategoria(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    categoria_id = db.Column(db.Integer, db.ForeignKey('categoria_usuario.id'), nullable=False)
    modulo = db.Column(db.String(50), nullable=False)  # lancamentos, clientes, fornecedores, estoque, vendas, compras, relatorios, configuracoes
    acao = db.Column(db.String(50), nullable=False)  # visualizar, criar, editar, deletar
    ativo = db.Column(db.Boolean, default=True)
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relacionamentos
    categoria = db.relationship('CategoriaUsuario', backref='permissoes')

class VinculoContador(db.Model):
    """Tabela para vincular contadores/BPO com empresas/pessoas físicas"""
    __tablename__ = 'vinculo_contador'
    
    id = db.Column(db.Integer, primary_key=True)
    contador_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)
    status = db.Column(db.String(20), default='pendente')  # pendente, autorizado, rejeitado
    data_solicitacao = db.Column(db.DateTime, default=datetime.utcnow)
    data_autorizacao = db.Column(db.DateTime, nullable=True)
    
    # Relacionamentos
    contador = db.relationship('Empresa', foreign_keys=[contador_id], backref='empresas_vinculadas')
    empresa = db.relationship('Empresa', foreign_keys=[empresa_id], backref='contadores_vinculados')

class SubUsuarioContador(db.Model):
    """Sub-usuários criados dentro de uma conta Contador/BPO"""
    __tablename__ = 'sub_usuario_contador'
    
    id = db.Column(db.Integer, primary_key=True)
    contador_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)
    nome = db.Column(db.String(100), nullable=False)  # Nome completo da pessoa
    usuario = db.Column(db.String(50), nullable=True)  # Nome de usuário para login (nullable temporariamente para migração)
    email = db.Column(db.String(120), nullable=True)  # Email opcional
    senha = db.Column(db.String(200), nullable=False)
    ativo = db.Column(db.Boolean, default=True)
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)
    
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        # Se usuario não foi fornecido, usar email ou nome como fallback
        if not self.usuario:
            self.usuario = self.email if self.email else (self.nome.lower().replace(' ', '_') if self.nome else f'usuario_{self.id}')
    
    # Relacionamentos
    contador = db.relationship('Empresa', backref='sub_usuarios')
    permissoes_empresas = db.relationship('PermissaoSubUsuario', backref='sub_usuario', lazy=True, cascade='all, delete-orphan')

class PermissaoSubUsuario(db.Model):
    """Controla quais empresas cada sub-usuário pode acessar"""
    __tablename__ = 'permissao_sub_usuario'

    id = db.Column(db.Integer, primary_key=True)
    sub_usuario_id = db.Column(db.Integer, db.ForeignKey('sub_usuario_contador.id'), nullable=False)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)

    # Relacionamentos
    empresa = db.relationship('Empresa', backref='permissoes_sub_usuarios')

class Voucher(db.Model):
    """Modelo para vouchers de assinatura"""
    __tablename__ = 'voucher'

    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(50), unique=True, nullable=False, index=True)
    dias_assinatura = db.Column(db.Integer, nullable=False)
    validade = db.Column(db.DateTime, nullable=False)
    ativo = db.Column(db.Boolean, default=True)
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)
    criado_por = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=True)

    # Relacionamentos
    criador = db.relationship('Usuario', foreign_keys=[criado_por], backref='vouchers_criados')
    usos = db.relationship('VoucherUso', backref='voucher', lazy=True, cascade='all, delete-orphan')

    def __repr__(self):
        return f'<Voucher {self.codigo}>'

    def esta_valido(self):
        """Verifica se o voucher está válido (ativo e não expirado)"""
        return self.ativo and datetime.utcnow() < self.validade

    def pode_usar(self):
        """Verifica se o voucher pode ser usado (válido e não foi usado)"""
        if not self.esta_valido():
            return False
        return len(self.usos) == 0

class VoucherUso(db.Model):
    """Modelo para auditoria de uso de vouchers"""
    __tablename__ = 'voucher_uso'

    id = db.Column(db.Integer, primary_key=True)
    voucher_id = db.Column(db.Integer, db.ForeignKey('voucher.id'), nullable=False)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)
    data_uso = db.Column(db.DateTime, default=datetime.utcnow)
    dias_creditados = db.Column(db.Integer, nullable=False)
    usuario_admin_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=True)
    observacoes = db.Column(db.Text, nullable=True)

    # Relacionamentos
    empresa = db.relationship('Empresa', backref='vouchers_usados')
    usuario_admin = db.relationship('Usuario', foreign_keys=[usuario_admin_id], backref='vouchers_aplicados')

    def __repr__(self):
        return f'<VoucherUso {self.voucher_id} -> {self.empresa_id}>'

class Plano(db.Model):
    """Modelo para planos de assinatura"""
    __tablename__ = 'plano'

    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    codigo = db.Column(db.String(50), unique=True, nullable=False)
    dias_assinatura = db.Column(db.Integer, nullable=False)
    valor = db.Column(db.Float, nullable=False)
    descricao = db.Column(db.Text, nullable=True)
    ativo = db.Column(db.Boolean, default=True)
    ordem_exibicao = db.Column(db.Integer, default=0)
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)

    pagamentos = db.relationship('Pagamento', backref='plano', lazy=True)

    def __repr__(self):
        return f'<Plano {self.nome} - R$ {self.valor}>'

class Pagamento(db.Model):
    """Modelo para registro de pagamentos"""
    __tablename__ = 'pagamento'

    id = db.Column(db.Integer, primary_key=True)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=True)
    plano_id = db.Column(db.Integer, db.ForeignKey('plano.id'), nullable=False)

    preference_id = db.Column(db.String(200), nullable=True)
    payment_id = db.Column(db.String(200), nullable=True)
    external_reference = db.Column(db.String(200), unique=True, nullable=False)

    valor = db.Column(db.Float, nullable=False)
    dias_assinatura = db.Column(db.Integer, nullable=False)

    status = db.Column(db.String(50), default='pending')
    status_detail = db.Column(db.String(200), nullable=True)

    payment_type = db.Column(db.String(50), nullable=True)
    payment_method = db.Column(db.String(50), nullable=True)

    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)
    data_aprovacao = db.Column(db.DateTime, nullable=True)
    data_expiracao = db.Column(db.DateTime, nullable=True)

    observacoes = db.Column(db.Text, nullable=True)
    dados_mp = db.Column(db.Text, nullable=True)

    empresa = db.relationship('Empresa', backref='pagamentos')

    def __repr__(self):
        return f'<Pagamento {self.external_reference} - {self.status}>'

class DreConfiguracao(db.Model):
    """Modelo para configuração da DRE - Demonstração do Resultado do Exercício"""
    __tablename__ = 'dre_configuracao'

    id = db.Column(db.Integer, primary_key=True)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)
    plano_conta_id = db.Column(db.Integer, db.ForeignKey('plano_conta.id'), nullable=True)

    # Configuração da linha
    codigo = db.Column(db.String(20), nullable=False)  # Código único da linha (ex: "010", "020")
    descricao = db.Column(db.String(200), nullable=False)  # Descrição customizada
    tipo_linha = db.Column(db.String(20), default='conta')  # 'conta', 'subtotal', 'total'
    formula = db.Column(db.String(500), nullable=True)  # Para linhas de total. Ex: "(+) 010 020 (-) 030"
    operacao = db.Column(db.String(10), nullable=True)  # '+', '-', '=' (para indicar se soma/subtrai/total)

    # Ordenação e hierarquia
    ordem = db.Column(db.Integer, nullable=False)  # Ordem de exibição
    nivel = db.Column(db.Integer, default=1)  # Nível de indentação (1, 2, 3...)

    # Formatação
    negrito = db.Column(db.Boolean, default=False)  # Se deve aparecer em negrito
    linha_acima = db.Column(db.Boolean, default=False)  # Se tem linha acima
    linha_abaixo = db.Column(db.Boolean, default=False)  # Se tem linha abaixo

    ativo = db.Column(db.Boolean, default=True)
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)

    # Relacionamentos
    empresa = db.relationship('Empresa', backref='dre_configuracoes')
    plano_conta = db.relationship('PlanoConta', backref='linhas_dre')

    def __repr__(self):
        return f'<DreConfiguracao {self.codigo} - {self.descricao}>'

# ===== INICIALIZAÇÃO DOS SERVIÇOS =====
def obter_modelos():
    """Retorna dicionário com todos os modelos para os serviços"""
    return {
        'Empresa': Empresa,
        'Usuario': Usuario,
        'Importacao': Importacao,
        'Cliente': Cliente,
        'Fornecedor': Fornecedor,
        'Produto': Produto,
        'Servico': Servico,
        'PlanoConta': PlanoConta,
        'ContaCaixa': ContaCaixa,
        'Venda': Venda,
        'Compra': Compra,
        'Parcela': Parcela,
        'Vinculo': Vinculo,
        'EventLog': EventLog,
        'Permissao': Permissao,
        'VinculoContador': VinculoContador,
        'SubUsuarioContador': SubUsuarioContador,
        'PermissaoSubUsuario': PermissaoSubUsuario
    }

# Função auxiliar para validação consistente de sessão
def validar_sessao_ativa():
    """
    Valida se há uma sessão ativa (usuário ou sub-usuário).
    Retorna (valido, usuario_ou_sub_usuario, tipo_usuario)
    """
    from flask import session, redirect, url_for

    # Verificar se há usuario_id ou sub_usuario_id na sessão
    if 'usuario_id' in session:
        usuario = db.session.get(Usuario, session['usuario_id'])
        if usuario and usuario.ativo:
            return True, usuario, 'usuario'
        else:
            session.clear()
            return False, None, None

    elif 'sub_usuario_id' in session:
        sub_usuario = db.session.get(SubUsuarioContador, session['sub_usuario_id'])
        if sub_usuario and sub_usuario.ativo:
            return True, sub_usuario, 'sub_usuario'
        else:
            session.clear()
            return False, None, None

    return False, None, None

# Função auxiliar para obter empresa_id correta da sessão
def obter_empresa_id_sessao(session, usuario):
    """
    Retorna a empresa_id correta baseada na sessão.
    Se acesso_contador estiver ativo, valida o vínculo e retorna a empresa vinculada.
    Caso contrário, retorna a empresa do usuário.
    """
    if session.get('acesso_contador') and session.get('empresa_id'):
        # SEGURANÇA: Validar que o vínculo ainda está autorizado
        empresa_vinculada_id = session.get('empresa_id')

        # Determinar contador_id baseado no tipo de usuário
        if session.get('usuario_tipo') == 'sub_contador':
            contador_id = session.get('contador_id')
        else:
            contador_id = session.get('empresa_id_original')

        if contador_id:
            # Verificar se o vínculo existe e está autorizado
            vinculo = VinculoContador.query.filter_by(
                contador_id=contador_id,
                empresa_id=empresa_vinculada_id,
                status='autorizado'
            ).first()

            if vinculo:
                # Vínculo válido - para sub-usuários, verificar permissão adicional
                if session.get('usuario_tipo') == 'sub_contador':
                    sub_usuario_id = session.get('sub_usuario_id')
                    permissao = PermissaoSubUsuario.query.filter_by(
                        sub_usuario_id=sub_usuario_id,
                        empresa_id=empresa_vinculada_id
                    ).first()

                    if not permissao:
                        # Sub-usuário sem permissão - limpar sessão
                        session.pop('acesso_contador', None)
                        session.pop('empresa_id', None)
                        return usuario.empresa_id if usuario and usuario.empresa_id else None

                return empresa_vinculada_id
            else:
                # Vínculo não autorizado ou removido - limpar sessão
                session.pop('acesso_contador', None)
                session.pop('empresa_id', None)
                return usuario.empresa_id if usuario and usuario.empresa_id else None

        # Se não tiver contador_id, retornar empresa_id do usuário
        return usuario.empresa_id if usuario and usuario.empresa_id else None

    return usuario.empresa_id if usuario and usuario.empresa_id else None

# Funções auxiliares para parcelamento
def criar_parcelas_automaticas(venda_ou_compra, tipo, usuario_id):
    """
    Cria parcelas automaticamente para vendas ou compras parceladas
    """
    try:
        if tipo == 'venda':
            entidade = venda_ou_compra
            entidade_id = entidade.id
        else:  # compra
            entidade = venda_ou_compra
            entidade_id = entidade.id
        
        # Verificar se é parcelado
        if entidade.tipo_pagamento != 'parcelado' or entidade.numero_parcelas <= 1:
            return True, "Pagamento à vista - não há parcelas para criar"
        
        # Verificar se já existem parcelas
        parcelas_existentes = Parcela.query.filter_by(
            venda_id=entidade_id if tipo == 'venda' else None,
            compra_id=entidade_id if tipo == 'compra' else None
        ).count()
        
        if parcelas_existentes > 0:
            return True, f"Parcelas já existem ({parcelas_existentes} parcelas)"
        
        # Calcular valor da parcela
        if hasattr(entidade, 'valor_final') and entidade.valor_final:
            valor_total = entidade.valor_final
        else:
            valor_total = entidade.valor

        # ✅ CORREÇÃO: Distribuir centavos corretamente para evitar diferenças de arredondamento
        import math
        numero_parcelas = entidade.numero_parcelas

        # Calcular valor base da parcela (arredondado para baixo com 2 casas decimais)
        valor_parcela_base = math.floor((valor_total / numero_parcelas) * 100) / 100

        # Calcular o resto (diferença devido ao arredondamento)
        resto = round(valor_total - (valor_parcela_base * numero_parcelas), 2)

        # Distribuir o resto nas primeiras parcelas (adicionar 0.01 em cada uma)
        parcelas_com_centavo_extra = int(round(resto * 100))  # Quantas parcelas receberão +0.01

        # Criar parcelas
        parcelas_criadas = []
        data_base = entidade.data_prevista

        for i in range(1, numero_parcelas + 1):
            # Calcular data de vencimento usando a função específica
            data_vencimento = calcular_data_vencimento_parcela(data_base, i, 'mensal')

            # Definir valor da parcela (primeiras parcelas recebem o centavo extra se houver resto)
            if i <= parcelas_com_centavo_extra:
                valor_parcela = round(valor_parcela_base + 0.01, 2)
            else:
                valor_parcela = valor_parcela_base

            parcela = Parcela(
                numero=i,
                valor=valor_parcela,
                data_vencimento=data_vencimento,
                realizado=False,
                venda_id=entidade_id if tipo == 'venda' else None,
                compra_id=entidade_id if tipo == 'compra' else None,
                usuario_id=usuario_id
            )

            db.session.add(parcela)
            parcelas_criadas.append(parcela)
        
        # Buscar o usuário para obter empresa_id
        usuario = db.session.get(Usuario, usuario_id)
        # Obter empresa_id correta da sessão (considera acesso_contador)
        from flask import session as flask_session
        empresa_id = obter_empresa_id_sessao(flask_session, usuario) if usuario else None
        
        # Determinar plano_conta_id base para as parcelas
        tipo_lancamento = 'entrada' if tipo == 'venda' else 'saida'
        categoria_nome = 'Vendas' if tipo == 'venda' else 'Compras'
        
        pc = buscar_plano_conta_automatico(usuario_id, tipo_lancamento, categoria_nome)
        plano_conta_id = pc.id if pc else None
        
        # Criar lançamento financeiro para cada parcela
        for parcela in parcelas_criadas:
            lancamento = Lancamento(
                descricao=f"{tipo.title()} - Parcela {parcela.numero}/{entidade.numero_parcelas} - {entidade.produto}",
                valor=parcela.valor,
                tipo=tipo_lancamento,
                categoria=categoria_nome,
                plano_conta_id=plano_conta_id,
                data_prevista=parcela.data_vencimento,
                realizado=False,
                usuario_id=usuario_id,
                empresa_id=empresa_id,
                venda_id=entidade_id if tipo == 'venda' else None,
                compra_id=entidade_id if tipo == 'compra' else None,
                cliente_id=entidade.cliente_id if tipo == 'venda' else None
            )
            
            db.session.add(lancamento)
        
        # Commit para gerar os IDs dos lançamentos
        db.session.commit()
        
        # Agora vincular as parcelas aos lançamentos
        for i, parcela in enumerate(parcelas_criadas):
            # Buscar o lançamento correspondente
            lancamento = Lancamento.query.filter_by(
                usuario_id=usuario_id,
                venda_id=entidade_id if tipo == 'venda' else None,
                compra_id=entidade_id if tipo == 'compra' else None,
                data_prevista=parcela.data_vencimento
            ).first()
            
            if lancamento:
                parcela.lancamento_id = lancamento.id
        
        db.session.commit()
        
        return True, f"{entidade.numero_parcelas} parcelas criadas com sucesso (R$ {valor_parcela:.2f} cada)"
        
    except Exception as e:
        db.session.rollback()
        return False, f"Erro ao criar parcelas: {str(e)}"

# Rotas
@app.route('/')
def landing():
    """Landing page pública - mostra para usuários não autenticados"""
    if 'usuario_id' in session:
        return redirect(url_for('dashboard'))
    return render_template('landing.html')


@app.route('/precos')
def precos():
    """Página de preços pública"""
    return render_template('precos.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    # Garantir que o admin existe
    try:
        empresa_admin = Empresa.query.filter_by(cnpj='00.000.000/0000-00').first()
        if not empresa_admin:
            empresa_admin = Empresa(
                cnpj='00.000.000/0000-00',
                razao_social='Sistema Administrativo',
                nome_fantasia='Admin',
                tipo_empresa='servicos',
                tipo_conta='admin',
                dias_assinatura=999999,
                ativo=True
            )
            db.session.add(empresa_admin)
            db.session.flush()
        
        if empresa_admin.tipo_conta != 'admin':
            empresa_admin.tipo_conta = 'admin'
            db.session.commit()
        
        usuario_admin = Usuario.query.filter_by(empresa_id=empresa_admin.id, usuario='admin').first()
        if not usuario_admin:
            usuario_admin = Usuario(
                nome='Administrador',
                usuario='admin',
                email='admin@sistema.com',
                senha=generate_password_hash('admin123'),
                tipo='admin',
                empresa_id=empresa_admin.id,
                ativo=True
            )
            db.session.add(usuario_admin)
            db.session.commit()
    except Exception as e:
        app.logger.error(f"Erro ao garantir existência do admin: {str(e)}")
        db.session.rollback()  # Rollback para limpar transação em estado failed
    
    if request.method == 'POST':
        tipo_acesso = request.form.get('tipo_acesso', '')
        usuario_login = request.form['usuario']
        senha = request.form['senha']
        
        # Logs de debug
        app.logger.info(f"Tentativa de login - Tipo: {tipo_acesso}, Usuário: '{usuario_login}'")
        
        empresa = None
        
        # LOGIN PESSOA FÍSICA (CPF)
        if tipo_acesso == 'pessoa_fisica':
            cpf = request.form.get('cpf', '').strip()
            
            if not cpf:
                flash('CPF é obrigatório.', 'error')
                return redirect(url_for('login'))
            
            if not validar_cpf(cpf):
                flash('CPF inválido.', 'error')
                return redirect(url_for('login'))
            
            empresa = Empresa.query.filter_by(cpf=cpf, tipo_conta='pessoa_fisica').first()
            
            if not empresa:
                flash('CPF não encontrado.', 'error')
                return redirect(url_for('login'))
        
        # LOGIN EMPRESA (CNPJ) - Inclui admin
        elif tipo_acesso == 'empresa':
            cnpj = request.form.get('cnpj', '').strip()
            
            if not cnpj:
                flash('CNPJ é obrigatório.', 'error')
                return redirect(url_for('login'))
            
            # Para o CNPJ admin, não validar formato
            if cnpj != '00.000.000/0000-00':
                if not validar_cnpj(cnpj):
                    flash('CNPJ inválido.', 'error')
                    return redirect(url_for('login'))
            
            # Buscar empresa - incluir admin também
            empresa = Empresa.query.filter(
                Empresa.cnpj == cnpj,
                or_(Empresa.tipo_conta == 'empresa', Empresa.tipo_conta == 'admin')
            ).first()
            
            if not empresa:
                flash('CNPJ não encontrado.', 'error')
                return redirect(url_for('login'))
        
        # LOGIN CONTADOR/BPO
        elif tipo_acesso == 'contador_bpo':
            contador_doc = request.form.get('contador_doc', '').strip()
            
            if not contador_doc:
                flash('CPF ou CNPJ é obrigatório.', 'error')
                return redirect(url_for('login'))
            
            # Remover formatação para validar
            documento_limpo = contador_doc.replace('.', '').replace('/', '').replace('-', '')
            
            # Validar se é CPF ou CNPJ
            if len(documento_limpo) == 11:
                if not validar_cpf(contador_doc):
                    flash('CPF inválido.', 'error')
                    return redirect(url_for('login'))
            elif len(documento_limpo) == 14:
                if not validar_cnpj(contador_doc):
                    flash('CNPJ inválido.', 'error')
                    return redirect(url_for('login'))
            else:
                flash('CPF ou CNPJ inválido.', 'error')
                return redirect(url_for('login'))
            
            # Primeiro, buscar a empresa contador pelo documento
            empresa = Empresa.query.filter(
                or_(Empresa.cpf == contador_doc, Empresa.cnpj == contador_doc),
                Empresa.tipo_conta == 'contador_bpo'
            ).first()
            
            # Verificar se é sub-usuário (usando o documento do contador para buscar o contador_id)
            if empresa:
                # Buscar sub-usuário que pertence a este contador e tem o usuário/senha informados
                sub_usuario = SubUsuarioContador.query.filter(
                    SubUsuarioContador.contador_id == empresa.id,
                    or_(
                        SubUsuarioContador.usuario == usuario_login,
                        SubUsuarioContador.email == usuario_login
                    )
                ).first()
                
                if sub_usuario and check_password_hash(sub_usuario.senha, senha):
                    if not sub_usuario.ativo:
                        flash('Conta desativada. Entre em contato com o administrador.', 'error')
                        return redirect(url_for('login'))
                    
                    # Login como sub-usuário
                    session['sub_usuario_id'] = sub_usuario.id
                    session['usuario_nome'] = sub_usuario.nome
                    session['usuario_tipo'] = 'sub_contador'
                    session['contador_id'] = sub_usuario.contador_id
                    session['tipo_conta'] = 'contador_bpo'  # Importante: definir tipo_conta
                    
                    # Buscar empresa do contador
                    empresa_contador = db.session.get(Empresa, sub_usuario.contador_id)
                    if empresa_contador:
                        session['empresa_id'] = empresa_contador.id
                        session['empresa_nome'] = empresa_contador.razao_social
                        session['empresa_cpf'] = empresa_contador.cpf
                        session['empresa_cnpj'] = empresa_contador.cnpj
                    
                    app.logger.info(f"Login realizado como sub-usuário - Nome: {sub_usuario.nome}, Contador: {empresa_contador.razao_social if empresa_contador else 'N/A'}")
                    return redirect(url_for('dashboard_contador'))
            
            # Se não for sub-usuário ou não encontrou empresa, tentar login normal
            if not empresa:
                flash('Contador/BPO não encontrado.', 'error')
                return redirect(url_for('login'))
        
        else:
            flash('Tipo de acesso inválido.', 'error')
            return redirect(url_for('login'))
        
        if not empresa:
            flash('Conta não encontrada.', 'error')
            return redirect(url_for('login'))
        
        if not empresa.ativo:
            flash('Conta desativada. Entre em contato com o administrador.', 'error')
            return redirect(url_for('login'))
        
        # Atualizar dias de assinatura antes de verificar
        if empresa.tipo_conta != 'admin':
            atualizar_dias_assinatura(empresa)

        # NÃO BLOQUEAR login para assinatura expirada - permitir entrar e mostrar modal dentro do dashboard
        
        # Buscar usuário pela empresa e usuario
        usuario = Usuario.query.filter_by(empresa_id=empresa.id, usuario=usuario_login).first()
        
        # Logs de debug
        if not usuario:
            app.logger.warning(f"Usuário não encontrado - Empresa ID: {empresa.id}, Usuário: '{usuario_login}'")
        elif not check_password_hash(usuario.senha, senha):
            app.logger.warning(f"Senha incorreta - Usuário: '{usuario_login}', Empresa: {empresa.razao_social}")
        
        if usuario and check_password_hash(usuario.senha, senha):
            if not usuario.ativo:
                flash('Conta desativada. Entre em contato com o administrador.', 'error')
                return redirect(url_for('login'))
            
            if usuario.pausado:
                flash('Conta pausada. Entre em contato com o administrador.', 'error')
                return redirect(url_for('login'))
            
            session['usuario_id'] = usuario.id
            session['usuario_nome'] = usuario.nome
            session['usuario_tipo'] = usuario.tipo
            session['empresa_id'] = empresa.id
            session['empresa_nome'] = empresa.razao_social
            session['empresa_tipo_pessoa'] = empresa.tipo_pessoa
            session['empresa_nome_fantasia'] = empresa.nome_fantasia
            session['empresa_cpf'] = empresa.cpf
            session['empresa_cnpj'] = empresa.cnpj
            session['tipo_conta'] = empresa.tipo_conta if hasattr(empresa, 'tipo_conta') else 'empresa'
            session['dias_assinatura'] = empresa.dias_assinatura if hasattr(empresa, 'dias_assinatura') else 30

            # Informações do plano ativo
            if hasattr(empresa, 'plano_ativo') and empresa.plano_ativo:
                session['plano_nome'] = empresa.plano_ativo.nome
                session['plano_codigo'] = empresa.plano_ativo.codigo
            else:
                session['plano_nome'] = None
                session['plano_codigo'] = None

            app.logger.info(f"Login realizado com sucesso - Usuário: {usuario.nome}, Empresa: {empresa.razao_social}")
            
            if usuario.tipo == 'admin':
                return redirect(url_for('admin_dashboard'))
            elif session['tipo_conta'] == 'contador_bpo':
                return redirect(url_for('dashboard_contador'))
            else:
                return redirect(url_for('dashboard'))
        else:
            flash('Usuário ou senha incorretos.', 'error')
    
    return render_template('login_novo.html')

@app.route('/registro', methods=['GET', 'POST'])
def registro():
    if request.method == 'POST':
        nome = request.form['nome'].strip()
        usuario = request.form['usuario'].strip()
        email = request.form['email'].strip()
        senha = request.form['senha']
        telefone = request.form['telefone'].strip()
        tipo_pessoa = request.form['tipo_pessoa']  # PF, PJ ou CONTADOR
        cpf = request.form.get('cpf', '').strip()
        cnpj = request.form.get('cnpj', '').strip()
        razao_social = request.form['razao_social'].strip()
        nome_fantasia = request.form['nome_fantasia'].strip()
        endereco_empresa = request.form['endereco_empresa'].strip()
        tipo_empresa = request.form.get('tipo_empresa', '')
        
        # Se for Contador/BPO, definir tipo_empresa como 'servicos' automaticamente
        if tipo_pessoa == 'CONTADOR':
            tipo_empresa = 'servicos'
        
        # Logs de debug
        app.logger.info(f"Tentativa de registro - Tipo: {tipo_pessoa}, CPF: '{cpf}', CNPJ: '{cnpj}'")
        
        # Validações
        erros = []
        
        if not nome or len(nome) < 3:
            erros.append('Nome deve ter pelo menos 3 caracteres')
        
        if not usuario or len(usuario) < 3:
            erros.append('Nome de usuário deve ter pelo menos 3 caracteres')
        
        if not validar_email(email):
            erros.append('Email inválido')
        
        if len(senha) < 6:
            erros.append('Senha deve ter pelo menos 6 caracteres')
        
        # Validação específica por tipo de pessoa
        if tipo_pessoa == 'PF':
            if not cpf:
                erros.append('CPF é obrigatório para pessoa física')
            else:
                app.logger.info(f"Validando CPF: '{cpf}'")
                cpf_valido = validar_cpf(cpf)
                app.logger.info(f"Resultado da validação de CPF: {cpf_valido}")
                if not cpf_valido:
                    erros.append('CPF inválido')
                else:
                    # Verificar se CPF já existe
                    if Empresa.query.filter_by(cpf=cpf).first():
                        erros.append('CPF já está cadastrado')
        
        elif tipo_pessoa == 'PJ':
            if not cnpj:
                erros.append('CNPJ é obrigatório para pessoa jurídica')
            else:
                app.logger.info(f"Validando CNPJ: '{cnpj}'")
                cnpj_valido = validar_cnpj(cnpj)
                app.logger.info(f"Resultado da validação de CNPJ: {cnpj_valido}")
                if not cnpj_valido:
                    erros.append('CNPJ inválido')
                else:
                    # Verificar se CNPJ já existe
                    if Empresa.query.filter_by(cnpj=cnpj).first():
                        erros.append('CNPJ já está cadastrado')
        
        elif tipo_pessoa == 'CONTADOR':
            # Para Contador/BPO, o documento pode ser CPF ou CNPJ
            documento = cnpj.strip()  # O campo cnpj é usado para receber CPF ou CNPJ
            if not documento:
                erros.append('CPF ou CNPJ é obrigatório para Contador/BPO')
            else:
                # Remover formatação para verificar tamanho
                documento_limpo = documento.replace('.', '').replace('/', '').replace('-', '')
                if len(documento_limpo) == 11:
                    # É um CPF
                    app.logger.info(f"Validando CPF de Contador: '{documento}'")
                    cpf_valido = validar_cpf(documento)
                    if not cpf_valido:
                        erros.append('CPF inválido')
                    else:
                        # Verificar se CPF já existe
                        if Empresa.query.filter_by(cpf=documento).first():
                            erros.append('CPF já está cadastrado')
                elif len(documento_limpo) == 14:
                    # É um CNPJ
                    app.logger.info(f"Validando CNPJ de Contador: '{documento}'")
                    cnpj_valido = validar_cnpj(documento)
                    if not cnpj_valido:
                        erros.append('CNPJ inválido')
                    else:
                        # Verificar se CNPJ já existe
                        if Empresa.query.filter_by(cnpj=documento).first():
                            erros.append('CNPJ já está cadastrado')
                else:
                    erros.append('CPF ou CNPJ inválido')
        
        if not razao_social or len(razao_social) < 3:
            erros.append('Razão social deve ter pelo menos 3 caracteres')
        
        # Verificar se usuário já existe na mesma empresa (mesmo CPF/CNPJ)
        if tipo_pessoa == 'PF':
            # Para pessoa física, verificar se já existe usuário com mesmo nome na empresa com este CPF
            empresa_existente = Empresa.query.filter_by(cpf=cpf).first()
            if empresa_existente:
                # Se a empresa já existe, verificar se já tem usuário com este nome
                if Usuario.query.filter_by(usuario=usuario, empresa_id=empresa_existente.id).first():
                    erros.append('Nome de usuário já está em uso nesta empresa (CPF)')
        elif tipo_pessoa == 'PJ':
            # Para pessoa jurídica, verificar se já existe usuário com mesmo nome na empresa com este CNPJ
            empresa_existente = Empresa.query.filter_by(cnpj=cnpj).first()
            if empresa_existente:
                # Se a empresa já existe, verificar se já tem usuário com este nome
                if Usuario.query.filter_by(usuario=usuario, empresa_id=empresa_existente.id).first():
                    erros.append('Nome de usuário já está em uso nesta empresa (CNPJ)')
        elif tipo_pessoa == 'CONTADOR':
            # Para Contador/BPO, verificar por CPF ou CNPJ
            documento = cnpj.strip()
            documento_limpo = documento.replace('.', '').replace('/', '').replace('-', '')
            if len(documento_limpo) == 11:
                empresa_existente = Empresa.query.filter_by(cpf=documento).first()
            else:
                empresa_existente = Empresa.query.filter_by(cnpj=documento).first()
            if empresa_existente:
                if Usuario.query.filter_by(usuario=usuario, empresa_id=empresa_existente.id).first():
                    erros.append('Nome de usuário já está em uso nesta conta')
        
        # Verificar se email já existe na mesma empresa (mesmo CPF/CNPJ)
        if tipo_pessoa == 'PF':
            empresa_existente = Empresa.query.filter_by(cpf=cpf).first()
            if empresa_existente:
                if Usuario.query.filter_by(email=email, empresa_id=empresa_existente.id).first():
                    erros.append('Email já está cadastrado nesta empresa (CPF)')
        elif tipo_pessoa == 'PJ':
            empresa_existente = Empresa.query.filter_by(cnpj=cnpj).first()
            if empresa_existente:
                if Usuario.query.filter_by(email=email, empresa_id=empresa_existente.id).first():
                    erros.append('Email já está cadastrado nesta empresa (CNPJ)')
        elif tipo_pessoa == 'CONTADOR':
            # Para Contador/BPO, verificar por CPF ou CNPJ
            documento = cnpj.strip()
            documento_limpo = documento.replace('.', '').replace('/', '').replace('-', '')
            if len(documento_limpo) == 11:
                empresa_existente = Empresa.query.filter_by(cpf=documento).first()
            else:
                empresa_existente = Empresa.query.filter_by(cnpj=documento).first()
            if empresa_existente:
                if Usuario.query.filter_by(email=email, empresa_id=empresa_existente.id).first():
                    erros.append('Email já está cadastrado nesta conta')
        
        if erros:
            app.logger.warning(f"Erros no registro: {erros}")
            for erro in erros:
                flash(erro, 'error')
            return render_template('registro.html')
        
        # Determinar tipo_conta e campos CPF/CNPJ baseado no tipo_pessoa
        # Mapeamento automático: PF -> pessoa_fisica, PJ -> empresa, CONTADOR -> contador_bpo
        if tipo_pessoa == 'PF':
            tipo_conta = 'pessoa_fisica'
            cpf_final = cpf
            cnpj_final = None
        elif tipo_pessoa == 'PJ':
            tipo_conta = 'empresa'
            cpf_final = None
            cnpj_final = cnpj
        elif tipo_pessoa == 'CONTADOR':
            tipo_conta = 'contador_bpo'
            # Para Contador/BPO, determinar se é CPF ou CNPJ
            documento = cnpj.strip()
            documento_limpo = documento.replace('.', '').replace('/', '').replace('-', '')
            if len(documento_limpo) == 11:
                # É CPF
                cpf_final = documento
                cnpj_final = None
            else:
                # É CNPJ
                cpf_final = None
                cnpj_final = documento
        
        # Buscar plano padrão (Básico 30 Dias)
        plano_padrao = Plano.query.filter_by(codigo='basico_30d', ativo=True).first()

        # Criar empresa
        nova_empresa = Empresa(
            tipo_pessoa=tipo_pessoa,
            cpf=cpf_final,
            cnpj=cnpj_final,
            razao_social=razao_social,
            nome_fantasia=nome_fantasia or razao_social,
            email=email,
            telefone=telefone,
            endereco=endereco_empresa,
            tipo_empresa=tipo_empresa,
            tipo_conta=tipo_conta,
            dias_assinatura=30,
            data_inicio_assinatura=datetime.utcnow(),  # Definir data de início ao criar
            plano_id=plano_padrao.id if plano_padrao else None  # Atribuir plano padrão
        )

        db.session.add(nova_empresa)
        db.session.flush()  # Para obter o ID da empresa
        
        # Criar usuário principal
        novo_usuario = Usuario(
            nome=nome,
            usuario=usuario,
            email=email,
            senha=generate_password_hash(senha),
            telefone=telefone,
            tipo='usuario_principal',
            empresa_id=nova_empresa.id
        )
        
        db.session.add(novo_usuario)
        db.session.commit()
        
        # Criar permissões padrão
        criar_permissoes_padrao(novo_usuario.id)
        
        app.logger.info(f"Empresa e usuário criados com sucesso - Tipo: {tipo_pessoa}, CPF: {cpf}, CNPJ: {cnpj}")
        flash('Empresa e conta criadas com sucesso! Faça login para continuar.', 'success')
        return redirect(url_for('login'))
    
    return render_template('registro.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('landing'))

@app.route('/dashboard')
def dashboard():
    # Verificar se está logado (pode ser usuario_id ou sub_usuario_id com acesso_contador)
    if 'usuario_id' not in session and 'sub_usuario_id' not in session:
        return redirect(url_for('login'))
    
    # Se for sub-usuário acessando empresa vinculada, buscar usuário da empresa
    if session.get('acesso_contador') and session.get('sub_usuario_id') and not session.get('usuario_id'):
        empresa_id = session.get('empresa_id')
        if empresa_id:
            # Buscar o primeiro usuário ativo da empresa vinculada
            usuario_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).first()
            if usuario_empresa:
                session['usuario_id'] = usuario_empresa.id
                session['usuario_id_temporario'] = True
    
    # Se ainda não tem usuario_id, verificar se é acesso_contador
    if 'usuario_id' not in session:
        if session.get('acesso_contador') and session.get('empresa_id'):
            # Buscar qualquer usuário da empresa vinculada
            empresa_id = session.get('empresa_id')
            usuario_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).first()
            if usuario_empresa:
                session['usuario_id'] = usuario_empresa.id
                session['usuario_id_temporario'] = True
            else:
                flash('Nenhum usuário encontrado na empresa vinculada.', 'error')
                return redirect(url_for('dashboard_contador'))
        else:
            return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario:
        # Se o usuário não existe e não é acesso temporário, limpar a sessão
        if not session.get('usuario_id_temporario'):
            session.clear()
            flash('Usuário não encontrado. Faça login novamente.', 'error')
            return redirect(url_for('login'))
        else:
            # Se é temporário, buscar novamente
            empresa_id = session.get('empresa_id')
            if empresa_id:
                usuario_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).first()
                if usuario_empresa:
                    session['usuario_id'] = usuario_empresa.id
                    usuario = usuario_empresa
                else:
                    flash('Nenhum usuário encontrado na empresa vinculada.', 'error')
                    return redirect(url_for('dashboard_contador'))
    
    # Se o usuário não tem empresa_id, criar uma empresa padrão
    # Nota: Se for acesso temporário (sub-usuário), a empresa_id correta está em session['empresa_id']
    if getattr(usuario, 'empresa_id', None) is None:
        # Criar empresa padrão para usuários existentes
        empresa_padrao = Empresa.query.filter_by(cnpj='00.000.000/0000-00').first()
        if not empresa_padrao:
            empresa_padrao = Empresa(
                cnpj='00.000.000/0000-00',
                razao_social='Empresa Padrão',
                nome_fantasia='Padrão',
                tipo_empresa='servicos',
                tipo_conta='admin' if usuario.tipo == 'admin' else 'empresa'
            )
            db.session.add(empresa_padrao)
            db.session.flush()
        
        # Atualizar usuário com empresa_id
        usuario.empresa_id = empresa_padrao.id
        db.session.commit()
    
    # Garantir que a empresa admin existe e está configurada corretamente
    empresa_admin = Empresa.query.filter_by(cnpj='00.000.000/0000-00').first()
    if empresa_admin:
        # Atualizar tipo_conta se necessário
        if not hasattr(empresa_admin, 'tipo_conta') or empresa_admin.tipo_conta not in ['admin', 'empresa']:
            empresa_admin.tipo_conta = 'admin'
            db.session.commit()
    else:
        # Criar empresa admin se não existir
        empresa_admin = Empresa(
            cnpj='00.000.000/0000-00',
            razao_social='Sistema Administrativo',
            nome_fantasia='Admin',
            tipo_empresa='servicos',
            tipo_conta='admin'
        )
        db.session.add(empresa_admin)
        db.session.flush()
        
        # Criar usuário admin se não existir
        admin_user = Usuario.query.filter_by(usuario='admin', tipo='admin').first()
        if not admin_user:
            admin_user = Usuario(
                nome='Administrador',
                usuario='admin',
                email='admin@sistema.com',
                senha=generate_password_hash('admin123'),
                tipo='admin',
                empresa_id=empresa_admin.id
            )
            db.session.add(admin_user)
        
        db.session.commit()
    
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))

    # Verificar se a assinatura está expirada (para mostrar modal)
    empresa_atual = db.session.get(Empresa, usuario.empresa_id)
    assinatura_expirada = False
    if empresa_atual and empresa_atual.tipo_conta != 'admin':
        assinatura_expirada = (empresa_atual.dias_assinatura <= 0)

    # Obter parâmetros de filtro
    filtro_tipo = request.args.get('filtro', 'mes')  # mes, ano
    ano = int(request.args.get('ano', datetime.now().year))
    mes = int(request.args.get('mes', datetime.now().month))
    
    # Definir período baseado no filtro
    if filtro_tipo == 'ano':
        inicio_periodo = datetime(ano, 1, 1).date()
        fim_periodo = datetime(ano, 12, 31).date()
        periodo_nome = f"Ano {ano}"
    else:  # filtro_tipo == 'mes'
        inicio_periodo = datetime(ano, mes, 1).date()
        if mes == 12:
            fim_periodo = datetime(ano + 1, 1, 1).date() - timedelta(days=1)
        else:
            fim_periodo = datetime(ano, mes + 1, 1).date() - timedelta(days=1)
        # Mapeamento de meses em português
        meses_nomes = {
            1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
            5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
            9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
        }
        periodo_nome = f"{meses_nomes[mes]} {ano}"
    
    # Dados para hoje (sempre atual)
    hoje = datetime.now().date()
    
    # Obter empresa_id correta da sessão (considera acesso_contador)
    empresa_id_correta = obter_empresa_id_sessao(session, usuario)
    if not empresa_id_correta:
        flash('Erro ao obter empresa associada', 'error')
        return redirect(url_for('login'))
    
    # Obter todos os usuários da empresa correta para sincronizar dados
    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id_correta, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]
    
    # Contas de entrada hoje (pendentes) - todos os usuários da empresa correta
    entrada_hoje = Lancamento.query.filter(
        Lancamento.empresa_id == empresa_id_correta,
        Lancamento.tipo == 'entrada',
        Lancamento.data_prevista == hoje,
        Lancamento.realizado == False
    ).all()
    
    # Contas de saida hoje (pendentes) - todos os usuários da empresa correta
    saida_hoje = Lancamento.query.filter(
        Lancamento.empresa_id == empresa_id_correta,
        Lancamento.tipo == 'saida',
        Lancamento.data_prevista == hoje,
        Lancamento.realizado == False
    ).all()
    
    # Contas de entrada do mês (pendentes) - todos os usuários da empresa correta
    entrada_mes = Lancamento.query.filter(
        Lancamento.empresa_id == empresa_id_correta,
        Lancamento.tipo == 'entrada',
        Lancamento.realizado == False,
        Lancamento.data_prevista >= inicio_periodo,
        Lancamento.data_prevista <= fim_periodo
    ).all()
    
    # Contas de saida do mês (pendentes) - todos os usuários da empresa correta
    saida_mes = Lancamento.query.filter(
        Lancamento.empresa_id == empresa_id_correta,
        Lancamento.tipo == 'saida',
        Lancamento.realizado == False,
        Lancamento.data_prevista >= inicio_periodo,
        Lancamento.data_prevista <= fim_periodo
    ).all()
    
    # Buscar lançamentos agendados para hoje (precisam de confirmação)
    # Estes são lançamentos que têm data_realizada = hoje mas ainda não foram confirmados como realizados
    # Ou seja: data_realizada = hoje E realizado = False (precisam de confirmação do usuário)
    hoje = datetime.now().date()
    lancamentos_agendados_hoje = Lancamento.query.filter(
        Lancamento.empresa_id == empresa_id_correta,
        Lancamento.data_realizada == hoje,
        Lancamento.realizado == False  # Apenas os que ainda não foram confirmados como realizados
    ).all()
    
    # Calcular totais
    total_entrada_hoje = sum(l.valor for l in entrada_hoje)
    total_saida_hoje = sum(l.valor for l in saida_hoje)
    total_entrada_mes = sum(l.valor for l in entrada_mes)
    total_saida_mes = sum(l.valor for l in saida_mes)
    
    # Calcular entradas e saidas realizadas no período - todos os usuários da empresa correta
    entradas_periodo = Lancamento.query.filter(
        Lancamento.empresa_id == empresa_id_correta,
        Lancamento.tipo == 'entrada',
        Lancamento.realizado == True,
        Lancamento.data_realizada >= inicio_periodo,
        Lancamento.data_realizada <= fim_periodo
    ).all()
    
    saidas_periodo = Lancamento.query.filter(
        Lancamento.empresa_id == empresa_id_correta,
        Lancamento.tipo == 'saida',
        Lancamento.realizado == True,
        Lancamento.data_realizada >= inicio_periodo,
        Lancamento.data_realizada <= fim_periodo
    ).all()
    
    # Calcular totais apenas com lançamentos financeiros realizados no período
    # Observação: O card exibe "Apenas lançamentos realizados", portanto usamos somente a tabela Lancamento
    total_entradas_periodo = sum(l.valor for l in entradas_periodo)
    total_saidas_periodo = sum(l.valor for l in saidas_periodo)
    
    saldo_periodo = total_entradas_periodo - total_saidas_periodo
    
    # Calcular totais por status no período
    # Realizado: já foi realizado (saldo = entradas - saidas)
    total_realizado_entradas = sum(l.valor for l in entradas_periodo)
    total_realizado_saidas = sum(l.valor for l in saidas_periodo)
    total_realizado = total_realizado_entradas - total_realizado_saidas
    
    # A vencer: não realizado e data_prevista no futuro
    a_vencer_entradas = Lancamento.query.filter(
        Lancamento.empresa_id == empresa_id_correta,
        Lancamento.tipo == 'entrada',
        Lancamento.realizado == False,
        Lancamento.data_prevista > hoje,
        Lancamento.data_prevista >= inicio_periodo,
        Lancamento.data_prevista <= fim_periodo
    ).all()
    a_vencer_saidas = Lancamento.query.filter(
        Lancamento.empresa_id == empresa_id_correta,
        Lancamento.tipo == 'saida',
        Lancamento.realizado == False,
        Lancamento.data_prevista > hoje,
        Lancamento.data_prevista >= inicio_periodo,
        Lancamento.data_prevista <= fim_periodo
    ).all()
    total_a_vencer = sum(l.valor for l in a_vencer_entradas) - sum(l.valor for l in a_vencer_saidas)

    # Vencido: não realizado e data_prevista no passado
    vencido_entradas = Lancamento.query.filter(
        Lancamento.empresa_id == empresa_id_correta,
        Lancamento.tipo == 'entrada',
        Lancamento.realizado == False,
        Lancamento.data_prevista < hoje,
        Lancamento.data_prevista >= inicio_periodo,
        Lancamento.data_prevista <= fim_periodo
    ).all()
    vencido_saidas = Lancamento.query.filter(
        Lancamento.empresa_id == empresa_id_correta,
        Lancamento.tipo == 'saida',
        Lancamento.realizado == False,
        Lancamento.data_prevista < hoje,
        Lancamento.data_prevista >= inicio_periodo,
        Lancamento.data_prevista <= fim_periodo
    ).all()
    total_vencido = sum(l.valor for l in vencido_entradas) - sum(l.valor for l in vencido_saidas)

    # Agendado: tem data_realizada futura
    agendado_entradas = Lancamento.query.filter(
        Lancamento.empresa_id == empresa_id_correta,
        Lancamento.tipo == 'entrada',
        Lancamento.realizado == False,
        Lancamento.data_realizada.isnot(None),
        Lancamento.data_realizada > hoje,
        Lancamento.data_realizada >= inicio_periodo,
        Lancamento.data_realizada <= fim_periodo
    ).all()
    agendado_saidas = Lancamento.query.filter(
        Lancamento.empresa_id == empresa_id_correta,
        Lancamento.tipo == 'saida',
        Lancamento.realizado == False,
        Lancamento.data_realizada.isnot(None),
        Lancamento.data_realizada > hoje,
        Lancamento.data_realizada >= inicio_periodo,
        Lancamento.data_realizada <= fim_periodo
    ).all()
    total_agendado = sum(l.valor for l in agendado_entradas) - sum(l.valor for l in agendado_saidas)

    # Buscar todas as contas de saida e entrada do período para os quadros
    todas_contas_entrada = Lancamento.query.filter(
        Lancamento.empresa_id == empresa_id_correta,
        Lancamento.tipo == 'entrada',
        Lancamento.data_prevista >= inicio_periodo,
        Lancamento.data_prevista <= fim_periodo
    ).order_by(Lancamento.data_prevista).all()
    
    todas_contas_saida = Lancamento.query.filter(
        Lancamento.empresa_id == empresa_id_correta,
        Lancamento.tipo == 'saida',
        Lancamento.data_prevista >= inicio_periodo,
        Lancamento.data_prevista <= fim_periodo
    ).order_by(Lancamento.data_prevista).all()
    
    # Obter alertas do usuário
    try:
        alertas = verificar_alertas_usuario(usuario.id)
        if alertas is None:
            alertas = []
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Erro ao verificar alertas: {str(e)}')
        alertas = []
    
    # Verificar se há vínculos pendentes (apenas para usuário principal)
    vinculos_pendentes = []
    if usuario.tipo in ['admin', 'usuario_principal']:
        vinculos_pendentes = VinculoContador.query.filter_by(
            empresa_id=usuario.empresa_id,
            status='pendente'
        ).all()
    
    # Verificar se é contador/BPO acessando dashboard de empresa vinculada
    empresas_vinculadas_resumo = []
    tipo_conta = session.get('tipo_conta')
    usuario_tipo = session.get('usuario_tipo')
    acesso_contador = session.get('acesso_contador')
    
    # Se for contador/BPO (não em acesso de empresa vinculada), buscar empresas vinculadas autorizadas
    if (tipo_conta == 'contador_bpo' or usuario_tipo == 'sub_contador') and not acesso_contador:
        contador_id = session.get('empresa_id') or session.get('contador_id')
        
        # Buscar empresas vinculadas autorizadas
        vinculos_autorizados = VinculoContador.query.filter_by(
            contador_id=contador_id,
            status='autorizado'
        ).all()
        
        # Para cada empresa vinculada, calcular resumo financeiro
        for vinculo in vinculos_autorizados:
            empresa_vinculada = vinculo.empresa
            
            # Obter usuários da empresa vinculada
            usuarios_empresa_vinculada = Usuario.query.filter_by(
                empresa_id=empresa_vinculada.id, 
                ativo=True
            ).all()
            usuarios_ids_vinculada = [u.id for u in usuarios_empresa_vinculada]
            
            # Entradas realizadas no período
            receitas_vinculada = Lancamento.query.filter(
                Lancamento.usuario_id.in_(usuarios_ids_vinculada),
                Lancamento.tipo == 'entrada',
                Lancamento.realizado == True,
                Lancamento.data_realizada >= inicio_periodo,
                Lancamento.data_realizada <= fim_periodo
            ).all()
            
            # Saídas realizadas no período
            despesas_vinculada = Lancamento.query.filter(
                Lancamento.usuario_id.in_(usuarios_ids_vinculada),
                Lancamento.tipo == 'saida',
                Lancamento.realizado == True,
                Lancamento.data_realizada >= inicio_periodo,
                Lancamento.data_realizada <= fim_periodo
            ).all()
            
            # Contas a receber pendentes no período
            contas_receber_vinculada = Lancamento.query.filter(
                Lancamento.usuario_id.in_(usuarios_ids_vinculada),
                Lancamento.tipo == 'entrada',
                Lancamento.realizado == False,
                Lancamento.data_prevista >= inicio_periodo,
                Lancamento.data_prevista <= fim_periodo
            ).all()
            
            # Contas a pagar pendentes no período
            contas_pagar_vinculada = Lancamento.query.filter(
                Lancamento.usuario_id.in_(usuarios_ids_vinculada),
                Lancamento.tipo == 'saida',
                Lancamento.realizado == False,
                Lancamento.data_prevista >= inicio_periodo,
                Lancamento.data_prevista <= fim_periodo
            ).all()
            
            # Calcular totais
            total_receitas_vinculada = sum(l.valor for l in receitas_vinculada)
            total_despesas_vinculada = sum(l.valor for l in despesas_vinculada)
            saldo_vinculada = total_receitas_vinculada - total_despesas_vinculada
            total_receber_vinculada = sum(l.valor for l in contas_receber_vinculada)
            total_pagar_vinculada = sum(l.valor for l in contas_pagar_vinculada)
            
            empresas_vinculadas_resumo.append({
                'empresa': empresa_vinculada,
                'saldo_periodo': saldo_vinculada,
                'total_entradas': total_receitas_vinculada,
                'total_saidas': total_despesas_vinculada,
                'total_entrada_pendente': total_receber_vinculada,
                'total_saida_pendente': total_pagar_vinculada
            })
    
    # Variáveis para os filtros
    anos_disponiveis = list(range(datetime.now().year - 5, datetime.now().year + 2))
    meses = [
        (1, 'Janeiro'), (2, 'Fevereiro'), (3, 'Março'), (4, 'Abril'),
        (5, 'Maio'), (6, 'Junho'), (7, 'Julho'), (8, 'Agosto'),
        (9, 'Setembro'), (10, 'Outubro'), (11, 'Novembro'), (12, 'Dezembro')
    ]
    
    return render_template('dashboard.html',
                         usuario=usuario,
                         assinatura_expirada=assinatura_expirada,
                         entrada_hoje=entrada_hoje,
                         saida_hoje=saida_hoje,
                         entrada_mes=entrada_mes,
                         saida_mes=saida_mes,
                         total_entrada_hoje=total_entrada_hoje,
                         total_saida_hoje=total_saida_hoje,
                         total_entrada_mes=total_entrada_mes,
                         total_saida_mes=total_saida_mes,
                         total_entradas_periodo=total_entradas_periodo,
                         total_saidas_periodo=total_saidas_periodo,
                         saldo_periodo=saldo_periodo,
                         periodo_nome=periodo_nome,
                         filtro_tipo=filtro_tipo,
                         ano=ano,
                         mes=mes,
                         alertas=alertas,
                         anos_disponiveis=anos_disponiveis,
                         meses=meses,
                         lancamentos_agendados_hoje=lancamentos_agendados_hoje,
                         vinculos_pendentes=vinculos_pendentes,
                         empresas_vinculadas_resumo=empresas_vinculadas_resumo,
                         total_realizado=total_realizado,
                         total_a_vencer=total_a_vencer,
                         total_vencido=total_vencido,
                         total_agendado=total_agendado,
                         todas_contas_entrada=todas_contas_entrada,
                         todas_contas_saida=todas_contas_saida,
                         hoje=hoje)

@app.route('/admin/dashboard')
def admin_dashboard():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))

    usuario = Usuario.query.options(db.joinedload(Usuario.empresa)).get(session['usuario_id'])
    if usuario.tipo != 'admin':
        return redirect(url_for('dashboard'))

    # Estatísticas para o admin
    total_usuarios = Usuario.query.filter(Usuario.tipo != 'admin').count()
    usuarios_ativos = Usuario.query.filter(Usuario.tipo != 'admin', Usuario.ativo == True).count()
    usuarios_pausados = Usuario.query.filter(Usuario.tipo != 'admin', Usuario.pausado == True).count()

    usuarios = Usuario.query.options(db.joinedload(Usuario.empresa)).filter(Usuario.tipo != 'admin').order_by(Usuario.data_criacao.desc()).all()

    # Dados para gráficos
    # 1. Distribuição por tipo de conta
    empresas = Empresa.query.filter(Empresa.tipo_conta != 'admin').all()
    contas_por_tipo = {
        'empresa': sum(1 for e in empresas if e.tipo_conta == 'empresa'),
        'pessoa_fisica': sum(1 for e in empresas if e.tipo_conta == 'pessoa_fisica'),
        'contador_bpo': sum(1 for e in empresas if e.tipo_conta == 'contador_bpo')
    }

    # 2. Distribuição por dias de assinatura
    dias_distribuicao = {
        'excelente': sum(1 for e in empresas if e.dias_assinatura and e.dias_assinatura > 90),
        'bom': sum(1 for e in empresas if e.dias_assinatura and 30 <= e.dias_assinatura <= 90),
        'alerta': sum(1 for e in empresas if e.dias_assinatura and 7 <= e.dias_assinatura < 30),
        'critico': sum(1 for e in empresas if e.dias_assinatura and 1 <= e.dias_assinatura < 7),
        'expirado': sum(1 for e in empresas if e.dias_assinatura is not None and e.dias_assinatura == 0)
    }

    # 3. Crescimento de contas nos últimos 12 meses
    from datetime import datetime, timedelta
    hoje = datetime.now()
    meses_crescimento = []
    contas_crescimento = []
    for i in range(11, -1, -1):  # Últimos 12 meses
        mes_ref = hoje - timedelta(days=30*i)
        mes_nome = mes_ref.strftime('%b/%y')
        total_ate_mes = Empresa.query.filter(
            Empresa.tipo_conta != 'admin',
            Empresa.data_criacao <= mes_ref
        ).count()
        meses_crescimento.append(mes_nome)
        contas_crescimento.append(total_ate_mes)

    return render_template('admin_dashboard.html',
                         usuario=usuario,
                         total_usuarios=total_usuarios,
                         usuarios_ativos=usuarios_ativos,
                         usuarios_pausados=usuarios_pausados,
                         usuarios=usuarios,
                         contas_por_tipo=contas_por_tipo,
                         dias_distribuicao=dias_distribuicao,
                         meses_crescimento=meses_crescimento,
                         contas_crescimento=contas_crescimento)

@app.route('/admin/usuarios')
def admin_usuarios():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = Usuario.query.options(db.joinedload(Usuario.empresa)).get(session['usuario_id'])
    if not usuario or usuario.tipo != 'admin':
        return redirect(url_for('dashboard'))
    
    try:
        # Buscar todas as empresas/pessoas cadastradas (exceto admin) com seus usuários
        try:
            empresas = Empresa.query.filter(Empresa.tipo_conta != 'admin').order_by(Empresa.id.desc()).all()
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Erro ao buscar empresas com ordenação: {str(e)}")
            empresas = Empresa.query.filter(Empresa.tipo_conta != 'admin').all()
        
        # Buscar usuários de cada empresa e atualizar dias de assinatura
        empresas_com_usuarios = []
        for empresa in empresas:
            try:
                # Atualizar dias de assinatura automaticamente
                atualizar_dias_assinatura(empresa)
                
                usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa.id).order_by(Usuario.nome).all()
            except Exception as e:
                db.session.rollback()
                app.logger.error(f"Erro ao buscar usuários da empresa {empresa.id}: {str(e)}")
                usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa.id).all()
            
            empresas_com_usuarios.append({
                'empresa': empresa,
                'usuarios': usuarios_empresa
            })
        
        # Estatísticas
        stats = {
            'total': len(empresas),
            'empresas': len([e for e in empresas if e.tipo_conta == 'empresa']),
            'pessoas_fisicas': len([e for e in empresas if e.tipo_conta == 'pessoa_fisica']),
            'contadores': len([e for e in empresas if e.tipo_conta == 'contador_bpo']),
            'ativas': len([e for e in empresas if e.ativo]),
            'inativas': len([e for e in empresas if not e.ativo])
        }
        
        return render_template('admin_usuarios.html', usuario=usuario, empresas_com_usuarios=empresas_com_usuarios, stats=stats)
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro na rota admin_usuarios: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        flash(f'Erro ao carregar dados: {str(e)}', 'error')
        # Retornar template com dados vazios em caso de erro
        return render_template('admin_usuarios.html', 
                             usuario=usuario, 
                             empresas_com_usuarios=[], 
                             stats={'total': 0, 'empresas': 0, 'pessoas_fisicas': 0, 'contadores': 0, 'ativas': 0, 'inativas': 0})

@app.route('/admin/usuario/<int:user_id>/toggle_status')
def toggle_usuario_status(user_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    admin = db.session.get(Usuario, session['usuario_id'])
    if admin.tipo != 'admin':
        return redirect(url_for('dashboard'))
    
    usuario = db.session.get(Usuario, user_id)
    if usuario:
        usuario.ativo = not usuario.ativo
        db.session.commit()
        flash(f'Status do usuário {usuario.nome} alterado com sucesso.', 'success')
    
    return redirect(url_for('admin_usuarios'))

@app.route('/admin/usuario/<int:user_id>/toggle_pause')
def toggle_usuario_pause(user_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    admin = db.session.get(Usuario, session['usuario_id'])
    if admin.tipo != 'admin':
        return redirect(url_for('dashboard'))
    
    usuario = db.session.get(Usuario, user_id)
    if usuario:
        usuario.pausado = not usuario.pausado
        db.session.commit()
        flash(f'Pausa do usuário {usuario.nome} alterada com sucesso.', 'success')
    
    return redirect(url_for('admin_usuarios'))

@app.route('/admin/usuario/<int:user_id>/alterar_senha', methods=['GET', 'POST'])
def alterar_senha_usuario(user_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    admin = db.session.get(Usuario, session['usuario_id'])
    if admin.tipo != 'admin':
        return redirect(url_for('dashboard'))
    
    usuario = db.session.get(Usuario, user_id)
    if not usuario:
        flash('Usuário não encontrado.', 'error')
        return redirect(url_for('admin_usuarios'))
    
    if request.method == 'POST':
        nova_senha = request.form['nova_senha']
        usuario.senha = generate_password_hash(nova_senha)
        db.session.commit()
        flash(f'Senha do usuário {usuario.nome} alterada com sucesso.', 'success')
        return redirect(url_for('admin_usuarios'))
    
    return render_template('admin_alterar_senha.html', usuario=usuario, admin=admin)

@app.route('/admin/usuario/<int:user_id>/deletar')
def deletar_usuario_admin(user_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    admin = db.session.get(Usuario, session['usuario_id'])
    if admin.tipo != 'admin':
        return redirect(url_for('dashboard'))
    
    usuario = db.session.get(Usuario, user_id)
    if not usuario:
        flash('Usuário não encontrado.', 'error')
        return redirect(url_for('admin_usuarios'))
    
    if usuario.tipo == 'admin':
        flash('Não é possível deletar um usuário administrador.', 'error')
        return redirect(url_for('admin_usuarios'))
    
    try:
        # Primeiro, atualizar registros que fazem referência ao usuário (antes de qualquer commit)
        with db.session.no_autoflush:
            # Atualizar categorias de usuário criadas por este usuário (definir criado_por como NULL)
            categorias_criadas = CategoriaUsuario.query.filter_by(criado_por=user_id).all()
            for categoria in categorias_criadas:
                categoria.criado_por = None
            
            # Deletar vínculos criados por este usuário (mais seguro que atualizar)
            vinculos_criados = Vinculo.query.filter_by(usuario_id=user_id).all()
            for vinculo in vinculos_criados:
                db.session.delete(vinculo)
        
        # Deletar todas as permissões associadas ao usuário
        permissoes = Permissao.query.filter_by(usuario_id=user_id).all()
        for permissao in permissoes:
            db.session.delete(permissao)
        
        # Deletar todas as parcelas associadas ao usuário
        parcelas = Parcela.query.filter_by(usuario_id=user_id).all()
        for parcela in parcelas:
            db.session.delete(parcela)
        
        # Deletar todos os lançamentos associados ao usuário
        lancamentos = Lancamento.query.filter_by(usuario_id=user_id).all()
        for lancamento in lancamentos:
            db.session.delete(lancamento)
        
        # Deletar todas as vendas associadas ao usuário
        vendas = Venda.query.filter_by(usuario_id=user_id).all()
        for venda in vendas:
            db.session.delete(venda)
        
        # Deletar todas as compras associadas ao usuário
        compras = Compra.query.filter_by(usuario_id=user_id).all()
        for compra in compras:
            db.session.delete(compra)
        
        # Deletar todos os clientes associados ao usuário
        clientes = Cliente.query.filter_by(usuario_id=user_id).all()
        for cliente in clientes:
            db.session.delete(cliente)
        
        # Deletar todos os fornecedores associados ao usuário
        fornecedores = Fornecedor.query.filter_by(usuario_id=user_id).all()
        for fornecedor in fornecedores:
            db.session.delete(fornecedor)
        
        # Deletar todos os produtos associados ao usuário
        produtos = Produto.query.filter_by(usuario_id=user_id).all()
        for produto in produtos:
            db.session.delete(produto)
        
        # Deletar todos os serviços associados ao usuário
        servicos = Servico.query.filter_by(usuario_id=user_id).all()
        for servico in servicos:
            db.session.delete(servico)
        
        # Deletar todos os planos de conta associados ao usuário
        planos_conta = PlanoConta.query.filter_by(usuario_id=user_id).all()
        for plano in planos_conta:
            db.session.delete(plano)
        
        # Deletar todas as contas caixa associadas ao usuário
        contas_caixa = ContaCaixa.query.filter_by(usuario_id=user_id).all()
        for conta in contas_caixa:
            db.session.delete(conta)
        
        # Deletar todos os logs de evento associados ao usuário
        event_logs = EventLog.query.filter_by(usuario_id=user_id).all()
        for event_log in event_logs:
            db.session.delete(event_log)
        
        # Deletar todas as importações associadas ao usuário
        importacoes = Importacao.query.filter_by(usuario_id=user_id).all()
        for importacao in importacoes:
            db.session.delete(importacao)
        
        
        # Salvar informações da empresa antes de deletar o usuário
        empresa_id = usuario.empresa_id
        empresa_nome = usuario.empresa.razao_social if usuario.empresa else "N/A"
        empresa_cnpj = usuario.empresa.cnpj if usuario.empresa else "N/A"
        
        app.logger.info(f"Iniciando exclusão do usuário {usuario.nome} da empresa {empresa_nome} (ID: {empresa_id})")
        
        # Deletar o usuário primeiro
        db.session.delete(usuario)
        app.logger.info(f"Usuário {usuario.nome} deletado da sessão.")
        
        # Fazer flush para garantir que a exclusão seja processada
        db.session.flush()
        
        # Verificar se ainda há outros usuários na empresa
        outros_usuarios = Usuario.query.filter_by(empresa_id=empresa_id).count()
        app.logger.info(f"Após exclusão: Empresa {empresa_id} tem {outros_usuarios} usuários restantes.")
        
        # Se não há outros usuários na empresa, deletar a empresa também
        if outros_usuarios == 0:
            empresa = db.session.get(Empresa, empresa_id)
            if empresa:
                app.logger.info(f"Empresa órfã encontrada: {empresa.razao_social} (CNPJ: {empresa.cnpj})")
                try:
                    db.session.delete(empresa)
                    app.logger.info(f"Empresa {empresa.razao_social} marcada para exclusão.")
                    flash(f'Usuário {usuario.nome} e empresa "{empresa_nome}" deletados com sucesso. CNPJ {empresa_cnpj} liberado para novo cadastro.', 'success')
                except Exception as empresa_error:
                    app.logger.error(f"Erro ao deletar empresa: {str(empresa_error)}")
                    flash(f'Usuário {usuario.nome} deletado, mas empresa não pôde ser removida: {str(empresa_error)}', 'warning')
            else:
                app.logger.warning(f"Empresa {empresa_id} não encontrada após exclusão do usuário.")
                flash(f'Usuário {usuario.nome} deletado com sucesso.', 'success')
        else:
            app.logger.info(f"Empresa {empresa_id} mantida - ainda tem {outros_usuarios} usuários.")
            flash(f'Usuário {usuario.nome} deletado com sucesso.', 'success')
        
        # Commit de todas as alterações
        db.session.commit()
        app.logger.info("Transação commitada com sucesso.")
        
        # Retornar resposta JSON de sucesso
        return jsonify({
            'success': True,
            'message': f'Usuário {usuario.nome} e empresa "{empresa_nome}" deletados com sucesso. CNPJ {empresa_cnpj} liberado para novo cadastro.',
            'redirect': False
        })
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Erro ao deletar usuário: {str(e)}')
        
        # Retornar resposta JSON de erro
        return jsonify({
            'success': False,
            'message': f'Erro ao deletar usuário: {str(e)}',
            'redirect': False
        }), 500

@app.route('/lancamentos')
def lancamentos():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    # Obter empresa_id correta da sessão (considera acesso_contador)
    empresa_id_correta = obter_empresa_id_sessao(session, usuario)
    if not empresa_id_correta:
        flash('Erro ao obter empresa associada', 'error')
        return redirect(url_for('dashboard'))
    
    # Buscar todos os usuários da empresa correta (pode ser empresa vinculada se acesso_contador estiver ativo)
    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id_correta, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]
    
    # Aplicar filtros
    query = Lancamento.query.options(
        db.joinedload(Lancamento.cliente),
        db.joinedload(Lancamento.fornecedor),
        db.joinedload(Lancamento.usuario)
    ).filter(Lancamento.empresa_id == empresa_id_correta)
    
    # Filtro por tipo
    tipo = request.args.get('tipo')
    if tipo:
        query = query.filter(Lancamento.tipo == tipo)
    
    # Filtro por categoria
    categoria = request.args.get('categoria')
    if categoria:
        query = query.filter(Lancamento.categoria == categoria)
    
    # Filtro por data (sempre em cima de data_prevista / vencimento)
    def _parse_data(valor: str):
        if not valor:
            return None
        valor = valor.strip()
        try:
            if '/' in valor:
                return datetime.strptime(valor, '%d/%m/%Y').date()
            return datetime.strptime(valor, '%Y-%m-%d').date()
        except ValueError:
            return None

    hoje_completo = datetime.now()
    hoje = hoje_completo.date()
    
    # Datas padrão: primeiro e último dia do mês corrente
    primeiro_dia_mes = hoje.replace(day=1)
    import calendar
    ultimo_dia_mes = hoje.replace(day=calendar.monthrange(hoje.year, hoje.month)[1])

    data_inicio_raw = request.args.get('data_inicio', '')
    data_fim_raw = request.args.get('data_fim', '')
    
    # Se não houver filtros de data e não houver busca, usar o mês corrente como padrão
    if not data_inicio_raw and not data_fim_raw and not request.args.get('busca') and not request.args.get('nota_fiscal'):
        data_inicio_obj = primeiro_dia_mes
        data_fim_obj = ultimo_dia_mes
        # Atualizar strings para exibir no template
        data_inicio_str = primeiro_dia_mes.strftime('%d/%m/%Y')
        data_fim_str = ultimo_dia_mes.strftime('%d/%m/%Y')
    else:
        data_inicio_obj = _parse_data(data_inicio_raw)
        data_fim_obj = _parse_data(data_fim_raw)
        data_inicio_str = data_inicio_raw
        data_fim_str = data_fim_raw

    # Se ambas as datas foram passadas e estão invertidas, corrigir a ordem
    if data_inicio_obj and data_fim_obj and data_inicio_obj > data_fim_obj:
        data_inicio_obj, data_fim_obj = data_fim_obj, data_inicio_obj

    if data_inicio_obj and data_fim_obj:
        query = query.filter(Lancamento.data_prevista.between(data_inicio_obj, data_fim_obj))
    elif data_inicio_obj:
        query = query.filter(Lancamento.data_prevista >= data_inicio_obj)
    elif data_fim_obj:
        query = query.filter(Lancamento.data_prevista <= data_fim_obj)
    
    # Filtro por status
    status = request.args.get('status')
    hoje = datetime.now().date()
    if status == 'realizado':
        query = query.filter(Lancamento.data_realizada.isnot(None), Lancamento.data_realizada <= hoje)
    elif status == 'pendente':
        query = query.filter(
            db.and_(
                Lancamento.data_realizada.is_(None),
                Lancamento.data_prevista >= hoje
            )
        )
    elif status == 'agendado':
        query = query.filter(
            db.and_(
                Lancamento.data_realizada.isnot(None),
                Lancamento.data_realizada > hoje
            )
        )
    elif status == 'vencido':
        query = query.filter(
            db.and_(
                Lancamento.data_realizada.is_(None),
                Lancamento.data_prevista < hoje
            )
        )
    
    # Filtro por Conta Caixa
    conta_caixa_id = request.args.get('conta_caixa_id')
    if conta_caixa_id:
        try:
            conta_caixa_id = int(conta_caixa_id)
            query = query.filter(Lancamento.conta_caixa_id == conta_caixa_id)
        except (ValueError, TypeError):
            pass

    # Filtro por Nota Fiscal (Suporte a múltiplos NFs separados por ponto e vírgula)
    nf_raw = request.args.get('nota_fiscal', '').strip()
    if nf_raw:
        nfs = [n.strip() for n in nf_raw.split(';') if n.strip()]
        if nfs:
            if len(nfs) == 1:
                query = query.filter(Lancamento.nota_fiscal.ilike(f'%{nfs[0]}%'))
            else:
                query = query.filter(Lancamento.nota_fiscal.in_(nfs))
    
    # Filtro por busca (descrição, cliente, fornecedor ou valor)
    busca = request.args.get('busca')
    if busca and busca.strip():
        busca_clean = busca.strip()
        query = query.join(Cliente, Lancamento.cliente_id == Cliente.id, isouter=True).join(
            Fornecedor, Lancamento.fornecedor_id == Fornecedor.id, isouter=True)
        
        # Verificar se a busca é um número (valor)
        try:
            valor_busca = float(busca_clean.replace(',', '.'))
            query = query.filter(
                db.or_(
                    Lancamento.descricao.ilike(f'%{busca_clean}%'),
                    Cliente.nome.ilike(f'%{busca_clean}%'),
                    Fornecedor.nome.ilike(f'%{busca_clean}%'),
                    db.func.abs(Lancamento.valor - valor_busca) < 0.01  # Busca aproximada para valores
                )
            )
        except ValueError:
            # Se não é um número, buscar apenas texto (busca mais precisa)
            query = query.filter(
                db.or_(
                    Lancamento.descricao.ilike(f'%{busca_clean}%'),
                    Cliente.nome.ilike(f'%{busca_clean}%'),
                    Fornecedor.nome.ilike(f'%{busca_clean}%')
                )
            )
    
    # Ordenar e executar query
    lancamentos = query.order_by(Lancamento.data_prevista.desc()).all()

    # Data atual para verificar status
    hoje = datetime.now().date()

    # Calcular totais financeiros dinâmicos baseados nos filtros aplicados
    entradas_totais = sum([l.valor for l in lancamentos if l.tipo == 'entrada'])
    saidas_totais = sum([l.valor for l in lancamentos if l.tipo == 'saida'])
    # Saldo atual considera apenas entradas e saidas REALIZADAS
    entradas_totais_realizadas = sum([l.valor for l in lancamentos if l.tipo == 'entrada' and ((l.data_realizada and l.data_realizada <= hoje) or (l.realizado and not l.data_realizada))])
    saidas_totais_realizadas = sum([l.valor for l in lancamentos if l.tipo == 'saida' and ((l.data_realizada and l.data_realizada <= hoje) or (l.realizado and not l.data_realizada))])
    saldo_atual = entradas_totais_realizadas - saidas_totais_realizadas
    
    # Totais por status (baseado na nova lógica)
    # Realizado: tem data_realizada <= hoje OU (realizado=True sem data_realizada)
    entradas_realizadas = sum([l.valor for l in lancamentos if l.tipo == 'entrada' and ((l.data_realizada and l.data_realizada <= hoje) or (l.realizado and not l.data_realizada))])
    saidas_realizadas = sum([l.valor for l in lancamentos if l.tipo == 'saida' and ((l.data_realizada and l.data_realizada <= hoje) or (l.realizado and not l.data_realizada))])
    # Agendado: tem data_realizada futura (data_realizada > hoje)
    entradas_agendadas = sum([l.valor for l in lancamentos if l.tipo == 'entrada' and l.data_realizada and l.data_realizada > hoje])
    saidas_agendadas = sum([l.valor for l in lancamentos if l.tipo == 'saida' and l.data_realizada and l.data_realizada > hoje])
    # A vencer: não realizado, sem data_realizada E data_prevista >= hoje
    entradas_a_vencer = sum([l.valor for l in lancamentos if l.tipo == 'entrada' and not l.realizado and not l.data_realizada and l.data_prevista >= hoje])
    saidas_a_vencer = sum([l.valor for l in lancamentos if l.tipo == 'saida' and not l.realizado and not l.data_realizada and l.data_prevista >= hoje])
    # Vencido: não realizado, sem data_realizada E data_prevista < hoje
    entradas_vencidas = sum([l.valor for l in lancamentos if l.tipo == 'entrada' and not l.realizado and not l.data_realizada and l.data_prevista < hoje])
    saidas_vencidas = sum([l.valor for l in lancamentos if l.tipo == 'saida' and not l.realizado and not l.data_realizada and l.data_prevista < hoje])

    # Saldos por status
    saldo_realizado = entradas_realizadas - saidas_realizadas
    saldo_a_vencer = abs(entradas_a_vencer - saidas_a_vencer)  # Valor absoluto
    saldo_vencido = abs(entradas_vencidas - saidas_vencidas)  # Valor absoluto
    saldo_agendado = abs(entradas_agendadas - saidas_agendadas)  # Valor absoluto
    
    # Buscar categorias para filtros (todas analíticas da empresa)
    categorias_plano = PlanoConta.query.filter(
        PlanoConta.empresa_id == empresa_id_correta,
        PlanoConta.natureza == 'analitica',
        PlanoConta.ativo == True
    ).order_by(PlanoConta.codigo).all()

    # Buscar contas caixa da empresa
    contas_caixa = ContaCaixa.query.filter(
        ContaCaixa.usuario_id.in_(usuarios_ids),
        ContaCaixa.ativo == True
    ).order_by(ContaCaixa.nome).all()

    # Criar lista de strings formatadas para compatibilidade ou objetos para o novo template
    # Para o template moderno, melhor passar os objetos

    return render_template('lancamentos_moderno.html', usuario=usuario, lancamentos=lancamentos,
                          total_entradas=entradas_totais, total_saidas=saidas_totais,
                          saldo_atual=saldo_atual, categorias=categorias_plano, contas_caixa=contas_caixa,
                          hoje=hoje, data_inicio=data_inicio_str, data_fim=data_fim_str,
                          # Valores separados para entradas
                          entrada_total=entradas_totais,
                          entrada_realizada=entradas_realizadas,
                          entrada_a_vencer=entradas_a_vencer,
                          entrada_vencida=entradas_vencidas,
                          entrada_agendada=entradas_agendadas,
                          # Valores separados para saidas
                          saida_total=saidas_totais,
                          saida_realizada=saidas_realizadas,
                          saida_a_vencer=saidas_a_vencer,
                          saida_vencida=saidas_vencidas,
                          saida_agendada=saidas_agendadas,
                          # Valores legados (manter compatibilidade)
                          entradas_realizadas=entradas_realizadas, saidas_realizadas=saidas_realizadas,
                          entradas_a_vencer=entradas_a_vencer, saidas_a_vencer=saidas_a_vencer,
                          entradas_vencidas=entradas_vencidas, saidas_vencidas=saidas_vencidas,
                          entradas_agendadas=entradas_agendadas, saidas_agendadas=saidas_agendadas,
                          saldo_realizado=saldo_realizado, saldo_a_vencer=saldo_a_vencer,
                          saldo_vencido=saldo_vencido, saldo_agendado=saldo_agendado)

@app.route('/lancamentos/excluir-lote', methods=['POST'])
def excluir_lancamentos_lote():
    if 'usuario_id' not in session:
        return jsonify({'success': False, 'message': 'Não autorizado'}), 401
    
    try:
        data = request.get_json()
        # Aceitar tanto 'lancamento_ids' quanto 'ids' para compatibilidade
        ids = data.get('lancamento_ids', data.get('ids', []))

        if not ids:
            return jsonify({'success': False, 'message': 'Nenhum lançamento selecionado'})
        
        # Verificar se o usuário tem permissão para excluir estes lançamentos
        usuario = db.session.get(Usuario, session['usuario_id'])
        empresa_id = obter_empresa_id_sessao(session, usuario)

        usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
        usuarios_ids = [u.id for u in usuarios_empresa]
        
        # Buscar lançamentos que pertencem à empresa do usuário
        lancamentos = Lancamento.query.filter(
            Lancamento.id.in_(ids),
            Lancamento.empresa_id == empresa_id
        ).all()

        if not lancamentos:
            return jsonify({'success': False, 'message': 'Nenhum lançamento válido encontrado'})

        # Verificar vínculos (parâmetros de impedimento)
        lancamentos_vinculados = []
        avisos = []

        for lancamento in lancamentos:
            # Verificar vínculos diretos (compras/vendas)
            if hasattr(lancamento, 'compra') and lancamento.compra:
                avisos.append(f'Lançamento #{lancamento.id}: vinculado à compra #{lancamento.compra.id}')
                lancamentos_vinculados.append(lancamento.id)
                continue
            if hasattr(lancamento, 'venda') and lancamento.venda:
                avisos.append(f'Lançamento #{lancamento.id}: vinculado à venda #{lancamento.venda.id}')
                lancamentos_vinculados.append(lancamento.id)
                continue

            # Verificar vínculos através da tabela Vinculo
            vinculo_existe = Vinculo.query.filter(
                db.or_(
                    db.and_(Vinculo.lado_a_tipo == 'lancamento', Vinculo.lado_a_id == lancamento.id),
                    db.and_(Vinculo.lado_b_tipo == 'lancamento', Vinculo.lado_b_id == lancamento.id)
                )
            ).first()

            if vinculo_existe:
                avisos.append(f'Lançamento #{lancamento.id}: vinculado a outra transação')
                lancamentos_vinculados.append(lancamento.id)

        # Excluir apenas lançamentos não vinculados
        excluidos = 0
        for lancamento in lancamentos:
            if lancamento.id not in lancamentos_vinculados:
                db.session.delete(lancamento)
                excluidos += 1

        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'{excluidos} lançamento(s) excluído(s) com sucesso',
            'avisos': avisos if avisos else []
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Erro ao excluir: {str(e)}'}), 500

# Rota removida - duplicada. Ver linha 11611 para a implementação completa

@app.route('/admin/verificar-empresas-orfas')
def verificar_empresas_orfas():
    """Rota administrativa para verificar e limpar empresas órfãs"""
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    admin = db.session.get(Usuario, session['usuario_id'])
    if admin.tipo != 'admin':
        flash('Acesso negado.', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        app.logger.info("Verificando empresas órfãs...")
        
        # Buscar todas as empresas
        empresas = Empresa.query.all()
        empresas_orfas = []
        
        for empresa in empresas:
            usuarios_count = Usuario.query.filter_by(empresa_id=empresa.id).count()
            if usuarios_count == 0:
                empresas_orfas.append(empresa)
                app.logger.info(f"Empresa órfã encontrada: {empresa.razao_social} (CNPJ: {empresa.cnpj})")
        
        if empresas_orfas:
            # Deletar empresas órfãs
            for empresa in empresas_orfas:
                try:
                    db.session.delete(empresa)
                    app.logger.info(f"Empresa órfã {empresa.razao_social} (CNPJ: {empresa.cnpj}) deletada.")
                except Exception as e:
                    db.session.rollback()
                    app.logger.error(f"Erro ao deletar empresa {empresa.razao_social}: {str(e)}")
            
            db.session.commit()
            flash(f'{len(empresas_orfas)} empresa(s) órfã(s) removida(s) com sucesso! CNPJs liberados para novo cadastro.', 'success')
        else:
            flash('Nenhuma empresa órfã encontrada.', 'info')
            
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao verificar empresas órfãs: {str(e)}")
        flash(f'Erro ao verificar empresas órfãs: {str(e)}', 'error')
    
    return redirect(url_for('admin_dashboard'))

@app.route('/lancamentos/novo', methods=['GET', 'POST'])
def novo_lancamento():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario:
        flash('Usuário não encontrado', 'error')
        return redirect(url_for('login'))
    
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    if request.method == 'POST':
        try:
            # Validar campos obrigatórios
            form_data = request.form.to_dict(flat=True)
            def render_form(errors=None):
                try:
                    empresa_id_correta_local = obter_empresa_id_sessao(session, usuario)
                    if empresa_id_correta_local:
                        usuarios_empresa_local = Usuario.query.filter_by(empresa_id=empresa_id_correta_local, ativo=True).all()
                        usuarios_ids_local = [u.id for u in usuarios_empresa_local]
                        contas_caixa_local = ContaCaixa.query.filter(ContaCaixa.usuario_id.in_(usuarios_ids_local), ContaCaixa.ativo==True).order_by(ContaCaixa.nome).all()
                    else:
                        contas_caixa_local = []
                        usuarios_ids_local = []
                except Exception:
                    db.session.rollback()
                    contas_caixa_local = []
                    usuarios_ids_local = []
                try:
                    if usuarios_ids_local:
                        clientes_local = Cliente.query.filter(Cliente.usuario_id.in_(usuarios_ids_local)).order_by(Cliente.nome).all()
                        fornecedores_local = Fornecedor.query.filter(Fornecedor.usuario_id.in_(usuarios_ids_local)).order_by(Fornecedor.nome).all()
                        produtos_local = Produto.query.filter(Produto.usuario_id.in_(usuarios_ids_local), Produto.ativo==True).order_by(Produto.nome).all()
                        servicos_local = Servico.query.filter(Servico.usuario_id.in_(usuarios_ids_local), Servico.ativo==True).order_by(Servico.nome).all()
                        categorias_entrada_local = PlanoConta.query.filter(PlanoConta.usuario_id.in_(usuarios_ids_local), PlanoConta.tipo=='entrada', PlanoConta.ativo==True).order_by(PlanoConta.nome).all()
                        categorias_saida_local = PlanoConta.query.filter(PlanoConta.usuario_id.in_(usuarios_ids_local), PlanoConta.tipo=='saida', PlanoConta.ativo==True).order_by(PlanoConta.nome).all()
                    else:
                        clientes_local = []
                        fornecedores_local = []
                        produtos_local = []
                        servicos_local = []
                        categorias_receita_local = []
                        categorias_despesa_local = []
                except Exception:
                    db.session.rollback()
                    clientes_local = []
                    fornecedores_local = []
                    produtos_local = []
                    servicos_local = []
                    categorias_receita_local = []
                    categorias_despesa_local = []
                return render_template('novo_lancamento.html', usuario=usuario, contas_caixa=contas_caixa_local, clientes=clientes_local, fornecedores=fornecedores_local, produtos=produtos_local, servicos=servicos_local, categorias_entrada=categorias_entrada_local, categorias_saida=categorias_saida_local, form_data=form_data, errors=errors or {})
            descricao = request.form.get('descricao', '').strip()
            if not descricao:
                return render_form({'descricao': 'Descrição é obrigatória'})
            
            # Coletar campos para cálculo do valor total
            # Verificar se o usuário marcou para usar carrinho
            usar_carrinho = request.form.get('usar_carrinho') == '1'
            
            # Verificar se há itens no carrinho (novo formato)
            item_nomes = request.form.getlist('item_nome[]')
            item_tipos = request.form.getlist('item_tipo[]')
            item_precos = request.form.getlist('item_preco[]')
            item_qtds = request.form.getlist('item_qtd[]')
            item_totais = request.form.getlist('item_total[]')
            item_descontos = request.form.getlist('item_desconto[]')
            
            desconto_str = request.form.get('desconto', '').strip()
            desconto = 0.0
            if desconto_str:
                try:
                    desconto = float(desconto_str.replace(',', '.'))
                except ValueError:
                    desconto = 0.0
            
            # Valores padrão
            valor_total_calculado = None
            tipo_produto_servico = None
            produto_servico_nome = None
            quantidade_itens = 1
            itens_carrinho_json = None  # Para armazenar itens do carrinho em JSON
            
            # Se usar carrinho e há itens no carrinho, processar
            if usar_carrinho and item_nomes and len(item_nomes) > 0 and any(nome.strip() for nome in item_nomes):
                # Processar itens do carrinho
                itens_validos = []
                for i, nome in enumerate(item_nomes):
                    nome_limpo = nome.strip() if nome else ''
                    if nome_limpo:
                        try:
                            preco = float(item_precos[i].replace(',', '.')) if i < len(item_precos) and item_precos[i] else 0
                        except (ValueError, AttributeError):
                            preco = 0
                        
                        try:
                            qtd = float(item_qtds[i]) if i < len(item_qtds) and item_qtds[i] else 1
                        except (ValueError, AttributeError):
                            qtd = 1
                        
                        try:
                            desc_item = float(item_descontos[i].replace(',', '.')) if i < len(item_descontos) and item_descontos[i] else 0
                        except (ValueError, AttributeError):
                            desc_item = 0
                        
                        try:
                            total_str = item_totais[i] if i < len(item_totais) and item_totais[i] else '0'
                            total = float(total_str.replace(',', '.')) if total_str else 0
                        except (ValueError, AttributeError):
                            total = 0
                        
                        # Se total for zero, calcular baseado em preço * quantidade - desconto
                        if total == 0 and preco > 0:
                            total = max(0, (preco * qtd) - desc_item)
                        
                        itens_validos.append({
                            'nome': nome_limpo,
                            'tipo': item_tipos[i] if i < len(item_tipos) else 'produto',
                            'preco': preco,
                            'qtd': qtd,
                            'desconto': desc_item,
                            'total': total
                        })
                
                if itens_validos:
                    # Se houver múltiplos itens, concatenar os nomes
                    if len(itens_validos) > 1:
                        nomes_itens = [item['nome'] for item in itens_validos]
                        produto_servico_nome = ', '.join(nomes_itens)
                    else:
                        produto_servico_nome = itens_validos[0]['nome']

                    # Usar primeiro item para tipo e quantidade
                    primeiro_item = itens_validos[0]
                    tipo_produto_servico = primeiro_item['tipo']
                    quantidade_itens = primeiro_item['qtd']

                    # Calcular valor total do carrinho
                    valor_total_calculado = sum(item['total'] for item in itens_validos) - desconto
                    valor_total_calculado = max(0.0, valor_total_calculado)

                    # Salvar itens do carrinho em JSON para recuperação posterior
                    import json
                    itens_carrinho_json = json.dumps(itens_validos, ensure_ascii=False)
            
            # Se não usar carrinho ou não há itens no carrinho, usar valor direto ou formato antigo
            if valor_total_calculado is None:
                # Tentar valor direto primeiro
                valor_direto_str = request.form.get('valor_direto', '').strip()
                if valor_direto_str:
                    try:
                        valor_direto = float(valor_direto_str.replace(',', '.'))
                        valor_total_calculado = max(0.0, (valor_direto - desconto))
                    except ValueError:
                        pass
                
                # Se ainda não tem valor, usar formato antigo
                if valor_total_calculado is None:
                    item_tipo = request.form.get('item_tipo', '').strip()  # '' | 'servico' | 'produto'
                    # Serviço
                    if item_tipo == 'servico':
                        tipo_produto_servico = 'servico'
                        produto_servico_nome = request.form.get('servico_nome', '').strip() or None
                        servico_valor_str = request.form.get('servico_valor', '').strip()
                        servico_valor = 0.0
                        if servico_valor_str:
                            try:
                                servico_valor = float(servico_valor_str.replace(',', '.'))
                            except ValueError:
                                servico_valor = 0.0
                        valor_total_calculado = max(0.0, (servico_valor - desconto))
                    # Produto
                    elif item_tipo == 'produto':
                        tipo_produto_servico = 'produto'
                        produto_servico_nome = request.form.get('produto_nome', '').strip() or None
                        quantidade_str = request.form.get('quantidade', '').strip()
                        valor_unitario_str = request.form.get('valor_unitario', '').strip()
                        try:
                            quantidade_itens = int(quantidade_str) if quantidade_str else 1
                        except ValueError:
                            quantidade_itens = 1
                        if quantidade_itens <= 0:
                            quantidade_itens = 1
                        valor_unitario = 0.0
                        if valor_unitario_str:
                            try:
                                valor_unitario = float(valor_unitario_str.replace(',', '.'))
                            except ValueError:
                                valor_unitario = 0.0
                        valor_total_calculado = max(0.0, ((quantidade_itens * valor_unitario) - desconto))
                    # Não especificado: usa valor digitado e aplica desconto se houver
                    valor_str = request.form.get('valor', '').strip()
                    if valor_total_calculado is None:
                        if not valor_str:
                            return render_form({'valor': 'Valor Total é obrigatório'})
                        try:
                            valor_digitado = float(valor_str.replace(',', '.'))
                        except ValueError:
                            return render_form({'valor': 'Valor Total inválido'})
                        valor_total_calculado = max(0.0, (valor_digitado - desconto))
            
            tipo = request.form.get('tipo', '').strip()
            if not tipo or tipo not in ['entrada', 'saida', 'transferencia']:
                return render_form({'tipo': 'Tipo deve ser entrada, saida ou transferência'})

            # Se for transferência, validar contas origem e destino
            if tipo == 'transferencia':
                conta_origem_id = request.form.get('conta_caixa_id', '').strip()
                conta_destino_id = request.form.get('conta_caixa_destino_id', '').strip()

                if not conta_origem_id or not conta_destino_id:
                    return render_form({'conta_caixa_id': 'Contas origem e destino são obrigatórias para transferência'})

                try:
                    conta_origem_id = int(conta_origem_id)
                    conta_destino_id = int(conta_destino_id)
                except ValueError:
                    return render_form({'conta_caixa_id': 'IDs de conta inválidos'})

                if conta_origem_id == conta_destino_id:
                    return render_form({'conta_caixa_id': 'Contas origem e destino devem ser diferentes'})

                # Processar transferência (criar 2 lançamentos vinculados)
                # Pular validações de categoria para transferência
                plano_conta_id = None
                categoria_nome = "Transferência entre contas"
            else:
                # Lógica normal para entrada/saida
                categoria_input = request.form.get('categoria', '').strip()
                if not categoria_input:
                    return render_form({'categoria': 'Categoria é obrigatória'})

                plano_conta_id = None
                categoria_nome = categoria_input

                # Tentar converter para ID se for numérico
                if categoria_input.isdigit():
                    plano_conta_id = int(categoria_input)
                    pc = db.session.get(PlanoConta, plano_conta_id)
                    if pc:
                        categoria_nome = pc.nome
                else:
                    # Caso venha como nome (compatibilidade), buscar o ID
                    empresa_id_sessao = obter_empresa_id_sessao(session, usuario)
                    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id_sessao, ativo=True).all()
                    u_ids = [u.id for u in usuarios_empresa]
                    pc = PlanoConta.query.filter(PlanoConta.usuario_id.in_(u_ids), PlanoConta.nome == categoria_input, PlanoConta.tipo == tipo).first()
                    if pc:
                        plano_conta_id = pc.id

                # Validação: só permitir analíticas (opcional, já filtrado no front)
                if plano_conta_id:
                    pc = db.session.get(PlanoConta, plano_conta_id)
                    if pc and pc.natureza == 'sintetica':
                        return render_form({'categoria': 'Não é permitido lançar em contas sintéticas'})

                # Se categoria não existe no plano de contas, criar automaticamente
                empresa_id_para_categoria = obter_empresa_id_sessao(session, usuario)
                usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id_para_categoria, ativo=True).all()
                usuarios_ids = [u.id for u in usuarios_empresa]

                if not plano_conta_id:
                    # Criar categoria automaticamente
                    nova_categoria = PlanoConta(
                        nome=categoria_nome,
                        tipo=tipo,
                        descricao=f"Categoria criada automaticamente para lançamento",
                        usuario_id=usuario.id
                    )
                    db.session.add(nova_categoria)
                    db.session.flush()  # Para obter o ID sem fazer commit ainda
                    plano_conta_id = nova_categoria.id
            
            data_prevista_str = request.form.get('data_prevista', '').strip()
            if not data_prevista_str:
                return render_form({'data_prevista': 'Data prevista é obrigatória'})
            
            try:
                # Tentar formato DD/MM/AAAA primeiro
                try:
                    data_prevista = datetime.strptime(data_prevista_str, '%d/%m/%Y').date()
                except ValueError:
                    # Fallback para formato AAAA-MM-DD
                    data_prevista = datetime.strptime(data_prevista_str, '%Y-%m-%d').date()

                # Validação: data não pode ser mais de 10 anos no futuro
                from datetime import date, timedelta
                data_limite = date.today() + timedelta(days=3650)  # 10 anos
                if data_prevista > data_limite:
                    return render_form({'data_prevista': 'Data prevista não pode ser superior a 10 anos no futuro'})

            except ValueError:
                return render_form({'data_prevista': 'Data prevista inválida'})
            
            data_realizada = request.form.get('data_realizada', '').strip()
            data_realizada_obj = None
            if data_realizada:
                try:
                    # Tentar formato DD/MM/AAAA primeiro
                    try:
                        data_realizada_obj = datetime.strptime(data_realizada, '%d/%m/%Y').date()
                    except ValueError:
                        # Fallback para formato AAAA-MM-DD
                        data_realizada_obj = datetime.strptime(data_realizada, '%Y-%m-%d').date()
                except ValueError:
                    return render_form({'data_realizada': 'Data realizada inválida'})
            # Regra: se data_realizada hoje ou passada -> realizado; futura -> agendado (pendente)
            from datetime import date
            realizado = False
            if data_realizada_obj:
                if data_realizada_obj <= date.today():
                    realizado = True
            conta_caixa_id = request.form.get('conta_caixa_id', '').strip()
            conta_caixa_id_obj = None
            if conta_caixa_id:
                try:
                    conta_caixa_id_obj = int(conta_caixa_id)
                except ValueError:
                    return render_form({'conta_caixa_id': 'Conta caixa inválida'})
            
            # Campos de cliente/fornecedor
            cliente_id = request.form.get('cliente_id', '').strip()
            cliente_id_obj = None
            if cliente_id:
                try:
                    cliente_id_obj = int(cliente_id)
                    cliente = db.session.get(Cliente, cliente_id_obj)
                    if not cliente or cliente.usuario_id != usuario.id:
                        return render_form({'cliente_fornecedor': 'Cliente inválido'})
                except ValueError:
                    return render_form({'cliente_fornecedor': 'ID do cliente inválido'})
            
            fornecedor_id = request.form.get('fornecedor_id', '').strip()
            fornecedor_id_obj = None
            if fornecedor_id:
                try:
                    fornecedor_id_obj = int(fornecedor_id)
                    fornecedor = db.session.get(Fornecedor, fornecedor_id_obj)
                    if not fornecedor or fornecedor.usuario_id != usuario.id:
                        return render_form({'cliente_fornecedor': 'Fornecedor inválido'})
                except ValueError:
                    return render_form({'cliente_fornecedor': 'ID do fornecedor inválido'})
            
            # Validar se empresa_id existe
            if not usuario.empresa_id:
                return render_form({'__all__': 'Usuário não possui empresa associada'})
            
            # Criar o lançamento
            # Processar novos campos adicionais
            observacoes = request.form.get('observacoes', '').strip() or None
            # Se não especificou produto/serviço via item_tipo, usa campo livre existente
            if not produto_servico_nome:
                produto_servico_nome = request.form.get('produto_servico', '').strip() or None
            
            # Log para debug
            app.logger.info(f"Salvando lançamento - produto_servico_nome: {produto_servico_nome}, tipo_produto_servico: {tipo_produto_servico}")
            
            # Obter empresa_id correta da sessão (considera acesso_contador)
            empresa_id_correta = obter_empresa_id_sessao(session, usuario)
            if not empresa_id_correta:
                return render_form({'__all__': 'Usuário não possui empresa associada'})

            nota_fiscal = request.form.get('nota_fiscal', '').strip() or None

            # Se for transferência, criar 2 lançamentos vinculados
            if tipo == 'transferencia':
                # Lançamento 1: SAÍDA (saída da conta origem)
                lancamento_saida = Lancamento(
                    descricao=f"{descricao} (Origem)",
                    valor=valor_total_calculado,
                    tipo='saida',
                    categoria=categoria_nome,
                    plano_conta_id=None,  # Transferências não têm categoria do plano de contas
                    data_prevista=data_prevista,
                    data_realizada=data_realizada_obj,
                    realizado=realizado,
                    nota_fiscal=nota_fiscal,
                    usuario_id=usuario.id,
                    empresa_id=empresa_id_correta,
                    conta_caixa_id=conta_origem_id,
                    cliente_id=None,
                    fornecedor_id=None,
                    observacoes=observacoes,
                    produto_servico=None,
                    tipo_produto_servico=None,
                    itens_carrinho=None,
                    usuario_criacao_id=usuario.id,
                    eh_transferencia=True,
                )
                db.session.add(lancamento_saida)
                db.session.flush()

                # Lançamento 2: ENTRADA (entrada na conta destino)
                lancamento_entrada = Lancamento(
                    descricao=f"{descricao} (Destino)",
                    valor=valor_total_calculado,
                    tipo='entrada',
                    categoria=categoria_nome,
                    plano_conta_id=None,
                    data_prevista=data_prevista,
                    data_realizada=data_realizada_obj,
                    realizado=realizado,
                    nota_fiscal=nota_fiscal,
                    usuario_id=usuario.id,
                    empresa_id=empresa_id_correta,
                    conta_caixa_id=conta_destino_id,
                    cliente_id=None,
                    fornecedor_id=None,
                    observacoes=observacoes,
                    produto_servico=None,
                    tipo_produto_servico=None,
                    itens_carrinho=None,
                    usuario_criacao_id=usuario.id,
                    eh_transferencia=True,
                )
                db.session.add(lancamento_entrada)
                db.session.flush()

                # Vincular os dois lançamentos usando o ID do lançamento de saída
                lancamento_saida.transferencia_id = lancamento_entrada.id
                lancamento_entrada.transferencia_id = lancamento_saida.id

                novo_lancamento = lancamento_saida  # Para compatibilidade com código abaixo
            else:
                # Lançamento normal (entrada ou saida)
                novo_lancamento = Lancamento(
                    descricao=descricao,
                    valor=valor_total_calculado,
                    tipo=tipo,
                    categoria=categoria_nome,
                    plano_conta_id=plano_conta_id,
                    data_prevista=data_prevista,
                    data_realizada=data_realizada_obj,
                    realizado=realizado,
                    nota_fiscal=nota_fiscal,
                    usuario_id=usuario.id,
                    empresa_id=empresa_id_correta,
                    conta_caixa_id=conta_caixa_id_obj,
                    cliente_id=cliente_id_obj,
                    fornecedor_id=fornecedor_id_obj,
                    observacoes=observacoes,
                    produto_servico=produto_servico_nome,
                    tipo_produto_servico=tipo_produto_servico,
                    itens_carrinho=itens_carrinho_json,
                    usuario_criacao_id=usuario.id,
                    eh_transferencia=False,
                )

                # REMOVIDO: Atualização manual de saldo - agora é calculado dinamicamente
                # O saldo_atual é calculado automaticamente pelo método calcular_saldo_atual()

                db.session.add(novo_lancamento)
                db.session.flush()  # Para obter o ID do lançamento antes do commit

            # Se usou carrinho e há cliente/fornecedor, criar venda/compra automaticamente
            # (não criar para transferências)
            if not tipo == 'transferencia' and itens_carrinho_json and (cliente_id_obj or fornecedor_id_obj):
                # Determinar se é venda ou compra
                # ENTRADA + CLIENTE = VENDA
                # SAÍDA + FORNECEDOR = COMPRA
                if tipo == 'entrada' and cliente_id_obj:
                    # Criar venda automaticamente
                    import json
                    itens = json.loads(itens_carrinho_json)

                    # Se múltiplos itens, mostrar "Diversos" na coluna produto
                    produto_descricao = "Diversos" if len(itens) > 1 else itens[0]['nome']
                    quantidade_total = sum(item['qtd'] for item in itens)

                    nova_venda = Venda(
                        produto=produto_descricao,
                        quantidade=quantidade_total,
                        valor=valor_total_calculado,
                        valor_final=valor_total_calculado,
                        tipo_venda=itens[0]['tipo'],
                        tipo_pagamento='a_vista',
                        numero_parcelas=1,
                        data_prevista=data_prevista,
                        data_realizada=data_realizada_obj,
                        realizado=realizado,
                        cliente_id=cliente_id_obj,
                        observacoes=f"Gerado automaticamente do lançamento financeiro",
                        usuario_id=usuario.id,
                        empresa_id=empresa_id_correta
                    )
                    db.session.add(nova_venda)
                    db.session.flush()

                    # Vincular lançamento à venda
                    novo_lancamento.venda_id = nova_venda.id

                    app.logger.info(f"✅ Venda ID {nova_venda.id} criada automaticamente: {quantidade_total}x {produto_descricao} = R$ {valor_total_calculado:.2f}")

                elif tipo == 'saida' and fornecedor_id_obj:
                    # Criar compra automaticamente
                    import json
                    itens = json.loads(itens_carrinho_json)

                    # Se múltiplos itens, mostrar "Diversos" na coluna produto
                    produto_descricao = "Diversos" if len(itens) > 1 else itens[0]['nome']
                    quantidade_total = sum(item['qtd'] for item in itens)

                    nova_compra = Compra(
                        produto=produto_descricao,
                        quantidade=quantidade_total,
                        preco_custo=valor_total_calculado / quantidade_total if quantidade_total > 0 else 0,
                        valor=valor_total_calculado,
                        tipo_compra=itens[0]['tipo'],
                        tipo_pagamento='a_vista',
                        numero_parcelas=1,
                        data_prevista=data_prevista,
                        data_realizada=data_realizada_obj,
                        realizado=realizado,
                        fornecedor_id=fornecedor_id_obj,
                        observacoes=f"Gerado automaticamente do lançamento financeiro",
                        usuario_id=usuario.id,
                        empresa_id=empresa_id_correta
                    )
                    db.session.add(nova_compra)
                    db.session.flush()

                    # Vincular lançamento à compra
                    novo_lancamento.compra_id = nova_compra.id

                    app.logger.info(f"✅ Compra ID {nova_compra.id} criada automaticamente: {quantidade_total}x {produto_descricao} = R$ {valor_total_calculado:.2f}")

            db.session.commit()

            if tipo == 'transferencia':
                flash('Transferência entre contas criada com sucesso!', 'success')
            else:
                flash('Lançamento criado com sucesso!', 'success')
            return redirect(url_for('lancamentos'))
            
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Erro ao processar lançamento: {str(e)}")
            form_data = request.form.to_dict(flat=True)
            return render_template('novo_lancamento.html', usuario=usuario, contas_caixa=[], clientes=[], fornecedores=[], form_data=form_data, errors={'__all__': f'Erro ao processar lançamento: {str(e)}'})
    
    # Buscar contas caixa ativas de todos os usuários da empresa correta
    try:
        empresa_id_correta = obter_empresa_id_sessao(session, usuario)
        if empresa_id_correta:
            usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id_correta, ativo=True).all()
            usuarios_ids = [u.id for u in usuarios_empresa]
            contas_caixa = ContaCaixa.query.filter(ContaCaixa.usuario_id.in_(usuarios_ids), ContaCaixa.ativo==True).order_by(ContaCaixa.nome).all()
        else:
            contas_caixa = []
            usuarios_ids = []
    except Exception as e:
        app.logger.error(f"Erro ao buscar contas caixa: {str(e)}")
        contas_caixa = []
        usuarios_ids = []

    # Buscar clientes, fornecedores, produtos, serviços e plano de contas da empresa vinculada
    try:
        if usuarios_ids:
            clientes = Cliente.query.filter(Cliente.empresa_id == empresa_id_correta).order_by(Cliente.nome).all()
            fornecedores = Fornecedor.query.filter(Fornecedor.empresa_id == empresa_id_correta).order_by(Fornecedor.nome).all()
            produtos = Produto.query.filter(Produto.usuario_id.in_(usuarios_ids), Produto.ativo==True).order_by(Produto.nome).all()
            servicos = Servico.query.filter(Servico.usuario_id.in_(usuarios_ids), Servico.ativo==True).order_by(Servico.nome).all()
            categorias_receita = PlanoConta.query.filter(PlanoConta.empresa_id == empresa_id_correta, PlanoConta.tipo=='entrada', PlanoConta.ativo==True).order_by(PlanoConta.nome).all()
            categorias_despesa = PlanoConta.query.filter(PlanoConta.empresa_id == empresa_id_correta, PlanoConta.tipo=='saida', PlanoConta.ativo==True).order_by(PlanoConta.nome).all()
        else:
            clientes = []
            fornecedores = []
            produtos = []
            servicos = []
            categorias_receita = []
            categorias_despesa = []
    except Exception as e:
        app.logger.error(f"Erro ao buscar dados: {str(e)}")
        clientes = []
        fornecedores = []
        produtos = []
        servicos = []
        categorias_receita = []
        categorias_despesa = []

    return render_template('novo_lancamento.html', usuario=usuario, contas_caixa=contas_caixa, clientes=clientes, fornecedores=fornecedores, produtos=produtos, servicos=servicos, categorias_receita=categorias_receita, categorias_despesa=categorias_despesa)

@app.route('/transferencia/nova', methods=['GET', 'POST'])
def nova_transferencia():
    """Criar nova transferência entre contas"""
    if 'usuario_id' not in session:
        return redirect(url_for('login'))

    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))

    empresa_id_correta = obter_empresa_id_sessao(session, usuario)
    if not empresa_id_correta:
        flash('Erro ao obter empresa associada', 'error')
        return redirect(url_for('dashboard'))

    # Buscar contas caixa da empresa
    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id_correta, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]
    contas_caixa = ContaCaixa.query.filter(ContaCaixa.usuario_id.in_(usuarios_ids), ContaCaixa.ativo==True).order_by(ContaCaixa.nome).all()

    if request.method == 'POST':
        try:
            # Validar dados
            descricao = request.form.get('descricao', '').strip()
            valor_str = request.form.get('valor', '').strip()
            conta_origem_id = request.form.get('conta_origem_id', '').strip()
            conta_destino_id = request.form.get('conta_destino_id', '').strip()
            data_prevista_str = request.form.get('data_prevista', '').strip()
            data_realizada_str = request.form.get('data_realizada', '').strip()
            observacoes = request.form.get('observacoes', '').strip()

            errors = {}

            if not descricao:
                errors['descricao'] = 'Descrição é obrigatória'
            if not valor_str:
                errors['valor'] = 'Valor é obrigatório'
            if not conta_origem_id:
                errors['conta_origem_id'] = 'Conta origem é obrigatória'
            if not conta_destino_id:
                errors['conta_destino_id'] = 'Conta destino é obrigatória'
            if not data_prevista_str:
                errors['data_prevista'] = 'Data prevista é obrigatória'

            if conta_origem_id == conta_destino_id:
                errors['conta_destino_id'] = 'Contas origem e destino devem ser diferentes'

            if errors:
                return render_template('nova_transferencia.html', usuario=usuario, contas_caixa=contas_caixa, errors=errors, form_data=request.form)

            # Converter valores
            try:
                valor = float(valor_str.replace(',', '.'))
            except ValueError:
                return render_template('nova_transferencia.html', usuario=usuario, contas_caixa=contas_caixa, errors={'valor': 'Valor inválido'}, form_data=request.form)

            try:
                data_prevista = datetime.strptime(data_prevista_str, '%d/%m/%Y').date()
            except ValueError:
                try:
                    data_prevista = datetime.strptime(data_prevista_str, '%Y-%m-%d').date()
                except ValueError:
                    return render_template('nova_transferencia.html', usuario=usuario, contas_caixa=contas_caixa, errors={'data_prevista': 'Data inválida'}, form_data=request.form)

            data_realizada_obj = None
            realizado = False
            if data_realizada_str:
                try:
                    data_realizada_obj = datetime.strptime(data_realizada_str, '%d/%m/%Y').date()
                    realizado = True
                except ValueError:
                    try:
                        data_realizada_obj = datetime.strptime(data_realizada_str, '%Y-%m-%d').date()
                        realizado = True
                    except ValueError:
                        pass

            # Criar lançamento de saída (saida)
            lancamento_saida = Lancamento(
                descricao=f"{descricao} (Origem)",
                valor=valor,
                tipo='saida',
                categoria="Transferência entre contas",
                plano_conta_id=None,
                data_prevista=data_prevista,
                data_realizada=data_realizada_obj,
                realizado=realizado,
                usuario_id=usuario.id,
                empresa_id=empresa_id_correta,
                conta_caixa_id=int(conta_origem_id),
                observacoes=observacoes,
                usuario_criacao_id=usuario.id,
                eh_transferencia=True
            )
            db.session.add(lancamento_saida)
            db.session.flush()

            # Criar lançamento de entrada (entrada)
            lancamento_entrada = Lancamento(
                descricao=f"{descricao} (Destino)",
                valor=valor,
                tipo='entrada',
                categoria="Transferência entre contas",
                plano_conta_id=None,
                data_prevista=data_prevista,
                data_realizada=data_realizada_obj,
                realizado=realizado,
                usuario_id=usuario.id,
                empresa_id=empresa_id_correta,
                conta_caixa_id=int(conta_destino_id),
                observacoes=observacoes,
                usuario_criacao_id=usuario.id,
                eh_transferencia=True
            )
            db.session.add(lancamento_entrada)
            db.session.flush()

            # Vincular os lançamentos
            lancamento_saida.transferencia_id = lancamento_entrada.id
            lancamento_entrada.transferencia_id = lancamento_saida.id

            db.session.commit()
            flash('Transferência criada com sucesso!', 'success')
            return redirect(url_for('lancamentos'))

        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Erro ao criar transferência: {str(e)}")
            return render_template('nova_transferencia.html', usuario=usuario, contas_caixa=contas_caixa, errors={'__all__': f'Erro ao criar transferência: {str(e)}'}, form_data=request.form)

    return render_template('nova_transferencia.html', usuario=usuario, contas_caixa=contas_caixa)

@app.route('/lancamentos/<int:lancamento_id>/toggle_realizado')
def toggle_lancamento_realizado_old(lancamento_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    # Obter empresa_id correta da sessão (considera acesso_contador)
    empresa_id_correta = obter_empresa_id_sessao(session, usuario)
    if not empresa_id_correta:
        flash('Erro ao obter empresa associada', 'error')
        return redirect(url_for('lancamentos'))
    
    # Buscar todos os usuários da empresa correta
    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id_correta, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]
    
    # Buscar lançamento pela empresa_id correta
    lancamento = Lancamento.query.filter(
        Lancamento.id == lancamento_id,
        Lancamento.empresa_id == empresa_id_correta
    ).first()
    
    if lancamento:
        # Salvar status anterior para cálculo de saldo
        realizado_anterior = lancamento.realizado
        
        lancamento.realizado = not lancamento.realizado
        if lancamento.realizado and not lancamento.data_realizada:
            lancamento.data_realizada = datetime.now().date()
        elif not lancamento.realizado:
            lancamento.data_realizada = None
        
        # REMOVIDO: Atualização manual de saldo - agora é calculado dinamicamente
        # O saldo_atual é calculado automaticamente pelo método calcular_saldo_atual()
        
        # Vincular com a compra correspondente (se existir)
        if lancamento.compra:
            lancamento.compra.realizado = lancamento.realizado
            lancamento.compra.data_realizada = lancamento.data_realizada
            
            # Atualizar estoque sempre que houver compra vinculada (independente do status)
            sucesso, mensagem = atualizar_estoque_compra(lancamento.compra, usuario.id)
            if sucesso and mensagem:
                flash(mensagem, 'info')
            elif not sucesso:
                flash(mensagem, 'error')
        
        # Vincular com a venda correspondente (se existir)
        if lancamento.venda:
            lancamento.venda.realizado = lancamento.realizado
            lancamento.venda.data_realizada = lancamento.data_realizada
            
            # Atualizar estoque sempre que houver venda vinculada (independente do status)
            sucesso, mensagem = atualizar_estoque_venda(lancamento.venda, usuario.id)
            if sucesso and mensagem:
                flash(mensagem, 'info')
            elif not sucesso:
                flash(mensagem, 'error')
        
        db.session.commit()
        flash('Status do lançamento alterado com sucesso.', 'success')
    else:
        flash('Lançamento não encontrado ou você não tem permissão para alterá-lo.', 'error')
    
    return redirect(url_for('lancamentos'))

@app.route('/lancamentos/<int:lancamento_id>/editar', methods=['GET', 'POST'])
def editar_lancamento(lancamento_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario:
        return redirect(url_for('login'))
    
    # Obter empresa_id correta da sessão (considera acesso_contador)
    empresa_id_correta = obter_empresa_id_sessao(session, usuario)
    if not empresa_id_correta:
        flash('Erro ao obter empresa associada', 'error')
        return redirect(url_for('lancamentos'))
    
    # Buscar todos os usuários da empresa correta
    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id_correta, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]
    
    # Buscar lançamento pela empresa_id correta
    lancamento = Lancamento.query.filter(
        Lancamento.id == lancamento_id,
        Lancamento.empresa_id == empresa_id_correta
    ).first()
    if not lancamento:
        flash('Lançamento não encontrado.', 'danger')
        return redirect(url_for('lancamentos'))
    
    # Buscar categorias para o template
    categorias_entrada = PlanoConta.query.filter(PlanoConta.empresa_id == empresa_id_correta, PlanoConta.tipo == 'entrada', PlanoConta.ativo == True).order_by(PlanoConta.nome).all()
    categorias_saida = PlanoConta.query.filter(PlanoConta.empresa_id == empresa_id_correta, PlanoConta.tipo == 'saida', PlanoConta.ativo == True).order_by(PlanoConta.nome).all()
    
    if request.method == 'POST':
        try:
            # Salvar valores antigos para cálculo de saldo
            valor_antigo = lancamento.valor
            tipo_antigo = lancamento.tipo
            conta_caixa_antiga_id = lancamento.conta_caixa_id
            realizado_antigo = lancamento.realizado
            
            lancamento.descricao = request.form['descricao']
            lancamento.tipo = request.form['tipo']
            
            categoria_input = request.form.get('categoria', '').strip()
            if categoria_input:
                # Tentar converter para ID se for numérico
                if categoria_input.isdigit():
                    plano_id = int(categoria_input)
                    pc = db.session.get(PlanoConta, plano_id)
                    if pc:
                        if pc.natureza == 'sintetica':
                            flash('Não é permitido lançar em contas sintéticas', 'error')
                            return redirect(url_for('editar_lancamento', lancamento_id=lancamento.id))
                        lancamento.plano_conta_id = plano_id
                        lancamento.categoria = pc.nome
                else:
                    # Caso venha como nome, buscar o ID
                    pc = PlanoConta.query.filter(PlanoConta.empresa_id == empresa_id_correta, PlanoConta.nome == categoria_input, PlanoConta.tipo == lancamento.tipo).first()
                    if pc:
                        if pc.natureza == 'sintetica':
                            flash('Não é permitido lançar em contas sintéticas', 'error')
                            return redirect(url_for('editar_lancamento', lancamento_id=lancamento.id))
                        lancamento.plano_conta_id = pc.id
                    lancamento.categoria = categoria_input
            
            # Processar valor: pode vir do carrinho ou valor direto
            usar_carrinho = request.form.get('usar_carrinho') == '1'
            item_nomes = request.form.getlist('item_nome[]')
            item_tipos = request.form.getlist('item_tipo[]')
            item_precos = request.form.getlist('item_preco[]')
            item_qtds = request.form.getlist('item_qtd[]')
            item_totais = request.form.getlist('item_total[]')
            item_descontos = request.form.getlist('item_desconto[]')
            
            valor_total_calculado = None
            
            # Se usar carrinho e há itens no carrinho, processar
            if usar_carrinho and item_nomes and len(item_nomes) > 0 and any(nome.strip() for nome in item_nomes):
                itens_validos = []
                for i, nome in enumerate(item_nomes):
                    nome_limpo = nome.strip() if nome else ''
                    if nome_limpo:
                        try:
                            preco = float(item_precos[i].replace(',', '.')) if i < len(item_precos) and item_precos[i] else 0
                        except (ValueError, AttributeError):
                            preco = 0
                        
                        try:
                            qtd = float(item_qtds[i]) if i < len(item_qtds) and item_qtds[i] else 1
                        except (ValueError, AttributeError):
                            qtd = 1
                        
                        try:
                            desc_item = float(item_descontos[i].replace(',', '.')) if i < len(item_descontos) and item_descontos[i] else 0
                        except (ValueError, AttributeError):
                            desc_item = 0
                        
                        try:
                            total_str = item_totais[i] if i < len(item_totais) and item_totais[i] else '0'
                            total = float(total_str.replace(',', '.')) if total_str else 0
                        except (ValueError, AttributeError):
                            total = 0
                        
                        if total == 0 and preco > 0:
                            total = max(0, (preco * qtd) - desc_item)
                        
                        itens_validos.append({
                            'nome': nome_limpo,
                            'tipo': item_tipos[i] if i < len(item_tipos) else 'produto',
                            'preco': preco,
                            'qtd': qtd,
                            'desconto': desc_item,
                            'total': total
                        })
                
                if itens_validos:
                    valor_total_calculado = sum(item['total'] for item in itens_validos)
                    valor_total_calculado = max(0.0, valor_total_calculado)

                    # Salvar itens do carrinho em JSON para recuperação posterior
                    import json
                    lancamento.itens_carrinho = json.dumps(itens_validos, ensure_ascii=False)

                    # Se houver múltiplos itens, concatenar os nomes
                    if len(itens_validos) > 1:
                        nomes_itens = [item['nome'] for item in itens_validos]
                        lancamento.produto_servico = ', '.join(nomes_itens)
                    else:
                        lancamento.produto_servico = itens_validos[0]['nome']

                    # Usar primeiro item para tipo
                    primeiro_item = itens_validos[0]
                    lancamento.tipo_produto_servico = primeiro_item['tipo']

                    # ATUALIZAR VENDA/COMPRA VINCULADA
                    if lancamento.venda and len(itens_validos) > 0:
                        # Atualizar venda vinculada com os novos dados do carrinho
                        venda = lancamento.venda
                        venda.produto = itens_validos[0]['nome'] if len(itens_validos) == 1 else ', '.join([item['nome'] for item in itens_validos])
                        venda.quantidade = sum(item['qtd'] for item in itens_validos)
                        # Calcular valor antes e depois do desconto
                        total_sem_desconto = sum((item['preco'] * item['qtd']) for item in itens_validos)
                        total_descontos = sum(item.get('desconto', 0) for item in itens_validos)
                        venda.valor = total_sem_desconto  # Total ANTES do desconto
                        venda.valor_final = valor_total_calculado  # Total DEPOIS do desconto
                        venda.desconto = total_descontos
                        venda.tipo_venda = itens_validos[0]['tipo']
                        app.logger.info(f"✅ Venda ID {venda.id} atualizada: {venda.quantidade}x {venda.produto} = R$ {venda.valor_final:.2f} (desconto: R$ {venda.desconto:.2f})")

                    elif lancamento.compra and len(itens_validos) > 0:
                        # Atualizar compra vinculada com os novos dados do carrinho
                        compra = lancamento.compra
                        compra.produto = itens_validos[0]['nome'] if len(itens_validos) == 1 else ', '.join([item['nome'] for item in itens_validos])
                        compra.quantidade = sum(item['qtd'] for item in itens_validos)
                        compra.preco_custo = itens_validos[0]['preco'] if len(itens_validos) == 1 else (valor_total_calculado / compra.quantidade if compra.quantidade > 0 else 0)
                        compra.valor = valor_total_calculado
                        compra.tipo_compra = itens_validos[0]['tipo']
                        app.logger.info(f"✅ Compra ID {compra.id} atualizada: {compra.quantidade}x {compra.produto} = R$ {compra.valor:.2f}")
            else:
                # Usar valor direto
                valor_direto_str = request.form.get('valor_direto', '').strip()
                if valor_direto_str:
                    try:
                        valor_total_calculado = float(valor_direto_str.replace(',', '.'))
                    except ValueError:
                        valor_total_calculado = None

                # Se não tem valor direto, tentar campo valor antigo
                if valor_total_calculado is None:
                    valor_str = request.form.get('valor', '').strip()
                    if valor_str:
                        try:
                            valor_total_calculado = float(valor_str.replace(',', '.'))
                        except ValueError:
                            valor_total_calculado = lancamento.valor
                    else:
                        valor_total_calculado = lancamento.valor

                # Limpar produto_servico e itens_carrinho se não usar carrinho
                lancamento.produto_servico = None
                lancamento.tipo_produto_servico = None
                lancamento.itens_carrinho = None
            
            if valor_total_calculado is not None:
                lancamento.valor = valor_total_calculado
            else:
                # Fallback: usar valor do formulário antigo
                lancamento.valor = float(request.form.get('valor', lancamento.valor))
            
            # Parsing da data prevista - suporte a DD/MM/AAAA e AAAA-MM-DD
            data_prevista_str = request.form['data_prevista']
            try:
                # Tentar formato DD/MM/AAAA primeiro
                try:
                    lancamento.data_prevista = datetime.strptime(data_prevista_str, '%d/%m/%Y').date()
                except ValueError:
                    # Fallback para formato AAAA-MM-DD
                    lancamento.data_prevista = datetime.strptime(data_prevista_str, '%Y-%m-%d').date()
            except ValueError:
                flash('Data prevista inválida', 'error')
                return redirect(url_for('editar_lancamento', lancamento_id=lancamento_id))
            
            lancamento.conta_caixa_id = request.form.get('conta_caixa_id') if request.form.get('conta_caixa_id') else None
            
            # Processar novos campos adicionais
            lancamento.observacoes = request.form.get('observacoes', '').strip() or None
            lancamento.nota_fiscal = request.form.get('nota_fiscal', '').strip() or None
            
            # Cliente/Fornecedor
            cliente_id = request.form.get('cliente_id', '').strip()
            fornecedor_id = request.form.get('fornecedor_id', '').strip()
            lancamento.cliente_id = int(cliente_id) if cliente_id else None
            lancamento.fornecedor_id = int(fornecedor_id) if fornecedor_id else None
            
            # Processar data realizada
            data_realizada_str = request.form.get('data_realizada', '').strip()
            if data_realizada_str:
                try:
                    # Tentar formato DD/MM/AAAA primeiro
                    try:
                        lancamento.data_realizada = datetime.strptime(data_realizada_str, '%d/%m/%Y').date()
                    except ValueError:
                        # Fallback para formato AAAA-MM-DD
                        lancamento.data_realizada = datetime.strptime(data_realizada_str, '%Y-%m-%d').date()
                    
                    # Regra: se data_realizada hoje ou passada -> realizado; futura -> agendado (pendente)
                    from datetime import date
                    if lancamento.data_realizada and lancamento.data_realizada <= date.today():
                        lancamento.realizado = True
                    else:
                        lancamento.realizado = False
                        
                except ValueError:
                    flash('Data realizada inválida', 'error')
                    return redirect(url_for('editar_lancamento', lancamento_id=lancamento_id))
            else:
                # Se não há data realizada, limpar data_realizada e ajustar status
                lancamento.data_realizada = None
                lancamento.realizado = False  # Resetar status baseado na nova lógica
                
                # Se este lançamento está vinculado a uma compra ou venda, atualizar o status deles também
                if lancamento.compra:
                    lancamento.compra.realizado = False
                    lancamento.compra.data_realizada = None
                if lancamento.venda:
                    lancamento.venda.realizado = False
                    lancamento.venda.data_realizada = None
            
            # Se o lançamento estava realizado e a data prevista mudou, resetar data_realizada
            data_prevista_anterior = lancamento.data_prevista
            if lancamento.realizado and lancamento.data_prevista != data_prevista_anterior:
                lancamento.data_realizada = None
                lancamento.realizado = False
                
                # Se este lançamento está vinculado a uma compra ou venda, atualizar o status deles também
                if lancamento.compra:
                    lancamento.compra.realizado = False
                    lancamento.compra.data_realizada = None
                if lancamento.venda:
                    lancamento.venda.realizado = False
                    lancamento.venda.data_realizada = None
            
            # REMOVIDO: Atualização manual de saldo - agora é calculado dinamicamente
            # O saldo_atual é calculado automaticamente pelo método calcular_saldo_atual()

            # Registrar quem editou o lançamento e quando
            lancamento.usuario_ultima_edicao_id = usuario.id
            lancamento.data_ultima_edicao = datetime.utcnow()

            # SINCRONIZAÇÃO REVERSA: Atualizar venda/compra vinculada com dados do lançamento
            if lancamento.venda_id:
                venda = db.session.get(Venda, lancamento.venda_id)
                if venda and lancamento.itens_carrinho:
                    try:
                        import json
                        itens = json.loads(lancamento.itens_carrinho)

                        # Calcular totais corretamente
                        # total_sem_desconto = soma(preco * qtd)
                        # total_com_desconto = soma(total) onde total já tem desconto aplicado
                        total_sem_desconto = sum((item.get('preco', 0) * item.get('qtd', 0)) for item in itens)
                        total_com_desconto = sum(item.get('total', 0) for item in itens)
                        total_descontos = total_sem_desconto - total_com_desconto

                        # Atualizar venda com valores do carrinho
                        venda.valor = total_sem_desconto  # Total ANTES do desconto
                        venda.valor_final = total_com_desconto  # Total COM desconto aplicado
                        venda.desconto = total_descontos  # Total de descontos
                        venda.quantidade = sum(item.get('qtd', 0) for item in itens)

                        app.logger.info(f"✅ Venda {venda.id} sincronizada: valor={venda.valor:.2f}, valor_final={venda.valor_final:.2f}, desconto={venda.desconto:.2f}")
                    except Exception as e:
                        app.logger.warning(f"⚠️ Erro ao sincronizar venda {lancamento.venda_id}: {e}")

            elif lancamento.compra_id:
                compra = db.session.get(Compra, lancamento.compra_id)
                if compra and lancamento.itens_carrinho:
                    try:
                        import json
                        itens = json.loads(lancamento.itens_carrinho)

                        # Calcular totais corretamente (total já tem desconto aplicado)
                        total_com_desconto = sum(item.get('total', 0) for item in itens)

                        # Atualizar compra com valores do carrinho
                        compra.valor = total_com_desconto  # Compra não tem campo desconto separado
                        compra.quantidade = sum(item.get('qtd', 0) for item in itens)
                        compra.preco_custo = itens[0].get('preco', 0) if len(itens) == 1 else (total_com_desconto / compra.quantidade if compra.quantidade > 0 else 0)

                        app.logger.info(f"✅ Compra {compra.id} sincronizada: valor={compra.valor:.2f}")
                    except Exception as e:
                        app.logger.warning(f"⚠️ Erro ao sincronizar compra {lancamento.compra_id}: {e}")

            db.session.commit()

            # RECALCULAR ESTOQUE se o lançamento já está realizado
            if lancamento.realizado:
                if lancamento.venda:
                    # Recalcular estoque da venda
                    estoque_ok, msg = atualizar_estoque_venda(lancamento.venda, usuario.id)
                    if not estoque_ok:
                        flash(msg, 'warning')
                    else:
                        flash(f'Lançamento atualizado com sucesso! {msg}', 'success')
                        return redirect(url_for('lancamentos'))
                elif lancamento.compra:
                    # Recalcular estoque da compra
                    estoque_ok, msg = atualizar_estoque_compra(lancamento.compra, usuario.id)
                    if not estoque_ok:
                        flash(msg, 'warning')
                    else:
                        flash(f'Lançamento atualizado com sucesso! {msg}', 'success')
                        return redirect(url_for('lancamentos'))

            flash('Lançamento atualizado com sucesso!', 'success')
            return redirect(url_for('lancamentos'))
            
        except ValueError:
            flash('Erro ao processar os dados. Verifique os valores informados.', 'danger')
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar lançamento: {str(e)}', 'danger')
    
    # Buscar contas caixa ativas de todos os usuários da empresa
    contas_caixa = ContaCaixa.query.filter(ContaCaixa.usuario_id.in_(usuarios_ids), ContaCaixa.ativo==True).order_by(ContaCaixa.nome).all()
    
    # Buscar clientes e fornecedores do usuário
    try:
        clientes = Cliente.query.filter_by(usuario_id=usuario.id).order_by(Cliente.nome).all()
        fornecedores = Fornecedor.query.filter_by(usuario_id=usuario.id).order_by(Fornecedor.nome).all()
    except Exception as e:
        app.logger.error(f"Erro ao buscar clientes/fornecedores: {str(e)}")
        clientes = []
        fornecedores = []
    
    # Data atual para verificar status "Agendado"
    hoje = datetime.now().date()

    # Carregar itens do carrinho
    itens_carrinho = []

    # PRIORIDADE 1: Se tem itens_carrinho em JSON, usar esses
    if lancamento.itens_carrinho:
        try:
            import json
            itens_carrinho = json.loads(lancamento.itens_carrinho)
            app.logger.info(f"✅ Carregados {len(itens_carrinho)} itens do carrinho do JSON")
        except (json.JSONDecodeError, ValueError) as e:
            app.logger.error(f"❌ Erro ao fazer parse do JSON do carrinho: {e}")
            itens_carrinho = []

    # PRIORIDADE 2: Se não tem JSON mas tem venda/compra vinculada, carregar desses
    if not itens_carrinho:
        if lancamento.compra:
            compra = lancamento.compra
            itens_carrinho.append({
                'tipo': 'produto',
                'nome': compra.produto,
                'preco': compra.preco_custo,
                'qtd': compra.quantidade,
                'desconto': 0,
                'total': compra.valor
            })
            app.logger.info(f"✅ Carregado 1 item da compra vinculada")
        elif lancamento.venda:
            venda = lancamento.venda
            if hasattr(venda, 'produto') and venda.produto:
                itens_carrinho.append({
                    'tipo': 'produto',
                    'nome': venda.produto,
                    'preco': venda.preco_venda if hasattr(venda, 'preco_venda') else venda.valor,
                    'qtd': venda.quantidade if hasattr(venda, 'quantidade') else 1,
                    'desconto': 0,
                    'total': venda.valor
                })
                app.logger.info(f"✅ Carregado 1 item da venda vinculada")

    return render_template('editar_lancamento.html', usuario=usuario, lancamento=lancamento, contas_caixa=contas_caixa, hoje=hoje, clientes=clientes, fornecedores=fornecedores, itens_carrinho=itens_carrinho, categorias_entrada=categorias_entrada, categorias_saida=categorias_saida)

@app.route('/lancamentos/<int:lancamento_id>/deletar', methods=['GET', 'POST'])
def deletar_lancamento(lancamento_id):
    print(f"🎯 Iniciando deletar_lancamento para ID: {lancamento_id}")
    
    if 'usuario_id' not in session:
        print("Usuário não autenticado")
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        print("Usuário é admin, redirecionando")
        return redirect(url_for('admin_dashboard'))
    
    print(f"✅ Usuário autenticado: {usuario.id} (empresa: {usuario.empresa_id})")
    
    # Obter empresa_id correta da sessão (considera acesso_contador)
    empresa_id_correta = obter_empresa_id_sessao(session, usuario)
    if not empresa_id_correta:
        print(f"❌ Erro ao obter empresa_id correta")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': 'Erro ao obter empresa associada'})
        flash('Erro ao obter empresa associada', 'error')
        return redirect(url_for('lancamentos'))
    
    # Buscar todos os usuários da empresa correta
    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id_correta, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]
    
    # Buscar lançamento pela empresa_id correta
    lancamento = Lancamento.query.filter(
        Lancamento.id == lancamento_id,
        Lancamento.empresa_id == empresa_id_correta
    ).first()
    if not lancamento:
        print(f"❌ Lançamento {lancamento_id} não encontrado")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': 'Lançamento não encontrado.'})
        flash('Lançamento não encontrado.', 'danger')
        return redirect(url_for('lancamentos'))
    
    print(f"✅ Lançamento encontrado: {lancamento.id} (usuário: {lancamento.usuario_id})")
    
    try:
        # PRIMEIRO: Verificar relacionamentos diretos (mais simples e direto)
        if lancamento.compra_id:
            mensagem = f'Este lançamento está vinculado a uma compra (ID: {lancamento.compra_id}). Para excluí-lo, você deve primeiro excluir a compra relacionada.'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': mensagem})
            flash(mensagem, 'warning')
            return redirect(url_for('lancamentos'))
        
        if lancamento.venda_id:
            mensagem = f'Este lançamento está vinculado a uma venda (ID: {lancamento.venda_id}). Para excluí-lo, você deve primeiro excluir a venda relacionada.'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': mensagem})
            flash(mensagem, 'warning')
            return redirect(url_for('lancamentos'))
        
        # SEGUNDO: Verificar vínculos na tabela Vinculo (mais complexos)
        vinculos = Vinculo.query.filter(
            db.or_(
                db.and_(Vinculo.lado_a_tipo == 'lancamento', Vinculo.lado_a_id == lancamento_id),
                db.and_(Vinculo.lado_b_tipo == 'lancamento', Vinculo.lado_b_id == lancamento_id)
            ),
            Vinculo.usuario_id.in_(usuarios_ids)
        ).all()
        
        # Verificar se há vínculos com vendas ou compras
        venda_vinculada = None
        compra_vinculada = None
        
        for vinculo in vinculos:
            if vinculo.lado_a_tipo == 'venda' and vinculo.lado_b_tipo == 'lancamento':
                venda_vinculada = vinculo.lado_a_id
            elif vinculo.lado_b_tipo == 'venda' and vinculo.lado_a_tipo == 'lancamento':
                venda_vinculada = vinculo.lado_b_id
            elif vinculo.lado_a_tipo == 'compra' and vinculo.lado_b_tipo == 'lancamento':
                compra_vinculada = vinculo.lado_a_id
            elif vinculo.lado_b_tipo == 'compra' and vinculo.lado_a_tipo == 'lancamento':
                compra_vinculada = vinculo.lado_b_id
        
        # Se tem vínculos, bloquear exclusão
        if venda_vinculada or compra_vinculada:
            if venda_vinculada and compra_vinculada:
                mensagem = 'Este lançamento está vinculado a uma venda e uma compra. Para excluí-lo, você deve primeiro excluir a venda e a compra relacionadas.'
            elif venda_vinculada:
                mensagem = f'Este lançamento está vinculado a uma venda. Para excluí-lo, você deve primeiro excluir a venda relacionada (ID: {venda_vinculada}).'
            elif compra_vinculada:
                mensagem = f'Este lançamento está vinculado a uma compra. Para excluí-lo, você deve primeiro excluir a compra relacionada (ID: {compra_vinculada}).'
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': mensagem})
            flash(mensagem, 'warning')
            return redirect(url_for('lancamentos'))
        
        # SE CHEGOU AQUI: Lançamento pode ser excluído (não tem vínculos)
        print(f"✅ Lançamento {lancamento_id} pode ser excluído (sem vínculos)")
        
        # Verificar se tem parcelas vinculadas
        if lancamento.parcelas:
            print(f"🔄 Removendo {len(lancamento.parcelas)} parcelas vinculadas")
            # Remover parcelas primeiro
            for parcela in lancamento.parcelas:
                db.session.delete(parcela)
        
        # Deletar o lançamento
        print(f"🗑️ Deletando lançamento {lancamento_id}")
        db.session.delete(lancamento)
        db.session.commit()
        print(f"✅ Lançamento {lancamento_id} excluído com sucesso!")
        
        # Verificar se é requisição AJAX
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            print(f"📡 Retornando resposta AJAX de sucesso")
            return jsonify({'success': True, 'message': 'Lançamento excluído com sucesso!'})
        
        flash('Lançamento excluído com sucesso!', 'success')
        
    except Exception as e:
        print(f"❌ Erro ao excluir lançamento {lancamento_id}: {str(e)}")
        db.session.rollback()
        
        # Verificar se é requisição AJAX
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            print(f"📡 Retornando resposta AJAX de erro")
            return jsonify({'success': False, 'message': f'Erro ao excluir lançamento: {str(e)}'})
        
        flash(f'Erro ao excluir lançamento: {str(e)}', 'danger')
    
    print(f"🔄 Redirecionando para página de lançamentos")
    return redirect(url_for('lancamentos'))

@app.route('/lancamentos/acao-em-lote', methods=['POST'])
def acao_em_lote_lancamentos():
    if 'usuario_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'})
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario:
        return jsonify({'success': False, 'message': 'Usuário não encontrado'})
    
    try:
        data = request.get_json()
        lancamentos_ids = data.get('lancamentos', [])
        acao = data.get('acao')
        
        if not lancamentos_ids:
            return jsonify({'success': False, 'message': 'Nenhum lançamento selecionado'})
        
        # Buscar todos os usuários da mesma empresa
        empresa_id = obter_empresa_id_sessao(session, usuario)

        usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
        usuarios_ids = [u.id for u in usuarios_empresa]
        
        # Buscar lançamentos da empresa
        lancamentos = Lancamento.query.filter(
            Lancamento.id.in_(lancamentos_ids),
            Lancamento.empresa_id == empresa_id
        ).all()
        
        if not lancamentos:
            return jsonify({'success': False, 'message': 'Nenhum lançamento válido encontrado'})
        
        for lancamento in lancamentos:
            if acao == 'realizar':
                lancamento.realizado = True
                if not lancamento.data_realizada:
                    lancamento.data_realizada = datetime.now().date()
                    
                # Vincular com a compra correspondente (se existir)
                if lancamento.compra:
                    lancamento.compra.realizado = True
                    lancamento.compra.data_realizada = lancamento.data_realizada
                    
                    # Atualizar estoque sempre que houver compra vinculada (independente do status)
                    sucesso, mensagem = atualizar_estoque_compra(lancamento.compra, usuario.id)
                    if not sucesso:
                        return jsonify({'success': False, 'message': mensagem})
                
                # Vincular com a venda correspondente (se existir)
                elif lancamento.venda:
                    lancamento.venda.realizado = True
                    lancamento.venda.data_realizada = lancamento.data_realizada
                    
                    # Atualizar estoque sempre que houver venda vinculada (independente do status)
                    sucesso, mensagem = atualizar_estoque_venda(lancamento.venda, usuario.id)
                    if not sucesso:
                        return jsonify({'success': False, 'message': mensagem})
                    
            elif acao == 'pendente':
                lancamento.realizado = False
                lancamento.data_realizada = None
                
                # Atualizar compra/venda vinculada e recalcular estoque (independente do status)
                if lancamento.compra:
                    lancamento.compra.realizado = False
                    lancamento.compra.data_realizada = None
                    sucesso, mensagem = atualizar_estoque_compra(lancamento.compra, usuario.id)
                    if not sucesso:
                        return jsonify({'success': False, 'message': mensagem})
                elif lancamento.venda:
                    lancamento.venda.realizado = False
                    lancamento.venda.data_realizada = None
                    sucesso, mensagem = atualizar_estoque_venda(lancamento.venda, usuario.id)
                    if not sucesso:
                        return jsonify({'success': False, 'message': mensagem})
                    
            elif acao == 'excluir':
                # Verificar se pode ser excluído (relacionamentos diretos)
                if lancamento.compra or lancamento.venda:
                    return jsonify({
                        'success': False, 
                        'message': f'Lançamento "{lancamento.descricao}" não pode ser excluído pois está vinculado a uma transação'
                    })
                
                # Verificar se há vínculos na tabela vinculo
                vinculos = Vinculo.query.filter(
                    db.or_(
                        db.and_(Vinculo.lado_a_tipo == 'lancamento', Vinculo.lado_a_id == lancamento.id),
                        db.and_(Vinculo.lado_b_tipo == 'lancamento', Vinculo.lado_b_id == lancamento.id)
                    ),
                    Vinculo.usuario_id.in_(usuarios_ids)
                ).all()
                
                if vinculos:
                    # Verificar se há vínculos com vendas ou compras
                    tem_vinculo_venda = False
                    tem_vinculo_compra = False
                    venda_id = None
                    compra_id = None
                    
                    for vinculo in vinculos:
                        if vinculo.lado_a_tipo == 'venda' and vinculo.lado_b_tipo == 'lancamento':
                            tem_vinculo_venda = True
                            venda_id = vinculo.lado_a_id
                        elif vinculo.lado_b_tipo == 'venda' and vinculo.lado_a_tipo == 'lancamento':
                            tem_vinculo_venda = True
                            venda_id = vinculo.lado_b_id
                        elif vinculo.lado_a_tipo == 'compra' and vinculo.lado_b_tipo == 'lancamento':
                            tem_vinculo_compra = True
                            compra_id = vinculo.lado_a_id
                        elif vinculo.lado_b_tipo == 'compra' and vinculo.lado_a_tipo == 'lancamento':
                            tem_vinculo_compra = True
                            compra_id = vinculo.lado_b_id
                    
                    # Orientar o usuário a apagar por lá
                    if tem_vinculo_venda and tem_vinculo_compra:
                        return jsonify({
                            'success': False, 
                            'message': f'Lançamento "{lancamento.descricao}" não pode ser excluído pois está vinculado a uma venda e uma compra. Para excluí-lo, você deve primeiro excluir a venda e a compra relacionadas.'
                        })
                    elif tem_vinculo_venda:
                        return jsonify({
                            'success': False, 
                            'message': f'Lançamento "{lancamento.descricao}" não pode ser excluído pois está vinculado a uma venda (ID: {venda_id}). Para excluí-lo, você deve primeiro excluir a venda relacionada.'
                        })
                    elif tem_vinculo_compra:
                        return jsonify({
                            'success': False, 
                            'message': f'Lançamento "{lancamento.descricao}" não pode ser excluído pois está vinculado a uma compra (ID: {compra_id}). Para excluí-lo, você deve primeiro excluir a compra relacionada.'
                        })
                
                db.session.delete(lancamento)
        
        db.session.commit()
        
        mensagem = f'{len(lancamentos)} lançamento(s) {"realizado(s)" if acao == "realizar" else "marcado(s) como pendente(s)" if acao == "pendente" else "excluído(s)"} com sucesso!'
        return jsonify({'success': True, 'message': mensagem})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Erro ao processar ação: {str(e)}'})


# Removida rota duplicada de toggle-status; lógica consolidada em toggle_lancamento_status

@app.route('/editar_lancamento/')
@app.route('/editar_lancamento/<int:lancamento_id>')
def editar_lancamento_redirect(lancamento_id=None):
    """Rota de redirecionamento para compatibilidade com URLs antigas"""
    if lancamento_id:
        return redirect(url_for('editar_lancamento', lancamento_id=lancamento_id))
    else:
        flash('ID do lançamento não especificado.', 'error')
        return redirect(url_for('lancamentos'))

@app.route('/clientes')
def clientes():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    # Buscar todos os usuários da mesma empresa
    empresa_id = obter_empresa_id_sessao(session, usuario)

    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]

    # Carregar clientes da empresa (filtro direto por empresa_id)
    clientes = Cliente.query.filter(Cliente.empresa_id == empresa_id).order_by(Cliente.nome).all()
    
    # Adicionar informação do usuário que criou cada cliente e calcular totais
    for cliente in clientes:
        cliente.usuario_criador = db.session.get(Usuario, cliente.usuario_id)
    
        # Calcular totais de entradas e saidas para cada cliente
        lancamentos = Lancamento.query.filter(
            Lancamento.cliente_id == cliente.id,
            Lancamento.empresa_id == empresa_id
        ).all()
        
        cliente.total_entradas = sum([l.valor for l in lancamentos if l.tipo == 'entrada'])
        cliente.total_saidas = sum([l.valor for l in lancamentos if l.tipo == 'saida'])
    
    return render_template('clientes_moderno.html', usuario=usuario, clientes=clientes)

@app.route('/clientes/novo', methods=['GET', 'POST'])
def novo_cliente():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        email = request.form.get('email', '').strip()
        telefone = request.form.get('telefone', '').strip()
        cpf_cnpj = request.form.get('cpf_cnpj', '').strip()
        endereco = request.form.get('endereco', '').strip()

        # Validação: nome é obrigatório
        if not nome:
            flash('Nome do cliente é obrigatório', 'danger')
            return redirect(url_for('novo_cliente'))

        # Obter empresa_id correto (considerando contexto de contador)
        empresa_id = obter_empresa_id_sessao(session, usuario)

        novo_cliente = Cliente(
            nome=nome,
            email=email,
            telefone=telefone,
            cpf_cnpj=cpf_cnpj,
            endereco=endereco,
            usuario_id=usuario.id,
            empresa_id=empresa_id
        )

        db.session.add(novo_cliente)
        db.session.commit()

        flash('Cliente cadastrado com sucesso!', 'success')
        return redirect(url_for('clientes'))
    
    return render_template('novo_cliente.html', usuario=usuario)

@app.route('/clientes/<int:cliente_id>/editar', methods=['GET', 'POST'])
def editar_cliente(cliente_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))

    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))

    # Obter empresa_id correta e validar acesso
    empresa_id = obter_empresa_id_sessao(session, usuario)

    # Validar acesso ao cliente (filtro direto por empresa_id)
    cliente = Cliente.query.filter(Cliente.id == cliente_id, Cliente.empresa_id == empresa_id).first()
    if not cliente:
        flash('Cliente não encontrado.', 'danger')
        return redirect(url_for('clientes'))
    
    if request.method == 'POST':
        cliente.nome = request.form['nome']
        cliente.email = request.form['email']
        cliente.telefone = request.form['telefone']
        cliente.cpf_cnpj = request.form['cpf_cnpj']
        cliente.endereco = request.form['endereco']
        
        db.session.commit()
        flash('Cliente atualizado com sucesso!', 'success')
        return redirect(url_for('clientes'))
    
    return render_template('editar_cliente.html', usuario=usuario, cliente=cliente)

@app.route('/clientes/<int:cliente_id>/deletar', methods=['GET', 'POST', 'DELETE'])
def deletar_cliente(cliente_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    # Buscar todos os usuários da mesma empresa
    empresa_id = obter_empresa_id_sessao(session, usuario)

    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]
    
    cliente = Cliente.query.filter(Cliente.id==cliente_id, Cliente.empresa_id == empresa_id).first()
    if not cliente:
        app.logger.warning(f"❌ Tentativa de excluir cliente inexistente: {cliente_id}")
        flash('Cliente não encontrado.', 'danger')
        return redirect(url_for('clientes'))

    app.logger.info(f"🗑️ Tentando excluir cliente: ID={cliente.id}, nome='{cliente.nome}', empresa_id={empresa_id}")

    try:
        # Verificar todas as dependências antes de excluir
        problemas = []

        # 1. Verificar vendas
        vendas = Venda.query.filter_by(cliente_id=cliente.id).all()
        app.logger.info(f"   Vendas encontradas: {len(vendas)}")
        if vendas:
            problemas.append(f"{len(vendas)} venda(s)")

        # 2. Verificar lançamentos que vieram de vendas (não podem ser desvinculados)
        lancamentos_com_venda = Lancamento.query.filter(
            Lancamento.cliente_id == cliente.id,
            Lancamento.venda_id != None
        ).all()
        app.logger.info(f"   Lançamentos com venda: {len(lancamentos_com_venda)}")
        if lancamentos_com_venda:
            problemas.append(f"{len(lancamentos_com_venda)} lançamento(s) de venda")

        if problemas:
            app.logger.warning(f"   ❌ Não pode excluir - dependências: {problemas}")
            flash(f'Não é possível excluir o cliente "{cliente.nome}" pois existem: {", ".join(problemas)}. Exclua as vendas primeiro.', 'warning')
            return redirect(url_for('clientes'))

        # Desvincular lançamentos manuais (que não vieram de vendas)
        lancamentos_manuais = Lancamento.query.filter(
            Lancamento.cliente_id == cliente.id,
            Lancamento.venda_id == None
        ).all()

        app.logger.info(f"   Lançamentos manuais para desvincular: {len(lancamentos_manuais)}")

        for lanc in lancamentos_manuais:
            lanc.cliente_id = None

        db.session.delete(cliente)
        db.session.commit()

        app.logger.info(f"✅ Cliente {cliente.id} excluído com sucesso!")

        msg = f'Cliente "{cliente.nome}" excluído com sucesso!'
        if lancamentos_manuais:
            msg += f' ({len(lancamentos_manuais)} lançamento(s) manual(is) desvinculado(s))'
        flash(msg, 'success')

    except Exception as e:
        db.session.rollback()
        import traceback
        erro_detalhado = traceback.format_exc()
        app.logger.error(f"Erro ao excluir cliente {cliente_id}: {erro_detalhado}")
        flash(f'Erro ao excluir cliente: {str(e)}', 'danger')
    
    return redirect(url_for('clientes'))

@app.route('/fornecedores')
def fornecedores():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    # Buscar todos os usuários da mesma empresa
    empresa_id = obter_empresa_id_sessao(session, usuario)

    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]

    # Carregar fornecedores da empresa (filtro direto por empresa_id)
    fornecedores = Fornecedor.query.filter(Fornecedor.empresa_id == empresa_id).order_by(Fornecedor.nome).all()
    
    # Adicionar informação do usuário que criou cada fornecedor e calcular totais
    for fornecedor in fornecedores:
        fornecedor.usuario_criador = db.session.get(Usuario, fornecedor.usuario_id)
    
        # Calcular totais de compras e pagamentos para cada fornecedor
        lancamentos = Lancamento.query.filter(
            Lancamento.fornecedor_id == fornecedor.id,
            Lancamento.empresa_id == empresa_id
        ).all()
        
        fornecedor.total_compras = sum([l.valor for l in lancamentos if l.tipo == 'saida'])
        fornecedor.total_pagamentos = sum([l.valor for l in lancamentos if l.tipo == 'entrada'])
    
    return render_template('fornecedores_moderno.html', usuario=usuario, fornecedores=fornecedores)

@app.route('/fornecedores/novo', methods=['GET', 'POST'])
def novo_fornecedor():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        email = request.form.get('email', '').strip()
        telefone = request.form.get('telefone', '').strip()
        cpf_cnpj = request.form.get('cpf_cnpj', '').strip()
        endereco = request.form.get('endereco', '').strip()

        # Validação: nome é obrigatório
        if not nome:
            flash('Nome do fornecedor é obrigatório', 'danger')
            return redirect(url_for('novo_fornecedor'))

        # Obter empresa_id correto (considerando contexto de contador)
        empresa_id = obter_empresa_id_sessao(session, usuario)

        novo_fornecedor = Fornecedor(
            nome=nome,
            email=email,
            telefone=telefone,
            cpf_cnpj=cpf_cnpj,
            endereco=endereco,
            usuario_id=usuario.id,
            empresa_id=empresa_id
        )

        db.session.add(novo_fornecedor)
        db.session.commit()

        flash('Fornecedor cadastrado com sucesso!', 'success')
        return redirect(url_for('fornecedores'))
    
    return render_template('novo_fornecedor.html', usuario=usuario)

@app.route('/fornecedores/<int:fornecedor_id>/editar', methods=['GET', 'POST'])
def editar_fornecedor(fornecedor_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))

    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))

    # Obter empresa_id correta e validar acesso
    empresa_id = obter_empresa_id_sessao(session, usuario)

    # Validar acesso ao fornecedor (filtro direto por empresa_id)
    fornecedor = Fornecedor.query.filter(Fornecedor.id == fornecedor_id, Fornecedor.empresa_id == empresa_id).first()
    if not fornecedor:
        flash('Fornecedor não encontrado.', 'danger')
        return redirect(url_for('fornecedores'))
    
    if request.method == 'POST':
        fornecedor.nome = request.form['nome']
        fornecedor.email = request.form['email']
        fornecedor.telefone = request.form['telefone']
        fornecedor.cpf_cnpj = request.form['cpf_cnpj']
        fornecedor.endereco = request.form['endereco']
        
        db.session.commit()
        flash('Fornecedor atualizado com sucesso!', 'success')
        return redirect(url_for('fornecedores'))
    
    return render_template('editar_fornecedor.html', usuario=usuario, fornecedor=fornecedor)

@app.route('/fornecedores/<int:fornecedor_id>/deletar')
def deletar_fornecedor(fornecedor_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    # Buscar todos os usuários da mesma empresa
    empresa_id = obter_empresa_id_sessao(session, usuario)

    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]
    
    fornecedor = Fornecedor.query.filter(Fornecedor.id==fornecedor_id, Fornecedor.empresa_id == empresa_id).first()
    if not fornecedor:
        flash('Fornecedor não encontrado.', 'danger')
        return redirect(url_for('fornecedores'))
    
    try:
        # Verificar se o fornecedor tem compras associadas
        tem_compras = Compra.query.filter_by(fornecedor_id=fornecedor.id).count() > 0
        if tem_compras:
            flash('Não é possível excluir este fornecedor pois existem compras associadas a ele.', 'danger')
            return redirect(url_for('fornecedores'))
        
        db.session.delete(fornecedor)
        db.session.commit()
        flash('Fornecedor excluído com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir fornecedor: {str(e)}', 'danger')
    
    return redirect(url_for('fornecedores'))

# Alias compatível com URL mencionada
@app.route('/excluir_fornecedor/<int:fornecedor_id>')
def excluir_fornecedor_alias(fornecedor_id):
    return redirect(url_for('deletar_fornecedor', fornecedor_id=fornecedor_id))

def buscar_produtos_empresa(empresa_id):
    """Busca todos os produtos de uma empresa, independente do usuário"""
    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]
    
    produtos = Produto.query.filter(Produto.usuario_id.in_(usuarios_ids)).order_by(Produto.nome).all()
    app.logger.info(f"Produtos encontrados para empresa {empresa_id}: {len(produtos)}")
    try:
        exemplo = [p.nome for p in produtos[:5]]
        app.logger.info(f"Exemplos de produtos: {exemplo}")
    except Exception:
        db.session.rollback()
        pass
    
    # Adicionar informação do usuário que criou cada produto
    for produto in produtos:
        produto.usuario_criador = db.session.get(Usuario, produto.usuario_id)
        # Calcular valor do estoque
        produto.valor_estoque = (produto.estoque or 0) * (produto.preco_venda or 0)
    
    return produtos

@app.route('/estoque')
def estoque():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    # Carregar produtos de toda a empresa
    produtos = buscar_produtos_empresa(usuario.empresa_id)
    app.logger.info(f"/estoque - usuário {usuario.id} (empresa {usuario.empresa_id}) - produtos: {len(produtos)}")
    
    # Buscar categorias para filtros (usando nome do produto como categoria)
    categorias = db.session.query(Produto.nome).distinct().limit(10).all()
    categorias = [cat[0] for cat in categorias]
    
    return render_template('produtos_moderno.html', usuario=usuario, produtos=produtos, categorias=categorias)

@app.route('/api/estoque/<int:produto_id>', methods=['GET'])
def get_estoque_produto(produto_id):
    if 'usuario_id' not in session:
        return jsonify({'error': 'Não autorizado'}), 401
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario:
        return jsonify({'error': 'Usuário não encontrado'}), 401
    
    produto = Produto.query.filter_by(id=produto_id, usuario_id=usuario.id).first()
    if not produto:
        return jsonify({'error': 'Produto não encontrado'}), 404
    
    return jsonify({
        'produto_id': produto.id,
        'nome': produto.nome,
        'estoque': produto.estoque,
        'preco_custo': produto.preco_custo,
        'preco_venda': produto.preco_venda
    })

@app.route('/api/vinculos', methods=['GET'])
def get_vinculos():
    if 'usuario_id' not in session:
        return jsonify({'error': 'Não autorizado'}), 401
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario:
        return jsonify({'error': 'Usuário não encontrado'}), 401
    
    vinculos = Vinculo.query.filter_by(usuario_id=usuario.id).all()
    vinculos_data = []
    
    for vinculo in vinculos:
        vinculo_data = {
            'id': vinculo.id,
            'lado_a_tipo': vinculo.lado_a_tipo,
            'lado_a_id': vinculo.lado_a_id,
            'lado_b_tipo': vinculo.lado_b_tipo,
            'lado_b_id': vinculo.lado_b_id,
            'created_at': vinculo.created_at.isoformat() if vinculo.created_at else None
        }
        
        # Adicionar informações relacionadas
        if vinculo.lado_a_tipo == 'lancamento' and vinculo.lado_a_id:
            lancamento = db.session.get(Lancamento, vinculo.lado_a_id)
            if lancamento:
                vinculo_data['lancamento'] = {
                    'id': lancamento.id,
                    'descricao': lancamento.descricao,
                    'valor': lancamento.valor,
                    'tipo': lancamento.tipo
                }
        
        if vinculo.lado_a_tipo == 'venda' and vinculo.lado_a_id:
            venda = db.session.get(Venda, vinculo.lado_a_id)
            if venda:
                vinculo_data['venda'] = {
                    'id': venda.id,
                    'produto': venda.produto,
                    'cliente_nome': venda.cliente.nome if venda.cliente else None
                }
        
        if vinculo.lado_a_tipo == 'compra' and vinculo.lado_a_id:
            compra = db.session.get(Compra, vinculo.lado_a_id)
            if compra:
                vinculo_data['compra'] = {
                    'id': compra.id,
                    'produto': compra.produto,
                    'fornecedor_nome': compra.fornecedor.nome if compra.fornecedor else None
                }
        
        vinculos_data.append(vinculo_data)
    
    return jsonify(vinculos_data)

@app.route('/api/produtos/sugestoes')
def api_produtos_sugestoes():
    if 'usuario_id' not in session:
        return jsonify({'error': 'Não autorizado'}), 401
    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario:
        return jsonify({'error': 'Usuário não encontrado'}), 401
    q = (request.args.get('q') or '').strip().lower()
    empresa_id = obter_empresa_id_sessao(session, usuario)

    usuarios_ids = [u.id for u in Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()]
    query = Produto.query.filter(Produto.usuario_id.in_(usuarios_ids))
    if q:
        query = query.filter(db.func.lower(Produto.nome).like(f"%{q}%"))
    produtos = query.order_by(Produto.nome).limit(10).all()
    return jsonify({
        'produtos': [
            {
                'id': p.id,
                'nome': p.nome,
                'estoque': p.estoque or 0,
                'preco_venda': p.preco_venda or 0
            } for p in produtos
        ]
    })

@app.route('/estoque/sincronizar')
def sincronizar_estoque():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    # Verificar o tipo da empresa do usuário
    if usuario.empresa.tipo_empresa not in ['comercio', 'industria']:
        flash('Funcionalidade disponível apenas para comércio e indústria.', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        # Primeiro consolidar produtos duplicados
        sucesso_consolidacao, mensagem_consolidacao = consolidar_produtos_duplicados(usuario.id)
        
        # Depois sincronizar estoque
        sucesso, mensagem = sincronizar_estoque_usuario(usuario.id)
        if sucesso:
            flash(f'Estoque sincronizado com sucesso! {mensagem}. {mensagem_consolidacao}', 'success')
        else:
            flash(f'Erro ao sincronizar estoque: {mensagem}', 'error')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao sincronizar estoque: {str(e)}', 'error')
    
    return redirect(url_for('estoque'))

@app.route('/estoque/novo', methods=['GET', 'POST'])
def novo_produto():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    # Verificar o tipo da empresa do usuário
    if usuario.empresa.tipo_empresa not in ['comercio', 'industria']:
        flash('Funcionalidade disponível apenas para comércio e indústria.', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        nome = request.form['nome']
        descricao = request.form['descricao']
        preco_custo = float(request.form['preco_custo'])
        preco_venda = float(request.form['preco_venda'])
        estoque = int(request.form['estoque'])
        
        # Se estoque > 0, ativar automaticamente
        ativo = estoque > 0
        
        novo_produto = Produto(
            nome=nome,
            descricao=descricao,
            preco_custo=preco_custo,
            preco_venda=preco_venda,
            estoque=estoque,
            ativo=ativo,
            usuario_id=usuario.id
        )
        
        db.session.add(novo_produto)
        db.session.commit()
        
        flash('Produto cadastrado com sucesso!', 'success')
        return redirect(url_for('estoque'))
    
    return render_template('novo_produto.html', usuario=usuario)

@app.route('/estoque/servicos')
def servicos():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    # Buscar todos os usuários da mesma empresa
    empresa_id = obter_empresa_id_sessao(session, usuario)

    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]
    
    # Carregar serviços de todos os usuários da empresa
    servicos = Servico.query.filter(Servico.usuario_id.in_(usuarios_ids)).order_by(Servico.nome).all()
    
    # Adicionar informação do usuário que criou cada serviço
    for servico in servicos:
        servico.usuario_criador = db.session.get(Usuario, servico.usuario_id)
    
    return render_template('servicos.html', usuario=usuario, servicos=servicos)

@app.route('/estoque/servicos/novo', methods=['GET', 'POST'])
def novo_servico():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    # Verificar o tipo da empresa do usuário
    if usuario.empresa.tipo_empresa not in ['comercio', 'industria']:
        flash('Funcionalidade disponível apenas para comércio e indústria.', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        nome = request.form['nome']
        descricao = request.form['descricao']
        preco = float(request.form['preco'])
        
        novo_servico = Servico(
            nome=nome,
            descricao=descricao,
            preco=preco,
            usuario_id=usuario.id
        )
        
        db.session.add(novo_servico)
        db.session.commit()
        
        flash('Serviço cadastrado com sucesso!', 'success')
        return redirect(url_for('servicos'))
    
    return render_template('novo_servico.html', usuario=usuario)

@app.route('/estoque/ajuste/<int:produto_id>', methods=['GET', 'POST'])
def ajustar_estoque(produto_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    # Verificar o tipo da empresa do usuário
    if usuario.empresa.tipo_empresa not in ['comercio', 'industria']:
        flash('Funcionalidade disponível apenas para comércio e indústria.', 'error')
        return redirect(url_for('dashboard'))
    
    produto = db.session.get(Produto, produto_id)
    if not produto or produto.usuario_id != usuario.id:
        flash('Produto não encontrado.', 'error')
        return redirect(url_for('estoque'))
    
    if request.method == 'POST':
        tipo_ajuste = request.form['tipo_ajuste']
        quantidade = int(request.form['quantidade'])
        motivo = request.form['motivo']
        
        if tipo_ajuste == 'entrada':
            produto.estoque += quantidade
        elif tipo_ajuste == 'saida':
            if produto.estoque >= quantidade:
                produto.estoque -= quantidade
            else:
                flash('Estoque insuficiente para saída.', 'error')
                return redirect(url_for('ajustar_estoque', produto_id=produto_id))
        
        # Se estoque > 0 após ajuste, ativar automaticamente
        if produto.estoque > 0:
            produto.ativo = True
        
        db.session.commit()
        flash('Estoque ajustado com sucesso!', 'success')
        return redirect(url_for('estoque'))
    
    return render_template('ajustar_estoque.html', usuario=usuario, produto=produto)

@app.route('/estoque/produto/<int:produto_id>/editar', methods=['GET', 'POST'])
def editar_produto(produto_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    # Buscar todos os usuários da mesma empresa
    empresa_id = obter_empresa_id_sessao(session, usuario)

    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]
    
    produto = Produto.query.filter(Produto.id==produto_id, Produto.usuario_id.in_(usuarios_ids)).first()
    if not produto:
        flash('Produto não encontrado.', 'danger')
        return redirect(url_for('estoque'))
    
    if request.method == 'POST':
        produto.nome = request.form['nome']
        produto.descricao = request.form['descricao']
        produto.preco_custo = float(request.form['preco_custo'])
        produto.preco_venda = float(request.form['preco_venda'])
        estoque_novo = int(request.form['estoque'])
        produto.estoque = estoque_novo
        
        # Processar campo ativo
        produto.ativo = 'ativo' in request.form
        
        # Se estoque > 0, ativar automaticamente
        if estoque_novo > 0:
            produto.ativo = True
        
        db.session.commit()
        flash('Produto atualizado com sucesso!', 'success')
        return redirect(url_for('estoque'))
    
    return render_template('editar_produto.html', usuario=usuario, produto=produto)

@app.route('/estoque/produto/<int:produto_id>/deletar')
def deletar_produto(produto_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    # Buscar todos os usuários da mesma empresa
    empresa_id = obter_empresa_id_sessao(session, usuario)

    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]
    
    produto = Produto.query.filter(Produto.id==produto_id, Produto.usuario_id.in_(usuarios_ids)).first()
    if not produto:
        flash('Produto não encontrado.', 'danger')
        return redirect(url_for('estoque'))
    
    try:
        # ATUALIZAÇÃO: Verificar se o produto tem QUALQUER venda associada (independente do status realizado)
        vendas = Venda.query.filter_by(produto=produto.nome).first()
        if vendas:
            flash('Não é possível excluir este produto pois existem vendas associadas a ele. Exclua primeiro as vendas relacionadas.', 'danger')
            return redirect(url_for('estoque'))
        
        # ATUALIZAÇÃO: Verificar se o produto tem QUALQUER compra associada (independente do status realizado)
        compras = Compra.query.filter_by(produto=produto.nome).first()
        if compras:
            flash('Não é possível excluir este produto pois existem compras associadas a ele. Exclua primeiro as compras relacionadas.', 'danger')
            return redirect(url_for('estoque'))
        
        db.session.delete(produto)
        db.session.commit()
        flash('Produto excluído com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir produto: {str(e)}', 'danger')
    
    return redirect(url_for('estoque'))

@app.route('/estoque/servico/<int:servico_id>/editar', methods=['GET', 'POST'])
def editar_servico(servico_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    # Buscar todos os usuários da mesma empresa
    empresa_id = obter_empresa_id_sessao(session, usuario)

    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]
    
    servico = Servico.query.filter(Servico.id==servico_id, Servico.usuario_id.in_(usuarios_ids)).first()
    if not servico:
        flash('Serviço não encontrado.', 'danger')
        return redirect(url_for('servicos'))
    
    if request.method == 'POST':
        servico.nome = request.form['nome']
        servico.descricao = request.form['descricao']
        servico.preco = float(request.form['preco'])
        servico.ativo = 'ativo' in request.form
        
        db.session.commit()
        flash('Serviço atualizado com sucesso!', 'success')
        return redirect(url_for('servicos'))
    
    return render_template('editar_servico.html', usuario=usuario, servico=servico)

@app.route('/estoque/servico/<int:servico_id>/deletar')
def deletar_servico(servico_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    # Buscar todos os usuários da mesma empresa
    empresa_id = obter_empresa_id_sessao(session, usuario)

    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]
    
    servico = Servico.query.filter(Servico.id==servico_id, Servico.usuario_id.in_(usuarios_ids)).first()
    if not servico:
        flash('Serviço não encontrado.', 'danger')
        return redirect(url_for('servicos'))
    
    try:
        # Verificar se o serviço tem vendas associadas
        vendas = Venda.query.filter_by(produto=servico.nome).first()
        if vendas:
            flash('Não é possível excluir este serviço pois existem vendas associadas a ele.', 'danger')
            return redirect(url_for('servicos'))
        
        db.session.delete(servico)
        db.session.commit()
        flash('Serviço excluído com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir serviço: {str(e)}', 'danger')
    
    return redirect(url_for('servicos'))

# API Routes
@app.route('/api/produto/<int:produto_id>')
def api_produto(produto_id):
    if 'usuario_id' not in session:
        return jsonify({'error': 'Não autorizado'}), 401
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    produto = db.session.get(Produto, produto_id)
    
    if produto and produto.usuario_id == usuario.id:
        return jsonify({
            'id': produto.id,
            'nome': produto.nome,
            'descricao': produto.descricao,
            'preco_custo': produto.preco_custo,
            'preco_venda': produto.preco_venda,
            'estoque': produto.estoque
        })
    else:
        return jsonify({'error': 'Produto não encontrado'}), 404

# Rotas para Plano de Contas
@app.route('/plano-contas')
def plano_contas():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))

    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))

    # Obter empresa_id correta (considerando acesso de contador)
    empresa_id = obter_empresa_id_sessao(session, usuario)

    # Buscar todos os usuários da mesma empresa
    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]

    # Carregar plano de contas da empresa (filtro correto para acesso contador)
    contas = PlanoConta.query.filter(
        PlanoConta.empresa_id == empresa_id,
        PlanoConta.ativo == True
    ).order_by(PlanoConta.tipo, PlanoConta.codigo, PlanoConta.nome).all()

    app.logger.info(f"📊 Plano de contas - empresa_id={empresa_id}, usuario_id={usuario.id}, contas encontradas: {len(contas)}")
    if contas:
        app.logger.info(f"   Todas as contas: {[(c.id, c.nome, c.codigo, c.pai_id, c.tipo) for c in contas]}")

    # Calcular saldo de cada conta (somando lançamentos das analíticas)
    for conta in contas:
        # Apenas contas analíticas têm lançamentos diretos
        if conta.natureza == 'analitica':
            lancamentos = Lancamento.query.filter(
                Lancamento.empresa_id == empresa_id,
                Lancamento.categoria == conta.nome,
                Lancamento.realizado == True
            ).all()

            if conta.tipo == 'entrada':
                conta.saldo = sum([l.valor for l in lancamentos if l.tipo == 'entrada'])
            else:
                conta.saldo = sum([l.valor for l in lancamentos if l.tipo == 'saida'])
        else:
            conta.saldo = 0.0

        conta.usuario_criador = db.session.get(Usuario, conta.usuario_id)

    # Função auxiliar para organizar hierarquicamente
    def organizar_hierarquia(lista_contas):
        """Organiza contas em estrutura hierárquica"""
        # Dicionário para armazenar contas por ID
        contas_dict = {c.id: c for c in lista_contas}

        # Criar atributo filhos_lista para cada conta
        for conta in lista_contas:
            conta.filhos_lista = []

        # Organizar hierarquia
        raizes = []
        for conta in lista_contas:
            if conta.pai_id and conta.pai_id in contas_dict:
                contas_dict[conta.pai_id].filhos_lista.append(conta)
            else:
                raizes.append(conta)

        # Calcular saldos das sintéticas (soma dos filhos)
        def calcular_saldo_sintetica(conta):
            if conta.natureza == 'sintetica' and conta.filhos_lista:
                conta.saldo = sum([calcular_saldo_sintetica(filho) for filho in conta.filhos_lista])
            return conta.saldo or 0.0

        for raiz in raizes:
            calcular_saldo_sintetica(raiz)

        return raizes

    # Separar e organizar contas por tipo
    contas_receita = organizar_hierarquia([c for c in contas if c.tipo == 'entrada'])
    contas_despesa = organizar_hierarquia([c for c in contas if c.tipo == 'saida'])

    # Calcular totais
    total_entradas = sum([c.saldo for c in contas if c.tipo == 'entrada'])
    total_saidas = sum([c.saldo for c in contas if c.tipo == 'saida'])
    saldo_geral = total_entradas - total_saidas

    return render_template('plano_contas.html',
                          usuario=usuario,
                          contas_entrada=contas_receita,  # Template usa contas_entrada
                          contas_saida=contas_despesa,    # Template usa contas_saida
                          todas_contas=contas,  # Todas as contas sem hierarquia para debug
                          total_entradas=total_entradas,
                          total_saidas=total_saidas,
                          saldo_geral=saldo_geral)

@app.route('/plano-contas/nova', methods=['GET', 'POST'])
def nova_conta():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    # Obter empresa_id correta (considerando acesso de contador)
    empresa_id = obter_empresa_id_sessao(session, usuario)
    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]

    if request.method == 'POST':
        nome = request.form['nome']
        tipo = request.form['tipo']
        descricao = request.form['descricao']
        natureza = request.form.get('natureza', 'analitica')
        codigo = request.form.get('codigo', '')
        pai_id = request.form.get('pai_id')
        
        pai_id = int(pai_id) if pai_id and pai_id != "" else None

        # VALIDAÇÃO: Conta analítica DEVE ter uma conta sintética como pai
        if natureza == 'analitica':
            if not pai_id:
                flash('❌ Conta analítica deve estar vinculada a uma conta sintética (pai)!', 'error')
                planos_sinteticos = PlanoConta.query.filter(
                    PlanoConta.empresa_id == empresa_id,
                    PlanoConta.natureza == 'sintetica',
                    PlanoConta.ativo == True
                ).order_by(PlanoConta.tipo, PlanoConta.codigo).all()
                return render_template('nova_conta.html', usuario=usuario, planos_sinteticos=planos_sinteticos)

            # Verificar se o pai é sintético
            pai = db.session.get(PlanoConta, pai_id)
            if not pai or pai.natureza != 'sintetica':
                flash('❌ Conta analítica só pode ser filha de uma conta sintética!', 'error')
                planos_sinteticos = PlanoConta.query.filter(
                    PlanoConta.empresa_id == empresa_id,
                    PlanoConta.natureza == 'sintetica',
                    PlanoConta.ativo == True
                ).order_by(PlanoConta.tipo, PlanoConta.codigo).all()
                return render_template('nova_conta.html', usuario=usuario, planos_sinteticos=planos_sinteticos)

        nivel = 1
        if pai_id:
            pai = db.session.get(PlanoConta, pai_id)
            if pai:
                nivel = (pai.nivel or 1) + 1
        
        nova_conta = PlanoConta(
            nome=nome,
            tipo=tipo,
            descricao=descricao,
            natureza=natureza,
            codigo=codigo,
            pai_id=pai_id,
            nivel=nivel,
            usuario_id=usuario.id,
            empresa_id=empresa_id,
            ativo=True
        )

        db.session.add(nova_conta)
        db.session.commit()
        db.session.refresh(nova_conta)  # Garantir que o ID foi gerado

        app.logger.info(f"✅ Conta criada: ID={nova_conta.id}, nome='{nome}', codigo='{codigo}', tipo={tipo}, natureza={natureza}, pai_id={pai_id}, nivel={nivel}, empresa_id={empresa_id}, usuario_id={usuario.id}, ativo={nova_conta.ativo}")

        flash(f'Conta "{nome}" criada com sucesso! (ID: {nova_conta.id})', 'success')
        return redirect(url_for('plano_contas'))
    
    planos_sinteticos = PlanoConta.query.filter(
        PlanoConta.empresa_id == empresa_id,
        PlanoConta.natureza == 'sintetica',
        PlanoConta.ativo == True
    ).order_by(PlanoConta.tipo, PlanoConta.codigo).all()
    
    return render_template('nova_conta.html', usuario=usuario, planos_sinteticos=planos_sinteticos)

@app.route('/plano-contas/<int:conta_id>/toggle')
def toggle_conta(conta_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    conta = db.session.get(PlanoConta, conta_id)
    if conta and conta.usuario_id == usuario.id:
        conta.ativo = not conta.ativo
        db.session.commit()
        flash(f'Status da conta {conta.nome} alterado com sucesso.', 'success')
    
    return redirect(url_for('plano_contas'))

@app.route('/plano-contas/<int:conta_id>/deletar', methods=['POST'])
def deletar_conta(conta_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))

    usuario = db.session.get(Usuario, session['usuario_id'])

    conta = db.session.get(PlanoConta, conta_id)
    if not conta or conta.usuario_id != usuario.id:
        flash('Conta não encontrada ou sem permissão para deletar.', 'error')
        return redirect(url_for('plano_contas'))

    lancamentos_vinculados = Lancamento.query.filter(
        db.or_(
            Lancamento.plano_conta_id == conta.id,
            Lancamento.categoria == conta.nome
        ),
        Lancamento.usuario_id == usuario.id
    ).count()

    if lancamentos_vinculados > 0:
        flash(f'Não é possível deletar a conta. Existem {lancamentos_vinculados} lançamento(s) vinculado(s) a ela.', 'error')
        return redirect(url_for('plano_contas'))

    try:
        nome_conta = conta.nome
        db.session.delete(conta)
        db.session.commit()
        flash(f'Conta {nome_conta} deletada com sucesso.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao deletar conta: {str(e)}', 'error')

    return redirect(url_for('plano_contas'))

@app.route('/plano-contas/<int:conta_id>/editar', methods=['GET', 'POST'])
def editar_conta(conta_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    conta = db.session.get(PlanoConta, conta_id)
    if not (conta and (conta.usuario_id == usuario.id or session.get('acesso_contador'))):
        flash('Conta não encontrada ou sem permissão para editar.', 'error')
        return redirect(url_for('plano_contas'))
    
    # Obter empresa_id correta (considerando acesso de contador)
    empresa_id = obter_empresa_id_sessao(session, usuario)
    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]

    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        descricao = request.form.get('descricao', '').strip()
        natureza = request.form.get('natureza', 'analitica')
        codigo = request.form.get('codigo', '')
        pai_id = request.form.get('pai_id')
        
        if not nome:
            flash('Nome da conta é obrigatório', 'error')
            return redirect(url_for('editar_conta', conta_id=conta_id))
        
        try:
            pai_id = int(pai_id) if pai_id and pai_id != "" else None

            # VALIDAÇÃO: Conta analítica DEVE ter uma conta sintética como pai
            if natureza == 'analitica':
                if not pai_id:
                    flash('❌ Conta analítica deve estar vinculada a uma conta sintética (pai)!', 'error')
                    planos_sinteticos = PlanoConta.query.filter(
                        PlanoConta.empresa_id == empresa_id,
                        PlanoConta.natureza == 'sintetica',
                        PlanoConta.ativo == True,
                        PlanoConta.id != conta.id
                    ).order_by(PlanoConta.tipo, PlanoConta.codigo).all()
                    return render_template('editar_conta.html', usuario=usuario, conta=conta, planos_sinteticos=planos_sinteticos)

                # Verificar se o pai é sintético
                pai = db.session.get(PlanoConta, pai_id)
                if not pai or pai.natureza != 'sintetica':
                    flash('❌ Conta analítica só pode ser filha de uma conta sintética!', 'error')
                    planos_sinteticos = PlanoConta.query.filter(
                        PlanoConta.empresa_id == empresa_id,
                        PlanoConta.natureza == 'sintetica',
                        PlanoConta.ativo == True,
                        PlanoConta.id != conta.id
                    ).order_by(PlanoConta.tipo, PlanoConta.codigo).all()
                    return render_template('editar_conta.html', usuario=usuario, conta=conta, planos_sinteticos=planos_sinteticos)

            nivel = 1
            if pai_id:
                pai = db.session.get(PlanoConta, pai_id)
                if pai:
                    nivel = (pai.nivel or 1) + 1
            
            conta.nome = nome
            conta.descricao = descricao
            conta.natureza = natureza
            conta.codigo = codigo
            conta.pai_id = pai_id
            conta.nivel = nivel
            
            db.session.commit()
            flash('Conta atualizada com sucesso!', 'success')
            return redirect(url_for('plano_contas'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar conta: {str(e)}', 'error')
            return redirect(url_for('editar_conta', conta_id=conta_id))
    
    planos_sinteticos = PlanoConta.query.filter(
        PlanoConta.empresa_id == empresa_id,
        PlanoConta.natureza == 'sintetica',
        PlanoConta.ativo == True,
        PlanoConta.id != conta.id  # Não pode ser pai de si mesmo
    ).order_by(PlanoConta.tipo, PlanoConta.codigo).all()
    
    return render_template('editar_conta.html', usuario=usuario, conta=conta, planos_sinteticos=planos_sinteticos)

# Rotas para Vendas
@app.route('/vendas')
def vendas():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))

    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))

    # Obter empresa_id correta (considerando acesso de contador)
    empresa_id = obter_empresa_id_sessao(session, usuario)

    # Buscar todos os usuários da mesma empresa
    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]
    
    # Aplicar filtros - carregar lançamentos vinculados para mostrar status financeiro correto (filtro direto por empresa_id)
    query = Venda.query.options(
        db.joinedload(Venda.cliente),
        db.joinedload(Venda.usuario)
    ).filter(Venda.empresa_id == empresa_id)
    
    # Filtro por data
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    if data_inicio:
        try:
            # Tentar formato DD/MM/AAAA primeiro, depois YYYY-MM-DD
            if '/' in data_inicio:
                data_inicio_obj = datetime.strptime(data_inicio, '%d/%m/%Y').date()
            else:
                data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            query = query.filter(Venda.data_prevista >= data_inicio_obj)
        except ValueError as e:
            print(f"Erro ao converter data_inicio '{data_inicio}': {e}")
            # Não aplicar filtro se data for inválida
    if data_fim:
        try:
            # Tentar formato DD/MM/AAAA primeiro, depois YYYY-MM-DD
            if '/' in data_fim:
                data_fim_obj = datetime.strptime(data_fim, '%d/%m/%Y').date()
            else:
                data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d').date()
            query = query.filter(Venda.data_prevista <= data_fim_obj)
        except ValueError as e:
            print(f"Erro ao converter data_fim '{data_fim}': {e}")
            # Não aplicar filtro se data for inválida
    
    # Filtro por Nota Fiscal (Suporte a múltiplos NFs separados por ponto e vírgula)
    nf_raw = request.args.get('nota_fiscal', '').strip()
    if nf_raw:
        nfs = [n.strip() for n in nf_raw.split(';') if n.strip()]
        if nfs:
            if len(nfs) == 1:
                query = query.filter(Venda.nota_fiscal.ilike(f'%{nfs[0]}%'))
            else:
                query = query.filter(Venda.nota_fiscal.in_(nfs))

    # Filtro por busca (cliente, produto, descrição ou valor)
    busca = request.args.get('busca')
    if busca and busca.strip():
        busca_clean = busca.strip()
        query = query.join(Cliente, Venda.cliente_id == Cliente.id, isouter=True)
        
        # Verificar se a busca é um número (valor)
        try:
            valor_busca = float(busca_clean.replace(',', '.'))
            query = query.filter(
                db.or_(
                    Cliente.nome.ilike(f'%{busca_clean}%'),
                    Venda.produto.ilike(f'%{busca_clean}%'),
                    Venda.observacoes.ilike(f'%{busca_clean}%'),
                    db.func.abs(Venda.valor - valor_busca) < 0.01  # Busca aproximada para valores
                )
            )
        except ValueError:
            # Se não é um número, buscar apenas texto (busca mais precisa)
            query = query.filter(
                db.or_(
                    Cliente.nome.ilike(f'%{busca_clean}%'),
                    Venda.produto.ilike(f'%{busca_clean}%'),
                    Venda.observacoes.ilike(f'%{busca_clean}%')
                )
            )
    
    # Ordenar e executar query
    vendas = query.order_by(Venda.data_prevista.desc()).all()
    
    # Adicionar informação do usuário que criou cada venda e carregar lançamentos
    for venda in vendas:
        venda.usuario_criador = db.session.get(Usuario, venda.usuario_id)

        # Carregar TODOS os lançamentos financeiros relacionados a esta venda
        # (pode haver múltiplos lançamentos para vendas parceladas)
        venda.lancamentos_financeiros = Lancamento.query.filter_by(venda_id=venda.id).all()

        # Se não encontrou lançamentos, buscar por vínculo através da tabela Vinculo (importação)
        if not venda.lancamentos_financeiros:
            vinculos = Vinculo.query.filter(
                db.or_(
                    db.and_(Vinculo.lado_a_tipo == 'venda', Vinculo.lado_a_id == venda.id, Vinculo.lado_b_tipo == 'lancamento'),
                    db.and_(Vinculo.lado_b_tipo == 'venda', Vinculo.lado_b_id == venda.id, Vinculo.lado_a_tipo == 'lancamento')
                )
            ).all()

            if vinculos:
                venda.lancamentos_financeiros = []
                for vinculo in vinculos:
                    if vinculo.lado_a_tipo == 'lancamento':
                        lanc = db.session.get(Lancamento, vinculo.lado_a_id)
                        if lanc:
                            venda.lancamentos_financeiros.append(lanc)
                    else:
                        lanc = db.session.get(Lancamento, vinculo.lado_b_id)
                        if lanc:
                            venda.lancamentos_financeiros.append(lanc)

        # Calcular status financeiro baseado nos lançamentos
        if venda.lancamentos_financeiros:
            total_lancamentos = len(venda.lancamentos_financeiros)
            lancamentos_realizados = len([l for l in venda.lancamentos_financeiros if l.realizado])
            venda.financeiro_realizados = lancamentos_realizados
            venda.financeiro_total = total_lancamentos
        else:
            # Se não há lançamentos, criar automaticamente
            try:
                lancamento_criado = criar_lancamento_financeiro_automatico(venda, 'venda', venda.usuario_id)
                if lancamento_criado:
                    venda.lancamentos_financeiros = [lancamento_criado]
                    venda.financeiro_realizados = 1 if lancamento_criado.realizado else 0
                    venda.financeiro_total = 1
                    print(f"✅ Lançamento financeiro criado automaticamente para venda {venda.id}")
                else:
                    venda.financeiro_realizados = 0
                    venda.financeiro_total = 1
                    print(f"❌ Falha ao criar lançamento financeiro para venda {venda.id}")
            except Exception as e:
                venda.financeiro_realizados = 0
                venda.financeiro_total = 1
                print(f"❌ Erro ao criar lançamento financeiro para venda {venda.id}: {str(e)}")
    
    return render_template('vendas_moderno.html', usuario=usuario, vendas=vendas)

@app.route('/vendas/nova', methods=['GET', 'POST'])
def nova_venda():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    if request.method == 'POST':
        try:
            cliente_id = request.form.get('cliente_id')
            if not cliente_id:
                flash('Cliente é obrigatório.', 'error')
                return redirect(url_for('nova_venda'))
            
            # NOVO: Processar itens do carrinho
            item_nomes = request.form.getlist('item_nome[]')  # Corrigido: usar item_nome[] em vez de produto
            item_precos = request.form.getlist('item_preco[]')
            item_qtds = request.form.getlist('item_qtd[]')
            item_totais = request.form.getlist('item_total[]')
            item_tipos = request.form.getlist('item_tipo[]')
            item_descontos = request.form.getlist('item_desconto[]')
            
            # Validar se há pelo menos um item no carrinho
            if not item_nomes or len(item_nomes) == 0 or all(not nome.strip() for nome in item_nomes):
                # Tentar formato antigo (campo único produto)
                produto = request.form.get('produto', '').strip()
                if not produto:
                    flash('Adicione pelo menos um item ao carrinho.', 'error')
                    return redirect(url_for('nova_venda'))
                else:
                    # Usar formato antigo
                    item_nomes = [produto]
                    item_precos = [request.form.get('valor', '0')]
                    item_qtds = [request.form.get('quantidade', '1')]
                    item_totais = []
                    item_tipos = ['mercadoria']
                    item_descontos = ['0']
            
            # Filtrar itens vazios e calcular totais
            itens_validos = []
            for i, nome in enumerate(item_nomes):
                nome_limpo = nome.strip() if nome else ''
                if nome_limpo:
                    # Parse de valores
                    try:
                        preco = float(item_precos[i].replace(',', '.')) if i < len(item_precos) and item_precos[i] else 0
                    except (ValueError, AttributeError):
                        preco = 0
                    
                    try:
                        qtd = float(item_qtds[i]) if i < len(item_qtds) and item_qtds[i] else 1
                    except (ValueError, AttributeError):
                        qtd = 1
                    
                    try:
                        desc_item = float(item_descontos[i].replace(',', '.')) if i < len(item_descontos) and item_descontos[i] else 0
                    except (ValueError, AttributeError):
                        desc_item = 0
                    
                    # Calcular total: se o campo total estiver vazio ou zero, calcular (preco * qtd)
                    try:
                        total_str = item_totais[i] if i < len(item_totais) and item_totais[i] else '0'
                        total = float(total_str.replace(',', '.')) if total_str else 0
                    except (ValueError, AttributeError):
                        total = 0
                    
                    # Se total for zero, calcular baseado em preço * quantidade - desconto
                    if total == 0 and preco > 0:
                        total = max(0, (preco * qtd) - desc_item)
                    
                    # Determinar tipo do item (padrão: produto/mercadoria)
                    tipo_item = item_tipos[i] if i < len(item_tipos) and item_tipos[i] else 'produto'
                    # Normalizar tipos
                    if tipo_item in ['produto', 'mercadoria']:
                        tipo_item = 'produto'
                    elif tipo_item in ['servico', 'serviço']:
                        tipo_item = 'servico'
                    
                    itens_validos.append({
                        'nome': nome_limpo,
                        'preco': preco,
                        'qtd': qtd,
                        'total': total,
                        'tipo': tipo_item,
                        'desconto': desc_item
                    })
            
            if not itens_validos:
                flash('Adicione pelo menos um item válido ao carrinho.', 'error')
                return redirect(url_for('nova_venda'))
            
            # Calcular valor total do carrinho usando itens válidos
            try:
                valor_total_carrinho = sum(item['total'] for item in itens_validos)
                if valor_total_carrinho <= 0:
                    flash('Valor total deve ser maior que zero.', 'error')
                    return redirect(url_for('nova_venda'))
            except (ValueError, TypeError) as e:
                app.logger.error(f"Erro ao calcular valor total: {str(e)}")
                flash('Valor inválido no carrinho.', 'error')
                return redirect(url_for('nova_venda'))
            
            # Se houver múltiplos itens, concatenar os nomes
            if len(itens_validos) > 1:
                nomes_itens = [item['nome'] for item in itens_validos]
                produto = ', '.join(nomes_itens)
            else:
                produto = itens_validos[0]['nome']

            # Calcular totais corretos para multi-item
            primeiro_item = itens_validos[0]
            valor = valor_total_carrinho  # Valor total do carrinho (soma de todos os itens)
            quantidade = sum(item['qtd'] for item in itens_validos)  # Quantidade total de todos os itens
            
            # Determinar tipo_venda baseado nos itens do carrinho
            tipos_itens = [item['tipo'] for item in itens_validos]
            tipo_venda = 'produto'  # Padrão
            
            # Se houver itens no carrinho, determinar o tipo baseado nos itens
            if tipos_itens:
                # Se todos os itens são produto, tipo é produto
                if all(tipo == 'produto' for tipo in tipos_itens):
                    tipo_venda = 'produto'
                # Se houver pelo menos um serviço, tipo é serviço
                elif any(tipo == 'servico' for tipo in tipos_itens):
                    tipo_venda = 'servico'
                # Se houver frete
                elif any(tipo == 'frete' for tipo in tipos_itens):
                    tipo_venda = 'frete'
                # Se houver saida
                elif any(tipo == 'saida' for tipo in tipos_itens):
                    tipo_venda = 'saida'
                else:
                    # Fallback: usar o tipo do primeiro item
                    tipo_venda = tipos_itens[0] if tipos_itens else 'produto'
            
            # Se ainda não tiver tipo_venda, tentar buscar do formulário (formato antigo)
            if not tipo_venda:
                tipo_venda = request.form.get('tipo_venda', 'produto')
            
            # Garantir que tipo_venda sempre tenha um valor válido
            if tipo_venda not in ['produto', 'servico', 'frete', 'saida']:
                tipo_venda = 'produto'  # Fallback final
        except Exception as e:
            app.logger.error(f"Erro na validação inicial: {str(e)}")
            flash('Erro na validação dos dados. Tente novamente.', 'error')
            return redirect(url_for('nova_venda'))
        # Validar data prevista
        data_prevista_str = request.form['data_prevista']
        try:
            # Tentar formato DD/MM/AAAA primeiro
            try:
                data_prevista = datetime.strptime(data_prevista_str, '%d/%m/%Y').date()
            except ValueError:
                # Fallback para formato AAAA-MM-DD
                data_prevista = datetime.strptime(data_prevista_str, '%Y-%m-%d').date()
        except ValueError:
            flash('Data prevista inválida', 'error')
            return redirect(url_for('nova_venda'))
        
        # Validar data realizada
        data_realizada = request.form.get('data_realizada')
        if data_realizada and data_realizada.strip():
            try:
                # Tentar formato DD/MM/AAAA primeiro
                try:
                    data_realizada = datetime.strptime(data_realizada, '%d/%m/%Y').date()
                except ValueError:
                    # Fallback para formato AAAA-MM-DD
                    data_realizada = datetime.strptime(data_realizada, '%Y-%m-%d').date()
            except ValueError:
                flash('Data realizada inválida', 'error')
                return redirect(url_for('nova_venda'))
        else:
            data_realizada = None
        observacoes = request.form.get('observacoes', '')
        tipo_pagamento = request.form.get('tipo_pagamento', 'a_vista')
        numero_parcelas = int(request.form.get('numero_parcelas', 1))
        desconto_geral = float(request.form.get('desconto', 0))  # Desconto geral (além dos descontos por item)
        
        # Buscar todos os usuários da mesma empresa
        empresa_id = obter_empresa_id_sessao(session, usuario)

        usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
        usuarios_ids = [u.id for u in usuarios_empresa]
        
        # Verificar se o cliente pertence à empresa
        cliente = Cliente.query.filter(Cliente.id==cliente_id, Cliente.empresa_id == empresa_id).first()
        if not cliente:
            flash('Cliente não encontrado.', 'danger')
            return redirect(url_for('vendas'))
        
        # Calcular valor final após desconto geral (descontos por item já foram aplicados no valor_total_carrinho)
        valor_final = valor_total_carrinho - desconto_geral
        valor_final = max(0.0, valor_final)  # Garantir que não seja negativo
        valor_parcela = valor_final / numero_parcelas if numero_parcelas > 0 else valor_final
        
        # Criar a venda
        # Obter empresa_id correto (considerando contexto de contador)
        empresa_id = obter_empresa_id_sessao(session, usuario)

        nota_fiscal = request.form.get('nota_fiscal', '').strip() or None
        
        nova_venda = Venda(
            cliente_id=cliente_id,
            produto=produto,
            valor=valor,  # Preço unitário do primeiro item (mantido para compatibilidade)
            quantidade=quantidade,  # Quantidade do primeiro item (mantido para compatibilidade)
            tipo_venda=tipo_venda,
            data_prevista=data_prevista,
            data_realizada=data_realizada,
            observacoes=observacoes,
            usuario_id=usuario.id,
            empresa_id=empresa_id,
            tipo_pagamento=tipo_pagamento,
            numero_parcelas=numero_parcelas,
            valor_parcela=valor_parcela,
            desconto=desconto_geral,  # Desconto geral aplicado
            valor_final=valor_final,
            nota_fiscal=nota_fiscal
        )
        
        db.session.add(nova_venda)
        db.session.commit()  # Commit primeiro para confirmar a venda
        
        # ===== PROCESSAMENTO AUTOMÁTICO =====
        try:
            app.logger.info(f"🚀 INICIANDO PROCESSAMENTO da venda {nova_venda.id}")
            
            # Criar parcelas se for pagamento parcelado
            if nova_venda.tipo_pagamento == 'parcelado' and nova_venda.numero_parcelas > 1:
                sucesso_parcelas, mensagem_parcelas = criar_parcelas_automaticas(nova_venda, 'venda', usuario.id)
                if sucesso_parcelas:
                    app.logger.info(f"✅ Parcelas criadas para venda {nova_venda.id}: {mensagem_parcelas}")
                    flash(f'Venda registrada com sucesso! {mensagem_parcelas}', 'success')
                else:
                    app.logger.error(f"❌ Erro ao criar parcelas para venda {nova_venda.id}: {mensagem_parcelas}")
                    flash(f'Venda criada, mas houve erro no parcelamento: {mensagem_parcelas}', 'warning')
            else:
                # Criar lançamento financeiro único se for à vista
                if not nova_venda.lancamento_financeiro:
                    lancamento_criado = criar_lancamento_financeiro_automatico(nova_venda, 'venda', usuario.id)
                    if lancamento_criado:
                        nova_venda.lancamento_financeiro = lancamento_criado
                        app.logger.info(f"✅ Lançamento financeiro criado para venda {nova_venda.id}")
                
                flash('Venda registrada com sucesso!', 'success')
            
            # Atualizar estoque se for produto
            if nova_venda.tipo_venda == 'produto':
                sucesso_estoque, mensagem_estoque = atualizar_estoque_venda(nova_venda, usuario.id)
                if sucesso_estoque:
                    app.logger.info(f"✅ Estoque atualizado para venda {nova_venda.id}: {mensagem_estoque}")
                else:
                    app.logger.warning(f"⚠️ Aviso no estoque para venda {nova_venda.id}: {mensagem_estoque}")
            
            app.logger.info(f"✅ PROCESSAMENTO CONCLUÍDO para venda {nova_venda.id}")
            
        except Exception as e:
            app.logger.error(f"❌ ERRO no processamento da venda {nova_venda.id}: {str(e)}")
            flash(f'Venda criada, mas houve erro no processamento: {str(e)}', 'warning')
        
        return redirect(url_for('vendas'))
    
    # Buscar clientes de todos os usuários da empresa
    empresa_id = obter_empresa_id_sessao(session, usuario)

    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]
    
    clientes = Cliente.query.filter(Cliente.empresa_id == empresa_id).order_by(Cliente.nome).all()
    produtos = Produto.query.filter(Produto.usuario_id.in_(usuarios_ids)).order_by(Produto.nome).all()
    servicos = Servico.query.filter(Servico.usuario_id.in_(usuarios_ids), Servico.ativo==True).order_by(Servico.nome).all()
    contas_caixa = ContaCaixa.query.filter(ContaCaixa.usuario_id.in_(usuarios_ids), ContaCaixa.ativo==True).order_by(ContaCaixa.nome).all()
    
    return render_template('nova_venda.html', usuario=usuario, clientes=clientes, produtos=produtos, servicos=servicos, contas_caixa=contas_caixa)

@app.route('/vendas/<int:venda_id>/toggle_realizado')
def toggle_venda_realizado(venda_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    venda = db.session.get(Venda, venda_id)
    if venda and venda.usuario_id == usuario.id:
        # Toggle do status da venda
        venda.realizado = not venda.realizado
        if venda.realizado and not venda.data_realizada:
            venda.data_realizada = datetime.now().date()
        elif not venda.realizado:
            venda.data_realizada = None
        
        # Vincular com o lançamento financeiro correspondente
        if venda.lancamento_financeiro:
            venda.lancamento_financeiro.realizado = venda.realizado
            venda.lancamento_financeiro.data_realizada = venda.data_realizada
        else:
            # Criar lançamento financeiro se não existir
            try:
                lancamento_criado = criar_lancamento_financeiro_automatico(venda, 'venda', usuario.id)
                if lancamento_criado:
                    print(f"✅ Lançamento financeiro criado automaticamente para venda {venda.id}")
                else:
                    print(f"❌ Erro ao criar lançamento financeiro para venda {venda.id}")
            except Exception as e:
                db.session.rollback()
                print(f"❌ Erro ao criar lançamento financeiro: {str(e)}")
        
        # ATUALIZAR ESTOQUE baseado no status da venda (sem criar produtos)
        if venda.tipo_venda == 'produto':
            # Buscar produto na empresa (não apenas no usuário)
            if not hasattr(usuario, 'empresa_id') or not usuario.empresa_id:
                app.logger.error(f"Usuário {usuario.id} não tem empresa_id definido")
                flash('Usuário não tem empresa associada.', 'error')
                return redirect(url_for('vendas'))
            
            empresa_id = obter_empresa_id_sessao(session, usuario)

            
            usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
            usuarios_ids = [u.id for u in usuarios_empresa]
            
            produto_estoque = Produto.query.filter(
                Produto.nome == venda.produto,
                Produto.usuario_id.in_(usuarios_ids)
            ).first()
            
            if produto_estoque:
                # Calcular estoque real baseado em todas as compras e vendas (independente do status realizado)
                estoque_real = calcular_estoque_produto(venda.produto, usuario.id)
                
                if estoque_real >= 0:
                    produto_estoque.estoque = estoque_real
                    flash(f'Estoque do produto "{venda.produto}" atualizado. Quantidade vendida: {venda.quantidade}. Restante: {estoque_real}', 'info')
                else:
                    flash(f'Erro: Estoque insuficiente para o produto "{venda.produto}". Estoque atual: {produto_estoque.estoque}, Quantidade solicitada: {venda.quantidade}', 'error')
            else:
                flash(f'Produto "{venda.produto}" não encontrado no estoque. O produto deve ser criado automaticamente ao registrar uma compra.', 'warning')
        elif venda.tipo_venda == 'servico':
            # Para serviços, não há verificação de estoque
            flash(f'Serviço "{venda.produto}" processado com sucesso.', 'info')
        
        db.session.commit()
        flash('Status da venda alterado com sucesso.', 'success')
    
    return redirect(url_for('vendas'))

@app.route('/vendas/<int:venda_id>/deletar')
def deletar_venda(venda_id):
    """Exclui uma venda usando o novo gerenciador de cancelamento"""
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    # Obter empresa_id correta da sessão (considera acesso_contador)
    empresa_id = obter_empresa_id_sessao(session, usuario)
    
    venda = db.session.get(Venda, venda_id)
    if not venda or venda.empresa_id != empresa_id:
        flash('Venda não encontrada ou você não tem permissão para excluí-la.', 'error')
        return redirect(url_for('vendas'))
    
    try:
        app.logger.info(f"🗑️ INICIANDO EXCLUSÃO da venda {venda_id}")
        
        # CORREÇÃO: Buscar e excluir TODOS os lançamentos financeiros vinculados à venda
        # 1. Lançamentos com venda_id direto
        lancamentos_diretos = Lancamento.query.filter_by(venda_id=venda_id).all()
        for lancamento in lancamentos_diretos:
            # Excluir parcelas do lançamento se existirem
            if lancamento.parcelas:
                for parcela in lancamento.parcelas:
                    db.session.delete(parcela)
            db.session.delete(lancamento)
            app.logger.info(f"✅ Lançamento direto {lancamento.id} excluído para venda {venda_id}")
        
        # 2. Lançamentos vinculados através da tabela Vinculo
        vinculos = Vinculo.query.filter(
            db.or_(
                db.and_(Vinculo.lado_a_tipo == 'venda', Vinculo.lado_a_id == venda_id),
                db.and_(Vinculo.lado_b_tipo == 'venda', Vinculo.lado_b_id == venda_id)
            )
        ).all()
        
        for vinculo in vinculos:
            # Buscar lançamento vinculado
            if vinculo.lado_a_tipo == 'lancamento':
                lancamento_vinculado = db.session.get(Lancamento, vinculo.lado_a_id)
            elif vinculo.lado_b_tipo == 'lancamento':
                lancamento_vinculado = db.session.get(Lancamento, vinculo.lado_b_id)
            else:
                lancamento_vinculado = None
            
            # Excluir lançamento vinculado se existir
            if lancamento_vinculado:
                # Excluir parcelas do lançamento se existirem
                if lancamento_vinculado.parcelas:
                    for parcela in lancamento_vinculado.parcelas:
                        db.session.delete(parcela)
                db.session.delete(lancamento_vinculado)
                app.logger.info(f"✅ Lançamento vinculado {lancamento_vinculado.id} excluído para venda {venda_id}")
            
            # Excluir o vínculo
            db.session.delete(vinculo)
        
        app.logger.info(f"✅ Todos os lançamentos financeiros e vínculos excluídos para venda {venda_id}")
        
        # Excluir a venda
        db.session.delete(venda)
        db.session.commit()
        
        app.logger.info(f"✅ VENDA EXCLUÍDA COM SUCESSO: {venda_id}")
        flash('Venda excluída com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"❌ ERRO ao excluir venda {venda_id}: {str(e)}")
        flash(f'Erro ao excluir venda: {str(e)}', 'error')
    
    return redirect(url_for('vendas'))

# Rotas para Compras
@app.route('/compras')
def compras():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))

    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))

    # Obter empresa_id correta (considerando acesso de contador)
    empresa_id = obter_empresa_id_sessao(session, usuario)

    # Buscar todos os usuários da mesma empresa
    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]
    
    # Aplicar filtros (filtro direto por empresa_id)
    query = Compra.query.options(
        db.joinedload(Compra.fornecedor),
        db.joinedload(Compra.usuario)
    ).filter(Compra.empresa_id == empresa_id)
    
    # Filtro por data
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    if data_inicio:
        try:
            # Tentar formato DD/MM/AAAA primeiro, depois YYYY-MM-DD
            if '/' in data_inicio:
                data_inicio_obj = datetime.strptime(data_inicio, '%d/%m/%Y').date()
            else:
                data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            query = query.filter(Compra.data_prevista >= data_inicio_obj)
        except ValueError as e:
            print(f"Erro ao converter data_inicio '{data_inicio}': {e}")
            # Não aplicar filtro se data for inválida
    if data_fim:
        try:
            # Tentar formato DD/MM/AAAA primeiro, depois YYYY-MM-DD
            if '/' in data_fim:
                data_fim_obj = datetime.strptime(data_fim, '%d/%m/%Y').date()
            else:
                data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d').date()
            query = query.filter(Compra.data_prevista <= data_fim_obj)
        except ValueError as e:
            print(f"Erro ao converter data_fim '{data_fim}': {e}")
            # Não aplicar filtro se data for inválida
    
    # Filtro por Nota Fiscal (Suporte a múltiplos NFs separados por ponto e vírgula)
    nf_raw = request.args.get('nota_fiscal', '').strip()
    if nf_raw:
        nfs = [n.strip() for n in nf_raw.split(';') if n.strip()]
        if nfs:
            if len(nfs) == 1:
                query = query.filter(Compra.nota_fiscal.ilike(f'%{nfs[0]}%'))
            else:
                query = query.filter(Compra.nota_fiscal.in_(nfs))

    # Filtro por busca (fornecedor, produto, descrição ou valor)
    busca = request.args.get('busca')
    if busca and busca.strip():
        busca_clean = busca.strip()
        query = query.join(Fornecedor, Compra.fornecedor_id == Fornecedor.id, isouter=True)
        
        # Verificar se a busca é um número (valor)
        try:
            valor_busca = float(busca_clean.replace(',', '.'))
            query = query.filter(
                db.or_(
                    Fornecedor.nome.ilike(f'%{busca_clean}%'),
                    Compra.produto.ilike(f'%{busca_clean}%'),
                    Compra.observacoes.ilike(f'%{busca_clean}%'),
                    db.func.abs(Compra.valor - valor_busca) < 0.01  # Busca aproximada para valores
                )
            )
        except ValueError:
            # Se não é um número, buscar apenas texto (busca mais precisa)
            query = query.filter(
                db.or_(
                    Fornecedor.nome.ilike(f'%{busca_clean}%'),
                    Compra.produto.ilike(f'%{busca_clean}%'),
                    Compra.observacoes.ilike(f'%{busca_clean}%')
                )
            )
    
    # Ordenar e executar query
    compras = query.order_by(Compra.data_prevista.desc()).all()
    
    # Adicionar informação do usuário que criou cada compra e carregar lançamentos
    for compra in compras:
        compra.usuario_criador = db.session.get(Usuario, compra.usuario_id)

        # Carregar TODOS os lançamentos financeiros relacionados a esta compra
        # (pode haver múltiplos lançamentos para compras parceladas)
        compra.lancamentos_financeiros = Lancamento.query.filter_by(compra_id=compra.id).all()

        # Se não encontrou lançamentos, buscar por vínculo através da tabela Vinculo (importação)
        if not compra.lancamentos_financeiros:
            vinculos = Vinculo.query.filter(
                db.or_(
                    db.and_(Vinculo.lado_a_tipo == 'compra', Vinculo.lado_a_id == compra.id, Vinculo.lado_b_tipo == 'lancamento'),
                    db.and_(Vinculo.lado_b_tipo == 'compra', Vinculo.lado_b_id == compra.id, Vinculo.lado_a_tipo == 'lancamento')
                )
            ).all()

            if vinculos:
                compra.lancamentos_financeiros = []
                for vinculo in vinculos:
                    if vinculo.lado_a_tipo == 'lancamento':
                        lanc = db.session.get(Lancamento, vinculo.lado_a_id)
                        if lanc:
                            compra.lancamentos_financeiros.append(lanc)
                    else:
                        lanc = db.session.get(Lancamento, vinculo.lado_b_id)
                        if lanc:
                            compra.lancamentos_financeiros.append(lanc)

        # Calcular status financeiro baseado nos lançamentos
        if compra.lancamentos_financeiros:
            total_lancamentos = len(compra.lancamentos_financeiros)
            lancamentos_realizados = len([l for l in compra.lancamentos_financeiros if l.realizado])
            compra.financeiro_realizados = lancamentos_realizados
            compra.financeiro_total = total_lancamentos
        else:
            # Se não há lançamentos, criar automaticamente
            try:
                lancamento_criado = criar_lancamento_financeiro_automatico(compra, 'compra', compra.usuario_id)
                if lancamento_criado:
                    compra.lancamentos_financeiros = [lancamento_criado]
                    compra.financeiro_realizados = 1 if lancamento_criado.realizado else 0
                    compra.financeiro_total = 1
                    print(f"✅ Lançamento financeiro criado automaticamente para compra {compra.id}")
                else:
                    compra.financeiro_realizados = 0
                    compra.financeiro_total = 1
                    print(f"❌ Falha ao criar lançamento financeiro para compra {compra.id}")
            except Exception as e:
                compra.financeiro_realizados = 0
                compra.financeiro_total = 1
                print(f"❌ Erro ao criar lançamento financeiro para compra {compra.id}: {str(e)}")
                import traceback
                traceback.print_exc()
    
    return render_template('compras_moderno.html', usuario=usuario, compras=compras)

@app.route('/compras/nova', methods=['GET', 'POST'])
def nova_compra():
    # Validação robusta de sessão
    if 'usuario_id' not in session:
        app.logger.warning("Tentativa de acessar nova_compra sem sessão válida")
        flash('Sessão expirada. Faça login novamente.', 'error')
        return redirect(url_for('login'))
    
    try:
        usuario = db.session.get(Usuario, session['usuario_id'])
        if not usuario:
            app.logger.error(f"Usuário {session['usuario_id']} não encontrado na sessão")
            session.clear()
            flash('Usuário não encontrado. Faça login novamente.', 'error')
            return redirect(url_for('login'))
        
        if not usuario.ativo:
            app.logger.warning(f"Usuário {usuario.id} inativo tentou acessar nova_compra")
            flash('Conta desativada. Entre em contato com o administrador.', 'error')
            return redirect(url_for('login'))
        
        if usuario.tipo == 'admin':
            return redirect(url_for('admin_dashboard'))
            
        app.logger.info(f"Usuário {usuario.nome} (ID: {usuario.id}) acessando nova_compra")
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao validar sessão do usuário: {str(e)}")
        session.clear()
        flash('Erro na validação da sessão. Faça login novamente.', 'error')
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        try:
            # Validação robusta dos dados de entrada
            fornecedor_id = request.form.get('fornecedor_id')
            fornecedor_nome = request.form.get('fornecedor_nome', '').strip()

            # Se fornecedor_id não foi preenchido mas há um nome digitado, tentar buscar ou criar
            if not fornecedor_id and fornecedor_nome:
                # Buscar fornecedor existente pelo nome
                fornecedor_existente = Fornecedor.query.filter(
                    Fornecedor.empresa_id == empresa_id,
                    Fornecedor.nome.ilike(f'%{fornecedor_nome}%')
                ).first()

                if fornecedor_existente:
                    fornecedor_id = fornecedor_existente.id
                else:
                    # Criar novo fornecedor automaticamente
                    novo_fornecedor = Fornecedor(
                        empresa_id=empresa_id,
                        nome=fornecedor_nome,
                        usuario_id=usuario.id
                    )
                    db.session.add(novo_fornecedor)
                    db.session.flush()  # Para obter o ID
                    fornecedor_id = novo_fornecedor.id
                    app.logger.info(f"Fornecedor '{fornecedor_nome}' criado automaticamente (ID: {fornecedor_id})")

            # Validar se fornecedor foi selecionado ou criado
            if not fornecedor_id:
                flash('Fornecedor é obrigatório. Digite o nome do fornecedor.', 'error')
                return redirect(url_for('nova_compra'))
            
            # NOVO: Processar itens do carrinho
            item_nomes = request.form.getlist('item_nome[]')  # Corrigido: usar item_nome[] em vez de item_produto[]
            item_precos = request.form.getlist('item_preco[]')
            item_qtds = request.form.getlist('item_qtd[]')
            item_totais = request.form.getlist('item_total[]')
            item_tipos = request.form.getlist('item_tipo[]')
            item_obs = request.form.getlist('item_obs[]')
            
            # Validar se há pelo menos um item no carrinho
            if not item_nomes or len(item_nomes) == 0 or all(not nome.strip() for nome in item_nomes):
                flash('Adicione pelo menos um item ao carrinho.', 'error')
                return redirect(url_for('nova_compra'))
            
            # Filtrar itens vazios e calcular totais
            itens_validos = []
            for i, nome in enumerate(item_nomes):
                nome_limpo = nome.strip() if nome else ''
                if nome_limpo:
                    # Parse de valores
                    try:
                        preco = float(item_precos[i].replace(',', '.')) if i < len(item_precos) and item_precos[i] else 0
                    except (ValueError, AttributeError):
                        preco = 0
                    
                    try:
                        qtd = float(item_qtds[i]) if i < len(item_qtds) and item_qtds[i] else 1
                    except (ValueError, AttributeError):
                        qtd = 1
                    
                    # Calcular total: se o campo total estiver vazio ou zero, calcular (preco * qtd)
                    try:
                        total_str = item_totais[i] if i < len(item_totais) and item_totais[i] else '0'
                        total = float(total_str.replace(',', '.')) if total_str else 0
                    except (ValueError, AttributeError):
                        total = 0
                    
                    # Se total for zero, calcular baseado em preço * quantidade
                    if total == 0 and preco > 0:
                        total = preco * qtd
                    
                    itens_validos.append({
                        'nome': nome_limpo,
                        'preco': preco,
                        'qtd': qtd,
                        'total': total,
                        'tipo': item_tipos[i] if i < len(item_tipos) else 'mercadoria',
                        'obs': item_obs[i] if i < len(item_obs) else ''
                    })
            
            if not itens_validos:
                flash('Adicione pelo menos um item válido ao carrinho.', 'error')
                return redirect(url_for('nova_compra'))
            
            # Usar primeiro item para compatibilidade (mantido para código legado)
            produto = itens_validos[0]['nome']
            
            # Calcular valor total do carrinho usando itens válidos
            try:
                valor = sum(item['total'] for item in itens_validos)
                if valor <= 0:
                    flash('Valor total deve ser maior que zero.', 'error')
                    return redirect(url_for('nova_compra'))
            except (ValueError, TypeError) as e:
                app.logger.error(f"Erro ao calcular valor total: {str(e)}")
                flash('Valor inválido no carrinho.', 'error')
                return redirect(url_for('nova_compra'))
            
            # Usar valores do primeiro item válido para compatibilidade
            primeiro_item = itens_validos[0]
            quantidade = primeiro_item['qtd']
            preco_custo = primeiro_item['preco']
            
            if quantidade <= 0:
                flash('Quantidade deve ser maior que zero.', 'error')
                return redirect(url_for('nova_compra'))
            
            if preco_custo <= 0:
                flash('Preço de custo deve ser maior que zero.', 'error')
                return redirect(url_for('nova_compra'))
            
            # Determinar tipo_compra baseado nos itens do carrinho
            tipos_itens = [item['tipo'] for item in itens_validos]
            tipo_compra = 'mercadoria'  # Padrão
            
            # Se houver itens no carrinho, determinar o tipo baseado nos itens
            if tipos_itens:
                # Se todos os itens são mercadoria, tipo é mercadoria
                # Se houver pelo menos um serviço, tipo é serviço (ou pode ser misto)
                if all(tipo in ['mercadoria', 'produto'] for tipo in tipos_itens):
                    tipo_compra = 'mercadoria'
                elif any(tipo in ['servico', 'serviço'] for tipo in tipos_itens):
                    # Se houver serviço, usar serviço (ou considerar criar um tipo misto)
                    tipo_compra = 'servico'
            else:
                # Fallback: tentar pegar do formulário antigo se ainda existir
                tipo_compra_form = request.form.get('tipo_compra')
                if tipo_compra_form:
                    tipo_compra = tipo_compra_form.lower()
                    if tipo_compra in ['produto', 'produtos', 'mercadoria', 'mercadorias']:
                        tipo_compra = 'mercadoria'
                    elif tipo_compra in ['servico', 'serviço', 'servicos', 'serviços']:
                        tipo_compra = 'servico'
            
            # Parsing da data prevista - suporte a DD/MM/AAAA e AAAA-MM-DD
            data_prevista_str = request.form.get('data_prevista')
            try:
                # Tentar formato DD/MM/AAAA primeiro
                try:
                    data_prevista = datetime.strptime(data_prevista_str, '%d/%m/%Y').date()
                except ValueError:
                    # Fallback para formato AAAA-MM-DD
                    data_prevista = datetime.strptime(data_prevista_str, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                flash('Data prevista inválida.', 'error')
                return redirect(url_for('nova_compra'))
            
            observacoes = request.form.get('observacoes', '').strip()
            tipo_pagamento = request.form.get('tipo_pagamento', 'a_vista')
            
            try:
                numero_parcelas = int(request.form.get('numero_parcelas', 1))
                if numero_parcelas <= 0:
                    flash('Número de parcelas deve ser maior que zero.', 'error')
                    return redirect(url_for('nova_compra'))
            except (ValueError, TypeError):
                flash('Número de parcelas inválido.', 'error')
                return redirect(url_for('nova_compra'))
            
            app.logger.info(f"Validação de dados bem-sucedida para compra: {produto}")
            
        except Exception as e:
            app.logger.error(f"Erro na validação de dados: {str(e)}")
            flash('Erro na validação dos dados. Tente novamente.', 'error')
            return redirect(url_for('nova_compra'))
        
        # Buscar todos os usuários da mesma empresa
        empresa_id = obter_empresa_id_sessao(session, usuario)

        usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
        usuarios_ids = [u.id for u in usuarios_empresa]
        
        # Verificar se o fornecedor pertence à empresa
        fornecedor = Fornecedor.query.filter(Fornecedor.id==fornecedor_id, Fornecedor.empresa_id == empresa_id).first()
        if not fornecedor:
            flash('Fornecedor não encontrado.', 'danger')
            return redirect(url_for('compras'))
        
        # O campo 'valor' já contém o valor total da compra
        valor_total_compra = valor
        
        # Calcular valor da parcela
        valor_parcela = valor_total_compra / numero_parcelas if numero_parcelas > 0 else valor_total_compra
        
        # Criar a compra
        # Obter empresa_id correto (considerando contexto de contador)
        empresa_id = obter_empresa_id_sessao(session, usuario)

        nota_fiscal = request.form.get('nota_fiscal', '').strip() or None
        
        nova_compra = Compra(
            fornecedor_id=fornecedor_id,
            produto=produto,
            valor=valor_total_compra,
            quantidade=quantidade,
            preco_custo=preco_custo,
            tipo_compra=tipo_compra,
            data_prevista=data_prevista,
            observacoes=observacoes,
            usuario_id=usuario.id,
            empresa_id=empresa_id,
            tipo_pagamento=tipo_pagamento,
            numero_parcelas=numero_parcelas,
            valor_parcela=valor_parcela,
            nota_fiscal=nota_fiscal
        )
        
        db.session.add(nova_compra)
        db.session.commit()  # Commit primeiro para confirmar a compra
        
        # PROCESSAMENTO AUTOMÁTICO IMEDIATO - Estoque + Financeiro + Vínculo
        try:
            # Verificar se a compra foi criada com sucesso
            if not nova_compra or not nova_compra.id:
                app.logger.error("Compra não foi criada com sucesso")
                flash('Erro ao criar a compra. Tente novamente.', 'error')
                return redirect(url_for('compras'))
            
            app.logger.info(f"🚀 INICIANDO PROCESSAMENTO V2 da compra {nova_compra.id}")
            
            # Processar compra automaticamente
            try:
                # Criar parcelas se for pagamento parcelado
                if nova_compra.tipo_pagamento == 'parcelado' and nova_compra.numero_parcelas > 1:
                    sucesso_parcelas, mensagem_parcelas = criar_parcelas_automaticas(nova_compra, 'compra', usuario.id)
                    if sucesso_parcelas:
                        app.logger.info(f"✅ Parcelas criadas para compra {nova_compra.id}: {mensagem_parcelas}")
                        flash(f'Compra registrada com sucesso! {mensagem_parcelas}', 'success')
                    else:
                        app.logger.error(f"❌ Erro ao criar parcelas para compra {nova_compra.id}: {mensagem_parcelas}")
                        flash(f'Compra criada, mas houve erro no parcelamento: {mensagem_parcelas}', 'warning')
                else:
                    # Criar lançamento financeiro único se for à vista
                    if not nova_compra.lancamento_financeiro:
                        lancamento_criado = criar_lancamento_financeiro_automatico(nova_compra, 'compra', usuario.id)
                        if lancamento_criado:
                            nova_compra.lancamento_financeiro = lancamento_criado
                            app.logger.info(f"✅ Lançamento financeiro criado para compra {nova_compra.id}")
                    
                    flash('Compra registrada com sucesso!', 'success')
                
                # Atualizar estoque se for mercadoria
                if nova_compra.tipo_compra == 'mercadoria':
                    sucesso_estoque, mensagem_estoque = atualizar_estoque_compra(nova_compra, usuario.id)
                    if sucesso_estoque:
                        app.logger.info(f"✅ Estoque atualizado para compra {nova_compra.id}: {mensagem_estoque}")
                    else:
                        app.logger.warning(f"⚠️ Aviso no estoque para compra {nova_compra.id}: {mensagem_estoque}")
                
                app.logger.info(f"✅ PROCESSAMENTO CONCLUÍDO para compra {nova_compra.id}")
                
            except Exception as e:
                app.logger.error(f"❌ ERRO no processamento da compra {nova_compra.id}: {str(e)}")
                flash(f'Compra criada, mas houve erro no processamento: {str(e)}', 'warning')
            
        except Exception as e:
            app.logger.error(f"❌ EXCEÇÃO no processamento V2 da compra {nova_compra.id}: {str(e)}")
            db.session.rollback()
            flash(f'Compra criada, mas houve erro no processamento automático: {str(e)}', 'warning')
        
        return redirect(url_for('compras'))
    
    # Buscar fornecedores de todos os usuários da empresa
    empresa_id = obter_empresa_id_sessao(session, usuario)

    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]
    
    fornecedores = Fornecedor.query.filter(Fornecedor.empresa_id == empresa_id).order_by(Fornecedor.nome).all()
    produtos = Produto.query.filter(Produto.usuario_id.in_(usuarios_ids)).order_by(Produto.nome).all()
    servicos = Servico.query.filter(Servico.usuario_id.in_(usuarios_ids), Servico.ativo==True).order_by(Servico.nome).all()
    contas_caixa = ContaCaixa.query.filter(ContaCaixa.usuario_id.in_(usuarios_ids), ContaCaixa.ativo==True).order_by(ContaCaixa.nome).all()
    
    return render_template('nova_compra.html', usuario=usuario, fornecedores=fornecedores, produtos=produtos, servicos=servicos, contas_caixa=contas_caixa)

@app.route('/compras/<int:compra_id>/toggle_realizado')
def toggle_compra_realizado(compra_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    compra = db.session.get(Compra, compra_id)
    if compra and compra.usuario_id == usuario.id:
        # Toggle do status da compra
        compra.realizado = not compra.realizado
        if compra.realizado and not compra.data_realizada:
            compra.data_realizada = datetime.now().date()
        elif not compra.realizado:
            compra.data_realizada = None
        
        # Vincular com o lançamento financeiro correspondente
        if compra.lancamento_financeiro:
            compra.lancamento_financeiro.realizado = compra.realizado
            compra.lancamento_financeiro.data_realizada = compra.data_realizada
        else:
            # Criar lançamento financeiro se não existir
            try:
                lancamento_criado = criar_lancamento_financeiro_automatico(compra, 'compra', usuario.id)
                if lancamento_criado:
                    print(f"✅ Lançamento financeiro criado automaticamente para compra {compra.id}")
                else:
                    print(f"❌ Erro ao criar lançamento financeiro para compra {compra.id}")
            except Exception as e:
                db.session.rollback()
                print(f"❌ Erro ao criar lançamento financeiro: {str(e)}")
        
        # ATUALIZAR ESTOQUE baseado no status da compra (sem criar produtos)
        if compra.tipo_compra == 'mercadoria':
            # Buscar produto na empresa (não apenas no usuário)
            if not hasattr(usuario, 'empresa_id') or not usuario.empresa_id:
                app.logger.error(f"Usuário {usuario.id} não tem empresa_id definido")
                flash('Usuário não tem empresa associada.', 'error')
                return redirect(url_for('compras'))
            
            empresa_id = obter_empresa_id_sessao(session, usuario)

            
            usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
            usuarios_ids = [u.id for u in usuarios_empresa]
            
            produto_estoque = Produto.query.filter(
                Produto.nome == compra.produto,
                Produto.usuario_id.in_(usuarios_ids)
            ).first()
            
            if produto_estoque:
                # Calcular estoque real baseado em todas as compras e vendas (independente do status realizado)
                estoque_real = calcular_estoque_produto(compra.produto, usuario.id)
                
                # Calcular preço médio real baseado em todas as compras (independente do status realizado)
                preco_medio_real = calcular_preco_medio_produto(compra.produto, usuario.id)
                
                produto_estoque.estoque = estoque_real
                produto_estoque.preco_custo = preco_medio_real
                
                flash(f'Estoque do produto "{compra.produto}" atualizado. Quantidade total: {estoque_real}. Preço médio: R$ {preco_medio_real:.2f}', 'info')
            else:
                flash(f'Produto "{compra.produto}" não encontrado no estoque. O produto deve ser criado automaticamente ao registrar a compra.', 'warning')
        
        db.session.commit()
        flash('Status da compra alterado com sucesso.', 'success')
    
    return redirect(url_for('compras'))

@app.route('/compras/<int:compra_id>/deletar')
def deletar_compra(compra_id):
    """Exclui uma compra usando o novo gerenciador de cancelamento"""
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    # Obter empresa_id correta da sessão (considera acesso_contador)
    empresa_id = obter_empresa_id_sessao(session, usuario)
    
    compra = db.session.get(Compra, compra_id)
    if not compra or compra.empresa_id != empresa_id:
        flash('Compra não encontrada ou você não tem permissão para excluí-la.', 'error')
        return redirect(url_for('compras'))
    
    try:
        app.logger.info(f"🗑️ INICIANDO EXCLUSÃO da compra {compra_id}")
        
        # CORREÇÃO: Buscar e excluir TODOS os lançamentos financeiros vinculados à compra
        # 1. Lançamentos com compra_id direto
        lancamentos_diretos = Lancamento.query.filter_by(compra_id=compra_id).all()
        for lancamento in lancamentos_diretos:
            # Excluir parcelas do lançamento se existirem
            if lancamento.parcelas:
                for parcela in lancamento.parcelas:
                    db.session.delete(parcela)
            db.session.delete(lancamento)
            app.logger.info(f"✅ Lançamento direto {lancamento.id} excluído para compra {compra_id}")
        
        # 2. Lançamentos vinculados através da tabela Vinculo
        vinculos = Vinculo.query.filter(
            db.or_(
                db.and_(Vinculo.lado_a_tipo == 'compra', Vinculo.lado_a_id == compra_id),
                db.and_(Vinculo.lado_b_tipo == 'compra', Vinculo.lado_b_id == compra_id)
            )
        ).all()
        
        for vinculo in vinculos:
            # Buscar lançamento vinculado
            if vinculo.lado_a_tipo == 'lancamento':
                lancamento_vinculado = db.session.get(Lancamento, vinculo.lado_a_id)
            elif vinculo.lado_b_tipo == 'lancamento':
                lancamento_vinculado = db.session.get(Lancamento, vinculo.lado_b_id)
            else:
                lancamento_vinculado = None
            
            # Excluir lançamento vinculado se existir
            if lancamento_vinculado:
                # Excluir parcelas do lançamento se existirem
                if lancamento_vinculado.parcelas:
                    for parcela in lancamento_vinculado.parcelas:
                        db.session.delete(parcela)
                db.session.delete(lancamento_vinculado)
                app.logger.info(f"✅ Lançamento vinculado {lancamento_vinculado.id} excluído para compra {compra_id}")
            
            # Excluir o vínculo
            db.session.delete(vinculo)
        
        app.logger.info(f"✅ Todos os lançamentos financeiros e vínculos excluídos para compra {compra_id}")
        
        # Reverter movimento de estoque se for mercadoria
        if compra.tipo_compra == 'mercadoria':
            sucesso_reversao, mensagem_reversao = reverter_movimento_estoque_compra(compra_id, usuario.id)
            if sucesso_reversao:
                app.logger.info(f"✅ Movimento de estoque revertido para compra {compra_id}")
            else:
                app.logger.warning(f"⚠️ Aviso na reversão de estoque: {mensagem_reversao}")
        
        # Excluir a compra
        db.session.delete(compra)
        db.session.commit()
        
        app.logger.info(f"✅ COMPRA EXCLUÍDA COM SUCESSO: {compra_id}")
        flash('Compra excluída com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"❌ ERRO ao excluir compra {compra_id}: {str(e)}")
        flash(f'Erro ao excluir compra: {str(e)}', 'error')
    
    return redirect(url_for('compras'))

# Rota para alterar dados do usuário
@app.route('/perfil', methods=['GET', 'POST'])
def perfil():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = Usuario.query.options(db.joinedload(Usuario.empresa)).get(session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    if request.method == 'POST':
        nome = request.form['nome']
        email = request.form['email']
        telefone = request.form.get('telefone', '')
        tipo_empresa = request.form['tipo_empresa']
        
        # Removida validação de email único global para permitir múltiplos perfis com mesmo email
        usuario.nome = nome
        usuario.email = email
        usuario.telefone = telefone
        if usuario.empresa:
            usuario.empresa.tipo_empresa = tipo_empresa
        db.session.commit()
        session['usuario_nome'] = nome
        # Recarregar relacionamento empresa após alteração
        db.session.refresh(usuario)
        flash('Dados atualizados com sucesso!', 'success')
        return redirect(url_for('perfil'))
    
    return render_template('perfil.html', usuario=usuario)

# Funções auxiliares para atualização de estoque
# DESABILITADO: Função antiga substituída pela versão na linha ~13490 que suporta múltiplos produtos
# def atualizar_estoque_compra(compra, usuario_id):
#     """Atualiza o estoque quando uma compra é registrada (independente do status)"""
#     if compra.tipo_compra == 'servico':
#         return True, None
#
#     # Verificar se o produto já existe no estoque
#     produto_estoque = Produto.query.filter_by(
#         nome=compra.produto,
#         usuario_id=usuario_id
#     ).first()
#
#     if produto_estoque:
#         # Calcular estoque real baseado em todas as compras e vendas realizadas
#         estoque_real = calcular_estoque_produto(compra.produto, usuario_id)
#
#         # Calcular preço médio real baseado em todas as compras realizadas
#         preco_medio_real = calcular_preco_medio_produto(compra.produto, usuario_id)
#
#         # Atualizar produto
#         produto_estoque.estoque = estoque_real
#         produto_estoque.preco_custo = preco_medio_real
#         db.session.commit()
#         return True, f'Estoque do produto "{compra.produto}" atualizado. Quantidade total: {estoque_real}. Preço médio: R$ {preco_medio_real:.2f}'
#     else:
#         # Criar novo produto no estoque
#         # Se estoque > 0, ativar automaticamente
#         ativo_produto = compra.quantidade > 0
#
#         novo_produto = Produto(
#             nome=compra.produto,
#             descricao=f'Produto adicionado via compra - {compra.observacoes}' if compra.observacoes else f'Produto adicionado via compra',
#             preco_custo=compra.preco_custo,
#             preco_venda=compra.preco_custo * 1.3,  # Margem de 30% por padrão
#             estoque=compra.quantidade,
#             ativo=ativo_produto,
#             usuario_id=usuario_id
#         )
#         db.session.add(novo_produto)
#         db.session.commit()
#         return True, f'Produto "{compra.produto}" criado no estoque com {compra.quantidade} unidades.'

# Função para normalizar nome de produto preservando informações
def normalizar_nome_produto(nome_produto):
    """
    Normaliza o nome de um produto removendo APENAS duplicações consecutivas.
    Preserva todos os itens únicos.
    Ex: "macbook, macbook" -> "macbook"
    Ex: "mouse, teclado" -> "mouse, teclado" (preserva ambos)
    """
    if not nome_produto:
        return nome_produto

    # Remover espaços extras
    nome = nome_produto.strip()

    # Se contém vírgula, processar lista de itens
    if ',' in nome:
        partes = [p.strip().lower() for p in nome.split(',') if p.strip()]

        # Remover duplicatas mantendo ordem, usando case-insensitive
        partes_unicas = []
        vistos = set()
        for parte in partes:
            if parte not in vistos:
                # Adicionar com capitalização original
                idx = [p.strip().lower() for p in nome.split(',')].index(parte)
                partes_unicas.append(nome.split(',')[idx].strip())
                vistos.add(parte)

        # Retornar todos os itens únicos separados por vírgula
        return ', '.join(partes_unicas)

    return nome

def consolidar_produtos_duplicados(usuario_id):
    """
    Consolida produtos duplicados ou com nomes similares
    Agrupa todos os produtos com o mesmo nome normalizado
    """
    try:
        usuario = db.session.get(Usuario, usuario_id)
        if not usuario:
            return False, "Usuário não encontrado"
        
        # Buscar todos os usuários da mesma empresa
        empresa_id = obter_empresa_id_sessao(session, usuario)

        usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
        usuarios_ids = [u.id for u in usuarios_empresa]
        
        # Buscar todos os produtos da empresa
        produtos = Produto.query.filter(Produto.usuario_id.in_(usuarios_ids)).all()
        
        # Agrupar por nome normalizado
        produtos_normalizados = {}
        for produto in produtos:
            nome_normalizado = normalizar_nome_produto(produto.nome)
            if nome_normalizado not in produtos_normalizados:
                produtos_normalizados[nome_normalizado] = []
            produtos_normalizados[nome_normalizado].append(produto)
        
        # Consolidar produtos duplicados
        consolidados = 0
        consolidacoes_detalhes = []  # Para auditoria

        for nome_normalizado, produtos_grupo in produtos_normalizados.items():
            if len(produtos_grupo) > 1:
                # Manter o primeiro produto e consolidar os outros
                produto_principal = produtos_grupo[0]
                produtos_consolidados = []

                for produto_duplicado in produtos_grupo[1:]:
                    # Registrar para auditoria
                    produtos_consolidados.append({
                        'id': produto_duplicado.id,
                        'nome': produto_duplicado.nome,
                        'estoque_anterior': produto_duplicado.estoque
                    })

                    # Atualizar todas as vendas e compras para usar o produto principal
                    Venda.query.filter_by(produto=produto_duplicado.nome).update(
                        {'produto': produto_principal.nome},
                        synchronize_session=False
                    )
                    Compra.query.filter_by(produto=produto_duplicado.nome).update(
                        {'produto': produto_principal.nome},
                        synchronize_session=False
                    )

                    # Deletar o produto duplicado
                    db.session.delete(produto_duplicado)
                    consolidados += 1
                    app.logger.info(f"Produto duplicado '{produto_duplicado.nome}' consolidado para '{produto_principal.nome}'")

                # Recalcular estoque do produto principal
                estoque_anterior = produto_principal.estoque
                estoque_real = calcular_estoque_produto(produto_principal.nome, usuario_id)
                produto_principal.estoque = estoque_real

                # Registrar consolidação para auditoria
                consolidacoes_detalhes.append({
                    'produto_principal': {
                        'id': produto_principal.id,
                        'nome': produto_principal.nome,
                        'estoque_anterior': estoque_anterior,
                        'estoque_novo': estoque_real
                    },
                    'produtos_consolidados': produtos_consolidados
                })

        # Criar registro de auditoria no EventLog
        if consolidados > 0:
            import json
            import hashlib

            detalhes_json = json.dumps(consolidacoes_detalhes, ensure_ascii=False)
            hash_evento = hashlib.sha256(
                f"consolidacao_produtos_{usuario_id}_{datetime.now().isoformat()}".encode()
            ).hexdigest()

            evento = EventLog(
                tipo_evento='consolidacao_produtos',
                descricao=f'Consolidação de {consolidados} produto(s) duplicado(s)',
                detalhes=detalhes_json,
                usuario_id=usuario_id,
                empresa_id=empresa_id,
                hash_evento=hash_evento
            )
            db.session.add(evento)
            app.logger.info(f"✅ Auditoria registrada: Consolidação de {consolidados} produtos")
        
        db.session.commit()
        app.logger.info(f"Consolidação concluída: {consolidados} produtos duplicados removidos")
        return True, f"Consolidação concluída: {consolidados} produtos duplicados removidos"
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao consolidar produtos: {str(e)}")
        return False, f"Erro ao consolidar produtos: {str(e)}"

def calcular_preco_medio_produto(nome_produto, usuario_id):
    """
    Calcula o preço médio ponderado de um produto baseado em todas as compras
    (independente do status "realizado")
    """
    # Buscar o usuário para obter a empresa_id
    usuario = db.session.get(Usuario, usuario_id)
    if not usuario:
        return 0
    
    # Buscar todos os usuários da mesma empresa
    empresa_id = obter_empresa_id_sessao(session, usuario)

    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]
    
    app.logger.info(f"Calculando preço médio para produto {nome_produto} - Usuários da empresa: {usuarios_ids}")
    
    # ATUALIZAÇÃO: Buscar todas as compras para este produto (independente do status realizado)
    compras_todas = Compra.query.filter(
        Compra.produto == nome_produto,
        Compra.empresa_id == empresa_id,
        Compra.tipo_compra == 'mercadoria'
    ).all()
    
    if not compras_todas:
        app.logger.info(f"Nenhuma compra encontrada para produto {nome_produto}")
        return 0
    
    # Calcular preço médio ponderado (todas as compras)
    valor_total = 0
    quantidade_total = 0
    
    for compra in compras_todas:
        valor_total += compra.quantidade * compra.preco_custo
        quantidade_total += compra.quantidade
    
    if quantidade_total > 0:
        preco_medio = valor_total / quantidade_total
        app.logger.info(f"Preço médio calculado para {nome_produto}: R$ {preco_medio:.2f} (baseado em {len(compras_todas)} compras)")
        return preco_medio
    else:
        app.logger.info(f"Nenhuma compra encontrada para produto {nome_produto}")
        return 0

def calcular_estoque_produto(nome_produto, usuario_id):
    """
    Calcula a quantidade em estoque de um produto baseado em todas as compras e vendas
    (independente do status "realizado")
    """
    # Buscar o usuário para obter a empresa_id
    usuario = db.session.get(Usuario, usuario_id)
    if not usuario:
        app.logger.error(f"Usuário {usuario_id} não encontrado para cálculo de estoque")
        return 0
    
    # Buscar todos os usuários da mesma empresa
    empresa_id = obter_empresa_id_sessao(session, usuario)

    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]

    # Normalizar nome do produto (remover duplicações como "macbook, macbook")
    # Pegar apenas o primeiro nome se houver vírgula
    nomes_produto = [n.strip() for n in nome_produto.split(',')]
    nome_produto_base = nomes_produto[0] if nomes_produto else nome_produto

    # ATUALIZAÇÃO: Buscar todas as compras para este produto (independente do status realizado)
    # Busca por nome exato apenas (não usa fuzzy match para evitar conflitos entre produtos similares)
    compras_todas = Compra.query.filter(
        db.or_(
            Compra.produto == nome_produto,
            Compra.produto == nome_produto_base
        ),
        Compra.empresa_id == empresa_id,
        Compra.tipo_compra == 'mercadoria'
    ).all()

    # ATUALIZAÇÃO: Buscar todas as vendas para este produto (independente do status realizado)
    # Busca por nome exato apenas (não usa fuzzy match para evitar conflitos entre produtos similares)
    vendas_todas = Venda.query.filter(
        db.or_(
            Venda.produto == nome_produto,
            Venda.produto == nome_produto_base
        ),
        Venda.empresa_id == empresa_id,
        Venda.tipo_venda == 'produto'
    ).all()
    
    # Calcular quantidade total comprada (todas as compras)
    quantidade_comprada = 0
    for compra in compras_todas:
        quantidade_comprada += compra.quantidade
    
    # Calcular quantidade total vendida (todas as vendas)
    quantidade_vendida = 0
    for venda in vendas_todas:
        quantidade_vendida += venda.quantidade
    
    # Estoque = Compras - Vendas
    estoque_calculado = quantidade_comprada - quantidade_vendida
    
    app.logger.info(f"Estoque calculado para {nome_produto}: {quantidade_comprada} compras - {quantidade_vendida} vendas = {estoque_calculado}")
    
    return estoque_calculado

def calcular_data_vencimento_parcela(data_base, numero_parcela, intervalo):
    """
    Calcula a data de vencimento de uma parcela baseada no intervalo especificado
    
    Args:
        data_base: Data base (primeira parcela)
        numero_parcela: Número da parcela (1, 2, 3, etc.)
        intervalo: 'semanal', 'quinzenal', 'mensal' ou 'personalizado'
    
    Returns:
        Data de vencimento da parcela
    """
    from datetime import date, timedelta
    
    if isinstance(data_base, str):
        try:
            data_base = datetime.strptime(data_base, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            # Se não conseguir converter, usar data atual
            data_base = date.today()
    
    if intervalo == 'semanal':
        # Adicionar semanas (7 dias por parcela)
        dias_para_adicionar = 7 * (numero_parcela - 1)
        return data_base + timedelta(days=dias_para_adicionar)
    
    elif intervalo == 'quinzenal':
        # Adicionar quinzenas (15 dias por parcela)
        dias_para_adicionar = 15 * (numero_parcela - 1)
        return data_base + timedelta(days=dias_para_adicionar)
    
    elif intervalo == 'mensal':
        # Adicionar meses
        ano = data_base.year
        mes = data_base.month + (numero_parcela - 1)
        
        # Ajustar ano se o mês passar de 12
        while mes > 12:
            ano += 1
            mes -= 12
        
        # Tentar criar a data, ajustando o dia se necessário
        try:
            return data_base.replace(year=ano, month=mes)
        except ValueError:
            # Se o dia não existir no mês (ex: 31 de fevereiro), usar o último dia do mês
            if mes == 2:
                if ano % 4 == 0 and (ano % 100 != 0 or ano % 400 == 0):  # Ano bissexto
                    return data_base.replace(year=ano, month=mes, day=29)
                else:
                    return data_base.replace(year=ano, month=mes, day=28)
            elif mes in [4, 6, 9, 11]:  # Meses com 30 dias
                return data_base.replace(year=ano, month=mes, day=30)
            else:  # Meses com 31 dias
                return data_base.replace(year=ano, month=mes, day=31)
    
    else:  # personalizado ou padrão
        # Usar mensal como padrão
        return calcular_data_vencimento_parcela(data_base, numero_parcela, 'mensal')

def sincronizar_estoque_usuario(usuario_id):
    """
    Sincroniza o estoque de todos os produtos de uma empresa com todos os dados de compras e vendas
    (independente do status "realizado")
    """
    # Verificar se o usuário existe
    usuario = db.session.get(Usuario, usuario_id)
    if not usuario:
        app.logger.error(f"Usuário {usuario_id} não encontrado para sincronização de estoque")
        return False, "Usuário não encontrado"
    
    # Carregar produtos de toda a empresa usando a função auxiliar
    produtos = buscar_produtos_empresa(usuario.empresa_id)
    app.logger.info(f"Sincronizando estoque para {len(produtos)} produtos da empresa {usuario.empresa_id}")
    
    for produto in produtos:
        try:
            # Calcular estoque real
            estoque_real = calcular_estoque_produto(produto.nome, usuario_id)
            
            # Calcular preço médio real
            preco_medio_real = calcular_preco_medio_produto(produto.nome, usuario_id)
            
            # Atualizar produto
            produto.estoque = estoque_real
            produto.preco_custo = preco_medio_real
            
            app.logger.info(f"Produto {produto.nome} sincronizado: estoque={estoque_real}, preço_custo={preco_medio_real:.2f}")
        except Exception as e:
            app.logger.error(f"Erro ao sincronizar produto {produto.nome}: {str(e)}")
    
    try:
        db.session.commit()
        app.logger.info(f"Estoque sincronizado com sucesso para {len(produtos)} produtos")
        return True, f"Estoque sincronizado para {len(produtos)} produtos"
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao fazer commit da sincronização de estoque: {str(e)}")
        return False, f"Erro ao salvar alterações: {str(e)}"

def resetar_banco():
    """Reseta completamente o banco de dados e cria apenas o administrador"""
    with app.app_context():
        try:
            # Deletar todas as tabelas
            db.drop_all()
            print("Banco de dados resetado - todas as tabelas removidas!")
            
            # Recriar todas as tabelas
            db.create_all()
            print("Tabelas recriadas com sucesso!")
            
            # Criar empresa administrativa
            empresa_admin = Empresa(
                cnpj='00.000.000/0000-00',
                razao_social='Sistema Administrativo',
                nome_fantasia='Admin',
                tipo_empresa='servicos',
                tipo_conta='admin'
            )
            db.session.add(empresa_admin)
            db.session.flush()
            
            # Criar usuário administrador
            admin = Usuario(
                nome='Administrador',
                usuario='admin',  # Campo usuario
                email='admin@sistema.com',
                senha=generate_password_hash('admin123'),
                tipo='admin',
                empresa_id=empresa_admin.id
            )
            db.session.add(admin)
            db.session.commit()
            
            print("=" * 50)
            print("SISTEMA DE GESTAO FINANCEIRA")
            print("=" * 50)
            print()
            print("Banco de dados resetado com sucesso!")
            print()
            print("Credenciais do Administrador:")
            print("CNPJ: 00.000.000/0000-00")
            print("Usuário: admin")
            print("Senha: admin123")
            print()
            print("=" * 50)
            
        except Exception as e:
            print(f"Erro ao resetar banco de dados: {e}")
            raise

def criar_banco():
    with app.app_context():
        try:
            db.create_all()
            
            # Criar usuário admin se não existir
            admin = Usuario.query.filter_by(tipo='admin').first()
            if not admin:
                # Criar empresa padrão para admin
                empresa_admin = Empresa(
                    cnpj='00.000.000/0000-00',
                    razao_social='Sistema Administrativo',
                    nome_fantasia='Admin',
                    tipo_empresa='servicos',
                    tipo_conta='admin'
                )
                db.session.add(empresa_admin)
                db.session.flush()
                
                admin = Usuario(
                    nome='Administrador',
                    usuario='admin',  # Campo usuario
                    email='admin@sistema.com',
                    senha=generate_password_hash('admin123'),
                    tipo='admin',
                    empresa_id=empresa_admin.id
                )
                db.session.add(admin)
                db.session.commit()
                print("Banco de dados criado com sucesso!")
        except Exception as e:
            print(f"Erro ao criar banco de dados: {e}")
            # Tentar recriar o banco
            try:
                db.drop_all()
                db.create_all()
                
                # Criar empresa padrão para admin
                empresa_admin = Empresa(
                    cnpj='00.000.000/0000-00',
                    razao_social='Sistema Administrativo',
                    nome_fantasia='Admin',
                    tipo_empresa='servicos',
                    tipo_conta='admin'
                )
                db.session.add(empresa_admin)
                db.session.flush()
                
                admin = Usuario(
                    nome='Administrador',
                    usuario='admin',  # Campo usuario
                    email='admin@sistema.com',
                    senha=generate_password_hash('admin123'),
                    tipo='admin',
                    empresa_id=empresa_admin.id
                )
                db.session.add(admin)
                db.session.commit()
                print("Banco de dados recriado com sucesso!")
            except Exception as e2:
                print(f"Erro fatal ao recriar banco de dados: {e2}")
                raise

def migrar_banco():
    with app.app_context():
        # Verificar se as colunas já existem usando SQL direto
        try:
            # Verificar se a tabela empresa existe
            result = db.session.execute(text("SELECT name FROM sqlite_master WHERE type='table' AND name='empresa'"))
            if not result.fetchone():
                # Criar a tabela empresa
                db.session.execute(text('''
                    CREATE TABLE empresa (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        cnpj VARCHAR(20) UNIQUE NOT NULL,
                        razao_social VARCHAR(200) NOT NULL,
                        nome_fantasia VARCHAR(200),
                        email VARCHAR(120),
                        telefone VARCHAR(20),
                        endereco TEXT,
                        tipo_empresa VARCHAR(20) DEFAULT 'servicos',
                        ativo BOOLEAN DEFAULT 1,
                        data_criacao DATETIME DEFAULT CURRENT_TIMESTAMP
                    )
                '''))
                print("Tabela empresa criada com sucesso!")
            else:
                print("Tabela empresa já existe!")
            
            # Verificar se a tabela usuario existe e se tem o campo usuario
            result = db.session.execute(text("SELECT name FROM sqlite_master WHERE type='table' AND name='usuario'"))
            if result.fetchone():
                # Verificar se o campo usuario existe
                result = db.session.execute(text('PRAGMA table_info(usuario)'))
                colunas_usuario = [row[1] for row in result.fetchall()]
                
                if 'usuario' not in colunas_usuario:
                    db.session.execute(text('ALTER TABLE usuario ADD COLUMN usuario VARCHAR(50)'))
                    print("Coluna usuario adicionada na tabela usuario com sucesso!")
                    
                    # Atualizar usuários existentes com um valor padrão
                    db.session.execute(text("UPDATE usuario SET usuario = 'usuario' || id WHERE usuario IS NULL OR usuario = ''"))
                    db.session.commit()
                    print("Usuários existentes atualizados com valor padrão para o campo usuario!")
                else:
                    print("Coluna usuario já existe na tabela usuario!")
            else:
                print("Tabela usuario não existe ainda!")
            
            # Verificar se compra_id existe na tabela lancamento
            result = db.session.execute(text('PRAGMA table_info(lancamento)'))
            colunas_lancamento = [row[1] for row in result.fetchall()]
            
            if 'compra_id' not in colunas_lancamento:
                db.session.execute(text('ALTER TABLE lancamento ADD COLUMN compra_id INTEGER REFERENCES compra(id)'))
                print("Coluna compra_id adicionada com sucesso!")
            else:
                print("Coluna compra_id já existe!")
            
            if 'venda_id' not in colunas_lancamento:
                db.session.execute(text('ALTER TABLE lancamento ADD COLUMN venda_id INTEGER REFERENCES venda(id)'))
                print("Coluna venda_id adicionada com sucesso!")
            else:
                print("Coluna venda_id já existe!")
            
            if 'conta_caixa_id' not in colunas_lancamento:
                db.session.execute(text('ALTER TABLE lancamento ADD COLUMN conta_caixa_id INTEGER REFERENCES conta_caixa(id)'))
                print("Coluna conta_caixa_id adicionada com sucesso!")
            else:
                print("Coluna conta_caixa_id já existe!")
            
            # Verificar se quantidade existe na tabela venda
            result = db.session.execute(text('PRAGMA table_info(venda)'))
            colunas_venda = [row[1] for row in result.fetchall()]
            
            if 'quantidade' not in colunas_venda:
                db.session.execute(text('ALTER TABLE venda ADD COLUMN quantidade INTEGER DEFAULT 1'))
                print("Coluna quantidade adicionada na tabela venda com sucesso!")
            else:
                print("Coluna quantidade já existe na tabela venda!")
            
            if 'tipo_venda' not in colunas_venda:
                db.session.execute(text('ALTER TABLE venda ADD COLUMN tipo_venda VARCHAR(20) DEFAULT "produto"'))
                print("Coluna tipo_venda adicionada na tabela venda com sucesso!")
            else:
                print("Coluna tipo_venda já existe na tabela venda!")
            
            # Novos campos para venda
            if 'tipo_pagamento' not in colunas_venda:
                db.session.execute(text('ALTER TABLE venda ADD COLUMN tipo_pagamento VARCHAR(20) DEFAULT "a_vista"'))
                print("Coluna tipo_pagamento adicionada na tabela venda com sucesso!")
            else:
                print("Coluna tipo_pagamento já existe na tabela venda!")
            
            if 'numero_parcelas' not in colunas_venda:
                db.session.execute(text('ALTER TABLE venda ADD COLUMN numero_parcelas INTEGER DEFAULT 1'))
                print("Coluna numero_parcelas adicionada na tabela venda com sucesso!")
            else:
                print("Coluna numero_parcelas já existe na tabela venda!")
            
            if 'valor_parcela' not in colunas_venda:
                db.session.execute(text('ALTER TABLE venda ADD COLUMN valor_parcela FLOAT'))
                print("Coluna valor_parcela adicionada na tabela venda com sucesso!")
            else:
                print("Coluna valor_parcela já existe na tabela venda!")
            
            if 'desconto' not in colunas_venda:
                db.session.execute(text('ALTER TABLE venda ADD COLUMN desconto FLOAT DEFAULT 0.0'))
                print("Coluna desconto adicionada na tabela venda com sucesso!")
            else:
                print("Coluna desconto já existe na tabela venda!")
            
            if 'valor_final' not in colunas_venda:
                db.session.execute(text('ALTER TABLE venda ADD COLUMN valor_final FLOAT'))
                print("Coluna valor_final adicionada na tabela venda com sucesso!")
            else:
                print("Coluna valor_final já existe na tabela venda!")
            
            # Verificar se quantidade existe na tabela compra
            result = db.session.execute(text('PRAGMA table_info(compra)'))
            colunas_compra = [row[1] for row in result.fetchall()]
            
            if 'quantidade' not in colunas_compra:
                db.session.execute(text('ALTER TABLE compra ADD COLUMN quantidade INTEGER DEFAULT 1'))
                print("Coluna quantidade adicionada na tabela compra com sucesso!")
            else:
                print("Coluna quantidade já existe na tabela compra!")
            
            if 'tipo_compra' not in colunas_compra:
                db.session.execute(text('ALTER TABLE compra ADD COLUMN tipo_compra VARCHAR(20) DEFAULT "mercadoria"'))
                print("Coluna tipo_compra adicionada na tabela compra com sucesso!")
            else:
                print("Coluna tipo_compra já existe na tabela compra!")
            
            # Novos campos para compra
            if 'tipo_pagamento' not in colunas_compra:
                db.session.execute(text('ALTER TABLE compra ADD COLUMN tipo_pagamento VARCHAR(20) DEFAULT "a_vista"'))
                print("Coluna tipo_pagamento adicionada na tabela compra com sucesso!")
            else:
                print("Coluna tipo_pagamento já existe na tabela compra!")
            
            if 'numero_parcelas' not in colunas_compra:
                db.session.execute(text('ALTER TABLE compra ADD COLUMN numero_parcelas INTEGER DEFAULT 1'))
                print("Coluna numero_parcelas adicionada na tabela compra com sucesso!")
            else:
                print("Coluna numero_parcelas já existe na tabela compra!")
            
            if 'valor_parcela' not in colunas_compra:
                db.session.execute(text('ALTER TABLE compra ADD COLUMN valor_parcela FLOAT'))
                print("Coluna valor_parcela adicionada na tabela compra com sucesso!")
            else:
                print("Coluna valor_parcela já existe na tabela compra!")
            
            # Verificar se a tabela conta_caixa existe
            result = db.session.execute(text("SELECT name FROM sqlite_master WHERE type='table' AND name='conta_caixa'"))
            if not result.fetchone():
                # Criar a tabela conta_caixa
                db.session.execute(text('''
                    CREATE TABLE conta_caixa (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        nome VARCHAR(200) NOT NULL,
                        tipo VARCHAR(50) NOT NULL,
                        banco VARCHAR(100),
                        agencia VARCHAR(20),
                        conta VARCHAR(20),
                        saldo_inicial FLOAT DEFAULT 0.0,
                        saldo_atual FLOAT DEFAULT 0.0,
                        ativo BOOLEAN DEFAULT 1,
                        descricao TEXT,
                        usuario_id INTEGER NOT NULL,
                        data_criacao DATETIME DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (usuario_id) REFERENCES usuario (id)
                    )
                '''))
                print("Tabela conta_caixa criada com sucesso!")
            else:
                print("Tabela conta_caixa já existe!")
            
            # Verificar se a tabela parcela existe
            result = db.session.execute(text("SELECT name FROM sqlite_master WHERE type='table' AND name='parcela'"))
            if not result.fetchone():
                # Criar a tabela parcela
                db.session.execute(text('''
                    CREATE TABLE parcela (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        numero INTEGER NOT NULL,
                        valor FLOAT NOT NULL,
                        data_vencimento DATE NOT NULL,
                        data_pagamento DATE,
                        realizado BOOLEAN DEFAULT 0,
                        venda_id INTEGER,
                        compra_id INTEGER,
                        lancamento_id INTEGER,
                        usuario_id INTEGER NOT NULL,
                        data_criacao DATETIME DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (venda_id) REFERENCES venda (id),
                        FOREIGN KEY (compra_id) REFERENCES compra (id),
                        FOREIGN KEY (lancamento_id) REFERENCES lancamento (id),
                        FOREIGN KEY (usuario_id) REFERENCES usuario (id)
                    )
                '''))
                print("Tabela parcela criada com sucesso!")
            else:
                print("Tabela parcela já existe!")
            
            # Verificar se a tabela permissao existe
            result = db.session.execute(text("SELECT name FROM sqlite_master WHERE type='table' AND name='permissao'"))
            if not result.fetchone():
                # Criar a tabela permissao
                db.session.execute(text('''
                    CREATE TABLE permissao (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        usuario_id INTEGER NOT NULL,
                        modulo VARCHAR(50) NOT NULL,
                        acao VARCHAR(50) NOT NULL,
                        ativo BOOLEAN DEFAULT 1,
                        data_criacao DATETIME DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (usuario_id) REFERENCES usuario (id)
                    )
                '''))
                print("Tabela permissao criada com sucesso!")
            else:
                print("Tabela permissao já existe!")
            
            # Verificar se empresa_id existe na tabela usuario
            result = db.session.execute(text('PRAGMA table_info(usuario)'))
            colunas_usuario = [row[1] for row in result.fetchall()]
            
            if 'empresa_id' not in colunas_usuario:
                db.session.execute(text('ALTER TABLE usuario ADD COLUMN empresa_id INTEGER REFERENCES empresa(id)'))
                print("Coluna empresa_id adicionada na tabela usuario com sucesso!")
            else:
                print("Coluna empresa_id já existe na tabela usuario!")
            
            if 'criado_por' not in colunas_usuario:
                db.session.execute(text('ALTER TABLE usuario ADD COLUMN criado_por INTEGER REFERENCES usuario(id)'))
                print("Coluna criado_por adicionada na tabela usuario com sucesso!")
            else:
                print("Coluna criado_por já existe na tabela usuario!")
            
            # Verificar se há usuários sem empresa_id e associá-los a uma empresa padrão
            usuarios_sem_empresa = db.session.execute(text('SELECT id FROM usuario WHERE empresa_id IS NULL')).fetchall()
            if usuarios_sem_empresa:
                # Criar empresa padrão se não existir
                empresa_padrao = db.session.execute(text("SELECT id FROM empresa WHERE cnpj = '00.000.000/0000-00'")).fetchone()
                if not empresa_padrao:
                    db.session.execute(text('''
                        INSERT INTO empresa (cnpj, razao_social, nome_fantasia, tipo_empresa, ativo, data_criacao)
                        VALUES ('00.000.000/0000-00', 'Empresa Padrão', 'Padrão', 'servicos', 1, CURRENT_TIMESTAMP)
                    '''))
                    db.session.commit()
                    empresa_padrao = db.session.execute(text("SELECT id FROM empresa WHERE cnpj = '00.000.000/0000-00'")).fetchone()
                
                # Associar usuários à empresa padrão
                for usuario in usuarios_sem_empresa:
                    db.session.execute(text(f'UPDATE usuario SET empresa_id = {empresa_padrao[0]} WHERE id = {usuario[0]}'))
                
                db.session.commit()
                print(f"{len(usuarios_sem_empresa)} usuários associados à empresa padrão!")
            
            db.session.commit()
            
        except Exception as e:
            print(f"Erro ao verificar/adicionar colunas: {e}")
            db.session.rollback()
            return

@app.route('/relatorios')
def relatorios():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    return render_template('relatorios.html', usuario=usuario)

@app.route('/dre/visualizar')
def dre_visualizar():
    """Visualizar DRE - Demonstração do Resultado do Exercício"""
    if 'usuario_id' not in session:
        return redirect(url_for('login'))

    usuario = db.session.get(Usuario, session['usuario_id'])
    empresa_id = obter_empresa_id_sessao(session, usuario)

    # Buscar configuração da DRE
    linhas_config = DreConfiguracao.query.filter_by(
        empresa_id=empresa_id,
        ativo=True
    ).order_by(DreConfiguracao.ordem).all()

    # Obter período do filtro (padrão: mês atual)
    data_fim = datetime.now().date()
    data_inicio = data_fim.replace(day=1)  # Primeiro dia do mês

    if request.args.get('data_inicio'):
        try:
            # Tentar formato dd/mm/aaaa primeiro
            data_inicio = datetime.strptime(request.args.get('data_inicio'), '%d/%m/%Y').date()
        except ValueError:
            # Fallback para formato AAAA-MM-DD
            try:
                data_inicio = datetime.strptime(request.args.get('data_inicio'), '%Y-%m-%d').date()
            except ValueError:
                pass  # Manter data padrão

    if request.args.get('data_fim'):
        try:
            # Tentar formato dd/mm/aaaa primeiro
            data_fim = datetime.strptime(request.args.get('data_fim'), '%d/%m/%Y').date()
        except ValueError:
            # Fallback para formato AAAA-MM-DD
            try:
                data_fim = datetime.strptime(request.args.get('data_fim'), '%Y-%m-%d').date()
            except ValueError:
                pass  # Manter data padrão

    # Calcular valores das linhas
    linhas_dre = []
    valores_acumulados = []  # Para calcular subtotais

    for linha_config in linhas_config:
        valor = 0

        if linha_config.tipo_linha == 'conta' and linha_config.plano_conta_id:
            # Buscar lançamentos da conta (REGIME DE COMPETÊNCIA - usa data_prevista)
            # Não filtra por 'realizado' pois regime de competência considera independente do pagamento
            # Exclui transferências, pois não representam entradas/saidas reais
            lancamentos = Lancamento.query.filter(
                Lancamento.plano_conta_id == linha_config.plano_conta_id,
                Lancamento.empresa_id == empresa_id,
                Lancamento.data_prevista.between(data_inicio, data_fim),
                db.or_(Lancamento.eh_transferencia == False, Lancamento.eh_transferencia.is_(None))
            ).all()

            # Somar valores (entradas positivas, saidas negativas)
            for lanc in lancamentos:
                if lanc.tipo == 'entrada':
                    valor += lanc.valor
                else:
                    valor -= lanc.valor

            # Adicionar aos valores acumulados para cálculo de subtotais
            valores_acumulados.append(valor)

        elif linha_config.tipo_linha in ['subtotal', 'total']:
            # Calcular subtotal somando todos os valores acumulados desde o último subtotal
            valor = sum(valores_acumulados)
            # Limpar valores acumulados após calcular subtotal
            valores_acumulados = []

        linhas_dre.append({
            'codigo': linha_config.codigo,
            'descricao': linha_config.descricao,
            'tipo_linha': linha_config.tipo_linha,
            'operacao': linha_config.operacao,
            'valor': valor,
            'nivel': linha_config.nivel,
            'negrito': linha_config.negrito or False,
            'linha_acima': linha_config.linha_acima if hasattr(linha_config, 'linha_acima') else False,
            'linha_abaixo': linha_config.linha_abaixo if hasattr(linha_config, 'linha_abaixo') else False
        })

    return render_template('dre_visualizar.html',
                         linhas_dre=linhas_dre,
                         data_inicio=data_inicio.strftime('%d/%m/%Y'),
                         data_fim=data_fim.strftime('%d/%m/%Y'))

@app.route('/dre/configurar', methods=['GET', 'POST'])
def dre_configurar():
    """Configurar DRE - adicionar/remover contas"""
    if 'usuario_id' not in session:
        return redirect(url_for('login'))

    usuario = db.session.get(Usuario, session['usuario_id'])
    empresa_id = obter_empresa_id_sessao(session, usuario)

    if request.method == 'POST':
        # Remover configurações antigas
        DreConfiguracao.query.filter_by(empresa_id=empresa_id).delete()

        # Processar linhas serializadas do formulário
        tipos = request.form.getlist('linha_tipo[]')
        ordens = request.form.getlist('linha_ordem[]')
        contas_ids = request.form.getlist('linha_conta_id[]')
        descricoes = request.form.getlist('linha_descricao[]')
        operacoes = request.form.getlist('linha_operacao[]')

        for i in range(len(tipos)):
            tipo = tipos[i]
            ordem = int(ordens[i])

            if tipo == 'subtotal':
                # Criar linha sintética (subtotal)
                descricao = descricoes[i]
                operacao = operacoes[i]

                nova_linha = DreConfiguracao(
                    empresa_id=empresa_id,
                    plano_conta_id=None,
                    codigo=f"SUBTOTAL_{ordem}",
                    descricao=descricao,
                    tipo_linha='subtotal',
                    operacao=operacao,
                    ordem=ordem,
                    nivel=1,
                    negrito=True
                )
            else:
                # Criar linha de conta
                conta_id_str = contas_ids[i]
                if conta_id_str and conta_id_str.isdigit():
                    plano_conta = db.session.get(PlanoConta, int(conta_id_str))
                    if plano_conta:
                        nova_linha = DreConfiguracao(
                            empresa_id=empresa_id,
                            plano_conta_id=plano_conta.id,
                            codigo=plano_conta.codigo or f"{ordem:03d}",
                            descricao=plano_conta.nome,
                            tipo_linha='conta',
                            operacao='(+)' if plano_conta.tipo == 'entrada' else '(-)',
                            ordem=ordem,
                            nivel=1,
                            negrito=False
                        )
                    else:
                        continue
                else:
                    continue

            db.session.add(nova_linha)

        db.session.commit()
        flash('Configuração da DRE salva com sucesso!', 'success')
        return redirect(url_for('dre_visualizar'))

    # GET - Exibir formulário de configuração
    # Buscar contas do plano de contas (apenas analíticas)
    # Permitir que a mesma conta seja adicionada múltiplas vezes
    contas_disponiveis = PlanoConta.query.filter_by(
        empresa_id=empresa_id,
        natureza='analitica'
    ).order_by(PlanoConta.codigo).all()

    # Buscar linhas já configuradas
    linhas_dre = DreConfiguracao.query.filter_by(
        empresa_id=empresa_id,
        ativo=True
    ).order_by(DreConfiguracao.ordem).all()

    return render_template('dre_configurar.html',
                         contas_disponiveis=contas_disponiveis,
                         linhas_dre=linhas_dre)

@app.route('/relatorios/saldos')
def relatorio_saldos():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))

    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))

    try:
        # Buscar todos os usuários da mesma empresa
        empresa_id = obter_empresa_id_sessao(session, usuario)

        usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
        usuarios_ids = [u.id for u in usuarios_empresa]

        # Obter filtro de data de referência do request
        data_referencia_str = request.args.get('data_referencia', '')
        hoje = datetime.now().date()

        # Converter string de data para objeto date
        data_referencia = hoje
        if data_referencia_str:
            try:
                data_referencia = datetime.strptime(data_referencia_str, '%d/%m/%Y').date()
            except ValueError:
                flash('Data de referência inválida. Usando data de hoje.', 'error')
                data_referencia = hoje
                data_referencia_str = data_referencia.strftime('%d/%m/%Y')
        else:
            data_referencia_str = hoje.strftime('%d/%m/%Y')

        # Buscar todos os lançamentos da empresa ATÉ a data de referência (inclusive não realizados)
        lancamentos_ate_data = Lancamento.query.filter(
            Lancamento.empresa_id == empresa_id,
            Lancamento.data_prevista <= data_referencia
        ).all()

        # Calcular totais globais até a data
        total_receitas_realizadas = sum([l.valor for l in lancamentos_ate_data if l.tipo == 'entrada' and l.realizado])
        total_despesas_realizadas = sum([l.valor for l in lancamentos_ate_data if l.tipo == 'saida' and l.realizado])
        
        total_receitas_vencidas = sum([l.valor for l in lancamentos_ate_data if l.tipo == 'entrada' and not l.realizado and l.data_prevista < hoje])
        total_despesas_vencidas = sum([l.valor for l in lancamentos_ate_data if l.tipo == 'saida' and not l.realizado and l.data_prevista < hoje])

        total_receitas_agendadas = sum([l.valor for l in lancamentos_ate_data if l.tipo == 'entrada' and not l.realizado and l.data_prevista == hoje])
        total_despesas_agendadas = sum([l.valor for l in lancamentos_ate_data if l.tipo == 'saida' and not l.realizado and l.data_prevista == hoje])
        
        total_receitas_a_vencer = sum([l.valor for l in lancamentos_ate_data if l.tipo == 'entrada' and not l.realizado and l.data_prevista > hoje])
        total_despesas_a_vencer = sum([l.valor for l in lancamentos_ate_data if l.tipo == 'saida' and not l.realizado and l.data_prevista > hoje])

        # Totais consolidados
        saldo_realizado = total_receitas_realizadas - total_despesas_realizadas
        total_receitas_pendentes = total_receitas_vencidas + total_receitas_agendadas + total_receitas_a_vencer
        total_despesas_pendentes = total_despesas_vencidas + total_despesas_agendadas + total_despesas_a_vencer
        saldo_projetado = saldo_realizado + total_receitas_pendentes - total_despesas_pendentes
        saldo_vencido = total_receitas_vencidas - total_despesas_vencidas

        # 1. SALDOS POR CONTA DO PLANO DE CONTAS
        contas_plano = PlanoConta.query.filter_by(empresa_id=empresa_id, ativo=True).order_by(PlanoConta.codigo).all()
        p_contas_map = {c.id: c for c in contas_plano}
        
        categorias_receitas = {}
        categorias_despesas = {}

        # Inicializar todas as contas analíticas
        for c in contas_plano:
            if c.natureza == 'analitica':
                nome_exibicao = f"{c.codigo} - {c.nome}" if c.codigo else c.nome
                item = {
                    'realizado': 0.0,
                    'a_vencer': 0.0,
                    'vencido': 0.0,
                    'agendado': 0.0,
                    'total': 0.0,
                    'codigo': c.codigo
                }
                if c.tipo == 'entrada':
                    categorias_receitas[nome_exibicao] = item
                else:
                    categorias_despesas[nome_exibicao] = item

        # Distribuir valores dos lançamentos
        for l in lancamentos_ate_data:
            if l.plano_conta_id and l.plano_conta_id in p_contas_map:
                pc = p_contas_map[l.plano_conta_id]
                nome_exibicao = f"{pc.codigo} - {pc.nome}" if pc.codigo else pc.nome
                dest = categorias_receitas if pc.tipo == 'entrada' else categorias_despesas
                
                if nome_exibicao not in dest:
                    # Conta existe mas não foi inicializada (ex: sintética), criar entrada
                    dest[nome_exibicao] = {
                        'realizado': 0.0,
                        'a_vencer': 0.0,
                        'vencido': 0.0,
                        'agendado': 0.0,
                        'total': 0.0,
                        'codigo': pc.codigo or ''
                    }
                
                if l.realizado:
                    dest[nome_exibicao]['realizado'] += l.valor
                elif l.data_prevista < hoje:
                    dest[nome_exibicao]['vencido'] += l.valor
                elif l.data_prevista == hoje:
                    dest[nome_exibicao]['agendado'] += l.valor
                else:
                    dest[nome_exibicao]['a_vencer'] += l.valor
                    
                dest[nome_exibicao]['total'] = (dest[nome_exibicao]['realizado'] + 
                                               dest[nome_exibicao]['vencido'] + 
                                               dest[nome_exibicao]['agendado'] + 
                                               dest[nome_exibicao]['a_vencer'])
            else:
                # Lançamento sem plano de contas - agrupar como "Sem Categoria"
                cat_nome = "Sem Categoria"
                dest = categorias_receitas if l.tipo == 'entrada' else categorias_despesas
                
                if cat_nome not in dest:
                    dest[cat_nome] = {
                        'realizado': 0.0,
                        'a_vencer': 0.0,
                        'vencido': 0.0,
                        'agendado': 0.0,
                        'total': 0.0,
                        'codigo': '99.SEM'
                    }
                
                if l.realizado:
                    dest[cat_nome]['realizado'] += l.valor
                elif l.data_prevista < hoje:
                    dest[cat_nome]['vencido'] += l.valor
                elif l.data_prevista == hoje:
                    dest[cat_nome]['agendado'] += l.valor
                else:
                    dest[cat_nome]['a_vencer'] += l.valor
                    
                dest[cat_nome]['total'] = (dest[cat_nome]['realizado'] + 
                                          dest[cat_nome]['vencido'] + 
                                          dest[cat_nome]['agendado'] + 
                                          dest[cat_nome]['a_vencer'])

        # 2. SALDOS DAS CONTAS CAIXA
        contas_caixa = ContaCaixa.query.filter(ContaCaixa.usuario_id.in_(usuarios_ids), ContaCaixa.ativo==True).all()

        contas_caixa_detalhadas = []
        for conta in contas_caixa:
            lancs_conta = [l for l in lancamentos_ate_data if l.conta_caixa_id == conta.id]
            
            # Para o saldo atual da conta (realizado acumulado), precisamos considerar TODOS os lançamentos realizados
            # até hoje, não apenas os do filtro. Mas o usuário quer ATÉ a data.
            # Então calculamos o saldo acumulado realizado até a data_referencia.
            # Se houver um saldo inicial na conta, deve ser considerado.
            
            saldo_inicial = getattr(conta, 'saldo_inicial', 0) or 0
            
            # Saldo realizado até a data de referência
            rec_realizado = sum(l.valor for l in lancs_conta if l.tipo == 'entrada' and l.realizado)
            desp_realizado = sum(l.valor for l in lancs_conta if l.tipo == 'saida' and l.realizado)
            saldo_atual_conta = saldo_inicial + rec_realizado - desp_realizado
            
            # Pendentes até a data de referência
            rec_pendente = sum(l.valor for l in lancs_conta if l.tipo == 'entrada' and not l.realizado)
            desp_pendente = sum(l.valor for l in lancs_conta if l.tipo == 'saida' and not l.realizado)
            
            saldo_projetado_conta = saldo_atual_conta + rec_pendente - desp_pendente
            
            contas_caixa_detalhadas.append({
                'conta': conta,
                'saldo_atual': saldo_atual_conta,
                'rec_pendente': rec_pendente,
                'desp_pendente': desp_pendente,
                'saldo_projetado': saldo_projetado_conta,
                # Campos para compatibilidade com template atual
                'rec_a_vencer': sum(l.valor for l in lancs_conta if l.tipo == 'entrada' and not l.realizado and l.data_prevista > hoje),
                'desp_a_vencer': sum(l.valor for l in lancs_conta if l.tipo == 'saida' and not l.realizado and l.data_prevista > hoje),
                'rec_vencido': sum(l.valor for l in lancs_conta if l.tipo == 'entrada' and not l.realizado and l.data_prevista < hoje),
                'desp_vencido': sum(l.valor for l in lancs_conta if l.tipo == 'saida' and not l.realizado and l.data_prevista < hoje),
                'rec_agendado': sum(l.valor for l in lancs_conta if l.tipo == 'entrada' and not l.realizado and l.data_prevista == hoje),
                'desp_agendado': sum(l.valor for l in lancs_conta if l.tipo == 'saida' and not l.realizado and l.data_prevista == hoje),
            })

        saldo_total_caixa = sum(c['saldo_atual'] for c in contas_caixa_detalhadas)

        print(f"=== SALDOS DEBUG === data_ref={data_referencia_str}, lancamentos={len(lancamentos_ate_data)}, entradas={total_receitas_realizadas}, saidas={total_despesas_realizadas}, contas_caixa={len(contas_caixa_detalhadas)}, cat_rec={len(categorias_receitas)}, cat_desp={len(categorias_despesas)}")
        print(f"=== CATEGORIAS RECEITAS === {dict(list(categorias_receitas.items())[:3])}")
        print(f"=== CATEGORIAS DESPESAS === {dict(list(categorias_despesas.items())[:3])}")
        for cc in contas_caixa_detalhadas[:2]:
            print(f"=== CONTA CAIXA === nome={cc['conta'].nome}, saldo={cc['saldo_atual']}, projetado={cc['saldo_projetado']}")

        return render_template('relatorio_saldos.html',
                             usuario=usuario,
                             data_referencia=data_referencia_str,
                             total_entradas=total_receitas_realizadas,
                             total_saidas=total_despesas_realizadas,
                             total_entradas_realizadas=total_receitas_realizadas,
                             total_saidas_realizadas=total_despesas_realizadas,
                             total_receitas_realizadas=total_receitas_realizadas,
                             total_despesas_realizadas=total_despesas_realizadas,
                             total_entradas_pendentes=total_receitas_pendentes,
                             total_saidas_pendentes=total_despesas_pendentes,
                             total_receitas_pendentes=total_receitas_pendentes,
                             total_despesas_pendentes=total_despesas_pendentes,
                             total_entradas_vencidas=total_receitas_vencidas,
                             total_saidas_vencidas=total_despesas_vencidas,
                             total_receitas_vencidas=total_receitas_vencidas,
                             total_despesas_vencidas=total_despesas_vencidas,
                             total_entradas_agendadas=total_receitas_agendadas,
                             total_saidas_agendadas=total_despesas_agendadas,
                             total_receitas_agendadas=total_receitas_agendadas,
                             total_despesas_agendadas=total_despesas_agendadas,
                             total_entradas_a_vencer=total_receitas_a_vencer,
                             total_saidas_a_vencer=total_despesas_a_vencer,
                             total_receitas_a_vencer=total_receitas_a_vencer,
                             total_despesas_a_vencer=total_despesas_a_vencer,
                             saldo_geral=saldo_realizado,
                             saldo_atual=saldo_realizado,
                             saldo_vencido=saldo_vencido,
                             saldo_projetado=saldo_projetado,
                             categorias_receitas=categorias_receitas,
                             categorias_despesas=categorias_despesas,
                             categorias_entradas=categorias_receitas,
                             categorias_saidas=categorias_despesas,
                             contas_caixa_detalhadas=contas_caixa_detalhadas,
                             saldo_total_caixa=saldo_total_caixa,
                             sum_rec_a_vencer=sum(c['rec_a_vencer'] for c in contas_caixa_detalhadas),
                             sum_desp_a_vencer=sum(c['desp_a_vencer'] for c in contas_caixa_detalhadas),
                             sum_rec_vencido=sum(c['rec_vencido'] for c in contas_caixa_detalhadas),
                             sum_desp_vencido=sum(c['desp_vencido'] for c in contas_caixa_detalhadas),
                             sum_saldo_projetado_contas=sum(c['saldo_projetado'] for c in contas_caixa_detalhadas))

    except Exception as e:
        import traceback
        traceback.print_exc()
        app.logger.error(f"Erro no relatório de saldos: {str(e)}")
        flash('Erro ao gerar relatório de saldos', 'error')
        return render_template('relatorio_saldos.html',
                             data_referencia='',
                             usuario=usuario,
                             total_entradas=0,
                             total_saidas=0,
                             total_receitas_realizadas=0,
                             total_despesas_realizadas=0,
                             total_receitas_a_vencer=0,
                             total_despesas_a_vencer=0,
                             total_receitas_vencidas=0,
                             total_despesas_vencidas=0,
                             total_receitas_pendentes=0,
                             total_despesas_pendentes=0,
                             total_receitas_agendadas=0,
                             total_despesas_agendadas=0,
                             # Aliases para compatibilidade com template
                             total_entradas_realizadas=0,
                             total_saidas_realizadas=0,
                             total_entradas_a_vencer=0,
                             total_saidas_a_vencer=0,
                             total_entradas_vencidas=0,
                             total_saidas_vencidas=0,
                             total_entradas_agendadas=0,
                             total_saidas_agendadas=0,
                             total_entradas_pendentes=0,
                             total_saidas_pendentes=0,
                             saldo_atual=0,
                             saldo_geral=0,
                             saldo_projetado=0,
                             saldo_vencido=0,
                             categorias_receitas={},
                             categorias_despesas={},
                             categorias_entradas={},
                             categorias_saidas={},
                             contas_caixa=[],
                             contas_caixa_detalhadas=[],
                             saldo_total_caixa=0,
                             sum_rec_a_vencer=0,
                             sum_desp_a_vencer=0,
                             sum_rec_vencido=0,
                             sum_desp_vencido=0,
                             sum_saldo_projetado_contas=0)

@app.route('/relatorios/saldos/exportar/<formato>')
def exportar_relatorio_saldos(formato):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    # Buscar todos os usuários da mesma empresa
    empresa_id = obter_empresa_id_sessao(session, usuario)

    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]
    
    # Obter filtros de data do request
    data_inicio_str = request.args.get('data_inicio', '')
    data_fim_str = request.args.get('data_fim', '')
    
    # Converter strings de data para objetos datetime
    data_inicio = None
    data_fim = None
    
    if data_inicio_str:
        try:
            data_inicio = datetime.strptime(data_inicio_str, '%d/%m/%Y').date()
        except ValueError:
            pass
    
    if data_fim_str:
        try:
            data_fim = datetime.strptime(data_fim_str, '%d/%m/%Y').date()
        except ValueError:
            pass
    
    # Construir query base
    query = Lancamento.query.filter(Lancamento.empresa_id == empresa_id)
    
    # Aplicar filtros de data se fornecidos
    if data_inicio:
        query = query.filter(Lancamento.data_prevista >= data_inicio)
    if data_fim:
        query = query.filter(Lancamento.data_prevista <= data_fim)
    
    # Carregar lançamentos
    lancamentos = query.all()

    hoje = datetime.now().date()

    # Totais por status
    total_receitas_realizadas = sum([l.valor for l in lancamentos if l.tipo == 'entrada' and l.realizado and (not l.data_realizada or l.data_realizada <= hoje)])
    total_despesas_realizadas = sum([l.valor for l in lancamentos if l.tipo == 'saida' and l.realizado and (not l.data_realizada or l.data_realizada <= hoje)])
    total_receitas_pendentes = sum([l.valor for l in lancamentos if l.tipo == 'entrada' and not l.realizado])
    total_despesas_pendentes = sum([l.valor for l in lancamentos if l.tipo == 'saida' and not l.realizado])
    # Agendado: tem data_realizada futura OU (tem data_prevista futura e não está realizado)
    total_receitas_agendadas = sum([l.valor for l in lancamentos if l.tipo == 'entrada' and (
        (l.data_realizada and l.data_realizada > hoje) or 
        (not l.realizado and l.data_prevista and l.data_prevista > hoje)
    )])
    total_despesas_agendadas = sum([l.valor for l in lancamentos if l.tipo == 'saida' and (
        (l.data_realizada and l.data_realizada > hoje) or 
        (not l.realizado and l.data_prevista and l.data_prevista > hoje)
    )])

    total_entradas = total_receitas_realizadas
    total_saidas = total_despesas_realizadas
    saldo_atual = total_entradas - total_saidas
    
    # Buscar categorias válidas do plano de contas
    categorias_validas = db.session.query(PlanoConta.nome).filter(PlanoConta.ativo == True).all()
    categorias_validas_nomes = [cat[0] for cat in categorias_validas]
    
    # Agrupar por categoria (apenas categorias válidas do plano de contas)
    categorias_receitas = {}
    categorias_despesas = {}
    
    for lancamento in lancamentos:
        categoria = lancamento.categoria
        # Só incluir se a categoria estiver no plano de contas
        if categoria and categoria in categorias_validas_nomes:
            if lancamento.tipo == 'entrada':
                if categoria not in categorias_receitas:
                    categorias_receitas[categoria] = {'realizado': 0, 'pendente': 0, 'agendado': 0}
                # Realizado: tem data_realizada <= hoje OU (realizado=True E sem data_realizada)
                if (lancamento.data_realizada and lancamento.data_realizada <= hoje) or (lancamento.realizado and not lancamento.data_realizada):
                    categorias_receitas[categoria]['realizado'] += lancamento.valor
                # Agendado: tem data_realizada futura OU (tem data_prevista futura e não está realizado)
                elif (lancamento.data_realizada and lancamento.data_realizada > hoje) or (not lancamento.realizado and lancamento.data_prevista and lancamento.data_prevista > hoje):
                    categorias_receitas[categoria]['agendado'] += lancamento.valor
                # Pendente: não realizado
                else:
                    categorias_receitas[categoria]['pendente'] += lancamento.valor
            else:  # saida
                if categoria not in categorias_despesas:
                    categorias_despesas[categoria] = {'realizado': 0, 'pendente': 0, 'agendado': 0}
                # Realizado: tem data_realizada <= hoje OU (realizado=True E sem data_realizada)
                if (lancamento.data_realizada and lancamento.data_realizada <= hoje) or (lancamento.realizado and not lancamento.data_realizada):
                    categorias_despesas[categoria]['realizado'] += lancamento.valor
                # Agendado: tem data_realizada futura OU (tem data_prevista futura e não está realizado)
                elif (lancamento.data_realizada and lancamento.data_realizada > hoje) or (not lancamento.realizado and lancamento.data_prevista and lancamento.data_prevista > hoje):
                    categorias_despesas[categoria]['agendado'] += lancamento.valor
                # Pendente: não realizado
                else:
                    categorias_despesas[categoria]['pendente'] += lancamento.valor
    
    # Preparar dados para exportação
    dados = {
        'total_entradas': total_entradas,
        'total_saidas': total_saidas,
        'total_receitas_pendentes': total_receitas_pendentes,
        'total_despesas_pendentes': total_despesas_pendentes,
        'total_receitas_agendadas': total_receitas_agendadas,
        'total_despesas_agendadas': total_despesas_agendadas,
        'saldo_geral': saldo_atual,
        'categorias_receitas': categorias_receitas,
        'categorias_despesas': categorias_despesas
    }
    
    titulo = f"Relatório de Saldos"
    
    # Preparar filtros para o cabeçalho
    filtros = {
        'data_inicio': data_inicio_str,
        'data_fim': data_fim_str
    }
    
    if formato.lower() == 'excel':
        arquivo = exportar_relatorio_excel(dados, 'relatorio_saldos', titulo, usuario, filtros)
        if arquivo:
            return send_file(arquivo, as_attachment=True, download_name=f"relatorio_saldos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    elif formato.lower() == 'pdf':
        arquivo = exportar_relatorio_pdf(dados, 'relatorio_saldos', titulo, usuario, filtros)
        if arquivo:
            return send_file(arquivo, as_attachment=True, download_name=f"relatorio_saldos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
    
    flash('Erro ao exportar relatório.', 'error')
    return redirect(url_for('relatorio_saldos'))

@app.route('/relatorios/lancamentos')
def relatorio_lancamentos():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    # Buscar todos os usuários da mesma empresa
    empresa_id = obter_empresa_id_sessao(session, usuario)

    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]
    
    # Obter filtros do request (conforme requisitos)
    tipo_filtro = request.args.get('tipo', '')
    categoria_filtro = request.args.get('categoria', '')
    status_filtro = request.args.get('status', '')
    cliente_filtro = request.args.get('cliente', '')
    fornecedor_filtro = request.args.get('fornecedor', '')
    descricao_filtro = request.args.get('descricao', '')
    valor_min_str = request.args.get('valor_min', '')
    valor_max_str = request.args.get('valor_max', '')
    data_inicio_str = request.args.get('data_inicio', '')
    data_fim_str = request.args.get('data_fim', '')
    
    # Converter strings de data para objetos datetime
    data_inicio = None
    data_fim = None
    
    if data_inicio_str:
        try:
            data_inicio = datetime.strptime(data_inicio_str, '%d/%m/%Y').date()
        except ValueError:
            flash('Data de início inválida. Use o formato DD/MM/AAAA.', 'error')
            data_inicio_str = ''
    
    if data_fim_str:
        try:
            data_fim = datetime.strptime(data_fim_str, '%d/%m/%Y').date()
        except ValueError:
            flash('Data de fim inválida. Use o formato DD/MM/AAAA.', 'error')
            data_fim_str = ''
    
    # Construir query base
    query = Lancamento.query.options(
        joinedload(Lancamento.compra).joinedload(Compra.fornecedor),
        joinedload(Lancamento.venda).joinedload(Venda.cliente)
    ).filter(Lancamento.empresa_id == empresa_id)
    
    # Aplicar filtros
    if tipo_filtro:
        query = query.filter(Lancamento.tipo == tipo_filtro)
    
    if categoria_filtro:
        query = query.filter(Lancamento.categoria == categoria_filtro)
    
    if status_filtro == 'realizado':
        query = query.filter(Lancamento.realizado == True)
    elif status_filtro == 'pendente':
        query = query.filter(Lancamento.realizado == False)
    
    if data_inicio:
        query = query.filter(Lancamento.data_prevista >= data_inicio)
    if data_fim:
        query = query.filter(Lancamento.data_prevista <= data_fim)
    
    # Aplicar filtros de cliente e fornecedor
    if cliente_filtro:
        # Buscar lançamentos que têm venda vinculada com cliente específico
        query = query.join(Venda, Lancamento.venda_id == Venda.id).join(Cliente, Venda.cliente_id == Cliente.id).filter(Cliente.nome.ilike(f'%{cliente_filtro}%'))

    if fornecedor_filtro:
        # Buscar lançamentos que têm compra vinculada com fornecedor específico
        query = query.join(Compra, Lancamento.compra_id == Compra.id).join(Fornecedor, Compra.fornecedor_id == Fornecedor.id).filter(Fornecedor.nome.ilike(f'%{fornecedor_filtro}%'))

    # Filtro por busca geral (descrição ou valor)
    busca = request.args.get('busca', '').strip()
    if busca:
        termo = f"%{busca}%"
        try:
            valor_busca = float(busca.replace(',', '.'))
            query = query.filter(or_(
                Lancamento.descricao.ilike(termo),
                Lancamento.valor == valor_busca,
                Lancamento.produto_servico.ilike(termo)
            ))
        except ValueError:
            query = query.filter(or_(
                Lancamento.descricao.ilike(termo),
                Lancamento.produto_servico.ilike(termo)
            ))

    # Aplicar filtro de descrição (conforme requisitos)
    if descricao_filtro:
        query = query.filter(Lancamento.descricao.ilike(f'%{descricao_filtro}%'))

    # Aplicar filtro de valor (conforme requisitos)
    if valor_min_str:
        try:
            valor_min = float(valor_min_str.replace(',', '.'))
            query = query.filter(Lancamento.valor >= valor_min)
        except ValueError:
            pass

    if valor_max_str:
        try:
            valor_max = float(valor_max_str.replace(',', '.'))
            query = query.filter(Lancamento.valor <= valor_max)
        except ValueError:
            pass
    
    # Carregar lançamentos
    lancamentos = query.order_by(Lancamento.data_prevista.desc()).all()
    
    # Calcular totais - apenas lançamentos realizados
    total_entradas = sum([l.valor for l in lancamentos if l.tipo == 'entrada' and l.realizado])
    total_saidas = sum([l.valor for l in lancamentos if l.tipo == 'saida' and l.realizado])
    saldo_atual = total_entradas - total_saidas
    
    # Extrair categorias únicas
    categorias = list(set([l.categoria for l in lancamentos]))
    categorias.sort()
    
    # Criar objeto de filtros para o template (com novos filtros)
    filtros = type('Filtros', (), {
        'tipo': tipo_filtro,
        'categoria': categoria_filtro,
        'data_inicio': data_inicio_str,
        'data_fim': data_fim_str,
        'status': status_filtro,
        'cliente': cliente_filtro,
        'fornecedor': fornecedor_filtro,
        'descricao': descricao_filtro,
        'valor_min': valor_min_str,
        'valor_max': valor_max_str
    })()
    
    return render_template('relatorio_lancamentos.html', 
                         usuario=usuario,
                         lancamentos=lancamentos,
                         total_entradas=total_entradas,
                         total_saidas=total_saidas,
                         saldo_atual=saldo_atual,
                         categorias=categorias,
                         filtros=filtros)

@app.route('/relatorios/lancamentos/exportar/<formato>')
def exportar_relatorio_lancamentos(formato):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    # Buscar todos os usuários da mesma empresa
    empresa_id = obter_empresa_id_sessao(session, usuario)

    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]
    
    # Obter filtros
    tipo_filtro = request.args.get('tipo', '')
    categoria_filtro = request.args.get('categoria', '')
    status_filtro = request.args.get('status', '')
    data_inicio_str = request.args.get('data_inicio', '')
    data_fim_str = request.args.get('data_fim', '')
    
    # Construir query base
    query = Lancamento.query.filter(Lancamento.empresa_id == empresa_id)
    
    # Aplicar filtros
    if tipo_filtro:
        query = query.filter(Lancamento.tipo == tipo_filtro)
    if categoria_filtro:
        query = query.filter(Lancamento.categoria == categoria_filtro)
    if status_filtro:
        if status_filtro == 'realizado':
            query = query.filter(Lancamento.realizado == True)
        elif status_filtro == 'pendente':
            query = query.filter(Lancamento.realizado == False)
    
    # Aplicar filtros de data
    if data_inicio_str:
        try:
            data_inicio = datetime.strptime(data_inicio_str, '%d/%m/%Y').date()
            query = query.filter(Lancamento.data_prevista >= data_inicio)
        except ValueError:
            pass
    
    if data_fim_str:
        try:
            data_fim = datetime.strptime(data_fim_str, '%d/%m/%Y').date()
            query = query.filter(Lancamento.data_prevista <= data_fim)
        except ValueError:
            pass
    
    # Carregar lançamentos com relacionamentos
    lancamentos = query.options(joinedload(Lancamento.cliente), joinedload(Lancamento.fornecedor)).all()
    
    # Preparar dados para exportação
    dados = {'lancamentos': lancamentos}
    
    titulo = f"Relatório de Lançamentos"
    
    # Preparar filtros para o cabeçalho
    filtros = {
        'data_inicio': data_inicio_str,
        'data_fim': data_fim_str,
        'tipo': tipo_filtro,
        'categoria': categoria_filtro,
        'status': status_filtro
    }
    
    if formato.lower() == 'excel':
        arquivo = exportar_relatorio_excel(dados, 'relatorio_lancamentos', titulo, usuario, filtros)
        if arquivo:
            return send_file(arquivo, as_attachment=True, download_name=f"relatorio_lancamentos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    elif formato.lower() == 'pdf':
        arquivo = exportar_relatorio_pdf(dados, 'relatorio_lancamentos', titulo, usuario, filtros)
        if arquivo:
            return send_file(arquivo, as_attachment=True, download_name=f"relatorio_lancamentos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
    
    flash('Erro ao exportar relatório.', 'error')
    return redirect(url_for('relatorio_lancamentos'))

@app.route('/relatorios/clientes')
def relatorio_clientes():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario:
        flash('Usuário não encontrado', 'error')
        return redirect(url_for('login'))
    
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    try:
        app.logger.info(f"Iniciando relatório de clientes para usuário {usuario.id}")
        
        # Obter parâmetros de filtro com validação
        filtro_status = request.args.get('status', 'todos')
        if filtro_status not in ['todos', 'com_saldo', 'sem_saldo']:
            filtro_status = 'todos'
            
        filtro_ordenacao = request.args.get('ordenacao', 'nome')
        if filtro_ordenacao not in ['nome', 'valor_total', 'saldo_aberto']:
            filtro_ordenacao = 'nome'
            
        filtro_periodo = request.args.get('periodo', 'todos')
        if filtro_periodo not in ['todos', 'mes_atual', 'trimestre', 'ano_atual', 'personalizado']:
            filtro_periodo = 'todos'
            
        # Obter datas personalizadas se período for personalizado
        filtro_data_inicio = request.args.get('data_inicio', '').strip()
        filtro_data_fim = request.args.get('data_fim', '').strip()
        # Novos filtros avançados
        filtro_categoria = request.args.get('categoria', '').strip()
        filtro_status_avancado = request.args.get('status_avancado', 'todos')
        filtro_busca = request.args.get('busca', '').strip()
        
        # Obter todos os usuários da empresa vinculada
            
        exportar = request.args.get('exportar', '')
        pagina = request.args.get('pagina', 1, type=int)
        # por_pagina=0 significa "mostrar todos"
        por_pagina = request.args.get('por_pagina', 0, type=int)
        
        # Validar parâmetros
        if pagina < 1:
            pagina = 1
        if por_pagina < 0 or por_pagina > 100000:
            por_pagina = 0
        
        app.logger.info(f"Parâmetros: status={filtro_status}, ordenacao={filtro_ordenacao}, periodo={filtro_periodo}")

        # Obter todos os usuários da empresa vinculada (considerando acesso contador)
        empresa_id = obter_empresa_id_sessao(session, usuario)
        usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
        usuarios_ids = [u.id for u in usuarios_empresa]
        
        if not usuarios_ids:
            flash('Nenhum usuário ativo encontrado na empresa', 'warning')
            return render_template('relatorio_clientes.html', 
                                 usuario=usuario, 
                                 clientes_dados=[],
                                 filtro_status=filtro_status,
                                 filtro_ordenacao=filtro_ordenacao,
                                 filtro_periodo=filtro_periodo,
                                 pagina=pagina,
                                 por_pagina=por_pagina,
                                 total_paginas=0,
                                 total_clientes=0)
        
        # Buscar clientes da empresa (filtro direto por empresa_id)
        clientes = Cliente.query.filter(Cliente.empresa_id == empresa_id).all()
        
        if not clientes:
            flash('Nenhum cliente encontrado', 'warning')
            return render_template('relatorio_clientes.html',
                                 usuario=usuario,
                                 clientes_dados=[],
                                 filtro_status=filtro_status,
                                 filtro_ordenacao=filtro_ordenacao,
                                 filtro_periodo=filtro_periodo,
                                 filtro_data_inicio=filtro_data_inicio,
                                 filtro_data_fim=filtro_data_fim,
                                 pagina=pagina,
                                 por_pagina=por_pagina,
                                 total_paginas=0,
                                 total_clientes=0,
                                 sum_total_vendas=0,
                                 sum_total_vendas_pendentes=0,
                                 sum_saldo_vencido=0,
                                 sum_total_agendado=0,
                                 sum_saldo_aberto=0,
                                 sum_total_geral=0,
                                 sum_num_vendas=0,
                                 sum_ticket_medio=0)
        
        # Lista para armazenar dados dos clientes
        clientes_dados = []
        
        # Definir período para filtros
        hoje = datetime.now().date()
        if filtro_periodo == 'mes_atual':
            inicio_periodo = datetime(hoje.year, hoje.month, 1).date()
            fim_periodo = hoje
        elif filtro_periodo == 'trimestre':
            trimestre = (hoje.month - 1) // 3
            inicio_periodo = datetime(hoje.year, trimestre * 3 + 1, 1).date()
            fim_periodo = hoje
        elif filtro_periodo == 'ano_atual':
            inicio_periodo = datetime(hoje.year, 1, 1).date()
            fim_periodo = hoje
        elif filtro_periodo == 'personalizado':
            # Processar datas personalizadas
            try:
                if filtro_data_inicio:
                    # Converter DD/MM/AAAA para AAAA-MM-DD
                    partes_inicio = filtro_data_inicio.split('/')
                    if len(partes_inicio) == 3:
                        inicio_periodo = datetime(int(partes_inicio[2]), int(partes_inicio[1]), int(partes_inicio[0])).date()
                    else:
                        inicio_periodo = None
                else:
                    inicio_periodo = None
                    
                if filtro_data_fim:
                    # Converter DD/MM/AAAA para AAAA-MM-DD
                    partes_fim = filtro_data_fim.split('/')
                    if len(partes_fim) == 3:
                        fim_periodo = datetime(int(partes_fim[2]), int(partes_fim[1]), int(partes_fim[0])).date()
                    else:
                        fim_periodo = None
                else:
                    fim_periodo = None
            except (ValueError, IndexError):
                flash('Datas inválidas no filtro personalizado', 'error')
                inicio_periodo = None
                fim_periodo = None
        else:
            inicio_periodo = None
            fim_periodo = None
        
        app.logger.info(f"Período: {inicio_periodo} a {fim_periodo}")
        
        # Calcular dados de cada cliente
        for cliente in clientes:
            try:
                # app.logger.info(f"Processando cliente {cliente.id}: {cliente.nome}")
                
                # Buscar lançamentos financeiros do cliente considerando também vínculos por venda
                # Inclui registros com Lancamento.cliente_id diretamente ou via Venda.cliente_id
                lancamentos_query = (
                    Lancamento.query
                    .outerjoin(Venda, Lancamento.venda_id == Venda.id)
                    .filter(
                        Lancamento.empresa_id == empresa_id,
                        or_(Lancamento.cliente_id == cliente.id, Venda.cliente_id == cliente.id)
                    )
                )
                
                # Aplicar filtro de período se especificado
                if inicio_periodo and fim_periodo:
                    lancamentos_query = lancamentos_query.filter(
                        Lancamento.data_prevista >= inicio_periodo,
                        Lancamento.data_prevista <= fim_periodo
                    )

                # Aplicar filtro por categoria (se fornecido)
                if filtro_categoria:
                    lancamentos_query = lancamentos_query.filter(Lancamento.categoria == filtro_categoria)

                # Se houver busca, filtrar estritamente por nome ou documento
                if filtro_busca:
                     termo_l = filtro_busca.lower()
                     combina_nome = (cliente.nome and termo_l in cliente.nome.lower()) or \
                                    (cliente.cpf_cnpj and termo_l in cliente.cpf_cnpj.lower())
                     if not combina_nome:
                         continue
                
                # Aplicar filtro de status avançado
                hoje_status = datetime.now().date()
                if filtro_status_avancado == 'realizado':
                    lancamentos_query = lancamentos_query.filter(Lancamento.realizado == True)
                elif filtro_status_avancado == 'pendente':
                    lancamentos_query = lancamentos_query.filter(Lancamento.realizado == False)
                elif filtro_status_avancado == 'agendado':
                    # Agendado: data_prevista no futuro e não realizado
                    lancamentos_query = lancamentos_query.filter(Lancamento.realizado == False, Lancamento.data_prevista > hoje_status)
                elif filtro_status_avancado == 'vencido':
                    # Vencido: não realizado e data_prevista no passado
                    lancamentos_query = lancamentos_query.filter(Lancamento.realizado == False, Lancamento.data_prevista < hoje_status)
                
                lancamentos_cliente = lancamentos_query.all()
                app.logger.info(f"CLIENTE_FILTER: {cliente.nome} - lancamentos: {len(lancamentos_cliente)} - busca: {filtro_busca}")
                
                app.logger.info(f"Lançamentos financeiros encontrados para cliente {cliente.id}: {len(lancamentos_cliente)}")
                
                # Calcular totais baseados na movimentação financeira real (apenas entradas)
                lancamentos_receita = [l for l in lancamentos_cliente if l.tipo == 'entrada']

                # Vendas realizadas = lançamentos de entrada marcados como realizados
                total_vendas = sum(l.valor or 0 for l in lancamentos_receita if l.realizado)
                # A vencer: não realizado e data_prevista futura (estritamente > hoje)
                total_a_vencer = sum(
                    (l.valor or 0) for l in lancamentos_receita
                    if (not l.realizado) and (l.data_prevista is not None and l.data_prevista > hoje)
                )
                # Saldo vencido: não realizado e data_prevista passada
                saldo_vencido = sum(
                    (l.valor or 0) for l in lancamentos_receita
                    if (not l.realizado) and (l.data_prevista is not None and l.data_prevista < hoje)
                )
                # Agendado: não realizado e data_prevista = hoje OU sem data_prevista
                total_agendado = sum(
                    (l.valor or 0) for l in lancamentos_receita
                    if (not l.realizado) and ((l.data_prevista is not None and l.data_prevista == hoje) or l.data_prevista is None)
                )
                # Número de transações realizadas
                num_vendas = len([l for l in lancamentos_receita if l.realizado])
                # Número de transações pendentes
                num_vendas_pendentes = len([l for l in lancamentos_receita if not l.realizado])
                # Saldo em aberto = a vencer + vencido + agendado
                saldo_aberto = total_a_vencer + saldo_vencido + total_agendado

                # Total geral (realizadas + pendentes)
                total_geral = total_vendas + saldo_aberto

                # Aplicar filtro de status
                if filtro_status == 'com_saldo' and saldo_aberto <= 0:
                    app.logger.info(f"Cliente {cliente.id} filtrado por status (sem saldo)")
                    continue
                elif filtro_status == 'sem_saldo' and saldo_aberto > 0:
                    app.logger.info(f"Cliente {cliente.id} filtrado por status (com saldo)")
                    continue

                # Calcular ticket médio (baseado no total geral, não apenas nas realizadas)
                ticket_medio = total_geral / (num_vendas + num_vendas_pendentes) if (num_vendas + num_vendas_pendentes) > 0 else 0

                cliente_data = {
                    'cliente': cliente,
                    'total_vendas': total_vendas,
                    'total_vendas_pendentes': total_a_vencer,
                    'saldo_vencido': saldo_vencido,
                    'total_agendado': total_agendado,
                    'num_vendas': num_vendas + num_vendas_pendentes,  # Total de transações
                    'num_vendas_pendentes': num_vendas_pendentes,
                    'total_geral': total_geral,
                    'saldo_aberto': saldo_aberto,
                    'ticket_medio': ticket_medio
                }
                clientes_dados.append(cliente_data)
                
            except Exception as e:
                app.logger.error(f"Erro ao processar cliente {cliente.id}: {str(e)}")
                continue
        
        app.logger.info(f"Clientes processados: {len(clientes_dados)}")
        
        # Aplicar ordenação com validação
        try:
            if filtro_ordenacao == 'nome':
                clientes_dados.sort(key=lambda x: x['cliente'].nome.lower() if x['cliente'].nome else '')
            elif filtro_ordenacao == 'valor_total':
                clientes_dados.sort(key=lambda x: x['total_geral'], reverse=True)
            elif filtro_ordenacao == 'saldo_aberto':
                clientes_dados.sort(key=lambda x: x['saldo_aberto'], reverse=True)
        except Exception as e:
            app.logger.error(f"Erro na ordenação: {str(e)}")
            # Ordenação padrão por nome em caso de erro
            clientes_dados.sort(key=lambda x: x['cliente'].nome.lower() if x['cliente'].nome else '')
        
        # Calcular paginação
        total_clientes = len(clientes_dados)
        if por_pagina == 0:
            # Sem paginação: mostrar todos
            total_paginas = 1 if total_clientes > 0 else 0
            clientes_paginados = clientes_dados
        else:
            total_paginas = (total_clientes + por_pagina - 1) // por_pagina if total_clientes > 0 else 0
            # Aplicar paginação
            inicio = (pagina - 1) * por_pagina
            fim = inicio + por_pagina
            clientes_paginados = clientes_dados[inicio:fim]
        
        app.logger.info(f"Paginação: {pagina}/{total_paginas}, {len(clientes_paginados)} clientes")

        # Verificar se é exportação
        if exportar == 'pdf':
            return exportar_relatorio_clientes_pdf(clientes_dados)
        elif exportar == 'excel':
            return exportar_relatorio_clientes_excel(clientes_dados)

        app.logger.info("Relatório de clientes gerado com sucesso")

        # Pré-calcular totais para o template (evita problemas com Jinja sum + dict)
        sum_total_vendas = sum(c.get('total_vendas', 0) for c in clientes_paginados)
        sum_total_vendas_pendentes = sum(c.get('total_vendas_pendentes', 0) for c in clientes_paginados)
        sum_saldo_vencido = sum(c.get('saldo_vencido', 0) for c in clientes_paginados)
        sum_total_agendado = sum(c.get('total_agendado', 0) for c in clientes_paginados)
        sum_saldo_aberto = sum(c.get('saldo_aberto', 0) for c in clientes_paginados)
        sum_total_geral = sum(c.get('total_geral', 0) for c in clientes_paginados)
        sum_num_vendas = sum(c.get('num_vendas', 0) for c in clientes_paginados)
        sum_ticket_medio = sum_total_geral / sum_num_vendas if sum_num_vendas > 0 else 0

        return render_template('relatorio_clientes.html',
                             usuario=usuario,
                             clientes_dados=clientes_paginados,
                             filtro_status=filtro_status,
                             filtro_ordenacao=filtro_ordenacao,
                             filtro_periodo=filtro_periodo,
                             filtro_data_inicio=filtro_data_inicio,
                             filtro_data_fim=filtro_data_fim,
                             pagina=pagina,
                             por_pagina=por_pagina,
                             total_paginas=total_paginas,
                             total_clientes=total_clientes,
                             sum_total_vendas=sum_total_vendas,
                             sum_total_vendas_pendentes=sum_total_vendas_pendentes,
                             sum_saldo_vencido=sum_saldo_vencido,
                             sum_total_agendado=sum_total_agendado,
                             sum_saldo_aberto=sum_saldo_aberto,
                             sum_total_geral=sum_total_geral,
                             sum_num_vendas=sum_num_vendas,
                             sum_ticket_medio=sum_ticket_medio)
    
    except Exception as e:
        app.logger.error(f"Erro no relatório de clientes: {str(e)}")
        flash('Erro ao gerar relatório de clientes', 'error')
        return render_template('relatorio_clientes.html',
                             usuario=usuario,
                             clientes_dados=[],
                             filtro_status='todos',
                             filtro_ordenacao='nome',
                             filtro_periodo='todos',
                             filtro_data_inicio='',
                             filtro_data_fim='',
                             pagina=1,
                             por_pagina=20,
                             total_paginas=0,
                             total_clientes=0,
                             sum_total_vendas=0,
                             sum_total_vendas_pendentes=0,
                             sum_saldo_vencido=0,
                             sum_total_agendado=0,
                             sum_saldo_aberto=0,
                             sum_total_geral=0,
                             sum_num_vendas=0,
                             sum_ticket_medio=0)

@app.route('/relatorios/clientes/exportar/<formato>')
def exportar_relatorio_clientes(formato):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    # Buscar todos os usuários da mesma empresa
    empresa_id = obter_empresa_id_sessao(session, usuario)

    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]
    
    # Obter filtros
    filtro_status = request.args.get('status', 'todos')
    filtro_periodo = request.args.get('periodo', 'todos')
    filtro_ordenacao = request.args.get('ordenacao', 'nome')
    
    # Buscar clientes
    query = Cliente.query.filter(Cliente.empresa_id == empresa_id, Cliente.ativo == True)
    
    # Aplicar filtro de status
    if filtro_status == 'com_saldo':
        # Filtrar apenas clientes com saldo > 0
        pass  # Será aplicado após calcular os saldos
    
    clientes = query.all()
    
    # Processar dados dos clientes
    clientes_dados = []
    for cliente in clientes:
        try:
            # Buscar lançamentos do cliente
            lancamentos = Lancamento.query.filter(
                Lancamento.cliente_id == cliente.id,
                Lancamento.empresa_id == empresa_id
            ).all()
            
            # Calcular totais
            total_entradas = sum([l.valor for l in lancamentos if l.tipo == 'entrada'])
            total_saidas = sum([l.valor for l in lancamentos if l.tipo == 'saida'])
            saldo_aberto = total_entradas - total_saidas
            
            # Aplicar filtro de status após calcular saldo
            if filtro_status == 'com_saldo' and saldo_aberto <= 0:
                continue
            elif filtro_status == 'sem_saldo' and saldo_aberto > 0:
                continue
            
            clientes_dados.append({
                'cliente': cliente,
                'total_entradas': total_entradas,
                'total_saidas': total_saidas,
                'total_geral': total_entradas + total_saidas,
                'saldo_aberto': saldo_aberto,
                'quantidade_lancamentos': len(lancamentos)
            })
            
        except Exception as e:
            app.logger.error(f"Erro ao processar cliente {cliente.id}: {str(e)}")
            continue
    
    # Preparar dados para exportação
    dados = {'clientes_dados': clientes_dados}
    
    titulo = f"Relatório de Clientes"
    
    # Preparar filtros para o cabeçalho
    filtros = {
        'status': filtro_status,
        'periodo': filtro_periodo,
        'ordenacao': filtro_ordenacao
    }
    
    if formato.lower() == 'excel':
        arquivo = exportar_relatorio_excel(dados, 'relatorio_clientes', titulo, usuario, filtros)
        if arquivo:
            return send_file(arquivo, as_attachment=True, download_name=f"relatorio_clientes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    elif formato.lower() == 'pdf':
        arquivo = exportar_relatorio_pdf(dados, 'relatorio_clientes', titulo, usuario, filtros)
        if arquivo:
            return send_file(arquivo, as_attachment=True, download_name=f"relatorio_clientes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
    
    flash('Erro ao exportar relatório.', 'error')
    return redirect(url_for('relatorio_clientes'))

@app.route('/relatorios/produtos')
def relatorio_produtos():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))

    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))

    # Verificar se a empresa é do tipo comércio (conforme requisitos)
    empresa = usuario.empresa
    if empresa.tipo_empresa != 'comercio':
        flash('Relatório de Produtos disponível apenas para empresas do tipo Comércio.', 'warning')
        return redirect(url_for('relatorios'))

    # Obter filtros do request (conforme requisitos)
    data_inicio_str = request.args.get('data_inicio', '')
    data_fim_str = request.args.get('data_fim', '')

    # Processar filtros de data
    data_inicio = None
    data_fim = None

    if data_inicio_str:
        try:
            data_inicio = datetime.strptime(data_inicio_str, '%d/%m/%Y').date()
        except ValueError:
            flash('Data de início inválida.', 'error')

    if data_fim_str:
        try:
            data_fim = datetime.strptime(data_fim_str, '%d/%m/%Y').date()
        except ValueError:
            flash('Data de fim inválida.', 'error')

    # Buscar todas as vendas da empresa
    query = db.session.query(Venda).filter(
        Venda.usuario_id == usuario.id
    )

    # Aplicar filtro de período (conforme requisitos)
    if data_inicio:
        query = query.filter(Venda.data_prevista >= data_inicio)
    if data_fim:
        query = query.filter(Venda.data_prevista <= data_fim)

    vendas = query.all()

    # Agrupar vendas por produto e calcular totais (conforme requisitos)
    produtos_vendidos = {}

    for venda in vendas:
        produto_nome = venda.produto
        quantidade = venda.quantidade or 1
        valor_total = venda.valor_final or venda.valor

        if produto_nome not in produtos_vendidos:
            produtos_vendidos[produto_nome] = {
                'nome': produto_nome,
                'total_vendas': 0,
                'quantidade_total': 0,
                'valor_total': 0
            }

        produtos_vendidos[produto_nome]['total_vendas'] += 1
        produtos_vendidos[produto_nome]['quantidade_total'] += quantidade
        produtos_vendidos[produto_nome]['valor_total'] += valor_total

    # Converter para lista e ordenar por total de vendas (conforme requisitos: mais vendidos primeiro)
    produtos_lista = list(produtos_vendidos.values())
    produtos_lista.sort(key=lambda x: x['total_vendas'], reverse=True)

    # Criar objeto de filtros para o template
    filtros = type('Filtros', (), {
        'data_inicio': data_inicio_str,
        'data_fim': data_fim_str
    })()

    return render_template('relatorio_produtos.html',
                         usuario=usuario,
                         produtos=produtos_lista,
                         filtros=filtros)

@app.route('/relatorios/produtos/exportar/<formato>')
def exportar_relatorio_produtos(formato):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    try:
        # Buscar todos os produtos da empresa
        produtos = buscar_produtos_empresa(usuario.empresa_id)
        
        # Preparar dados para exportação
        dados = {'produtos': produtos}
        
        titulo = f"Relatório de Produtos/Estoque"
        
        # Preparar filtros para o cabeçalho
        filtros = {
            'periodo': 'todos'
        }
        
        if formato.lower() == 'excel':
            arquivo = exportar_relatorio_excel(dados, 'relatorio_produtos', titulo, usuario, filtros)
            if arquivo:
                return send_file(arquivo, as_attachment=True, download_name=f"relatorio_produtos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        elif formato.lower() == 'pdf':
            arquivo = exportar_relatorio_pdf(dados, 'relatorio_produtos', titulo, usuario, filtros)
            if arquivo:
                return send_file(arquivo, as_attachment=True, download_name=f"relatorio_produtos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao exportar relatório de produtos: {str(e)}")
        flash('Erro ao exportar relatório.', 'error')
        return redirect(url_for('relatorio_produtos'))
    
    flash('Erro ao exportar relatório.', 'error')
    return redirect(url_for('relatorio_produtos'))

@app.route('/relatorios/estoque')
def relatorio_estoque():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))

    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))

    # Verificar se a empresa é do tipo comércio (conforme requisitos)
    empresa = usuario.empresa
    if empresa.tipo_empresa != 'comercio':
        flash('Relatório de Estoque disponível apenas para empresas do tipo Comércio.', 'warning')
        return redirect(url_for('relatorios'))

    # Obter filtro de data (conforme requisitos: data específica para cálculo temporal)
    data_filtro_str = request.args.get('data_filtro', '')

    # Processar filtro de data
    data_filtro = None
    usar_data_atual = True

    if data_filtro_str:
        try:
            data_filtro = datetime.strptime(data_filtro_str, '%d/%m/%Y').date()
            usar_data_atual = False
        except ValueError:
            flash('Data inválida.', 'error')

    # Se não há filtro de data ou data é hoje, usar estoque atual
    hoje = datetime.now().date()
    if not data_filtro or data_filtro >= hoje:
        data_filtro = hoje
        usar_data_atual = True

    # Buscar todos os produtos da empresa
    produtos = Produto.query.filter_by(usuario_id=usuario.id).all()

    # Buscar todos os usuários da empresa para multi-tenancy
    empresa_id = obter_empresa_id_sessao(session, usuario)
    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]

    # Calcular estoque para cada produto (temporal ou atual)
    estoque_lista = []

    for produto in produtos:
        # Buscar todas as compras deste produto (para calcular preço médio)
        compras_produto = Compra.query.filter(
            Compra.empresa_id == empresa_id,
            Compra.produto == produto.nome,
            Compra.tipo_compra == 'mercadoria'
        )
        if not usar_data_atual and data_filtro:
            compras_produto = compras_produto.filter(Compra.data_prevista <= data_filtro)
        compras_produto = compras_produto.all()

        # Calcular preço médio de compra ponderado por quantidade
        total_custo_compras = sum((c.preco_custo or c.valor or 0) * (c.quantidade or 1) for c in compras_produto)
        total_qtd_compras = sum(c.quantidade or 1 for c in compras_produto)
        preco_medio_compra = total_custo_compras / total_qtd_compras if total_qtd_compras > 0 else (produto.preco_custo or 0)

        if usar_data_atual:
            # Usar estoque atual do produto (já calculado)
            estoque_atual = produto.estoque or 0
        else:
            quantidade_comprada = sum([c.quantidade or 0 for c in compras_produto])

            # Vendas (saída): somar quantidade de vendas até data_filtro
            vendas = Venda.query.filter(
                Venda.empresa_id == empresa_id,
                Venda.produto == produto.nome,
                Venda.tipo_venda == 'produto',
                Venda.data_prevista <= data_filtro
            ).all()

            quantidade_vendida = sum([v.quantidade or 0 for v in vendas])

            # Estoque temporal = compras - vendas até a data
            estoque_atual = quantidade_comprada - quantidade_vendida

        # Adicionar à lista apenas se ativo ou tiver estoque
        if produto.ativo or estoque_atual > 0:
            valor_estoque_venda = estoque_atual * (produto.preco_venda or 0)
            valor_estoque_custo = estoque_atual * preco_medio_compra
            margem = ((produto.preco_venda or 0) - preco_medio_compra) / preco_medio_compra * 100 if preco_medio_compra > 0 else 0

            estoque_lista.append({
                'id': produto.id,
                'nome': produto.nome,
                'descricao': produto.descricao,
                'estoque': estoque_atual,
                'preco_custo': produto.preco_custo or 0,
                'preco_medio_compra': preco_medio_compra,
                'preco_venda': produto.preco_venda or 0,
                'valor_estoque_custo': valor_estoque_custo,
                'valor_estoque': valor_estoque_venda,
                'margem': margem,
                'qtd_compras': total_qtd_compras,
                'ativo': produto.ativo
            })

    # Ordenar por nome
    estoque_lista.sort(key=lambda x: x['nome'])

    # Calcular totais
    total_produtos = len(estoque_lista)
    total_quantidade = sum([e['estoque'] for e in estoque_lista])
    total_valor_custo = sum([e['valor_estoque_custo'] for e in estoque_lista])
    total_valor_venda = sum([e['valor_estoque'] for e in estoque_lista])
    total_margem = ((total_valor_venda - total_valor_custo) / total_valor_custo * 100) if total_valor_custo > 0 else 0

    # Criar objeto de filtros para o template
    filtros = type('Filtros', (), {
        'data_filtro': data_filtro_str
    })()

    return render_template('relatorio_estoque.html',
                         usuario=usuario,
                         estoque=estoque_lista,
                         total_produtos=total_produtos,
                         total_quantidade=total_quantidade,
                         total_valor_custo=total_valor_custo,
                         total_valor_venda=total_valor_venda,
                         total_margem=total_margem,
                         data_referencia=data_filtro.strftime('%d/%m/%Y'),
                         usar_data_atual=usar_data_atual,
                         filtros=filtros)

@app.route('/relatorios/estoque/exportar/<formato>')
def exportar_relatorio_estoque(formato):
    """Exporta relatório de estoque em Excel"""
    if 'usuario_id' not in session:
        return redirect(url_for('login'))

    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))

    # Verificar se a empresa é do tipo comércio
    empresa = usuario.empresa
    if empresa.tipo_empresa != 'comercio':
        flash('Relatório de Estoque disponível apenas para empresas do tipo Comércio.', 'warning')
        return redirect(url_for('relatorios'))

    # Obter filtro de data
    data_filtro_str = request.args.get('data_filtro', '')
    data_filtro = None
    usar_data_atual = True

    if data_filtro_str:
        try:
            data_filtro = datetime.strptime(data_filtro_str, '%d/%m/%Y').date()
            usar_data_atual = False
        except ValueError:
            pass

    hoje = datetime.now().date()
    if not data_filtro or data_filtro >= hoje:
        data_filtro = hoje
        usar_data_atual = True

    # Buscar produtos e calcular estoque (mesmo código do relatório)
    produtos = Produto.query.filter_by(usuario_id=usuario.id).all()
    empresa_id = obter_empresa_id_sessao(session, usuario)
    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]

    estoque_lista = []

    for produto in produtos:
        compras_produto = Compra.query.filter(
            Compra.empresa_id == empresa_id,
            Compra.produto == produto.nome,
            Compra.tipo_compra == 'mercadoria'
        )
        if not usar_data_atual and data_filtro:
            compras_produto = compras_produto.filter(Compra.data_prevista <= data_filtro)
        compras_produto = compras_produto.all()

        total_custo_compras = sum((c.preco_custo or c.valor or 0) * (c.quantidade or 1) for c in compras_produto)
        total_qtd_compras = sum(c.quantidade or 1 for c in compras_produto)
        preco_medio_compra = total_custo_compras / total_qtd_compras if total_qtd_compras > 0 else (produto.preco_custo or 0)

        if usar_data_atual:
            estoque_atual = produto.estoque or 0
        else:
            quantidade_comprada = sum([c.quantidade or 0 for c in compras_produto])
            vendas = Venda.query.filter(
                Venda.empresa_id == empresa_id,
                Venda.produto == produto.nome,
                Venda.tipo_venda == 'produto',
                Venda.data_prevista <= data_filtro
            ).all()
            quantidade_vendida = sum([v.quantidade or 0 for v in vendas])
            estoque_atual = quantidade_comprada - quantidade_vendida

        if produto.ativo or estoque_atual > 0:
            valor_estoque_venda = estoque_atual * (produto.preco_venda or 0)
            valor_estoque_custo = estoque_atual * preco_medio_compra
            margem = ((produto.preco_venda or 0) - preco_medio_compra) / preco_medio_compra * 100 if preco_medio_compra > 0 else 0

            estoque_lista.append({
                'nome': produto.nome,
                'descricao': produto.descricao or '',
                'estoque': estoque_atual,
                'preco_custo': produto.preco_custo or 0,
                'preco_medio_compra': preco_medio_compra,
                'preco_venda': produto.preco_venda or 0,
                'valor_estoque_custo': valor_estoque_custo,
                'valor_estoque': valor_estoque_venda,
                'margem': margem,
                'qtd_compras': total_qtd_compras,
            })

    estoque_lista.sort(key=lambda x: x['nome'])

    # Calcular totais
    total_produtos = len(estoque_lista)
    total_quantidade = sum([e['estoque'] for e in estoque_lista])
    total_valor_custo = sum([e['valor_estoque_custo'] for e in estoque_lista])
    total_valor_venda = sum([e['valor_estoque'] for e in estoque_lista])
    total_margem = ((total_valor_venda - total_valor_custo) / total_valor_custo * 100) if total_valor_custo > 0 else 0

    # Gerar Excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Estoque"

    # Cabeçalho
    ws.merge_cells('A1:J1')
    ws['A1'] = f"RELATÓRIO DE ESTOQUE - {empresa.nome_fantasia or empresa.razao_social}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:J2')
    ws['A2'] = f"Data de Referência: {data_filtro.strftime('%d/%m/%Y')} {'(Atual)' if usar_data_atual else '(Histórico)'}"
    ws['A2'].font = Font(size=10)
    ws['A2'].alignment = Alignment(horizontal='center')

    # Resumo
    ws['A4'] = "Total de Produtos:"
    ws['B4'] = total_produtos
    ws['D4'] = "Quantidade Total:"
    ws['E4'] = total_quantidade
    ws['G4'] = "Valor Total (Custo):"
    ws['H4'] = total_valor_custo
    ws['I4'] = "Valor Total (Venda):"
    ws['J4'] = total_valor_venda

    for cell in ['A4', 'D4', 'G4', 'I4']:
        ws[cell].font = Font(bold=True)

    # Cabeçalho da tabela
    headers = ['Produto', 'Descrição', 'Qtd. Estoque', 'Qtd. Compras', 'P. Custo Unit.',
               'P. Médio Compra', 'P. Venda', 'Margem (%)', 'Vlr. Est. (Custo)', 'Vlr. Est. (Venda)']

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Dados
    row = 7
    for item in estoque_lista:
        ws.cell(row=row, column=1).value = item['nome']
        ws.cell(row=row, column=2).value = item['descricao']
        ws.cell(row=row, column=3).value = item['estoque']
        ws.cell(row=row, column=4).value = item['qtd_compras']
        ws.cell(row=row, column=5).value = item['preco_custo']
        ws.cell(row=row, column=6).value = item['preco_medio_compra']
        ws.cell(row=row, column=7).value = item['preco_venda']
        ws.cell(row=row, column=8).value = round(item['margem'], 2)
        ws.cell(row=row, column=9).value = item['valor_estoque_custo']
        ws.cell(row=row, column=10).value = item['valor_estoque']

        # Alinhamento
        for col in [3, 4, 5, 6, 7, 8, 9, 10]:
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')

        # Formato de moeda
        for col in [5, 6, 7, 9, 10]:
            ws.cell(row=row, column=col).number_format = 'R$ #,##0.00'

        row += 1

    # Linha de total
    ws.cell(row=row, column=1).value = "TOTAL GERAL"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3).value = total_quantidade
    ws.cell(row=row, column=8).value = round(total_margem, 2)
    ws.cell(row=row, column=9).value = total_valor_custo
    ws.cell(row=row, column=10).value = total_valor_venda

    for col in [1, 3, 8, 9, 10]:
        ws.cell(row=row, column=col).font = Font(bold=True)

    for col in [9, 10]:
        ws.cell(row=row, column=col).number_format = 'R$ #,##0.00'

    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20

    # Salvar em BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"relatorio_estoque_{data_filtro.strftime('%Y%m%d')}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/relatorios/fornecedores')
def relatorio_fornecedores():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario:
        flash('Usuário não encontrado', 'error')
        return redirect(url_for('login'))
    
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    try:
        app.logger.info(f"Iniciando relatório de fornecedores para usuário {usuario.id}")
        
        # Obter parâmetros de filtro com validação
        filtro_status = request.args.get('status', 'todos')
        if filtro_status not in ['todos', 'com_saldo', 'sem_saldo']:
            filtro_status = 'todos'
            
        filtro_ordenacao = request.args.get('ordenacao', 'nome')
        if filtro_ordenacao not in ['nome', 'valor_total', 'saldo_aberto']:
            filtro_ordenacao = 'nome'
            
        filtro_periodo = request.args.get('periodo', 'todos')
        if filtro_periodo not in ['todos', 'mes_atual', 'trimestre', 'ano_atual', 'personalizado']:
            filtro_periodo = 'todos'
            
        # Obter datas personalizadas se período for personalizado
        filtro_data_inicio = request.args.get('data_inicio', '').strip()
        filtro_data_fim = request.args.get('data_fim', '').strip()
            
        exportar = request.args.get('exportar', '')
        pagina = request.args.get('pagina', 1, type=int)
        # por_pagina=0 significa "mostrar todos"
        por_pagina = request.args.get('por_pagina', 0, type=int)
        filtro_busca = request.args.get('busca', '').strip()
        
        # Validar parâmetros
        if pagina < 1:
            pagina = 1
        if por_pagina < 0 or por_pagina > 100000:
            por_pagina = 0
        
        app.logger.info(f"Parâmetros: status={filtro_status}, ordenacao={filtro_ordenacao}, periodo={filtro_periodo}")

        # Obter todos os usuários da empresa vinculada (considerando acesso contador)
        empresa_id = obter_empresa_id_sessao(session, usuario)
        usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
        usuarios_ids = [u.id for u in usuarios_empresa]
        
        app.logger.info(f"Usuários da empresa: {len(usuarios_ids)}")
        
        if not usuarios_ids:
            flash('Nenhum usuário ativo encontrado na empresa', 'warning')
            return render_template('relatorio_fornecedores.html', 
                                 usuario=usuario, 
                                 fornecedores_dados=[],
                                 filtro_status=filtro_status,
                                 filtro_ordenacao=filtro_ordenacao,
                                 filtro_periodo=filtro_periodo,
                                 pagina=pagina,
                                 por_pagina=por_pagina,
                                 total_paginas=0,
                                 total_fornecedores=0)
        
        # Buscar fornecedores da empresa (filtro direto por empresa_id)
        fornecedores = Fornecedor.query.filter(Fornecedor.empresa_id == empresa_id).all()
        
        app.logger.info(f"Fornecedores encontrados: {len(fornecedores)}")
        
        if not fornecedores:
            flash('Nenhum fornecedor encontrado', 'warning')
            return render_template('relatorio_fornecedores.html',
                                 usuario=usuario,
                                 fornecedores_dados=[],
                                 filtro_status=filtro_status,
                                 filtro_ordenacao=filtro_ordenacao,
                                 filtro_periodo=filtro_periodo,
                                 filtro_data_inicio=filtro_data_inicio,
                                 filtro_data_fim=filtro_data_fim,
                                 pagina=pagina,
                                 por_pagina=por_pagina,
                                 total_paginas=0,
                                 total_fornecedores=0,
                                 sum_total_compras=0,
                                 sum_total_compras_pendentes=0,
                                 sum_saldo_vencido=0,
                                 sum_total_agendado=0,
                                 sum_saldo_aberto=0,
                                 sum_total_geral=0,
                                 sum_num_compras=0,
                                 sum_ticket_medio=0)
        
        app.logger.info(f"Fornecedores encontrados para processar: {len(fornecedores)}")
        
        # Lista para armazenar dados dos fornecedores
        fornecedores_dados = []
        
        # Definir período para filtros
        hoje = datetime.now().date()
        if filtro_periodo == 'mes_atual':
            inicio_periodo = datetime(hoje.year, hoje.month, 1).date()
            fim_periodo = hoje
        elif filtro_periodo == 'trimestre':
            trimestre = (hoje.month - 1) // 3
            inicio_periodo = datetime(hoje.year, trimestre * 3 + 1, 1).date()
            fim_periodo = hoje
        elif filtro_periodo == 'ano_atual':
            inicio_periodo = datetime(hoje.year, 1, 1).date()
            fim_periodo = hoje
        elif filtro_periodo == 'personalizado':
            # Processar datas personalizadas
            try:
                if filtro_data_inicio:
                    # Converter DD/MM/AAAA para AAAA-MM-DD
                    partes_inicio = filtro_data_inicio.split('/')
                    if len(partes_inicio) == 3:
                        inicio_periodo = datetime(int(partes_inicio[2]), int(partes_inicio[1]), int(partes_inicio[0])).date()
                    else:
                        inicio_periodo = None
                else:
                    inicio_periodo = None
                    
                if filtro_data_fim:
                    # Converter DD/MM/AAAA para AAAA-MM-DD
                    partes_fim = filtro_data_fim.split('/')
                    if len(partes_fim) == 3:
                        fim_periodo = datetime(int(partes_fim[2]), int(partes_fim[1]), int(partes_fim[0])).date()
                    else:
                        fim_periodo = None
                else:
                    fim_periodo = None
            except (ValueError, IndexError):
                flash('Datas inválidas no filtro personalizado', 'error')
                inicio_periodo = None
                fim_periodo = None
        else:
            inicio_periodo = None
            fim_periodo = None
        
        app.logger.info(f"Período: {inicio_periodo} a {fim_periodo}")
        
        # Calcular dados de cada fornecedor
        for fornecedor in fornecedores:
            try:
                app.logger.info(f"Processando fornecedor {fornecedor.id}: {fornecedor.nome}")
                
                # Buscar lançamentos financeiros do fornecedor considerando também vínculos por compra
                # Inclui registros com Lancamento.fornecedor_id diretamente ou via Compra.fornecedor_id
                lancamentos_query = (
                    Lancamento.query
                    .outerjoin(Compra, Lancamento.compra_id == Compra.id)
                    .filter(
                        Lancamento.empresa_id == empresa_id,
                        or_(Lancamento.fornecedor_id == fornecedor.id, Compra.fornecedor_id == fornecedor.id)
                    )
                )
                
                # Aplicar filtro de período se especificado
                if inicio_periodo and fim_periodo:
                    lancamentos_query = lancamentos_query.filter(
                        Lancamento.data_prevista >= inicio_periodo,
                        Lancamento.data_prevista <= fim_periodo
                    )
                
                lancamentos_fornecedor = lancamentos_query.all()
                app.logger.info(f"FORNECEDOR_FILTER: {fornecedor.nome} - lancamentos: {len(lancamentos_fornecedor)} - busca: {filtro_busca}")

                # Se houver busca, filtrar estritamente por nome ou documento
                if filtro_busca:
                     termo_l = filtro_busca.lower()
                     combina_nome = (fornecedor.nome and termo_l in fornecedor.nome.lower()) or \
                                    (fornecedor.cpf_cnpj and termo_l in fornecedor.cpf_cnpj.lower())
                     if not combina_nome:
                         continue
                
                app.logger.info(f"Lançamentos financeiros encontrados para fornecedor {fornecedor.id}: {len(lancamentos_fornecedor)}")
                
                # Calcular totais baseados na movimentação financeira real (apenas saidas)
                lancamentos_despesa = [l for l in lancamentos_fornecedor if l.tipo == 'saida']

                # Compras realizadas = lançamentos de saida marcados como realizados
                total_compras = sum(l.valor or 0 for l in lancamentos_despesa if l.realizado)
                # A vencer: não realizado e data_prevista futura (> hoje)
                total_a_vencer = sum(
                    (l.valor or 0) for l in lancamentos_despesa
                    if not l.realizado and l.data_prevista and l.data_prevista > hoje
                )
                # Saldo vencido = não realizado e data_prevista < hoje
                saldo_vencido = sum(
                    (l.valor or 0) for l in lancamentos_despesa
                    if not l.realizado and l.data_prevista and l.data_prevista < hoje
                )
                # Agendado: não realizado e data_prevista = hoje OU sem data_prevista
                total_agendado = sum(
                    (l.valor or 0) for l in lancamentos_despesa
                    if not l.realizado and ((l.data_prevista and l.data_prevista == hoje) or not l.data_prevista)
                )
                # Número de transações realizadas
                num_compras = len([l for l in lancamentos_despesa if l.realizado])
                # Número de transações pendentes
                num_compras_pendentes = len([l for l in lancamentos_despesa if not l.realizado])
                # Total geral = realizadas + a vencer + vencidas + agendado
                total_geral = total_compras + total_a_vencer + saldo_vencido + total_agendado
                # Saldo em aberto = a vencer + vencido + agendado
                saldo_aberto = total_a_vencer + saldo_vencido + total_agendado


                # Aplicar filtro de status
                if filtro_status == 'com_saldo' and saldo_aberto <= 0:
                    app.logger.info(f"Fornecedor {fornecedor.id} filtrado por status (sem saldo)")
                    continue
                elif filtro_status == 'sem_saldo' and saldo_aberto > 0:
                    app.logger.info(f"Fornecedor {fornecedor.id} filtrado por status (com saldo)")
                    continue

                # Calcular ticket médio (baseado no total geral, não apenas nas realizadas)
                ticket_medio = total_geral / (num_compras + num_compras_pendentes) if (num_compras + num_compras_pendentes) > 0 else 0

                fornecedor_data = {
                    'fornecedor': fornecedor,
                    'total_compras': total_compras,
                    'total_compras_pendentes': total_a_vencer,
                    'saldo_vencido': saldo_vencido,
                    'total_agendado': total_agendado,
                    'num_compras': num_compras + num_compras_pendentes,  # Total de transações
                    'num_compras_pendentes': num_compras_pendentes,
                    'total_geral': total_geral,
                    'saldo_aberto': saldo_aberto,
                    'ticket_medio': ticket_medio
                }
                fornecedores_dados.append(fornecedor_data)
                
                app.logger.info(f"Fornecedor {fornecedor.id} processado com sucesso")
                
            except Exception as e:
                app.logger.error(f"Erro ao processar fornecedor {fornecedor.id}: {str(e)}")
                continue
        
        app.logger.info(f"Fornecedores processados: {len(fornecedores_dados)}")
        
        # Aplicar ordenação com validação
        try:
            if filtro_ordenacao == 'nome':
                fornecedores_dados.sort(key=lambda x: x['fornecedor'].nome.lower() if x['fornecedor'].nome else '')
            elif filtro_ordenacao == 'valor_total':
                fornecedores_dados.sort(key=lambda x: x['total_geral'], reverse=True)
            elif filtro_ordenacao == 'saldo_aberto':
                fornecedores_dados.sort(key=lambda x: x['saldo_aberto'], reverse=True)
        except Exception as e:
            app.logger.error(f"Erro na ordenação: {str(e)}")
            # Ordenação padrão por nome em caso de erro
            fornecedores_dados.sort(key=lambda x: x['fornecedor'].nome.lower() if x['fornecedor'].nome else '')
        
        # Calcular paginação
        total_fornecedores = len(fornecedores_dados)
        if por_pagina == 0:
            # Sem paginação: mostrar todos
            total_paginas = 1 if total_fornecedores > 0 else 0
            fornecedores_paginados = fornecedores_dados
        else:
            total_paginas = (total_fornecedores + por_pagina - 1) // por_pagina if total_fornecedores > 0 else 0
            # Aplicar paginação
            inicio = (pagina - 1) * por_pagina
            fim = inicio + por_pagina
            fornecedores_paginados = fornecedores_dados[inicio:fim]
        
        app.logger.info(f"Paginação: {pagina}/{total_paginas}, {len(fornecedores_paginados)} fornecedores")

        app.logger.info("Relatório de fornecedores gerado com sucesso")

        # Pré-calcular totais para o template (evita problemas com Jinja sum + dict)
        sum_total_compras = sum(f.get('total_compras', 0) for f in fornecedores_paginados)
        sum_total_compras_pendentes = sum(f.get('total_compras_pendentes', 0) for f in fornecedores_paginados)
        sum_saldo_vencido_f = sum(f.get('saldo_vencido', 0) for f in fornecedores_paginados)
        sum_total_agendado_f = sum(f.get('total_agendado', 0) for f in fornecedores_paginados)
        sum_saldo_aberto_f = sum(f.get('saldo_aberto', 0) for f in fornecedores_paginados)
        sum_total_geral_f = sum(f.get('total_geral', 0) for f in fornecedores_paginados)
        sum_num_compras = sum(f.get('num_compras', 0) for f in fornecedores_paginados)
        sum_ticket_medio_f = sum_total_geral_f / sum_num_compras if sum_num_compras > 0 else 0

        return render_template('relatorio_fornecedores.html',
                             usuario=usuario,
                             fornecedores_dados=fornecedores_paginados,
                             filtro_status=filtro_status,
                             filtro_ordenacao=filtro_ordenacao,
                             filtro_periodo=filtro_periodo,
                             filtro_data_inicio=filtro_data_inicio,
                             filtro_data_fim=filtro_data_fim,
                             pagina=pagina,
                             por_pagina=por_pagina,
                             total_paginas=total_paginas,
                             total_fornecedores=total_fornecedores,
                             sum_total_compras=sum_total_compras,
                             sum_total_compras_pendentes=sum_total_compras_pendentes,
                             sum_saldo_vencido=sum_saldo_vencido_f,
                             sum_total_agendado=sum_total_agendado_f,
                             sum_saldo_aberto=sum_saldo_aberto_f,
                             sum_total_geral=sum_total_geral_f,
                             sum_num_compras=sum_num_compras,
                             sum_ticket_medio=sum_ticket_medio_f)
    
    except Exception as e:
        import traceback
        app.logger.error(f"Erro no relatório de fornecedores: {str(e)}")
        app.logger.error(traceback.format_exc())
        flash('Erro ao gerar relatório de fornecedores. Verifique os logs para mais detalhes.', 'error')
        return redirect(url_for('relatorios'))

# Rotas para Configurações
@app.route('/configuracoes')
def configuracoes():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    return render_template('configuracoes.html', usuario=usuario)

# Rotas para Contas Caixa
@app.route('/configuracoes/contas-caixa')
def contas_caixa():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    # Buscar todos os usuários da mesma empresa
    empresa_id = obter_empresa_id_sessao(session, usuario)

    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]
    
    # Carregar contas caixa de todos os usuários da empresa
    contas_caixa = ContaCaixa.query.filter(ContaCaixa.usuario_id.in_(usuarios_ids)).order_by(ContaCaixa.nome).all()
    
    # Adicionar informação do usuário que criou cada conta
    for conta in contas_caixa:
        conta.usuario_criador = db.session.get(Usuario, conta.usuario_id)
    
    return render_template('contas_caixa.html', usuario=usuario, contas_caixa=contas_caixa)

@app.route('/configuracoes/contas-caixa/nova', methods=['GET', 'POST'])
def nova_conta_caixa():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    if request.method == 'POST':
        nome = request.form['nome']
        tipo = request.form['tipo']
        banco = request.form.get('banco', '')
        agencia = request.form.get('agencia', '')
        conta = request.form.get('conta', '')
        saldo_inicial = float(request.form.get('saldo_inicial', 0))
        descricao = request.form.get('descricao', '')
        
        if not nome:
            flash('Nome da conta é obrigatório', 'error')
            return redirect(url_for('nova_conta_caixa'))
        
        nova_conta = ContaCaixa(
            nome=nome,
            tipo=tipo,
            banco=banco,
            agencia=agencia,
            conta=conta,
            saldo_inicial=saldo_inicial,
            saldo_atual=saldo_inicial,
            descricao=descricao,
            usuario_id=usuario.id
        )
        
        try:
            db.session.add(nova_conta)
            db.session.commit()
            flash('Conta caixa criada com sucesso!', 'success')
            return redirect(url_for('contas_caixa'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao criar conta caixa: {str(e)}', 'error')
            return redirect(url_for('nova_conta_caixa'))
    
    return render_template('nova_conta_caixa.html', usuario=usuario)

@app.route('/configuracoes/contas-caixa/<int:conta_id>/editar', methods=['GET', 'POST'])
def editar_conta_caixa(conta_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    conta = ContaCaixa.query.filter_by(id=conta_id, usuario_id=usuario.id).first()
    if not conta:
        flash('Conta caixa não encontrada', 'error')
        return redirect(url_for('contas_caixa'))
    
    if request.method == 'POST':
        nome = request.form['nome']
        tipo = request.form['tipo']
        banco = request.form.get('banco', '')
        agencia = request.form.get('agencia', '')
        conta_num = request.form.get('conta', '')
        saldo_inicial = float(request.form.get('saldo_inicial', 0))
        descricao = request.form.get('descricao', '')
        
        if not nome:
            flash('Nome da conta é obrigatório', 'error')
            return redirect(url_for('editar_conta_caixa', conta_id=conta_id))
        
        # REMOVIDO: Atualização manual de saldo - agora é calculado dinamicamente
        # O saldo_atual é calculado automaticamente pelo método calcular_saldo_atual()

        conta.nome = nome
        conta.tipo = tipo
        conta.banco = banco
        conta.agencia = agencia
        conta.conta = conta_num
        conta.saldo_inicial = saldo_inicial
        # conta.saldo_atual - NÃO atualiza mais manualmente
        conta.descricao = descricao
        
        try:
            db.session.commit()
            flash('Conta caixa atualizada com sucesso!', 'success')
            return redirect(url_for('contas_caixa'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar conta caixa: {str(e)}', 'error')
            return redirect(url_for('editar_conta_caixa', conta_id=conta_id))
    
    return render_template('editar_conta_caixa.html', usuario=usuario, conta=conta)

@app.route('/configuracoes/contas-caixa/<int:conta_id>/toggle')
def toggle_conta_caixa(conta_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    conta = ContaCaixa.query.filter_by(id=conta_id, usuario_id=usuario.id).first()
    if not conta:
        flash('Conta caixa não encontrada', 'error')
        return redirect(url_for('contas_caixa'))
    
    conta.ativo = not conta.ativo
    
    try:
        db.session.commit()
        status = 'ativada' if conta.ativo else 'desativada'
        flash(f'Conta caixa {status} com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao alterar status da conta caixa: {str(e)}', 'error')
    
    return redirect(url_for('contas_caixa'))

@app.route('/configuracoes/contas-caixa/<int:conta_id>/deletar')
def deletar_conta_caixa(conta_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    conta = ContaCaixa.query.filter_by(id=conta_id, usuario_id=usuario.id).first()
    if not conta:
        flash('Conta caixa não encontrada', 'error')
        return redirect(url_for('contas_caixa'))
    
    # Verificar se há lançamentos vinculados
    lancamentos_vinculados = Lancamento.query.filter_by(conta_caixa_id=conta.id).count()
    if lancamentos_vinculados > 0:
        flash(f'Não é possível deletar a conta caixa. Existem {lancamentos_vinculados} lançamento(s) vinculado(s) a ela.', 'error')
        return redirect(url_for('contas_caixa'))
    
    try:
        db.session.delete(conta)
        db.session.commit()
        flash('Conta caixa deletada com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao deletar conta caixa: {str(e)}', 'error')
    
    return redirect(url_for('contas_caixa'))

@app.route('/vendas/<int:venda_id>/parcelas')
def parcelas_venda(venda_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    venda = db.session.get(Venda, venda_id)
    if not venda or venda.usuario_id != usuario.id:
        flash('Venda não encontrada.', 'error')
        return redirect(url_for('vendas'))
    
    parcelas = Parcela.query.filter_by(venda_id=venda_id).order_by(Parcela.numero).all()
    
    return render_template('parcelas_venda.html', venda=venda, parcelas=parcelas, usuario=usuario)

@app.route('/compras/<int:compra_id>/parcelas')
def parcelas_compra(compra_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    compra = db.session.get(Compra, compra_id)
    if not compra or compra.usuario_id != usuario.id:
        flash('Compra não encontrada.', 'error')
        return redirect(url_for('compras'))
    
    parcelas = Parcela.query.filter_by(compra_id=compra_id).order_by(Parcela.numero).all()
    
    return render_template('parcelas_compra.html', compra=compra, parcelas=parcelas, usuario=usuario)

# Rotas para gerenciamento de usuários da empresa
@app.route('/empresa/usuarios')
def empresa_usuarios():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario or usuario.tipo not in ['usuario_principal', 'admin']:
        flash('Acesso negado.', 'error')
        return redirect(url_for('dashboard'))
    
    # Buscar todos os usuários da empresa
    empresa_id = obter_empresa_id_sessao(session, usuario)

    usuarios = Usuario.query.filter_by(empresa_id=empresa_id).order_by(Usuario.nome).all()
    
    return render_template('empresa_usuarios.html', usuarios=usuarios)

@app.route('/empresa/usuarios/novo', methods=['GET', 'POST'])
def novo_usuario_empresa():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario_atual = db.session.get(Usuario, session['usuario_id'])
    if not usuario_atual or usuario_atual.tipo not in ['usuario_principal', 'admin']:
        flash('Acesso negado.', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        nome = request.form['nome']
        usuario = request.form['usuario']  # Novo campo usuario
        email = request.form['email']
        senha = request.form['senha']
        telefone = request.form.get('telefone', '')
        tipo = request.form.get('tipo', 'usuario')
        categoria_id = request.form.get('categoria_id')
        
        # Verificar se usuario já existe na empresa
        if Usuario.query.filter_by(empresa_id=usuario_atual.empresa_id, usuario=usuario).first():
            flash('Nome de usuário já cadastrado para esta empresa.', 'error')
            return redirect(url_for('novo_usuario_empresa'))
        
        # Removida validação de email único por empresa
        # Validar categoria personalizada se selecionada
        
        # Validar categoria personalizada se selecionada
        if tipo == 'categoria_personalizada':
            if not categoria_id:
                flash('Categoria personalizada é obrigatória quando selecionada.', 'error')
                return redirect(url_for('novo_usuario_empresa'))
            
            categoria = CategoriaUsuario.query.filter_by(
                id=categoria_id, 
                empresa_id=usuario_atual.empresa_id,
                ativo=True
            ).first()
            
            if not categoria:
                flash('Categoria selecionada não encontrada.', 'error')
                return redirect(url_for('novo_usuario_empresa'))
        
        novo_usuario = Usuario(
            nome=nome,
            usuario=usuario,  # Novo campo usuario
            email=email,
            senha=generate_password_hash(senha),
            telefone=telefone,
            tipo=tipo,
            categoria_id=categoria_id if tipo == 'categoria_personalizada' else None,
            empresa_id=usuario_atual.empresa_id,
            criado_por=usuario_atual.id
        )
        
        db.session.add(novo_usuario)
        db.session.flush()  # Para obter o ID do usuário
        
        # Criar permissões baseadas na categoria ou padrão
        if tipo == 'categoria_personalizada' and categoria_id:
            criar_permissoes_por_categoria(novo_usuario.id, categoria_id)
        else:
            criar_permissoes_padrao(novo_usuario.id)
        
        db.session.commit()
        
        flash('Usuário criado com sucesso!', 'success')
        return redirect(url_for('empresa_usuarios'))
    
    return render_template('novo_usuario_empresa.html')

@app.route('/empresa/usuarios/<int:user_id>/editar', methods=['GET', 'POST'])
def editar_usuario_empresa(user_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario_atual = db.session.get(Usuario, session['usuario_id'])
    if not usuario_atual or usuario_atual.tipo not in ['usuario_principal', 'admin']:
        flash('Acesso negado.', 'error')
        return redirect(url_for('dashboard'))
    
    usuario = Usuario.query.filter_by(id=user_id, empresa_id=usuario_atual.empresa_id).first()
    if not usuario:
        flash('Usuário não encontrado.', 'error')
        return redirect(url_for('empresa_usuarios'))
    
    if request.method == 'POST':
        usuario.nome = request.form['nome']
        usuario.usuario = request.form['usuario']  # Novo campo usuario
        usuario.email = request.form['email']
        usuario.telefone = request.form.get('telefone', '')
        usuario.tipo = request.form.get('tipo', 'usuario')
        
        # Verificar se usuario já existe (exceto para o próprio usuário)
        usuario_existente = Usuario.query.filter_by(empresa_id=usuario_atual.empresa_id, usuario=usuario.usuario).first()
        if usuario_existente and usuario_existente.id != usuario.id:
            flash('Nome de usuário já cadastrado para esta empresa.', 'error')
            return redirect(url_for('editar_usuario_empresa', user_id=user_id))
        
        # Removida validação de email único por empresa
        # Se uma nova senha foi fornecida
        
        # Se uma nova senha foi fornecida
        if request.form.get('nova_senha'):
            usuario.senha = generate_password_hash(request.form['nova_senha'])
        
        db.session.commit()
        flash('Usuário atualizado com sucesso!', 'success')
        return redirect(url_for('empresa_usuarios'))
    
    return render_template('editar_usuario_empresa.html', usuario=usuario)

@app.route('/empresa/usuarios/<int:user_id>/toggle_status')
def toggle_usuario_empresa_status(user_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario_atual = db.session.get(Usuario, session['usuario_id'])
    if not usuario_atual or usuario_atual.tipo not in ['usuario_principal', 'admin']:
        flash('Acesso negado.', 'error')
        return redirect(url_for('dashboard'))
    
    usuario = Usuario.query.filter_by(id=user_id, empresa_id=usuario_atual.empresa_id).first()
    if not usuario:
        flash('Usuário não encontrado.', 'error')
        return redirect(url_for('empresa_usuarios'))
    
    if usuario.id == usuario_atual.id:
        flash('Você não pode desativar sua própria conta.', 'error')
        return redirect(url_for('empresa_usuarios'))
    
    usuario.ativo = not usuario.ativo
    db.session.commit()
    
    status = 'ativada' if usuario.ativo else 'desativada'
    flash(f'Conta {status} com sucesso!', 'success')
    return redirect(url_for('empresa_usuarios'))

@app.route('/empresa/usuarios/<int:user_id>/deletar')
def deletar_usuario_empresa(user_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario_atual = db.session.get(Usuario, session['usuario_id'])
    if not usuario_atual or usuario_atual.tipo not in ['usuario_principal', 'admin']:
        flash('Acesso negado.', 'error')
        return redirect(url_for('dashboard'))
    
    usuario = Usuario.query.filter_by(id=user_id, empresa_id=usuario_atual.empresa_id).first()
    if not usuario:
        flash('Usuário não encontrado.', 'error')
        return redirect(url_for('empresa_usuarios'))
    
    if usuario.id == usuario_atual.id:
        flash('Você não pode deletar sua própria conta.', 'error')
        return redirect(url_for('empresa_usuarios'))
    
    if usuario.tipo == 'usuario_principal':
        flash('Não é possível deletar o usuário principal da empresa.', 'error')
        return redirect(url_for('empresa_usuarios'))
    
    # Verificar se o usuário tem dados associados
    lancamentos = Lancamento.query.filter_by(usuario_id=usuario.id).count()
    if lancamentos > 0:
        flash('Não é possível deletar um usuário que possui lançamentos associados.', 'error')
        return redirect(url_for('empresa_usuarios'))
    
    try:
        # Primeiro, atualizar registros que fazem referência ao usuário (antes de qualquer commit)
        with db.session.no_autoflush:
            # Atualizar categorias de usuário criadas por este usuário (definir criado_por como NULL)
            categorias_criadas = CategoriaUsuario.query.filter_by(criado_por=usuario.id).all()
            for categoria in categorias_criadas:
                categoria.criado_por = None
            
            # Deletar vínculos criados por este usuário (mais seguro que atualizar)
            vinculos_criados = Vinculo.query.filter_by(usuario_id=usuario.id).all()
            for vinculo in vinculos_criados:
                db.session.delete(vinculo)
        
        # Deletar todas as permissões associadas ao usuário
        permissoes = Permissao.query.filter_by(usuario_id=usuario.id).all()
        for permissao in permissoes:
            db.session.delete(permissao)
        
        # Deletar todos os logs de evento associados ao usuário
        event_logs = EventLog.query.filter_by(usuario_id=usuario.id).all()
        for event_log in event_logs:
            db.session.delete(event_log)
        
        # Deletar todas as importações associadas ao usuário
        importacoes = Importacao.query.filter_by(usuario_id=usuario.id).all()
        for importacao in importacoes:
            db.session.delete(importacao)
        
        # Salvar informações da empresa antes de deletar o usuário
        empresa_id = usuario.empresa_id
        empresa_nome = usuario.empresa.razao_social if usuario.empresa else "N/A"
        empresa_cnpj = usuario.empresa.cnpj if usuario.empresa else "N/A"
        
        app.logger.info(f"Iniciando exclusão do usuário {usuario.nome} da empresa {empresa_nome} (ID: {empresa_id})")
        
        # Deletar o usuário primeiro
        db.session.delete(usuario)
        app.logger.info(f"Usuário {usuario.nome} deletado da sessão.")
        
        # Fazer flush para garantir que a exclusão seja processada
        db.session.flush()
        
        # Verificar se ainda há outros usuários na empresa
        outros_usuarios = Usuario.query.filter_by(empresa_id=empresa_id).count()
        app.logger.info(f"Após exclusão: Empresa {empresa_id} tem {outros_usuarios} usuários restantes.")
        
        # Se não há outros usuários na empresa, deletar a empresa também
        if outros_usuarios == 0:
            empresa = db.session.get(Empresa, empresa_id)
            if empresa:
                app.logger.info(f"Empresa órfã encontrada: {empresa.razao_social} (CNPJ: {empresa.cnpj})")
                
                # Atualizar categorias de usuário da empresa (definir empresa_id como NULL)
                with db.session.no_autoflush:
                    categorias_empresa = CategoriaUsuario.query.filter_by(empresa_id=empresa_id).all()
                    for categoria in categorias_empresa:
                        categoria.empresa_id = None
                        categoria.ativo = False  # Desativar categorias órfãs
                
                # Verificar se há dependências que impedem a exclusão
                try:
                    db.session.delete(empresa)
                    app.logger.info(f"Empresa {empresa.razao_social} marcada para exclusão.")
                    flash(f'Usuário e empresa "{empresa_nome}" deletados com sucesso! CNPJ {empresa_cnpj} liberado para novo cadastro.', 'success')
                except Exception as empresa_error:
                    app.logger.error(f"Erro ao deletar empresa: {str(empresa_error)}")
                    flash(f'Usuário deletado, mas empresa não pôde ser removida: {str(empresa_error)}', 'warning')
            else:
                app.logger.warning(f"Empresa {empresa_id} não encontrada após exclusão do usuário.")
                flash('Usuário deletado com sucesso!', 'success')
        else:
            app.logger.info(f"Empresa {empresa_id} mantida - ainda tem {outros_usuarios} usuários.")
            flash('Usuário deletado com sucesso!', 'success')
        
        # Commit de todas as alterações
        db.session.commit()
        app.logger.info("Transação commitada com sucesso.")
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao deletar usuário: {str(e)}', 'error')
    
    return redirect(url_for('empresa_usuarios'))

# Funções para gerenciar permissões
def verificar_permissao(usuario_id, modulo, acao):
    """
    Verifica se um usuário tem permissão para uma ação específica em um módulo
    """
    if not usuario_id:
        return False
    
    usuario = db.session.get(Usuario, usuario_id)
    if not usuario:
        return False
    
    # Usuários principais e admins têm todas as permissões
    if usuario.tipo in ['usuario_principal', 'admin']:
        return True
    
    # Verificar permissão específica
    permissao = Permissao.query.filter_by(
        usuario_id=usuario_id,
        modulo=modulo,
        acao=acao,
        ativo=True
    ).first()
    
    return permissao is not None

def criar_permissoes_padrao(usuario_id):
    """
    Cria permissões padrão para um novo usuário
    """
    modulos = ['lancamentos', 'clientes', 'fornecedores', 'estoque', 'vendas', 'compras', 'relatorios']
    acoes = ['visualizar']  # Por padrão, apenas visualizar
    
    for modulo in modulos:
        for acao in acoes:
            permissao = Permissao(
                usuario_id=usuario_id,
                modulo=modulo,
                acao=acao,
                ativo=True
            )
            db.session.add(permissao)
    
    db.session.commit()

def criar_permissoes_por_categoria(usuario_id, categoria_id):
    """
    Cria permissões para um usuário baseado em uma categoria personalizada
    """
    # Obter permissões da categoria
    permissoes_categoria = PermissaoCategoria.query.filter_by(
        categoria_id=categoria_id,
        ativo=True
    ).all()
    
    # Criar permissões individuais para o usuário
    for perm_cat in permissoes_categoria:
        permissao = Permissao(
            usuario_id=usuario_id,
            modulo=perm_cat.modulo,
            acao=perm_cat.acao,
            ativo=True
        )
        db.session.add(permissao)
    
    db.session.commit()

def atualizar_permissoes_usuario(usuario_id, permissoes_dict):
    """
    Atualiza as permissões de um usuário
    permissoes_dict: {'modulo': ['acao1', 'acao2'], ...}
    """
    # Remover permissões existentes
    Permissao.query.filter_by(usuario_id=usuario_id).delete()
    
    # Criar novas permissões
    for modulo, acoes in permissoes_dict.items():
        for acao in acoes:
            permissao = Permissao(
                usuario_id=usuario_id,
                modulo=modulo,
                acao=acao,
                ativo=True
            )
            db.session.add(permissao)
    
    db.session.commit()

def obter_permissoes_usuario(usuario_id):
    """
    Retorna as permissões de um usuário em formato de dicionário
    """
    permissoes = Permissao.query.filter_by(usuario_id=usuario_id, ativo=True).all()
    permissoes_dict = {}
    
    for permissao in permissoes:
        if permissao.modulo not in permissoes_dict:
            permissoes_dict[permissao.modulo] = []
        permissoes_dict[permissao.modulo].append(permissao.acao)
    
    return permissoes_dict

@app.route('/empresa/usuarios/<int:user_id>/permissoes', methods=['GET', 'POST'])
def gerenciar_permissoes_usuario(user_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario_atual = db.session.get(Usuario, session['usuario_id'])
    if not usuario_atual or usuario_atual.tipo not in ['usuario_principal', 'admin']:
        flash('Acesso negado.', 'error')
        return redirect(url_for('dashboard'))
    
    usuario = Usuario.query.filter_by(id=user_id, empresa_id=usuario_atual.empresa_id).first()
    if not usuario:
        flash('Usuário não encontrado.', 'error')
        return redirect(url_for('empresa_usuarios'))
    
    if request.method == 'POST':
        # Processar permissões enviadas
        permissoes_dict = {}
        modulos = ['lancamentos', 'clientes', 'fornecedores', 'estoque', 'vendas', 'compras', 'relatorios', 'configuracoes']
        acoes = ['visualizar', 'criar', 'editar', 'deletar']
        
        for modulo in modulos:
            permissoes_dict[modulo] = []
            for acao in acoes:
                if request.form.get(f'{modulo}_{acao}'):
                    permissoes_dict[modulo].append(acao)
        
        # Atualizar permissões
        atualizar_permissoes_usuario(usuario.id, permissoes_dict)
        
        flash('Permissões atualizadas com sucesso!', 'success')
        return redirect(url_for('empresa_usuarios'))
    
    # Obter permissões atuais
    permissoes_atuais = obter_permissoes_usuario(usuario.id)
    
    return render_template('gerenciar_permissoes.html', usuario=usuario, permissoes=permissoes_atuais)

# ===== ROTAS PARA CATEGORIAS DE USUÁRIOS =====

@app.route('/empresa/categorias-usuarios')
def categorias_usuarios():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario_atual = db.session.get(Usuario, session['usuario_id'])
    if not usuario_atual or usuario_atual.tipo not in ['usuario_principal', 'admin']:
        flash('Acesso negado.', 'error')
        return redirect(url_for('dashboard'))
    
    categorias = CategoriaUsuario.query.filter_by(empresa_id=usuario_atual.empresa_id, ativo=True).all()
    return render_template('categorias_usuarios.html', categorias=categorias)

@app.route('/empresa/categorias-usuarios/nova', methods=['GET', 'POST'])
def nova_categoria_usuario():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario_atual = db.session.get(Usuario, session['usuario_id'])
    if not usuario_atual or usuario_atual.tipo not in ['usuario_principal', 'admin']:
        flash('Acesso negado.', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        nome = request.form.get('nome')
        descricao = request.form.get('descricao')
        
        if not nome:
            flash('Nome da categoria é obrigatório.', 'error')
            return render_template('nova_categoria_usuario.html')
        
        # Verificar se já existe categoria com esse nome
        categoria_existente = CategoriaUsuario.query.filter_by(
            nome=nome, 
            empresa_id=usuario_atual.empresa_id,
            ativo=True
        ).first()
        
        if categoria_existente:
            flash('Já existe uma categoria com esse nome.', 'error')
            return render_template('nova_categoria_usuario.html')
        
        # Criar nova categoria
        nova_categoria = CategoriaUsuario(
            nome=nome,
            descricao=descricao,
            empresa_id=usuario_atual.empresa_id,
            criado_por=usuario_atual.id
        )
        
        db.session.add(nova_categoria)
        db.session.commit()
        
        # Processar permissões
        modulos = ['lancamentos', 'clientes', 'fornecedores', 'estoque', 'vendas', 'compras', 'relatorios', 'configuracoes']
        acoes = ['visualizar', 'criar', 'editar', 'deletar']
        
        for modulo in modulos:
            for acao in acoes:
                if request.form.get(f'{modulo}_{acao}'):
                    permissao = PermissaoCategoria(
                        categoria_id=nova_categoria.id,
                        modulo=modulo,
                        acao=acao
                    )
                    db.session.add(permissao)
        
        db.session.commit()
        flash('Categoria criada com sucesso!', 'success')
        return redirect(url_for('categorias_usuarios'))
    
    return render_template('nova_categoria_usuario.html')

@app.route('/empresa/categorias-usuarios/<int:categoria_id>/editar', methods=['GET', 'POST'])
def editar_categoria_usuario(categoria_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario_atual = db.session.get(Usuario, session['usuario_id'])
    if not usuario_atual or usuario_atual.tipo not in ['usuario_principal', 'admin']:
        flash('Acesso negado.', 'error')
        return redirect(url_for('dashboard'))
    
    categoria = CategoriaUsuario.query.filter_by(
        id=categoria_id, 
        empresa_id=usuario_atual.empresa_id
    ).first()
    
    if not categoria:
        flash('Categoria não encontrada.', 'error')
        return redirect(url_for('categorias_usuarios'))
    
    if request.method == 'POST':
        nome = request.form.get('nome')
        descricao = request.form.get('descricao')
        
        if not nome:
            flash('Nome da categoria é obrigatório.', 'error')
            return render_template('editar_categoria_usuario.html', categoria=categoria)
        
        # Verificar se já existe outra categoria com esse nome
        categoria_existente = CategoriaUsuario.query.filter(
            CategoriaUsuario.nome == nome,
            CategoriaUsuario.empresa_id == usuario_atual.empresa_id,
            CategoriaUsuario.id != categoria_id,
            CategoriaUsuario.ativo == True
        ).first()
        
        if categoria_existente:
            flash('Já existe uma categoria com esse nome.', 'error')
            return render_template('editar_categoria_usuario.html', categoria=categoria)
        
        # Atualizar categoria
        categoria.nome = nome
        categoria.descricao = descricao
        
        # Remover permissões existentes
        PermissaoCategoria.query.filter_by(categoria_id=categoria.id).delete()
        
        # Adicionar novas permissões
        modulos = ['lancamentos', 'clientes', 'fornecedores', 'estoque', 'vendas', 'compras', 'relatorios', 'configuracoes']
        acoes = ['visualizar', 'criar', 'editar', 'deletar']
        
        for modulo in modulos:
            for acao in acoes:
                if request.form.get(f'{modulo}_{acao}'):
                    permissao = PermissaoCategoria(
                        categoria_id=categoria.id,
                        modulo=modulo,
                        acao=acao
                    )
                    db.session.add(permissao)
        
        db.session.commit()
        flash('Categoria atualizada com sucesso!', 'success')
        return redirect(url_for('categorias_usuarios'))
    
    # Obter permissões atuais da categoria
    permissoes_atuais = {}
    permissoes = PermissaoCategoria.query.filter_by(categoria_id=categoria.id, ativo=True).all()
    
    for permissao in permissoes:
        if permissao.modulo not in permissoes_atuais:
            permissoes_atuais[permissao.modulo] = []
        permissoes_atuais[permissao.modulo].append(permissao.acao)
    
    return render_template('editar_categoria_usuario.html', categoria=categoria, permissoes=permissoes_atuais)

@app.route('/empresa/categorias-usuarios/<int:categoria_id>/deletar')
def deletar_categoria_usuario(categoria_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario_atual = db.session.get(Usuario, session['usuario_id'])
    if not usuario_atual or usuario_atual.tipo not in ['usuario_principal', 'admin']:
        flash('Acesso negado.', 'error')
        return redirect(url_for('dashboard'))
    
    categoria = CategoriaUsuario.query.filter_by(
        id=categoria_id, 
        empresa_id=usuario_atual.empresa_id
    ).first()
    
    if not categoria:
        flash('Categoria não encontrada.', 'error')
        return redirect(url_for('categorias_usuarios'))
    
    # Verificar se há usuários usando esta categoria
    usuarios_categoria = Usuario.query.filter_by(categoria_id=categoria_id).count()
    if usuarios_categoria > 0:
        flash(f'Não é possível excluir esta categoria pois {usuarios_categoria} usuário(s) estão usando ela.', 'error')
        return redirect(url_for('categorias_usuarios'))
    
    # Marcar como inativa ao invés de deletar
    categoria.ativo = False
    
    # Desativar permissões da categoria
    PermissaoCategoria.query.filter_by(categoria_id=categoria.id).update({'ativo': False})
    
    db.session.commit()
    flash('Categoria removida com sucesso!', 'success')
    return redirect(url_for('categorias_usuarios'))

# ===== API PARA CATEGORIAS DE USUÁRIOS =====

@app.route('/api/categorias-usuarios')
def api_categorias_usuarios():
    if 'usuario_id' not in session:
        return {'error': 'Não autenticado'}, 401
    
    usuario_atual = db.session.get(Usuario, session['usuario_id'])
    if not usuario_atual:
        return {'error': 'Usuário não encontrado'}, 404
    
    categorias = CategoriaUsuario.query.filter_by(
        empresa_id=usuario_atual.empresa_id,
        ativo=True
    ).all()
    
    categorias_data = []
    for categoria in categorias:
        categorias_data.append({
            'id': categoria.id,
            'nome': categoria.nome,
            'descricao': categoria.descricao
        })
    
    return {'categorias': categorias_data}

# Funções de validação
def validar_cnpj(cnpj):
    """
    Valida CNPJ brasileiro
    Aceita formatos: 00.000.000/0000-00, 00000000000000, 00.000.000/0000-00
    Permite CNPJs de teste e qualquer CNPJ válido
    """
    if not cnpj:
        return False

    # Remove caracteres não numéricos
    cnpj_limpo = ''.join(filter(str.isdigit, cnpj))

    # Verifica se tem 14 dígitos
    if len(cnpj_limpo) != 14:
        return False

    # Aceita qualquer CNPJ com 14 dígitos (validação flexível)
    return True

def validar_cpf(cpf):
    """
    Valida CPF brasileiro
    Aceita formatos: 000.000.000-00, 00000000000
    Permite CPFs de teste e qualquer CPF válido
    """
    if not cpf:
        return False

    # Remove caracteres não numéricos
    cpf_limpo = ''.join(filter(str.isdigit, cpf))

    # Verifica se tem 11 dígitos
    if len(cpf_limpo) != 11:
        return False

    # Aceita qualquer CPF com 11 dígitos (validação flexível)
    return True

def validar_email(email):
    """
    Valida formato de email
    """
    import re
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None

def formatar_moeda(valor):
    """
    Formata valor para exibição em moeda brasileira
    """
    return f"R$ {valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

def calcular_idade(data_nascimento):
    """
    Calcula idade a partir da data de nascimento
    """
    hoje = datetime.now().date()
    return hoje.year - data_nascimento.year - ((hoje.month, hoje.day) < (data_nascimento.month, data_nascimento.day))

def fazer_backup_banco():
    """
    Faz backup do banco de dados
    """
    try:
        import os
        
        # Criar diretório de backups se não existir
        backup_dir = 'backups'
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        
        # Nome do arquivo de backup com timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_filename = f'backup_saas_financeiro_{timestamp}.db'
        backup_path = os.path.join(backup_dir, backup_filename)
        
        # Copiar arquivo do banco
        db_path = 'instance/saas_financeiro_v2.db'
        if os.path.exists(db_path):
            copy2(db_path, backup_path)
            app.logger.info(f'Backup criado com sucesso: {backup_path}')
            return backup_path
        else:
            app.logger.warning('Arquivo do banco não encontrado para backup')
            return None
            
    except Exception as e:
        app.logger.error(f'Erro ao criar backup: {str(e)}')
        return None

def limpar_backups_antigos(dias=30):
    """
    Remove backups mais antigos que o número de dias especificado
    """
    try:
        import os
        from datetime import datetime, timedelta
        
        backup_dir = 'backups'
        if not os.path.exists(backup_dir):
            return
        
        # Data limite para remoção
        data_limite = datetime.now() - timedelta(days=dias)
        
        # Listar arquivos de backup
        for arquivo in os.listdir(backup_dir):
            if arquivo.startswith('backup_saas_financeiro_') and arquivo.endswith('.db'):
                arquivo_path = os.path.join(backup_dir, arquivo)
                data_arquivo = datetime.fromtimestamp(os.path.getctime(arquivo_path))
                
                if data_arquivo < data_limite:
                    os.remove(arquivo_path)
                    app.logger.info(f'Backup antigo removido: {arquivo}')
                    
    except Exception as e:
        app.logger.error(f'Erro ao limpar backups antigos: {str(e)}')

def verificar_alertas_usuario(usuario_id):
    """
    Verifica e retorna alertas importantes para o usuário
    """
    alertas = []
    
    try:
        # Verificar lançamentos vencidos
        hoje = datetime.now().date()
        lancamentos_vencidos = Lancamento.query.filter_by(
            usuario_id=usuario_id,
            realizado=False
        ).filter(
            Lancamento.data_prevista < hoje
        ).count()
        
        if lancamentos_vencidos > 0:
            alertas.append({
                'tipo': 'warning',
                'titulo': 'Lançamentos Vencidos',
                'mensagem': f'Você tem {lancamentos_vencidos} lançamento(s) vencido(s)',
                'icone': 'fas fa-exclamation-triangle'
            })
        
        # Verificar estoque baixo (apenas para comércio/indústria)
        usuario = db.session.get(Usuario, usuario_id)
        if usuario and usuario.empresa and hasattr(usuario.empresa, 'tipo_empresa') and usuario.empresa.tipo_empresa in ['comercio', 'industria']:
            # Buscar todos os usuários da mesma empresa
            empresa_id = obter_empresa_id_sessao(session, usuario)

            usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
            usuarios_ids = [u.id for u in usuarios_empresa]
            
            # Filtrar apenas produtos ativos (produtos inativos não são considerados nos alertas)
            produtos_estoque_baixo = Produto.query.filter(
                Produto.usuario_id.in_(usuarios_ids),
                Produto.ativo == True,
                Produto.estoque <= 5
            ).count()
            
            if produtos_estoque_baixo > 0:
                alertas.append({
                    'tipo': 'info',
                    'titulo': 'Estoque Baixo',
                    'mensagem': f'{produtos_estoque_baixo} produto(s) ativo(s) com estoque baixo (≤ 5 unidades)',
                    'icone': 'fas fa-boxes'
                })
        
        # Verificar contas a receber hoje
        contas_receber_hoje = Lancamento.query.filter_by(
            usuario_id=usuario_id,
            tipo='entrada',
            realizado=False
        ).filter(
            Lancamento.data_prevista == hoje
        ).count()
        
        if contas_receber_hoje > 0:
            alertas.append({
                'tipo': 'success',
                'titulo': 'Contas a Receber Hoje',
                'mensagem': f'{contas_receber_hoje} conta(s) a receber hoje',
                'icone': 'fas fa-calendar-check'
            })
        
        # Verificar contas a pagar hoje
        contas_pagar_hoje = Lancamento.query.filter_by(
            usuario_id=usuario_id,
            tipo='saida',
            realizado=False
        ).filter(
            Lancamento.data_prevista == hoje
        ).count()
        
        if contas_pagar_hoje > 0:
            alertas.append({
                'tipo': 'danger',
                'titulo': 'Contas a Pagar Hoje',
                'mensagem': f'{contas_pagar_hoje} conta(s) a pagar hoje',
                'icone': 'fas fa-calendar-times'
            })
            
    except Exception as e:
        app.logger.error(f'Erro ao verificar alertas: {str(e)}')
    
    return alertas

@app.route('/admin/backup', methods=['GET', 'POST'])
def admin_backup():
    """
    Rota para administradores fazerem backup manual do banco
    """
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario or usuario.tipo != 'admin':
        flash('Acesso negado. Apenas administradores podem fazer backup.', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        try:
            backup_path = fazer_backup_banco()
            if backup_path:
                flash(f'Backup criado com sucesso: {backup_path}', 'success')
            else:
                flash('Erro ao criar backup.', 'error')
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao criar backup: {str(e)}', 'error')
        
        return redirect(url_for('admin_backup'))
    
    # Listar backups existentes
    backups = []
    backup_dir = 'backups'
    if os.path.exists(backup_dir):
        for arquivo in os.listdir(backup_dir):
            if arquivo.startswith('backup_saas_financeiro_') and arquivo.endswith('.db'):
                arquivo_path = os.path.join(backup_dir, arquivo)
                stat = os.stat(arquivo_path)
                backups.append({
                    'nome': arquivo,
                    'tamanho': stat.st_size,
                    'data': datetime.fromtimestamp(stat.st_ctime),
                    'caminho': arquivo_path
                })
    
    # Ordenar por data (mais recente primeiro)
    backups.sort(key=lambda x: x['data'], reverse=True)
    
    return render_template('admin_backup.html', backups=backups)

@app.route('/admin/recalcular-saldos', methods=['GET', 'POST'])
def admin_recalcular_saldos():
    """
    Rota para administradores recalcularem os saldos das contas caixa
    baseado nos lançamentos financeiros realizados
    """
    if 'usuario_id' not in session:
        return redirect(url_for('login'))

    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario or usuario.tipo != 'admin':
        flash('Acesso negado. Apenas administradores podem recalcular saldos.', 'error')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        try:
            # Buscar todas as contas caixa
            contas = ContaCaixa.query.all()
            contas_atualizadas = 0

            for conta in contas:
                # Buscar todos os lançamentos REALIZADOS desta conta
                lancamentos_realizados = Lancamento.query.filter_by(
                    conta_caixa_id=conta.id,
                    realizado=True
                ).all()

                # Calcular saldo: saldo_inicial + entradas - saidas
                total_entradas = sum([l.valor for l in lancamentos_realizados if l.tipo == 'entrada'])
                total_saidas = sum([l.valor for l in lancamentos_realizados if l.tipo == 'saida'])

                saldo_calculado = conta.saldo_inicial + total_entradas - total_saidas

                # Atualizar apenas se houver diferença
                if abs(conta.saldo_atual - saldo_calculado) > 0.01:  # Tolerância de 1 centavo
                    app.logger.info(f"💰 Conta '{conta.nome}': Saldo antigo R$ {conta.saldo_atual:.2f} → Novo R$ {saldo_calculado:.2f}")
                    conta.saldo_atual = saldo_calculado
                    contas_atualizadas += 1

            db.session.commit()

            if contas_atualizadas > 0:
                flash(f'✅ Saldos recalculados com sucesso! {contas_atualizadas} conta(s) atualizada(s).', 'success')
            else:
                flash('✅ Todos os saldos já estão corretos. Nenhuma atualização necessária.', 'info')

        except Exception as e:
            db.session.rollback()
            app.logger.error(f"❌ Erro ao recalcular saldos: {str(e)}")
            flash(f'Erro ao recalcular saldos: {str(e)}', 'error')

        return redirect(url_for('admin_recalcular_saldos'))

    # GET: Mostrar página com informações das contas
    contas = ContaCaixa.query.all()
    informacoes_contas = []

    for conta in contas:
        # Buscar lançamentos realizados
        lancamentos_realizados = Lancamento.query.filter_by(
            conta_caixa_id=conta.id,
            realizado=True
        ).all()

        total_entradas = sum([l.valor for l in lancamentos_realizados if l.tipo == 'entrada'])
        total_saidas = sum([l.valor for l in lancamentos_realizados if l.tipo == 'saida'])
        saldo_calculado = conta.saldo_inicial + total_entradas - total_saidas

        diferenca = conta.saldo_atual - saldo_calculado

        informacoes_contas.append({
            'conta': conta,
            'saldo_atual': conta.saldo_atual,
            'saldo_calculado': saldo_calculado,
            'diferenca': diferenca,
            'total_entradas': total_entradas,
            'total_saidas': total_saidas,
            'qtd_lancamentos': len(lancamentos_realizados)
        })

    return render_template('admin_recalcular_saldos.html', informacoes_contas=informacoes_contas)

@app.route('/vendas/<int:venda_id>/editar', methods=['GET', 'POST'])
def editar_venda(venda_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))

    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))

    # Obter empresa_id correta e validar acesso (filtro direto por empresa_id)
    empresa_id = obter_empresa_id_sessao(session, usuario)
    venda = Venda.query.filter_by(id=venda_id, empresa_id=empresa_id).first()
    if not venda:
        flash('Venda não encontrada.', 'danger')
        return redirect(url_for('vendas'))
    
    if request.method == 'POST':
        try:
            app.logger.info(f'Editando venda ID: {venda_id}')
            
            # Salvar valores antigos para cálculo de saldo
            valor_antigo = venda.valor_final if venda.valor_final else venda.valor
            realizado_antigo = venda.realizado
            tipo_pagamento_antigo = venda.tipo_pagamento
            numero_parcelas_antigo = venda.numero_parcelas or 1
            
            venda.cliente_id = request.form['cliente_id']
            
            # NOVO: Processar itens do carrinho (mesma lógica de nova_venda)
            item_nomes = request.form.getlist('item_nome[]')
            item_precos = request.form.getlist('item_preco[]')
            item_qtds = request.form.getlist('item_qtd[]')
            item_totais = request.form.getlist('item_total[]')
            item_tipos = request.form.getlist('item_tipo[]')
            item_descontos = request.form.getlist('item_desconto[]')
            
            # Validar se há pelo menos um item no carrinho
            if not item_nomes or len(item_nomes) == 0 or all(not nome.strip() for nome in item_nomes):
                # Tentar formato antigo (campo único produto)
                produto = request.form.get('produto', '').strip()
                if not produto:
                    flash('Adicione pelo menos um item ao carrinho.', 'error')
                    return redirect(url_for('editar_venda', venda_id=venda_id))
                else:
                    # Usar formato antigo
                    item_nomes = [produto]
                    item_precos = [request.form.get('valor', '0')]
                    item_qtds = [request.form.get('quantidade', '1')]
                    item_totais = []
                    item_tipos = ['mercadoria']
                    item_descontos = ['0']
            
            # Filtrar itens vazios e calcular totais
            itens_validos = []
            for i, nome in enumerate(item_nomes):
                nome_limpo = nome.strip() if nome else ''
                if nome_limpo:
                    # Parse de valores
                    try:
                        preco = float(item_precos[i].replace(',', '.')) if i < len(item_precos) and item_precos[i] else 0
                    except (ValueError, AttributeError):
                        preco = 0
                    
                    try:
                        qtd = float(item_qtds[i]) if i < len(item_qtds) and item_qtds[i] else 1
                    except (ValueError, AttributeError):
                        qtd = 1
                    
                    try:
                        desc_item = float(item_descontos[i].replace(',', '.')) if i < len(item_descontos) and item_descontos[i] else 0
                    except (ValueError, AttributeError):
                        desc_item = 0
                    
                    # Calcular total
                    try:
                        total_str = item_totais[i] if i < len(item_totais) and item_totais[i] else '0'
                        total = float(total_str.replace(',', '.')) if total_str else 0
                    except (ValueError, AttributeError):
                        total = 0
                    
                    if total == 0 and preco > 0:
                        total = max(0, (preco * qtd) - desc_item)
                    
                    tipo_item = item_tipos[i] if i < len(item_tipos) and item_tipos[i] else 'produto'
                    if tipo_item in ['produto', 'mercadoria']:
                        tipo_item = 'produto'
                    elif tipo_item in ['servico', 'serviço']:
                        tipo_item = 'servico'
                    
                    itens_validos.append({
                        'nome': nome_limpo,
                        'preco': preco,
                        'qtd': qtd,
                        'total': total,
                        'tipo': tipo_item,
                        'desconto': desc_item
                    })
            
            if not itens_validos:
                flash('Adicione pelo menos um item válido ao carrinho.', 'error')
                return redirect(url_for('editar_venda', venda_id=venda_id))
            
            # Calcular valor total do carrinho
            valor_total_carrinho = sum(item['total'] for item in itens_validos)
            if valor_total_carrinho <= 0:
                flash('Valor total deve ser maior que zero.', 'error')
                return redirect(url_for('editar_venda', venda_id=venda_id))
            
            # Se houver múltiplos itens, concatenar os nomes
            if len(itens_validos) > 1:
                nomes_itens = [item['nome'] for item in itens_validos]
                produto = ', '.join(nomes_itens)
            else:
                produto = itens_validos[0]['nome']
            
            # Usar primeiro item para quantidade (compatibilidade)
            primeiro_item = itens_validos[0]
            quantidade = sum(item['qtd'] for item in itens_validos)  # Total de itens
            # valor deve ser o total do carrinho ANTES do desconto geral
            valor = valor_total_carrinho
            
            # Determinar tipo_venda baseado nos itens
            tipos_itens = [item['tipo'] for item in itens_validos]
            tipo_venda = 'produto'
            if tipos_itens:
                if all(tipo == 'produto' for tipo in tipos_itens):
                    tipo_venda = 'produto'
                elif any(tipo == 'servico' for tipo in tipos_itens):
                    tipo_venda = 'servico'
                elif any(tipo == 'frete' for tipo in tipos_itens):
                    tipo_venda = 'frete'
                elif any(tipo == 'saida' for tipo in tipos_itens):
                    tipo_venda = 'saida'
                else:
                    tipo_venda = tipos_itens[0] if tipos_itens else 'produto'
            
            if tipo_venda not in ['produto', 'servico', 'frete', 'saida']:
                tipo_venda = 'produto'
            
            venda.produto = produto
            venda.valor = valor
            venda.quantidade = quantidade
            venda.tipo_venda = tipo_venda
            
            # UI envia no formato DD/MM/AAAA (com fallback para YYYY-MM-DD)
            try:
                venda.data_prevista = datetime.strptime(request.form['data_prevista'], '%d/%m/%Y').date()
            except ValueError:
                venda.data_prevista = datetime.strptime(request.form['data_prevista'], '%Y-%m-%d').date()
            venda.observacoes = request.form.get('observacoes', '')
            venda.nota_fiscal = request.form.get('nota_fiscal', '').strip() or None
            
            # Novos campos
            venda.tipo_pagamento = request.form.get('tipo_pagamento', 'a_vista')
            venda.numero_parcelas = int(request.form.get('numero_parcelas', 1))
            desconto_geral = float(request.form.get('desconto', 0.0))
            venda.desconto = desconto_geral
            
            # Calcular valor final
            venda.valor_final = valor_total_carrinho - desconto_geral
            venda.valor_final = max(0.0, venda.valor_final)
            venda.valor_parcela = venda.valor_final / venda.numero_parcelas if venda.numero_parcelas > 1 else venda.valor_final
            
            # Verificar se deve ser marcada como realizada
            data_realizada = request.form.get('data_realizada')
            if data_realizada and data_realizada.strip():
                # Tentar formato DD/MM/AAAA primeiro, depois AAAA-MM-DD
                try:
                    venda.data_realizada = datetime.strptime(data_realizada, '%d/%m/%Y').date()
                except ValueError:
                    venda.data_realizada = datetime.strptime(data_realizada, '%Y-%m-%d').date()
                from datetime import date
                venda.realizado = venda.data_realizada <= date.today()
            else:
                venda.data_realizada = None
                venda.realizado = False
            
            # Recriar lançamentos/parcelas quando alterar condição de pagamento
            alterou_condicao = (
                (tipo_pagamento_antigo != venda.tipo_pagamento) or 
                (venda.tipo_pagamento == 'parcelado' and numero_parcelas_antigo != venda.numero_parcelas)
            )

            if alterou_condicao:
                # Apagar parcelas e lançamentos vinculados a esta venda
                parcelas = Parcela.query.filter_by(venda_id=venda.id).all()
                for parcela in parcelas:
                    if parcela.lancamento_id:
                        lanc = db.session.get(Lancamento, parcela.lancamento_id)
                        if lanc:
                            db.session.delete(lanc)
                    db.session.delete(parcela)
                # Apagar lançamento único, se existir
                lancs = Lancamento.query.filter_by(venda_id=venda.id).all()
                for l in lancs:
                    db.session.delete(l)
                venda.lancamento_financeiro = None
                db.session.flush()

                # Criar novamente conforme a condição nova
                if venda.tipo_pagamento == 'parcelado' and venda.numero_parcelas > 1:
                    ok, msg = criar_parcelas_automaticas(venda, 'venda', usuario.id)
                    app.logger.info(f"Parcelamento refeito na edição: {ok} - {msg}")
                else:
                    lanc = criar_lancamento_financeiro_automatico(venda, 'venda', usuario.id)
                    if lanc:
                        venda.lancamento_financeiro = lanc
            else:
                # Atualizar TODOS os lançamentos financeiros vinculados (pode haver múltiplos para parcelado)
                # 1. Buscar lançamentos com venda_id direto
                lancamentos_vinculados = Lancamento.query.filter_by(venda_id=venda.id).all()

                # 2. Buscar lançamentos através da tabela Vinculo
                if not lancamentos_vinculados:
                    vinculos = Vinculo.query.filter(
                        db.or_(
                            db.and_(Vinculo.lado_a_tipo == 'venda', Vinculo.lado_a_id == venda.id, Vinculo.lado_b_tipo == 'lancamento'),
                            db.and_(Vinculo.lado_b_tipo == 'venda', Vinculo.lado_b_id == venda.id, Vinculo.lado_a_tipo == 'lancamento')
                        )
                    ).all()

                    for vinculo in vinculos:
                        if vinculo.lado_a_tipo == 'lancamento':
                            lanc = db.session.get(Lancamento, vinculo.lado_a_id)
                            if lanc:
                                lancamentos_vinculados.append(lanc)
                        else:
                            lanc = db.session.get(Lancamento, vinculo.lado_b_id)
                            if lanc:
                                lancamentos_vinculados.append(lanc)

                # Atualizar cada lançamento
                conta_caixa_id = request.form.get('conta_caixa_id')

                if len(lancamentos_vinculados) == 1:
                    # Apenas um lançamento (à vista ou primeira parcela)
                    lanc = lancamentos_vinculados[0]
                    lanc.descricao = f'{venda.cliente.nome if venda.cliente else "Cliente"} - {venda.produto}'
                    lanc.valor = venda.valor_final
                    lanc.data_prevista = venda.data_prevista
                    lanc.data_realizada = venda.data_realizada
                    lanc.realizado = venda.realizado
                    lanc.conta_caixa_id = conta_caixa_id if conta_caixa_id else None
                    lanc.usuario_ultima_edicao_id = usuario.id
                    lanc.data_ultima_edicao = datetime.utcnow()

                    # Atualizar itens_carrinho JSON com os itens da venda editada
                    import json
                    # Calcular desconto real a partir da diferença entre soma dos itens e valor_final
                    total_itens = sum(item['total'] for item in itens_validos)
                    desconto_real = total_itens - venda.valor_final

                    # Distribuir desconto proporcionalmente entre os itens
                    if desconto_real > 0 and len(itens_validos) > 0:
                        for item in itens_validos:
                            # Calcular desconto proporcional deste item
                            if total_itens > 0:
                                desconto_proporcional = (item['total'] / total_itens) * desconto_real
                                item['desconto'] += desconto_proporcional
                                item['total'] -= desconto_proporcional

                    lanc.itens_carrinho = json.dumps(itens_validos, ensure_ascii=False)

                    app.logger.info(f"✅ Lançamento único {lanc.id} atualizado para venda {venda.id}")

                elif len(lancamentos_vinculados) > 1:
                    # Múltiplos lançamentos (parcelado)
                    # Atualizar valor de cada parcela
                    num_lancamentos = len(lancamentos_vinculados)
                    valor_por_parcela = venda.valor_final / num_lancamentos

                    for i, lanc in enumerate(lancamentos_vinculados, 1):
                        lanc.descricao = f'{venda.cliente.nome if venda.cliente else "Cliente"} - {venda.produto} - Parcela {i}/{num_lancamentos}'
                        lanc.valor = valor_por_parcela
                        # Manter data prevista e realizada de cada parcela (não sobrescrever)
                        lanc.conta_caixa_id = conta_caixa_id if conta_caixa_id else None
                        lanc.usuario_ultima_edicao_id = usuario.id
                        lanc.data_ultima_edicao = datetime.utcnow()
                        app.logger.info(f"✅ Lançamento parcela {i}/{num_lancamentos} (ID {lanc.id}) atualizado para venda {venda.id}")

            db.session.commit()
            
            app.logger.info(f'Venda {venda_id} atualizada com sucesso!')
            flash('Venda atualizada com sucesso!', 'success')
            return redirect(url_for('vendas'))
            
        except Exception as e:
            app.logger.error(f'Erro ao editar venda {venda_id}: {str(e)}')
            flash(f'Erro ao atualizar venda: {str(e)}', 'error')
            db.session.rollback()
    
    # Buscar clientes para o select
    clientes = Cliente.query.filter_by(usuario_id=usuario.id).all()
    
    # Buscar produtos e serviços para seleção dinâmica
    produtos = Produto.query.filter_by(usuario_id=usuario.id).all()
    servicos = Servico.query.filter_by(usuario_id=usuario.id).all()
    
    # Buscar contas caixa para seleção
    contas_caixa = ContaCaixa.query.filter_by(usuario_id=usuario.id, ativo=True).order_by(ContaCaixa.nome).all()

    # Converter para dicionários para JSON serialization
    produtos_dict = [{'id': p.id, 'nome': p.nome, 'preco_venda': p.preco_venda, 'estoque': p.estoque} for p in produtos]
    servicos_dict = [{'id': s.id, 'nome': s.nome, 'preco': s.preco} for s in servicos]

    # Buscar TODOS os lançamentos vinculados a esta venda (pode ter múltiplos por causa de parcelamento)
    itens_carrinho_json = '[]'  # String JSON vazia por padrão
    lancamentos_vinculados_ids = []
    try:
        # Buscar TODOS os lançamentos vinculados a esta venda
        lancamentos_vinculados = Lancamento.query.filter_by(venda_id=venda.id).all()
        if lancamentos_vinculados:
            lancamentos_vinculados_ids = [lanc.id for lanc in lancamentos_vinculados]
            # Usar o primeiro lançamento para carregar itens_carrinho
            primeiro_lancamento = lancamentos_vinculados[0]
            if primeiro_lancamento.itens_carrinho:
                itens_carrinho_json = primeiro_lancamento.itens_carrinho
                import json
                itens_count = len(json.loads(itens_carrinho_json))
                app.logger.info(f"✅ Carregados {itens_count} itens do carrinho para venda {venda.id} ({len(lancamentos_vinculados)} lançamentos vinculados)")
    except Exception as e:
        app.logger.warning(f"⚠️ Erro ao carregar itens_carrinho da venda {venda.id}: {e}")
        itens_carrinho_json = '[]'

    return render_template('editar_venda.html',
                         venda=venda,
                         usuario=usuario,
                         clientes=clientes,
                         produtos=produtos_dict,
                         servicos=servicos_dict,
                         contas_caixa=contas_caixa,
                         itens_carrinho_json=itens_carrinho_json,
                         lancamentos_vinculados_ids=lancamentos_vinculados_ids)

@app.route('/compras/<int:compra_id>/editar', methods=['GET', 'POST'])
def editar_compra(compra_id):
    if 'usuario_id' not in session:
        return redirect(url_for('login'))

    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))

    # Obter empresa_id correta e validar acesso (filtro direto por empresa_id)
    empresa_id = obter_empresa_id_sessao(session, usuario)
    compra = Compra.query.filter_by(id=compra_id, empresa_id=empresa_id).first()
    if not compra:
        flash('Compra não encontrada.', 'danger')
        return redirect(url_for('compras'))
    
    if request.method == 'POST':
        try:
            app.logger.info(f'Editando compra ID: {compra_id}')
            
            # Salvar valores antigos para cálculo de saldo
            valor_antigo = compra.valor
            realizado_antigo = compra.realizado
            tipo_pagamento_antigo = compra.tipo_pagamento
            numero_parcelas_antigo = compra.numero_parcelas or 1
            
            compra.fornecedor_id = request.form['fornecedor_id']
            
            # NOVO: Processar itens do carrinho (mesma lógica de nova_compra)
            item_nomes = request.form.getlist('item_nome[]')
            item_precos = request.form.getlist('item_preco[]')
            item_qtds = request.form.getlist('item_qtd[]')
            item_totais = request.form.getlist('item_total[]')
            item_tipos = request.form.getlist('item_tipo[]')
            item_descontos = request.form.getlist('item_desconto[]')
            
            # Validar se há pelo menos um item no carrinho
            if not item_nomes or len(item_nomes) == 0 or all(not nome.strip() for nome in item_nomes):
                produto = request.form.get('produto', '').strip()
                if not produto:
                    flash('Adicione pelo menos um item ao carrinho.', 'error')
                    return redirect(url_for('editar_compra', compra_id=compra_id))
                else:
                    item_nomes = [produto]
                    item_precos = [request.form.get('valor', '0')]
                    item_qtds = [request.form.get('quantidade', '1')]
                    item_totais = []
                    item_tipos = ['mercadoria']
                    item_descontos = ['0']
            
            # Filtrar itens vazios e calcular totais
            itens_validos = []
            for i, nome in enumerate(item_nomes):
                nome_limpo = nome.strip() if nome else ''
                if nome_limpo:
                    try:
                        preco = float(item_precos[i].replace(',', '.')) if i < len(item_precos) and item_precos[i] else 0
                    except (ValueError, AttributeError):
                        preco = 0
                    
                    try:
                        qtd = float(item_qtds[i]) if i < len(item_qtds) and item_qtds[i] else 1
                    except (ValueError, AttributeError):
                        qtd = 1
                    
                    try:
                        desc_item = float(item_descontos[i].replace(',', '.')) if i < len(item_descontos) and item_descontos[i] else 0
                    except (ValueError, AttributeError):
                        desc_item = 0
                    
                    try:
                        total_str = item_totais[i] if i < len(item_totais) and item_totais[i] else '0'
                        total = float(total_str.replace(',', '.')) if total_str else 0
                    except (ValueError, AttributeError):
                        total = 0
                    
                    if total == 0 and preco > 0:
                        total = max(0, (preco * qtd) - desc_item)
                    
                    tipo_item = item_tipos[i] if i < len(item_tipos) and item_tipos[i] else 'mercadoria'
                    if tipo_item in ['produto', 'mercadoria']:
                        tipo_item = 'mercadoria'
                    elif tipo_item in ['servico', 'serviço']:
                        tipo_item = 'servico'
                    
                    itens_validos.append({
                        'nome': nome_limpo,
                        'preco': preco,
                        'qtd': qtd,
                        'total': total,
                        'tipo': tipo_item,
                        'desconto': desc_item
                    })
            
            if not itens_validos:
                flash('Adicione pelo menos um item válido ao carrinho.', 'error')
                return redirect(url_for('editar_compra', compra_id=compra_id))
            
            # Calcular valor total do carrinho
            valor_total_carrinho = sum(item['total'] for item in itens_validos)
            if valor_total_carrinho <= 0:
                flash('Valor total deve ser maior que zero.', 'error')
                return redirect(url_for('editar_compra', compra_id=compra_id))
            
            # Se houver múltiplos itens, concatenar os nomes
            if len(itens_validos) > 1:
                nomes_itens = [item['nome'] for item in itens_validos]
                produto = ', '.join(nomes_itens)
            else:
                produto = itens_validos[0]['nome']
            
            primeiro_item = itens_validos[0]
            valor = valor_total_carrinho  # Usar total do carrinho
            preco_custo = primeiro_item['preco']  # Preço unitário
            quantidade = sum(item['qtd'] for item in itens_validos)  # Total de itens
            
            # Determinar tipo_compra
            tipos_itens = [item['tipo'] for item in itens_validos]
            tipo_compra = 'mercadoria'
            if tipos_itens:
                if all(tipo == 'mercadoria' for tipo in tipos_itens):
                    tipo_compra = 'mercadoria'
                elif any(tipo == 'servico' for tipo in tipos_itens):
                    tipo_compra = 'servico'
                else:
                    tipo_compra = tipos_itens[0] if tipos_itens else 'mercadoria'
            
            compra.produto = produto
            compra.valor = valor
            compra.preco_custo = preco_custo
            compra.quantidade = quantidade
            compra.tipo_compra = tipo_compra
            
            # UI envia no formato DD/MM/AAAA
            try:
                compra.data_prevista = datetime.strptime(request.form['data_prevista'], '%d/%m/%Y').date()
            except ValueError:
                compra.data_prevista = datetime.strptime(request.form['data_prevista'], '%Y-%m-%d').date()
            
            compra.observacoes = request.form.get('observacoes', '')
            compra.nota_fiscal = request.form.get('nota_fiscal', '').strip() or None
            
            # Novos campos
            compra.tipo_pagamento = request.form.get('tipo_pagamento', 'a_vista')
            compra.numero_parcelas = int(request.form.get('numero_parcelas', 1))
            compra.valor_parcela = compra.valor / compra.numero_parcelas if compra.numero_parcelas > 1 else compra.valor
            
            # Verificar se deve ser marcada como realizada
            data_realizada = request.form.get('data_realizada')
            if data_realizada and data_realizada.strip():
                # Tentar formato DD/MM/AAAA primeiro, depois AAAA-MM-DD
                try:
                    compra.data_realizada = datetime.strptime(data_realizada, '%d/%m/%Y').date()
                except ValueError:
                    compra.data_realizada = datetime.strptime(data_realizada, '%Y-%m-%d').date()
                from datetime import date
                compra.realizado = compra.data_realizada <= date.today()
            else:
                compra.data_realizada = None
                compra.realizado = False
            
            # Recriar lançamentos/parcelas quando alterar condição de pagamento
            alterou_condicao = (
                (tipo_pagamento_antigo != compra.tipo_pagamento) or 
                (compra.tipo_pagamento == 'parcelado' and numero_parcelas_antigo != compra.numero_parcelas)
            )

            if alterou_condicao:
                # Apagar parcelas e lançamentos vinculados a esta compra
                parcelas = Parcela.query.filter_by(compra_id=compra.id).all()
                for parcela in parcelas:
                    if parcela.lancamento_id:
                        lanc = db.session.get(Lancamento, parcela.lancamento_id)
                        if lanc:
                            db.session.delete(lanc)
                    db.session.delete(parcela)
                # Apagar lançamento único, se existir
                lancs = Lancamento.query.filter_by(compra_id=compra.id).all()
                for l in lancs:
                    db.session.delete(l)
                compra.lancamento_financeiro = None
                db.session.flush()

                # Criar novamente conforme a condição nova
                if compra.tipo_pagamento == 'parcelado' and compra.numero_parcelas > 1:
                    ok, msg = criar_parcelas_automaticas(compra, 'compra', usuario.id)
                    app.logger.info(f"Parcelamento refeito na edição: {ok} - {msg}")
                else:
                    lanc = criar_lancamento_financeiro_automatico(compra, 'compra', usuario.id)
                    if lanc:
                        compra.lancamento_financeiro = lanc
            else:
                # Atualizar TODOS os lançamentos financeiros vinculados (pode haver múltiplos para parcelado)
                # 1. Buscar lançamentos com compra_id direto
                lancamentos_vinculados = Lancamento.query.filter_by(compra_id=compra.id).all()

                # 2. Buscar lançamentos através da tabela Vinculo
                if not lancamentos_vinculados:
                    vinculos = Vinculo.query.filter(
                        db.or_(
                            db.and_(Vinculo.lado_a_tipo == 'compra', Vinculo.lado_a_id == compra.id, Vinculo.lado_b_tipo == 'lancamento'),
                            db.and_(Vinculo.lado_b_tipo == 'compra', Vinculo.lado_b_id == compra.id, Vinculo.lado_a_tipo == 'lancamento')
                        )
                    ).all()

                    for vinculo in vinculos:
                        if vinculo.lado_a_tipo == 'lancamento':
                            lanc = db.session.get(Lancamento, vinculo.lado_a_id)
                            if lanc:
                                lancamentos_vinculados.append(lanc)
                        else:
                            lanc = db.session.get(Lancamento, vinculo.lado_b_id)
                            if lanc:
                                lancamentos_vinculados.append(lanc)

                # Atualizar cada lançamento
                conta_caixa_id = request.form.get('conta_caixa_id')

                if len(lancamentos_vinculados) == 1:
                    # Apenas um lançamento (à vista ou primeira parcela)
                    lanc = lancamentos_vinculados[0]
                    lanc.descricao = f'{compra.fornecedor.nome if compra.fornecedor else "Fornecedor"} - {compra.produto}'
                    lanc.valor = compra.valor
                    lanc.data_prevista = compra.data_prevista
                    lanc.data_realizada = compra.data_realizada
                    lanc.realizado = compra.realizado
                    lanc.fornecedor_id = compra.fornecedor_id
                    lanc.conta_caixa_id = conta_caixa_id if conta_caixa_id else None
                    lanc.usuario_ultima_edicao_id = usuario.id
                    lanc.data_ultima_edicao = datetime.utcnow()

                    # Atualizar itens_carrinho JSON com os itens da compra editada
                    import json
                    # Calcular desconto real a partir da diferença entre soma dos itens e valor
                    total_itens = sum(item['total'] for item in itens_validos)
                    desconto_real = total_itens - compra.valor

                    # Distribuir desconto proporcionalmente entre os itens
                    if desconto_real > 0 and len(itens_validos) > 0:
                        for item in itens_validos:
                            # Calcular desconto proporcional deste item
                            if total_itens > 0:
                                desconto_proporcional = (item['total'] / total_itens) * desconto_real
                                item['desconto'] += desconto_proporcional
                                item['total'] -= desconto_proporcional

                    lanc.itens_carrinho = json.dumps(itens_validos, ensure_ascii=False)

                    app.logger.info(f"✅ Lançamento único {lanc.id} atualizado para compra {compra.id}")

                elif len(lancamentos_vinculados) > 1:
                    # Múltiplos lançamentos (parcelado)
                    # Atualizar valor de cada parcela
                    num_lancamentos = len(lancamentos_vinculados)
                    valor_por_parcela = compra.valor / num_lancamentos

                    for i, lanc in enumerate(lancamentos_vinculados, 1):
                        lanc.descricao = f'{compra.fornecedor.nome if compra.fornecedor else "Fornecedor"} - {compra.produto} - Parcela {i}/{num_lancamentos}'
                        lanc.valor = valor_por_parcela
                        # Manter data prevista e realizada de cada parcela (não sobrescrever)
                        lanc.fornecedor_id = compra.fornecedor_id
                        lanc.conta_caixa_id = conta_caixa_id if conta_caixa_id else None
                        lanc.usuario_ultima_edicao_id = usuario.id
                        lanc.data_ultima_edicao = datetime.utcnow()
                        app.logger.info(f"✅ Lançamento parcela {i}/{num_lancamentos} (ID {lanc.id}) atualizado para compra {compra.id}")

            db.session.commit()
            
            app.logger.info(f'Compra {compra_id} atualizada com sucesso!')
            flash('Compra atualizada com sucesso!', 'success')
            return redirect(url_for('compras'))
            
        except Exception as e:
            app.logger.error(f'Erro ao editar compra {compra_id}: {str(e)}')
            flash(f'Erro ao atualizar compra: {str(e)}', 'error')
            db.session.rollback()
    
    # Buscar fornecedores para o select
    fornecedores = Fornecedor.query.filter_by(usuario_id=usuario.id).all()

    # Buscar contas caixa para seleção
    contas_caixa = ContaCaixa.query.filter_by(usuario_id=usuario.id, ativo=True).order_by(ContaCaixa.nome).all()

    # Buscar TODOS os lançamentos vinculados a esta compra (pode ter múltiplos por causa de parcelamento)
    itens_carrinho_json = '[]'  # String JSON vazia por padrão
    lancamentos_vinculados_ids = []
    try:
        # Buscar TODOS os lançamentos vinculados a esta compra
        lancamentos_vinculados = Lancamento.query.filter_by(compra_id=compra.id).all()
        if lancamentos_vinculados:
            lancamentos_vinculados_ids = [lanc.id for lanc in lancamentos_vinculados]
            # Usar o primeiro lançamento para carregar itens_carrinho
            primeiro_lancamento = lancamentos_vinculados[0]
            if primeiro_lancamento.itens_carrinho:
                itens_carrinho_json = primeiro_lancamento.itens_carrinho
                import json
                itens_count = len(json.loads(itens_carrinho_json))
                app.logger.info(f"✅ Carregados {itens_count} itens do carrinho para compra {compra.id} ({len(lancamentos_vinculados)} lançamentos vinculados)")
    except Exception as e:
        app.logger.warning(f"⚠️ Erro ao carregar itens_carrinho da compra {compra.id}: {e}")
        itens_carrinho_json = '[]'

    return render_template('editar_compra.html',
                         compra=compra,
                         usuario=usuario,
                         fornecedores=fornecedores,
                         contas_caixa=contas_caixa,
                         itens_carrinho_json=itens_carrinho_json,
                         lancamentos_vinculados_ids=lancamentos_vinculados_ids)

def exportar_relatorio_clientes_pdf(clientes_dados):
    """Exporta relatório completo de clientes em PDF com todos os detalhes da tela"""
    try:
        from reportlab.lib.pagesizes import letter, A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from io import BytesIO
        from flask import send_file

        # Criar buffer para o PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=30, bottomMargin=30)
        elements = []

        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, spaceAfter=20, alignment=1, textColor=colors.HexColor('#2c3e50'))
        subtitle_style = ParagraphStyle('SubTitle', parent=styles['Normal'], fontSize=10, spaceAfter=15, alignment=1, textColor=colors.grey)

        # Título
        elements.append(Paragraph("Relatório Completo de Clientes", title_style))
        elements.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}", subtitle_style))
        elements.append(Spacer(1, 15))

        # Calcular totais para o resumo
        total_clientes = len(clientes_dados)
        sum_realizado = sum(c.get('total_vendas', 0) for c in clientes_dados)
        sum_a_vencer = sum(c.get('total_vendas_pendentes', 0) for c in clientes_dados)
        sum_vencido = sum(c.get('saldo_vencido', 0) for c in clientes_dados)
        sum_agendado = sum(c.get('total_agendado', 0) for c in clientes_dados)
        sum_saldo_aberto = sum(c.get('saldo_aberto', 0) for c in clientes_dados)
        sum_total_geral = sum(c.get('total_geral', 0) for c in clientes_dados)
        sum_transacoes = sum(c.get('num_vendas', 0) for c in clientes_dados)

        # Tabela de resumo
        resumo_data = [['RESUMO GERAL', '', '', '', '', ''], ['Total Clientes', 'Realizado', 'A Vencer', 'Vencido', 'Agendado', 'Saldo Aberto'], [str(total_clientes), f"R$ {sum_realizado:.2f}", f"R$ {sum_a_vencer:.2f}", f"R$ {sum_vencido:.2f}", f"R$ {sum_agendado:.2f}", f"R$ {sum_saldo_aberto:.2f}"]]
        resumo_table = Table(resumo_data, colWidths=[100, 100, 100, 100, 100, 100])
        resumo_table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke), ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, 0), 12), ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#ecf0f1')), ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'), ('FONTSIZE', (0, 1), (-1, 1), 9), ('BACKGROUND', (0, 2), (-1, 2), colors.white), ('FONTSIZE', (0, 2), (-1, 2), 10), ('GRID', (0, 0), (-1, -1), 1, colors.grey), ('SPAN', (0, 0), (-1, 0))]))
        elements.append(resumo_table)
        elements.append(Spacer(1, 20))

        # Dados da tabela principal - TODAS as colunas como na tela
        data = [['Cliente', 'Email', 'Telefone', 'CPF/CNPJ', 'Realizado', 'A Vencer', 'Vencido', 'Agendado', 'Saldo Aberto', 'Total Geral', 'Trans.', 'Ticket Médio', 'Status']]

        for cliente_data in clientes_dados:
            cliente = cliente_data['cliente']
            # Determinar status
            if cliente_data.get('saldo_vencido', 0) > 0:
                status = 'Vencido'
            elif cliente_data.get('saldo_aberto', 0) > 0:
                status = 'Em Aberto'
            else:
                status = 'Em Dia'

            data.append([(cliente.nome[:20] + '...') if cliente.nome and len(cliente.nome) > 20 else (cliente.nome or '-'), (cliente.email[:25] + '...') if cliente.email and len(cliente.email) > 25 else (cliente.email or '-'), cliente.telefone or '-', cliente.cpf_cnpj or '-', f"R$ {cliente_data.get('total_vendas', 0):.2f}", f"R$ {cliente_data.get('total_vendas_pendentes', 0):.2f}", f"R$ {cliente_data.get('saldo_vencido', 0):.2f}", f"R$ {cliente_data.get('total_agendado', 0):.2f}", f"R$ {cliente_data.get('saldo_aberto', 0):.2f}", f"R$ {cliente_data.get('total_geral', 0):.2f}", str(cliente_data.get('num_vendas', 0)), f"R$ {cliente_data.get('ticket_medio', 0):.2f}", status])

        # Adicionar linha de totais
        ticket_medio_geral = sum_total_geral / sum_transacoes if sum_transacoes > 0 else 0
        data.append(['TOTAL GERAL', '', '', '', f"R$ {sum_realizado:.2f}", f"R$ {sum_a_vencer:.2f}", f"R$ {sum_vencido:.2f}", f"R$ {sum_agendado:.2f}", f"R$ {sum_saldo_aberto:.2f}", f"R$ {sum_total_geral:.2f}", str(sum_transacoes), f"R$ {ticket_medio_geral:.2f}", ''])

        # Criar tabela com larguras ajustadas
        col_widths = [60, 70, 55, 60, 50, 50, 50, 50, 55, 55, 35, 55, 45]
        table = Table(data, colWidths=col_widths)
        table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke), ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, 0), 7), ('BOTTOMPADDING', (0, 0), (-1, 0), 8), ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor('#f8f9fa')), ('FONTSIZE', (0, 1), (-1, -2), 7), ('GRID', (0, 0), (-1, -1), 0.5, colors.grey), ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#34495e')), ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke), ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'), ('FONTSIZE', (0, -1), (-1, -1), 8)]))

        elements.append(table)
        doc.build(elements)

        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name=f'relatorio_clientes_completo_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf', mimetype='application/pdf')

    except ImportError:
        flash('Biblioteca reportlab não instalada. Instale com: pip install reportlab', 'error')
        return redirect(url_for('relatorio_clientes'))
    except Exception as e:
        app.logger.error(f"Erro ao exportar PDF: {str(e)}")
        flash('Erro ao exportar PDF', 'error')
        return redirect(url_for('relatorio_clientes'))

def exportar_relatorio_clientes_excel(clientes_dados):
    """Exporta relatório completo de clientes em Excel com todos os detalhes da tela"""
    try:
        import pandas as pd
        from io import BytesIO
        from flask import send_file

        # Preparar dados com TODAS as colunas
        dados = []
        for cliente_data in clientes_dados:
            cliente = cliente_data['cliente']

            # Determinar status
            if cliente_data.get('saldo_vencido', 0) > 0:
                status = 'Vencido'
            elif cliente_data.get('saldo_aberto', 0) > 0:
                status = 'Em Aberto'
            else:
                status = 'Em Dia'

            dados.append({
                'Cliente': cliente.nome or '',
                'Email': cliente.email or '',
                'Telefone': cliente.telefone or '',
                'CPF/CNPJ': cliente.cpf_cnpj or '',
                'Endereço': cliente.endereco or '',
                'Realizado': cliente_data.get('total_vendas', 0),
                'A Vencer': cliente_data.get('total_vendas_pendentes', 0),
                'Vencido': cliente_data.get('saldo_vencido', 0),
                'Agendado': cliente_data.get('total_agendado', 0),
                'Saldo em Aberto': cliente_data.get('saldo_aberto', 0),
                'Total Geral': cliente_data.get('total_geral', 0),
                'Nº Transações': cliente_data.get('num_vendas', 0),
                'Ticket Médio': cliente_data.get('ticket_medio', 0),
                'Status': status
            })

        # Criar DataFrame
        df = pd.DataFrame(dados)

        # Calcular totais
        total_clientes = len(clientes_dados)
        sum_realizado = sum(c.get('total_vendas', 0) for c in clientes_dados)
        sum_a_vencer = sum(c.get('total_vendas_pendentes', 0) for c in clientes_dados)
        sum_vencido = sum(c.get('saldo_vencido', 0) for c in clientes_dados)
        sum_agendado = sum(c.get('total_agendado', 0) for c in clientes_dados)
        sum_saldo_aberto = sum(c.get('saldo_aberto', 0) for c in clientes_dados)
        sum_total_geral = sum(c.get('total_geral', 0) for c in clientes_dados)
        sum_transacoes = sum(c.get('num_vendas', 0) for c in clientes_dados)
        ticket_medio_geral = sum_total_geral / sum_transacoes if sum_transacoes > 0 else 0

        # Adicionar linha de totais
        totais = pd.DataFrame([{
            'Cliente': 'TOTAL GERAL',
            'Email': '',
            'Telefone': '',
            'CPF/CNPJ': '',
            'Endereço': '',
            'Realizado': sum_realizado,
            'A Vencer': sum_a_vencer,
            'Vencido': sum_vencido,
            'Agendado': sum_agendado,
            'Saldo em Aberto': sum_saldo_aberto,
            'Total Geral': sum_total_geral,
            'Nº Transações': sum_transacoes,
            'Ticket Médio': ticket_medio_geral,
            'Status': ''
        }])

        df = pd.concat([df, totais], ignore_index=True)

        # Criar buffer para o Excel
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Escrever resumo
            resumo_df = pd.DataFrame([{
                'Total Clientes': total_clientes,
                'Realizado': sum_realizado,
                'A Vencer': sum_a_vencer,
                'Vencido': sum_vencido,
                'Agendado': sum_agendado,
                'Saldo Aberto': sum_saldo_aberto
            }])
            resumo_df.to_excel(writer, sheet_name='Relatório de Clientes', startrow=0, index=False)

            # Escrever dados principais
            df.to_excel(writer, sheet_name='Relatório de Clientes', startrow=3, index=False)

            # Formatar
            workbook = writer.book
            worksheet = writer.sheets['Relatório de Clientes']

            # Ajustar larguras de colunas
            worksheet.column_dimensions['A'].width = 30
            worksheet.column_dimensions['B'].width = 30
            worksheet.column_dimensions['C'].width = 15
            worksheet.column_dimensions['D'].width = 18
            worksheet.column_dimensions['E'].width = 35
            for col in ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
                worksheet.column_dimensions[col].width = 15

        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name=f'relatorio_clientes_completo_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except ImportError:
        flash('Biblioteca pandas não instalada. Instale com: pip install pandas openpyxl', 'error')
        return redirect(url_for('relatorio_clientes'))
    except Exception as e:
        app.logger.error(f"Erro ao exportar Excel: {str(e)}")
        flash('Erro ao exportar Excel', 'error')
        return redirect(url_for('relatorio_clientes'))

def exportar_relatorio_fornecedores_pdf(fornecedores_dados):
    """Exporta relatório completo de fornecedores em PDF com todos os detalhes da tela"""
    try:
        from reportlab.lib.pagesizes import letter, A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from io import BytesIO
        from flask import send_file

        # Criar buffer para o PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=30, bottomMargin=30)
        elements = []

        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, spaceAfter=20, alignment=1, textColor=colors.HexColor('#2c3e50'))
        subtitle_style = ParagraphStyle('SubTitle', parent=styles['Normal'], fontSize=10, spaceAfter=15, alignment=1, textColor=colors.grey)

        # Título
        elements.append(Paragraph("Relatório Completo de Fornecedores", title_style))
        elements.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}", subtitle_style))
        elements.append(Spacer(1, 15))

        # Calcular totais para o resumo
        total_fornecedores = len(fornecedores_dados)
        sum_realizado = sum(f.get('total_compras', 0) for f in fornecedores_dados)
        sum_a_vencer = sum(f.get('total_compras_pendentes', 0) for f in fornecedores_dados)
        sum_vencido = sum(f.get('saldo_vencido', 0) for f in fornecedores_dados)
        sum_agendado = sum(f.get('total_agendado', 0) for f in fornecedores_dados)
        sum_saldo_aberto = sum(f.get('saldo_aberto', 0) for f in fornecedores_dados)
        sum_total_geral = sum(f.get('total_geral', 0) for f in fornecedores_dados)
        sum_transacoes = sum(f.get('num_compras', 0) for f in fornecedores_dados)

        # Tabela de resumo
        resumo_data = [['RESUMO GERAL', '', '', '', '', ''], ['Total Fornecedores', 'Realizado', 'A Vencer', 'Vencido', 'Agendado', 'Saldo Aberto'], [str(total_fornecedores), f"R$ {sum_realizado:.2f}", f"R$ {sum_a_vencer:.2f}", f"R$ {sum_vencido:.2f}", f"R$ {sum_agendado:.2f}", f"R$ {sum_saldo_aberto:.2f}"]]
        resumo_table = Table(resumo_data, colWidths=[100, 100, 100, 100, 100, 100])
        resumo_table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e74c3c')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke), ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, 0), 12), ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#ecf0f1')), ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'), ('FONTSIZE', (0, 1), (-1, 1), 9), ('BACKGROUND', (0, 2), (-1, 2), colors.white), ('FONTSIZE', (0, 2), (-1, 2), 10), ('GRID', (0, 0), (-1, -1), 1, colors.grey), ('SPAN', (0, 0), (-1, 0))]))
        elements.append(resumo_table)
        elements.append(Spacer(1, 20))

        # Dados da tabela principal - TODAS as colunas como na tela
        data = [['Fornecedor', 'Email', 'Telefone', 'CPF/CNPJ', 'Realizado', 'A Vencer', 'Vencido', 'Agendado', 'Saldo Aberto', 'Total Geral', 'Trans.', 'Ticket Médio', 'Status']]

        for fornecedor_data in fornecedores_dados:
            fornecedor = fornecedor_data['fornecedor']
            # Determinar status
            if fornecedor_data.get('saldo_vencido', 0) > 0:
                status = 'Vencido'
            elif fornecedor_data.get('saldo_aberto', 0) > 0:
                status = 'Em Aberto'
            else:
                status = 'Em Dia'

            data.append([(fornecedor.nome[:20] + '...') if fornecedor.nome and len(fornecedor.nome) > 20 else (fornecedor.nome or '-'), (fornecedor.email[:25] + '...') if fornecedor.email and len(fornecedor.email) > 25 else (fornecedor.email or '-'), fornecedor.telefone or '-', fornecedor.cpf_cnpj or '-', f"R$ {fornecedor_data.get('total_compras', 0):.2f}", f"R$ {fornecedor_data.get('total_compras_pendentes', 0):.2f}", f"R$ {fornecedor_data.get('saldo_vencido', 0):.2f}", f"R$ {fornecedor_data.get('total_agendado', 0):.2f}", f"R$ {fornecedor_data.get('saldo_aberto', 0):.2f}", f"R$ {fornecedor_data.get('total_geral', 0):.2f}", str(fornecedor_data.get('num_compras', 0)), f"R$ {fornecedor_data.get('ticket_medio', 0):.2f}", status])

        # Adicionar linha de totais
        ticket_medio_geral = sum_total_geral / sum_transacoes if sum_transacoes > 0 else 0
        data.append(['TOTAL GERAL', '', '', '', f"R$ {sum_realizado:.2f}", f"R$ {sum_a_vencer:.2f}", f"R$ {sum_vencido:.2f}", f"R$ {sum_agendado:.2f}", f"R$ {sum_saldo_aberto:.2f}", f"R$ {sum_total_geral:.2f}", str(sum_transacoes), f"R$ {ticket_medio_geral:.2f}", ''])

        # Criar tabela com larguras ajustadas
        col_widths = [60, 70, 55, 60, 50, 50, 50, 50, 55, 55, 35, 55, 45]
        table = Table(data, colWidths=col_widths)
        table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke), ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, 0), 7), ('BOTTOMPADDING', (0, 0), (-1, 0), 8), ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor('#f8f9fa')), ('FONTSIZE', (0, 1), (-1, -2), 7), ('GRID', (0, 0), (-1, -1), 0.5, colors.grey), ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#34495e')), ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke), ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'), ('FONTSIZE', (0, -1), (-1, -1), 8)]))

        elements.append(table)
        doc.build(elements)

        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name=f'relatorio_fornecedores_completo_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf', mimetype='application/pdf')

    except ImportError:
        flash('Biblioteca reportlab não instalada. Instale com: pip install reportlab', 'error')
        return redirect(url_for('relatorio_fornecedores'))
    except Exception as e:
        app.logger.error(f"Erro ao exportar PDF: {str(e)}")
        flash('Erro ao exportar PDF', 'error')
        return redirect(url_for('relatorio_fornecedores'))

def exportar_relatorio_fornecedores_excel(fornecedores_dados):
    """Exporta relatório completo de fornecedores em Excel com todos os detalhes da tela"""
    try:
        import pandas as pd
        from io import BytesIO
        from flask import send_file

        # Preparar dados com TODAS as colunas
        dados = []
        for fornecedor_data in fornecedores_dados:
            fornecedor = fornecedor_data['fornecedor']

            # Determinar status
            if fornecedor_data.get('saldo_vencido', 0) > 0:
                status = 'Vencido'
            elif fornecedor_data.get('saldo_aberto', 0) > 0:
                status = 'Em Aberto'
            else:
                status = 'Em Dia'

            dados.append({
                'Fornecedor': fornecedor.nome or '',
                'Email': fornecedor.email or '',
                'Telefone': fornecedor.telefone or '',
                'CPF/CNPJ': fornecedor.cpf_cnpj or '',
                'Endereço': fornecedor.endereco or '',
                'Realizado': fornecedor_data.get('total_compras', 0),
                'A Vencer': fornecedor_data.get('total_compras_pendentes', 0),
                'Vencido': fornecedor_data.get('saldo_vencido', 0),
                'Agendado': fornecedor_data.get('total_agendado', 0),
                'Saldo em Aberto': fornecedor_data.get('saldo_aberto', 0),
                'Total Geral': fornecedor_data.get('total_geral', 0),
                'Nº Transações': fornecedor_data.get('num_compras', 0),
                'Ticket Médio': fornecedor_data.get('ticket_medio', 0),
                'Status': status
            })
        
        # Criar DataFrame
        df = pd.DataFrame(dados)

        # Calcular totais
        total_fornecedores = len(fornecedores_dados)
        sum_realizado = sum(f.get('total_compras', 0) for f in fornecedores_dados)
        sum_a_vencer = sum(f.get('total_compras_pendentes', 0) for f in fornecedores_dados)
        sum_vencido = sum(f.get('saldo_vencido', 0) for f in fornecedores_dados)
        sum_agendado = sum(f.get('total_agendado', 0) for f in fornecedores_dados)
        sum_saldo_aberto = sum(f.get('saldo_aberto', 0) for f in fornecedores_dados)
        sum_total_geral = sum(f.get('total_geral', 0) for f in fornecedores_dados)
        sum_transacoes = sum(f.get('num_compras', 0) for f in fornecedores_dados)
        ticket_medio_geral = sum_total_geral / sum_transacoes if sum_transacoes > 0 else 0

        # Adicionar linha de totais
        totais = pd.DataFrame([{
            'Fornecedor': 'TOTAL GERAL',
            'Email': '',
            'Telefone': '',
            'CPF/CNPJ': '',
            'Endereço': '',
            'Realizado': sum_realizado,
            'A Vencer': sum_a_vencer,
            'Vencido': sum_vencido,
            'Agendado': sum_agendado,
            'Saldo em Aberto': sum_saldo_aberto,
            'Total Geral': sum_total_geral,
            'Nº Transações': sum_transacoes,
            'Ticket Médio': ticket_medio_geral,
            'Status': ''
        }])

        df = pd.concat([df, totais], ignore_index=True)

        # Criar buffer para o Excel
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Escrever resumo
            resumo_df = pd.DataFrame([{
                'Total Fornecedores': total_fornecedores,
                'Realizado': sum_realizado,
                'A Vencer': sum_a_vencer,
                'Vencido': sum_vencido,
                'Agendado': sum_agendado,
                'Saldo Aberto': sum_saldo_aberto
            }])
            resumo_df.to_excel(writer, sheet_name='Relatório de Fornecedores', startrow=0, index=False)

            # Escrever dados principais
            df.to_excel(writer, sheet_name='Relatório de Fornecedores', startrow=3, index=False)

            # Formatar
            workbook = writer.book
            worksheet = writer.sheets['Relatório de Fornecedores']

            # Ajustar larguras de colunas
            worksheet.column_dimensions['A'].width = 30
            worksheet.column_dimensions['B'].width = 30
            worksheet.column_dimensions['C'].width = 15
            worksheet.column_dimensions['D'].width = 18
            worksheet.column_dimensions['E'].width = 35
            for col in ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
                worksheet.column_dimensions[col].width = 15

        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name=f'relatorio_fornecedores_completo_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except ImportError:
        flash('Biblioteca pandas não instalada. Instale com: pip install pandas openpyxl', 'error')
        return redirect(url_for('relatorio_fornecedores'))
    except Exception as e:
        app.logger.error(f"Erro ao exportar Excel: {str(e)}")
        flash('Erro ao exportar Excel', 'error')
        return redirect(url_for('relatorio_fornecedores'))

@app.route('/api/categorias/<tipo>')
def api_categorias(tipo):
    """API para buscar categorias baseadas no tipo (entrada/saida)"""
    if 'usuario_id' not in session:
        return jsonify({'error': 'Usuário não autenticado'}), 401
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario:
        return jsonify({'error': 'Usuário não encontrado'}), 404
    
    if usuario.tipo == 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    
    if tipo not in ['entrada', 'saida']:
        return jsonify({'error': 'Tipo inválido'}), 400
    
    try:
        # Buscar categorias de todos os usuários da empresa vinculada (considerando acesso contador)
        empresa_id = obter_empresa_id_sessao(session, usuario)
        usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
        usuarios_ids = [u.id for u in usuarios_empresa]

        categorias = PlanoConta.query.filter(
            PlanoConta.empresa_id == empresa_id,
            PlanoConta.tipo == tipo,
            PlanoConta.natureza == 'analitica',
            PlanoConta.ativo == True
        ).order_by(PlanoConta.codigo).all()

        categorias_list = [{'id': cat.id, 'nome': cat.nome, 'codigo': cat.codigo} for cat in categorias]

        return jsonify({'categorias': categorias_list})

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao buscar categorias: {str(e)}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

# Endpoints para autocomplete de clientes e fornecedores
@app.route('/api/clientes/buscar', methods=['GET'])
def api_buscar_clientes():
    """API para buscar clientes com autocomplete"""
    if 'usuario_id' not in session:
        return jsonify({'error': 'Usuário não autenticado'}), 401
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario or usuario.tipo == 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    
    try:
        query = request.args.get('q', '').strip()

        # Buscar clientes da empresa vinculada (considerando acesso contador)
        empresa_id = obter_empresa_id_sessao(session, usuario)

        if query:
            clientes = Cliente.query.filter(
                Cliente.empresa_id == empresa_id,
                Cliente.nome.ilike(f'%{query}%')
            ).order_by(Cliente.nome).limit(10).all()
        else:
            clientes = Cliente.query.filter(
                Cliente.empresa_id == empresa_id
            ).order_by(Cliente.nome).limit(20).all()
        
        clientes_list = [{'id': c.id, 'nome': c.nome, 'email': c.email, 'telefone': c.telefone} for c in clientes]
        
        return jsonify({'clientes': clientes_list})
    
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao buscar clientes: {str(e)}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/fornecedores/buscar', methods=['GET'])
def api_buscar_fornecedores():
    """API para buscar fornecedores com autocomplete"""
    if 'usuario_id' not in session:
        return jsonify({'error': 'Usuário não autenticado'}), 401
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario or usuario.tipo == 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    
    try:
        query = request.args.get('q', '').strip()

        # Buscar fornecedores da empresa vinculada (considerando acesso contador)
        empresa_id = obter_empresa_id_sessao(session, usuario)

        if query:
            fornecedores = Fornecedor.query.filter(
                Fornecedor.empresa_id == empresa_id,
                Fornecedor.nome.ilike(f'%{query}%')
            ).order_by(Fornecedor.nome).limit(10).all()
        else:
            fornecedores = Fornecedor.query.filter(
                Fornecedor.empresa_id == empresa_id
            ).order_by(Fornecedor.nome).limit(20).all()
        
        fornecedores_list = [{'id': f.id, 'nome': f.nome, 'email': f.email, 'telefone': f.telefone} for f in fornecedores]
        
        return jsonify({'fornecedores': fornecedores_list})
    
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao buscar fornecedores: {str(e)}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/clientes/criar', methods=['POST'])
def api_criar_cliente():
    """API para criar novo cliente"""
    if 'usuario_id' not in session:
        return jsonify({'error': 'Usuário não autenticado'}), 401
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario or usuario.tipo == 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    
    try:
        data = request.get_json()
        nome = data.get('nome', '').strip()
        
        if not nome:
            return jsonify({'error': 'Nome é obrigatório'}), 400
        
        # Verificar se já existe cliente com mesmo nome
        cliente_existente = Cliente.query.filter(
            Cliente.usuario_id == usuario.id,
            Cliente.nome.ilike(nome)
        ).first()
        
        if cliente_existente:
            return jsonify({'error': 'Cliente com este nome já existe'}), 400
        
        # Obter empresa_id correto (considerando contexto de contador)
        empresa_id = obter_empresa_id_sessao(session, usuario)

        # Criar novo cliente
        novo_cliente = Cliente(
            nome=nome,
            email=data.get('email', '').strip(),
            telefone=data.get('telefone', '').strip(),
            cpf_cnpj=data.get('cpf_cnpj', '').strip(),
            endereco=data.get('endereco', '').strip(),
            usuario_id=usuario.id,
            empresa_id=empresa_id
        )
        
        db.session.add(novo_cliente)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'cliente': {
                'id': novo_cliente.id,
                'nome': novo_cliente.nome,
                'email': novo_cliente.email,
                'telefone': novo_cliente.telefone
            }
        })
    
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao criar cliente: {str(e)}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/produtos/criar', methods=['POST'])
def api_criar_produto():
    """API para criar novo produto (usada pelo autocomplete do carrinho)."""
    if 'usuario_id' not in session:
        return jsonify({'error': 'Usuário não autenticado'}), 401
    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario or usuario.tipo == 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    try:
        data = request.get_json(force=True, silent=True) or {}
        nome = (data.get('nome') or '').strip()
        if not nome:
            return jsonify({'error': 'Nome é obrigatório'}), 400

        # Evitar duplicados (mesmo usuário)
        existente = Produto.query.filter(
            Produto.usuario_id == usuario.id,
            db.func.lower(Produto.nome) == db.func.lower(nome)
        ).first()
        if existente:
            return jsonify({
                'success': True,
                'produto': {
                    'id': existente.id,
                    'nome': existente.nome,
                    'preco_venda': existente.preco_venda or 0,
                    'estoque': existente.estoque or 0
                }
            })

        preco_venda = float(data.get('preco_venda') or 0)
        estoque = int(data.get('estoque') or 0)
        descricao = (data.get('descricao') or '').strip()

        novo_produto = Produto(
            nome=nome,
            descricao=descricao,
            preco_custo=0,
            preco_venda=preco_venda,
            estoque=estoque,
            usuario_id=usuario.id
        )
        db.session.add(novo_produto)
        db.session.commit()

        return jsonify({
            'success': True,
            'produto': {
                'id': novo_produto.id,
                'nome': novo_produto.nome,
                'preco_venda': novo_produto.preco_venda or 0,
                'estoque': novo_produto.estoque or 0
            }
        })
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao criar produto: {str(e)}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/fornecedores/criar', methods=['POST'])
def api_criar_fornecedor():
    """API para criar novo fornecedor"""
    if 'usuario_id' not in session:
        return jsonify({'error': 'Usuário não autenticado'}), 401
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario or usuario.tipo == 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    
    try:
        data = request.get_json()
        nome = data.get('nome', '').strip()
        
        if not nome:
            return jsonify({'error': 'Nome é obrigatório'}), 400
        
        # Verificar se já existe fornecedor com mesmo nome
        fornecedor_existente = Fornecedor.query.filter(
            Fornecedor.usuario_id == usuario.id,
            Fornecedor.nome.ilike(nome)
        ).first()
        
        if fornecedor_existente:
            return jsonify({'error': 'Fornecedor com este nome já existe'}), 400
        
        # Obter empresa_id correto (considerando contexto de contador)
        empresa_id = obter_empresa_id_sessao(session, usuario)

        # Criar novo fornecedor
        novo_fornecedor = Fornecedor(
            nome=nome,
            email=data.get('email', '').strip(),
            telefone=data.get('telefone', '').strip(),
            cpf_cnpj=data.get('cpf_cnpj', '').strip(),
            endereco=data.get('endereco', '').strip(),
            usuario_id=usuario.id,
            empresa_id=empresa_id
        )
        
        db.session.add(novo_fornecedor)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'fornecedor': {
                'id': novo_fornecedor.id,
                'nome': novo_fornecedor.nome,
                'email': novo_fornecedor.email,
                'telefone': novo_fornecedor.telefone
            }
        })
    
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao criar fornecedor: {str(e)}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

# APIs para ações em lote
@app.route('/api/lancamentos/marcar-status-lote', methods=['POST'])
def api_marcar_status_lote():
    """API para marcar status de lançamentos em lote"""
    if 'usuario_id' not in session:
        return jsonify({'error': 'Usuário não autenticado'}), 401
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario or usuario.tipo == 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    
    try:
        data = request.get_json()
        lancamento_ids = data.get('lancamento_ids', [])
        realizado = data.get('realizado', True)
        
        if not lancamento_ids:
            return jsonify({'error': 'Nenhum lançamento selecionado'}), 400
        
        # Buscar lançamentos do usuário
        lancamentos = Lancamento.query.filter(
            Lancamento.id.in_(lancamento_ids),
            Lancamento.usuario_id == usuario.id
        ).all()
        
        if not lancamentos:
            return jsonify({'error': 'Nenhum lançamento válido encontrado'}), 404
        
        # Atualizar status
        count = 0
        for lancamento in lancamentos:
            lancamento.realizado = realizado
            count += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'count': count,
            'message': f'{count} lançamento(s) atualizado(s) com sucesso'
        })
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao marcar status em lote: {str(e)}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/importacao')
def importacao():
    """Página de importação de dados"""
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    return render_template('importacao.html', usuario=usuario)

@app.route('/api/backup/exportar-geral', methods=['GET'])
def api_exportar_backup_geral():
    """Exporta backup completo de todos os dados em CSV"""
    if 'usuario_id' not in session:
        return jsonify({'error': 'Usuário não autenticado'}), 401
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario:
        return jsonify({'error': 'Usuário não encontrado'}), 403
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io
        from datetime import datetime
        
        # Criar workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remover planilha padrão
        
        # Buscar todos os usuários da empresa
        empresa_id = obter_empresa_id_sessao(session, usuario)

        usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
        usuarios_ids = [u.id for u in usuarios_empresa]
        
        # ===== ABA 1: RESUMO GERAL =====
        ws_resumo = wb.create_sheet("📊 RESUMO GERAL", 0)
        
        # Estilos
        titulo_font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
        titulo_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        titulo_align = Alignment(horizontal="center", vertical="center")
        
        subtitulo_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        subtitulo_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        subtitulo_align = Alignment(horizontal="left", vertical="center")
        
        texto_font = Font(name='Arial', size=10)
        texto_align = Alignment(horizontal="left", vertical="center")
        
        # Título principal
        ws_resumo.merge_cells('A1:F1')
        cell_titulo = ws_resumo.cell(row=1, column=1, value=f"BACKUP COMPLETO - {usuario.empresa.razao_social if usuario.empresa else 'Sistema'}")
        cell_titulo.font = titulo_font
        cell_titulo.fill = titulo_fill
        cell_titulo.alignment = titulo_align
        
        # Data do backup
        ws_resumo.merge_cells('A2:F2')
        cell_data = ws_resumo.cell(row=2, column=1, value=f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}")
        cell_data.font = texto_font
        cell_data.alignment = Alignment(horizontal="center", vertical="center")
        
        # Estatísticas gerais
        ws_resumo.cell(row=4, column=1, value="ESTATÍSTICAS GERAIS").font = subtitulo_font
        ws_resumo.cell(row=4, column=1).fill = subtitulo_fill
        ws_resumo.cell(row=4, column=1).alignment = subtitulo_align
        
        # Contar registros
        total_lancamentos = Lancamento.query.filter(Lancamento.empresa_id == empresa_id).count()
        total_vendas = Venda.query.filter(Venda.empresa_id == empresa_id).count()
        total_compras = Compra.query.filter(Compra.empresa_id == empresa_id).count()
        total_clientes = Cliente.query.filter(Cliente.empresa_id == empresa_id).count()
        total_fornecedores = Fornecedor.query.filter(Fornecedor.empresa_id == empresa_id).count()
        total_contas_caixa = ContaCaixa.query.filter(ContaCaixa.usuario_id.in_(usuarios_ids)).count()
        total_produtos = Produto.query.filter(Produto.usuario_id.in_(usuarios_ids)).count()
        
        # Calcular totais financeiros
        lancamentos_receita = Lancamento.query.filter(
            Lancamento.empresa_id == empresa_id,
            Lancamento.tipo == 'entrada',
            Lancamento.realizado == True
        ).all()
        total_entradas = sum(l.valor for l in lancamentos_receita)
        
        lancamentos_despesa = Lancamento.query.filter(
            Lancamento.empresa_id == empresa_id,
            Lancamento.tipo == 'saida',
            Lancamento.realizado == True
        ).all()
        total_saidas = sum(l.valor for l in lancamentos_despesa)
        
        saldo_geral = total_entradas - total_saidas
        
        # Adicionar estatísticas
        stats = [
            ("Total de Lançamentos", total_lancamentos),
            ("Total de Vendas", total_vendas),
            ("Total de Compras", total_compras),
            ("Total de Clientes", total_clientes),
            ("Total de Fornecedores", total_fornecedores),
            ("Total de Contas Caixa", total_contas_caixa),
            ("Total de Produtos", total_produtos),
            ("", ""),
            ("TOTAL ENTRADAS", f"R$ {total_entradas:,.2f}"),
            ("TOTAL SAÍDAS", f"R$ {total_saidas:,.2f}"),
            ("SALDO GERAL", f"R$ {saldo_geral:,.2f}")
        ]
        
        for i, (descricao, valor) in enumerate(stats, 5):
            ws_resumo.cell(row=i, column=1, value=descricao).font = texto_font
            ws_resumo.cell(row=i, column=2, value=valor).font = texto_font
            if descricao.startswith("TOTAL") or descricao.startswith("SALDO"):
                ws_resumo.cell(row=i, column=1).font = Font(name='Arial', size=10, bold=True)
                ws_resumo.cell(row=i, column=2).font = Font(name='Arial', size=10, bold=True)
        
        # Ajustar larguras
        ws_resumo.column_dimensions['A'].width = 25
        ws_resumo.column_dimensions['B'].width = 20
        
        # ===== ABA 2: LANÇAMENTOS FINANCEIROS =====
        ws_lancamentos = wb.create_sheet("💰 LANÇAMENTOS")
        
        # Cabeçalhos
        headers_lancamentos = [
            'ID', 'Descrição', 'Valor', 'Tipo', 'Categoria', 'Data Prevista', 
            'Data Realizada', 'Realizado', 'Conta Caixa', 'Cliente', 'Fornecedor',
            'Observações', 'Tags', 'Centro Custo', 'Projeto', 'Documento', 'Usuário'
        ]
        
        for col, header in enumerate(headers_lancamentos, 1):
            cell = ws_lancamentos.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Buscar lançamentos
        lancamentos = Lancamento.query.filter(Lancamento.empresa_id == empresa_id).order_by(Lancamento.data_prevista.desc()).all()
        
        for row, lancamento in enumerate(lancamentos, 2):
            ws_lancamentos.cell(row=row, column=1, value=lancamento.id)
            ws_lancamentos.cell(row=row, column=2, value=lancamento.descricao)
            ws_lancamentos.cell(row=row, column=3, value=lancamento.valor)
            ws_lancamentos.cell(row=row, column=4, value=lancamento.tipo.upper())
            ws_lancamentos.cell(row=row, column=5, value=lancamento.categoria)
            ws_lancamentos.cell(row=row, column=6, value=lancamento.data_prevista.strftime('%d/%m/%Y') if lancamento.data_prevista else '')
            ws_lancamentos.cell(row=row, column=7, value=lancamento.data_realizada.strftime('%d/%m/%Y') if lancamento.data_realizada else '')
            ws_lancamentos.cell(row=row, column=8, value='SIM' if lancamento.realizado else 'NÃO')
            ws_lancamentos.cell(row=row, column=9, value=lancamento.conta_caixa.nome if lancamento.conta_caixa else '')
            ws_lancamentos.cell(row=row, column=10, value=lancamento.cliente.nome if lancamento.cliente else '')
            ws_lancamentos.cell(row=row, column=11, value=lancamento.fornecedor.nome if lancamento.fornecedor else '')
            ws_lancamentos.cell(row=row, column=12, value=lancamento.observacoes or '')
            ws_lancamentos.cell(row=row, column=13, value=lancamento.tags or '')
            ws_lancamentos.cell(row=row, column=14, value=lancamento.centro_custo or '')
            ws_lancamentos.cell(row=row, column=15, value=lancamento.projeto or '')
            ws_lancamentos.cell(row=row, column=16, value=lancamento.documento or '')
            ws_lancamentos.cell(row=row, column=17, value=lancamento.usuario.nome)
            
            # Formatação de valor
            ws_lancamentos.cell(row=row, column=3).number_format = 'R$ #,##0.00'
        
        # Ajustar larguras
        column_widths_lancamentos = [8, 30, 12, 10, 15, 12, 12, 8, 18, 20, 20, 25, 20, 15, 15, 15, 20]
        for col, width in enumerate(column_widths_lancamentos, 1):
            ws_lancamentos.column_dimensions[get_column_letter(col)].width = width
        
        # Congelar primeira linha
        ws_lancamentos.freeze_panes = 'A2'
        
        # ===== ABA 3: VENDAS =====
        ws_vendas = wb.create_sheet("🛒 VENDAS")
        
        headers_vendas = [
            'ID', 'Produto', 'Quantidade', 'Preço Unitário', 'Preço Total', 
            'Tipo Venda', 'Cliente', 'Realizado', 'Data Realizada', 'Observações', 'Usuário'
        ]
        
        for col, header in enumerate(headers_vendas, 1):
            cell = ws_vendas.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Buscar vendas
        vendas = Venda.query.filter(Venda.empresa_id == empresa_id).order_by(Venda.data_criacao.desc()).all()
        
        for row, venda in enumerate(vendas, 2):
            ws_vendas.cell(row=row, column=1, value=venda.id)
            ws_vendas.cell(row=row, column=2, value=venda.produto)
            ws_vendas.cell(row=row, column=3, value=venda.quantidade)
            ws_vendas.cell(row=row, column=4, value=venda.preco_unitario)
            ws_vendas.cell(row=row, column=5, value=venda.preco_total)
            ws_vendas.cell(row=row, column=6, value=venda.tipo_venda.upper())
            ws_vendas.cell(row=row, column=7, value=venda.cliente.nome if venda.cliente else '')
            ws_vendas.cell(row=row, column=8, value='SIM' if venda.realizado else 'NÃO')
            ws_vendas.cell(row=row, column=9, value=venda.data_realizada.strftime('%d/%m/%Y') if venda.data_realizada else '')
            ws_vendas.cell(row=row, column=10, value=venda.observacoes or '')
            ws_vendas.cell(row=row, column=11, value=venda.usuario.nome)
            
            # Formatação de valores
            ws_vendas.cell(row=row, column=4).number_format = 'R$ #,##0.00'
            ws_vendas.cell(row=row, column=5).number_format = 'R$ #,##0.00'
        
        # Ajustar larguras
        column_widths_vendas = [8, 30, 12, 15, 15, 12, 20, 8, 12, 25, 20]
        for col, width in enumerate(column_widths_vendas, 1):
            ws_vendas.column_dimensions[get_column_letter(col)].width = width
        
        ws_vendas.freeze_panes = 'A2'
        
        # ===== ABA 4: COMPRAS =====
        ws_compras = wb.create_sheet("🛍️ COMPRAS")
        
        headers_compras = [
            'ID', 'Produto', 'Quantidade', 'Preço Unitário', 'Preço Total', 
            'Tipo Compra', 'Fornecedor', 'Realizado', 'Data Realizada', 'Observações', 'Usuário'
        ]
        
        for col, header in enumerate(headers_compras, 1):
            cell = ws_compras.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Buscar compras
        compras = Compra.query.filter(Compra.empresa_id == empresa_id).order_by(Compra.data_criacao.desc()).all()
        
        for row, compra in enumerate(compras, 2):
            ws_compras.cell(row=row, column=1, value=compra.id)
            ws_compras.cell(row=row, column=2, value=compra.produto)
            ws_compras.cell(row=row, column=3, value=compra.quantidade)
            ws_compras.cell(row=row, column=4, value=compra.preco_unitario)
            ws_compras.cell(row=row, column=5, value=compra.preco_total)
            ws_compras.cell(row=row, column=6, value=compra.tipo_compra.upper())
            ws_compras.cell(row=row, column=7, value=compra.fornecedor.nome if compra.fornecedor else '')
            ws_compras.cell(row=row, column=8, value='SIM' if compra.realizado else 'NÃO')
            ws_compras.cell(row=row, column=9, value=compra.data_realizada.strftime('%d/%m/%Y') if compra.data_realizada else '')
            ws_compras.cell(row=row, column=10, value=compra.observacoes or '')
            ws_compras.cell(row=row, column=11, value=compra.usuario.nome)
            
            # Formatação de valores
            ws_compras.cell(row=row, column=4).number_format = 'R$ #,##0.00'
            ws_compras.cell(row=row, column=5).number_format = 'R$ #,##0.00'
        
        # Ajustar larguras
        column_widths_compras = [8, 30, 12, 15, 15, 12, 20, 8, 12, 25, 20]
        for col, width in enumerate(column_widths_compras, 1):
            ws_compras.column_dimensions[get_column_letter(col)].width = width
        
        ws_compras.freeze_panes = 'A2'
        
        # ===== ABA 5: CONTAS CAIXA =====
        ws_contas = wb.create_sheet("🏦 CONTAS CAIXA")
        
        headers_contas = [
            'ID', 'Nome', 'Tipo', 'Saldo Atual', 'Ativo', 'Usuário'
        ]
        
        for col, header in enumerate(headers_contas, 1):
            cell = ws_contas.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="17A2B8", end_color="17A2B8", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Buscar contas caixa
        contas = ContaCaixa.query.filter(ContaCaixa.usuario_id.in_(usuarios_ids)).all()
        
        for row, conta in enumerate(contas, 2):
            ws_contas.cell(row=row, column=1, value=conta.id)
            ws_contas.cell(row=row, column=2, value=conta.nome)
            ws_contas.cell(row=row, column=3, value=conta.tipo.upper())
            ws_contas.cell(row=row, column=4, value=conta.saldo_atual)
            ws_contas.cell(row=row, column=5, value='SIM' if conta.ativo else 'NÃO')
            ws_contas.cell(row=row, column=6, value=conta.usuario.nome)
            
            # Formatação de saldo
            ws_contas.cell(row=row, column=4).number_format = 'R$ #,##0.00'
        
        # Ajustar larguras
        column_widths_contas = [8, 25, 12, 15, 8, 20]
        for col, width in enumerate(column_widths_contas, 1):
            ws_contas.column_dimensions[get_column_letter(col)].width = width
        
        ws_contas.freeze_panes = 'A2'
        
        # ===== ABA 6: CLIENTES =====
        ws_clientes = wb.create_sheet("👥 CLIENTES")
        
        headers_clientes = [
            'ID', 'Nome', 'Email', 'Telefone', 'Endereço', 'CPF/CNPJ', 'Usuário'
        ]
        
        for col, header in enumerate(headers_clientes, 1):
            cell = ws_clientes.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="6F42C1", end_color="6F42C1", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Buscar clientes
        clientes = Cliente.query.filter(Cliente.empresa_id == empresa_id).all()
        
        for row, cliente in enumerate(clientes, 2):
            ws_clientes.cell(row=row, column=1, value=cliente.id)
            ws_clientes.cell(row=row, column=2, value=cliente.nome)
            ws_clientes.cell(row=row, column=3, value=cliente.email or '')
            ws_clientes.cell(row=row, column=4, value=cliente.telefone or '')
            ws_clientes.cell(row=row, column=5, value=cliente.endereco or '')
            ws_clientes.cell(row=row, column=6, value=cliente.cpf_cnpj or '')
            ws_clientes.cell(row=row, column=7, value=cliente.usuario.nome)
        
        # Ajustar larguras
        column_widths_clientes = [8, 30, 25, 15, 30, 18, 20]
        for col, width in enumerate(column_widths_clientes, 1):
            ws_clientes.column_dimensions[get_column_letter(col)].width = width
        
        ws_clientes.freeze_panes = 'A2'
        
        # ===== ABA 7: FORNECEDORES =====
        ws_fornecedores = wb.create_sheet("🏭 FORNECEDORES")
        
        headers_fornecedores = [
            'ID', 'Nome', 'Email', 'Telefone', 'Endereço', 'CPF/CNPJ', 'Usuário'
        ]
        
        for col, header in enumerate(headers_fornecedores, 1):
            cell = ws_fornecedores.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Buscar fornecedores
        fornecedores = Fornecedor.query.filter(Fornecedor.empresa_id == empresa_id).all()
        
        for row, fornecedor in enumerate(fornecedores, 2):
            ws_fornecedores.cell(row=row, column=1, value=fornecedor.id)
            ws_fornecedores.cell(row=row, column=2, value=fornecedor.nome)
            ws_fornecedores.cell(row=row, column=3, value=fornecedor.email or '')
            ws_fornecedores.cell(row=row, column=4, value=fornecedor.telefone or '')
            ws_fornecedores.cell(row=row, column=5, value=fornecedor.endereco or '')
            ws_fornecedores.cell(row=row, column=6, value=fornecedor.cpf_cnpj or '')
            ws_fornecedores.cell(row=row, column=7, value=fornecedor.usuario.nome)
        
        # Ajustar larguras
        column_widths_fornecedores = [8, 30, 25, 15, 30, 18, 20]
        for col, width in enumerate(column_widths_fornecedores, 1):
            ws_fornecedores.column_dimensions[get_column_letter(col)].width = width
        
        ws_fornecedores.freeze_panes = 'A2'
        
        # ===== ABA 6: PRODUTOS/ESTOQUE =====
        ws_produtos = wb.create_sheet("📦 PRODUTOS/ESTOQUE")
        
        headers_produtos = [
            'ID', 'Nome', 'Descrição', 'Preço Custo', 'Preço Venda', 
            'Estoque Atual', 'Categoria', 'Data Criação', 'Usuário'
        ]
        
        for col, header in enumerate(headers_produtos, 1):
            cell = ws_produtos.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Buscar produtos
        produtos = Produto.query.filter(Produto.usuario_id.in_(usuarios_ids)).order_by(Produto.nome).all()
        
        for row, produto in enumerate(produtos, 2):
            ws_produtos.cell(row=row, column=1, value=produto.id)
            ws_produtos.cell(row=row, column=2, value=produto.nome)
            ws_produtos.cell(row=row, column=3, value=produto.descricao or '')
            ws_produtos.cell(row=row, column=4, value=produto.preco_custo or 0)
            ws_produtos.cell(row=row, column=5, value=produto.preco_venda or 0)
            ws_produtos.cell(row=row, column=6, value=produto.estoque)
            ws_produtos.cell(row=row, column=7, value=produto.categoria or '')
            ws_produtos.cell(row=row, column=8, value=produto.data_criacao.strftime('%d/%m/%Y') if produto.data_criacao else '')
            ws_produtos.cell(row=row, column=9, value=produto.usuario.nome)
            
            # Formatação de valores
            ws_produtos.cell(row=row, column=4).number_format = 'R$ #,##0.00'
            ws_produtos.cell(row=row, column=5).number_format = 'R$ #,##0.00'
        
        # Ajustar larguras
        column_widths_produtos = [8, 30, 40, 15, 15, 12, 20, 12, 20]
        for col, width in enumerate(column_widths_produtos, 1):
            ws_produtos.column_dimensions[get_column_letter(col)].width = width
        
        ws_produtos.freeze_panes = 'A2'
        
        # ===== ABA 7: SALDOS DE CONTAS CAIXA =====
        ws_saldos = wb.create_sheet("💰 SALDOS CONTAS")
        
        headers_saldos = [
            'ID', 'Nome da Conta', 'Saldo Atual', 'Tipo', 'Descrição', 
            'Data Criação', 'Ativo', 'Usuário'
        ]
        
        for col, header in enumerate(headers_saldos, 1):
            cell = ws_saldos.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="6F42C1", end_color="6F42C1", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Buscar contas caixa
        contas_caixa = ContaCaixa.query.filter(ContaCaixa.usuario_id.in_(usuarios_ids)).order_by(ContaCaixa.nome).all()
        
        for row, conta in enumerate(contas_caixa, 2):
            ws_saldos.cell(row=row, column=1, value=conta.id)
            ws_saldos.cell(row=row, column=2, value=conta.nome)
            ws_saldos.cell(row=row, column=3, value=conta.saldo_atual)
            ws_saldos.cell(row=row, column=4, value=conta.tipo or '')
            ws_saldos.cell(row=row, column=5, value=conta.descricao or '')
            ws_saldos.cell(row=row, column=6, value=conta.data_criacao.strftime('%d/%m/%Y') if conta.data_criacao else '')
            ws_saldos.cell(row=row, column=7, value='SIM' if conta.ativo else 'NÃO')
            ws_saldos.cell(row=row, column=8, value=conta.usuario.nome)
            
            # Formatação de saldo
            ws_saldos.cell(row=row, column=3).number_format = 'R$ #,##0.00'
        
        # Ajustar larguras
        column_widths_saldos = [8, 25, 15, 12, 30, 12, 8, 20]
        for col, width in enumerate(column_widths_saldos, 1):
            ws_saldos.column_dimensions[get_column_letter(col)].width = width
        
        ws_saldos.freeze_panes = 'A2'
        
        # Salvar workbook em memória com tratamento de erro
        try:
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Verificar se o arquivo foi criado corretamente
            if output.getvalue():
                # Retornar arquivo Excel
                from flask import send_file
                nome_arquivo = f"backup_completo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                return send_file(
                    output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=nome_arquivo
                )
            else:
                app.logger.error("Arquivo de backup vazio gerado")
                return jsonify({'error': 'Erro ao gerar arquivo de backup'}), 500
                
        except Exception as save_error:
            app.logger.error(f"Erro ao salvar workbook: {str(save_error)}")
            return jsonify({'error': 'Erro ao salvar arquivo de backup'}), 500
        
    except Exception as e:
        app.logger.error(f"Erro ao exportar backup geral: {str(e)}")
        return jsonify({'error': 'Erro ao gerar backup'}), 500

@app.route('/api/importacao/exportar-modelo', methods=['GET'])
def api_exportar_modelo():
    """Exporta planilha modelo para importação"""
    if 'usuario_id' not in session:
        return jsonify({'error': 'Usuário não autenticado'}), 401
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario or usuario.tipo == 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
        from openpyxl.utils import get_column_letter
        import io
        
        # Criar workbook
        wb = Workbook()

        # Usar a planilha ativa como aba de dados
        ws_dados = wb.active
        ws_dados.title = "DADOS"
        
        # Cabeçalhos (sem coluna Valor - será calculado automaticamente pelo sistema)
        headers = [
            'Tipo', 'Descrição', 'Quantidade', 'Valor Unitário', 'Desconto', 'Categoria', 'Data Prevista',
            'Data Realizada', 'Conta Caixa', 'Tipo Cliente/Fornecedor', 'Cliente/Fornecedor', 'Observações',
            'Tipo Produto/Serviço', 'Nome Produto/Serviço', 'Parcelado?', 'Quantidade de Parcelas'
        ]
        
        # Definir estilo de borda (usado em toda a planilha)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Adicionar cabeçalhos com formatação
        for col, header in enumerate(headers, 1):
            cell = ws_dados.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        # Adicionar dados de exemplo (valor será calculado automaticamente pelo sistema na importação)
        # Formato: [Tipo, Descrição, Quantidade, Valor Unitário, Desconto, Categoria, Data Prevista, Data Realizada, Conta Caixa, Tipo Cliente/Fornecedor, Cliente/Fornecedor, Observações, Tipo Produto/Serviço, Nome Produto/Serviço, Parcelado?, Quantidade de Parcelas]
        exemplos = [
            ['entrada', 'Venda de produto eletrônico', 3, 500.00, 0.00, 'Vendas', '13/09/2025', '13/09/2025', 'nubank', 'cliente', 'João Silva', 'Venda realizada com desconto', 'produto', 'iphone 15', 'sim', 3],
            ['saida', 'Compra de material de escritório', 10, 30.00, 0.00, 'Compras', '14/09/2025', '', 'nubank', 'fornecedor', 'Fornecedor ABC', '', 'produto', 'Material Escritório', 'não', ''],
            ['entrada', 'Serviço de consultoria', 1, 800.00, 0.00, 'Serviços', '15/09/2025', '18/09/2025', 'nubank', 'cliente', 'Maria Santos', '', 'serviço', 'Consultoria', 'não', ''],
            ['saida', 'Pagamento de aluguel', 1, 1200.00, 0.00, 'Saídas Fixas', '16/09/2025', '', 'nubank', '', '', 'Aluguel do escritório', '', '', 'não', ''],
            ['entrada', 'Recebimento de comissão', 1, 500.00, 0.00, 'Comissões', '17/09/2025', '', 'nubank', 'cliente', 'Pedro Costa', 'Comissão de vendas', '', '', 'não', ''],
            ['saida', 'Compra de equipamento', 2, 1600.00, 200.00, 'Investimentos', '18/09/2025', '14/09/2025', 'nubank', 'fornecedor', 'Equipamentos XYZ', 'Equipamento parcelado', 'produto', 'iphone 16', 'sim', 2]
        ]

        # Adicionar exemplos com formatação
        for row, exemplo in enumerate(exemplos, 2):
            for col, valor in enumerate(exemplo, 1):
                cell = ws_dados.cell(row=row, column=col, value=valor)

                # Aplicar formato numérico nas colunas numéricas
                if col == 3:  # Coluna C - Quantidade
                    cell.number_format = '0'  # Formato inteiro
                    cell.alignment = Alignment(horizontal="right")
                elif col in [4, 5]:  # Colunas D e E - Valor Unitário e Desconto
                    cell.number_format = '#,##0.00'  # Formato decimal
                    cell.alignment = Alignment(horizontal="right")
                elif col == 16:  # Coluna P - Quantidade de Parcelas
                    if valor != '':
                        cell.number_format = '0'  # Formato inteiro
                        cell.alignment = Alignment(horizontal="right")

                cell.border = thin_border
        
        # Ajustar largura das colunas (sem coluna Valor)
        column_widths = [12, 30, 12, 15, 12, 18, 15, 15, 12, 20, 20, 25, 20, 12, 12, 18]
        for col, width in enumerate(column_widths, 1):
            ws_dados.column_dimensions[get_column_letter(col)].width = width
        
        # Congelar primeira linha
        ws_dados.freeze_panes = 'A2'

        # Adicionar filtros
        ws_dados.auto_filter.ref = f'A1:{get_column_letter(len(headers))}50000'

        # Criar arquivo em memória
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Retornar arquivo
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='modelo_importacao_lancamentos.xlsx'
        )
        
    except Exception as e:
        app.logger.error(f"Erro ao exportar modelo: {str(e)}")
        return jsonify({'error': 'Erro ao gerar modelo'}), 500

@app.route('/api/importacao/preview', methods=['POST'])
def api_preview_importacao():
    """Gera prévia da importação sem salvar no banco"""
    if 'usuario_id' not in session:
        return jsonify({'error': 'Usuário não autenticado'}), 401
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario or usuario.tipo == 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    
    try:
        if 'arquivo' not in request.files:
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400
        
        arquivo = request.files['arquivo']
        if arquivo.filename == '':
            return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
        
        if not arquivo.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Formato de arquivo não suportado. Use .xlsx ou .xls'}), 400
        
        # Ler arquivo Excel
        # data_only=True garante que, se houver fórmulas, seja lido o valor calculado
        from openpyxl import load_workbook
        wb = load_workbook(arquivo, data_only=True)
        
        # Detectar automaticamente a aba com dados
        ws = None
        
        # 1. Tentar encontrar aba específica de dados (planilha modelo)
        if "📊 DADOS" in wb.sheetnames:
            ws = wb["📊 DADOS"]
        elif "DADOS" in wb.sheetnames:
            ws = wb["DADOS"]
        else:
            # 2. Procurar por aba que contenha cabeçalhos de dados
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                if sheet.max_row > 0:
                    # Verificar se a primeira linha contém cabeçalhos de dados
                    first_row = [str(cell.value).strip() if cell.value else '' for cell in sheet[1]]
                    if (len(first_row) >= 2 and 
                        first_row[0].lower() in ['tipo'] and 
                        first_row[1].lower() in ['descrição', 'descricao', 'description']):
                        ws = sheet
                        break
            
            # 3. Se não encontrou, usar aba ativa
            if not ws:
                ws = wb.active
        
        # Encontrar automaticamente a linha com os cabeçalhos
        header_row = 1
        headers = []
        
        # Detectar cabeçalhos automaticamente
        headers = []
        header_row = 1
        
        # Procurar pela linha com cabeçalhos de dados
        for row_num in range(1, min(10, ws.max_row + 1)):  # Verificar até 10 linhas
            row_headers = [str(cell.value).strip() if cell.value else '' for cell in ws[row_num]]
            
            # Verificar se esta linha contém cabeçalhos de dados
            if (len(row_headers) >= 2 and 
                row_headers[0].lower() in ['tipo'] and 
                row_headers[1].lower() in ['descrição', 'descricao', 'description']):
                headers = row_headers
                header_row = row_num
                break
        
        # Se não encontrou, usar a primeira linha como fallback
        if not headers:
            headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
        
        app.logger.info(f"Aba selecionada: {ws.title}")
        app.logger.info(f"Total de linhas na aba: {ws.max_row}")
        app.logger.info(f"Cabeçalhos encontrados na linha {header_row}: {headers}")
        
        # Mapear títulos para nomes internos
        header_mapping = {
            'Tipo': 'TIPO',
            'Descrição': 'DESCRICAO', 
            'Valor': 'VALOR',
            'Quantidade': 'QUANTIDADE',
            'Qtd': 'QUANTIDADE',
            'Qtd.': 'QUANTIDADE',
            'Valor Unitário': 'VALOR_UNITARIO',
            'Valor Unitario': 'VALOR_UNITARIO',
            'Preço Unitário': 'VALOR_UNITARIO',
            'Preco Unitario': 'VALOR_UNITARIO',
            'Preço': 'VALOR_UNITARIO',
            'Preco': 'VALOR_UNITARIO',
            'Desconto': 'DESCONTO',
            'Desc': 'DESCONTO',
            'Categoria': 'CATEGORIA',
            'Data Prevista': 'DATA_PREVISTA',
            'Data Realizada': 'DATA_REALIZADA',
            'Conta Caixa': 'CONTA_CAIXA',
            'Conta': 'CONTA_CAIXA',
            'Caixa': 'CONTA_CAIXA',
            'Conta/Caixa': 'CONTA_CAIXA',
            'Conta - Caixa': 'CONTA_CAIXA',
            'Conta_Caixa': 'CONTA_CAIXA',
            'CONTA_CAIXA': 'CONTA_CAIXA',
            'conta_caixa': 'CONTA_CAIXA',
            'conta caixa': 'CONTA_CAIXA',
            'conta': 'CONTA_CAIXA',
            'caixa': 'CONTA_CAIXA',
            'Tipo Cliente/Fornecedor': 'TIPO_CLIENTE_FORNECEDOR',
            'Cliente/Fornecedor': 'CLIENTE_FORNECEDOR',
            'Observações': 'OBSERVACOES',
            'Tipo Produto/Serviço': 'TIPO_PRODUTO_SERVICO',
            'Nome Produto/Serviço': 'NOME_PRODUTO_SERVICO',
            'Tipo Produto/Servico': 'TIPO_PRODUTO_SERVICO',
            'Nome Produto/Servico': 'NOME_PRODUTO_SERVICO',
            'Parcelado': 'PARCELADO',
            'Quantidade de Parcelas': 'QUANTIDADE_PARCELAS',
            'Nota Fiscal': 'NOTA_FISCAL',
            'NF': 'NOTA_FISCAL',
            'Nota_Fiscal': 'NOTA_FISCAL',
            # Manter compatibilidade com títulos antigos
            'TIPO': 'TIPO',
            'DESCRICAO': 'DESCRICAO',
            'VALOR': 'VALOR',
            'CATEGORIA': 'CATEGORIA',
            'DATA_PREVISTA': 'DATA_PREVISTA',
            'DATA_REALIZADA': 'DATA_REALIZADA',
            'CONTA_CAIXA': 'CONTA_CAIXA',
            'CLIENTE': 'CLIENTE',
            'FORNECEDOR': 'FORNECEDOR',
            'OBSERVACOES': 'OBSERVACOES',
            'TAGS': 'TAGS',
            'CENTRO_CUSTO': 'CENTRO_CUSTO',
            'PROJETO': 'PROJETO',
            'DOCUMENTO': 'DOCUMENTO',
            'NOTA_FISCAL': 'NOTA_FISCAL',
            # Variações adicionais para garantir compatibilidade
            'Tipo Produto/Serviço': 'TIPO_PRODUTO_SERVICO',
            'Nome Produto/Serviço': 'NOME_PRODUTO_SERVICO',
            'Parcelado?': 'PARCELADO',
            'TIPO_PRODUTO_SERVICO': 'TIPO_PRODUTO_SERVICO',
            'NOME_PRODUTO_SERVICO': 'NOME_PRODUTO_SERVICO',
            'PARCELADO': 'PARCELADO',
            'QUANTIDADE_PARCELAS': 'QUANTIDADE_PARCELAS'
        }
        
        # Normalizar cabeçalhos usando mapeamento direto
        normalized_headers = []
        
        # Garantir que headers não está vazio
        if not headers:
            headers = ['Tipo', 'Descrição', 'Quantidade', 'Valor Unitário', 'Desconto', 'Categoria', 'Data Prevista']
        
        for header in headers:
            # Usar o mapeamento existente ou converter para maiúsculas
            normalized_header = header_mapping.get(header, header.upper())
            normalized_headers.append(normalized_header)
        
        app.logger.info(f"Cabeçalhos normalizados: {normalized_headers}")
        
        # Não validar cabeçalhos obrigatórios - aceitar qualquer planilha
        # Apenas verificar se há pelo menos uma linha de dados
        if ws.max_row < 2:
            return jsonify({
                'error': 'Planilha não contém dados para importar'
            }), 400
        
        # Verificar se há dados após os cabeçalhos
        if header_row >= ws.max_row:
            return jsonify({
                'error': 'Planilha não contém dados após os cabeçalhos'
            }), 400
        
        # Buscar todos os usuários da empresa para validações
        empresa_id = obter_empresa_id_sessao(session, usuario)

        usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
        usuarios_ids = [u.id for u in usuarios_empresa]
        
        # Cache de dados para otimização
        contas_caixa_cache = {c.nome.lower(): c for c in ContaCaixa.query.filter(
            ContaCaixa.usuario_id.in_(usuarios_ids), ContaCaixa.ativo==True
        ).all()}
        
        clientes_cache = {c.nome.lower(): c for c in Cliente.query.filter(
            Cliente.empresa_id == empresa_id
        ).all()}
        
        fornecedores_cache = {f.nome.lower(): f for f in Fornecedor.query.filter(
            Fornecedor.empresa_id == empresa_id
        ).all()}
        
        # Processar linhas para prévia
        preview_data = []
        erros = []
        
        for row_num in range(header_row + 1, ws.max_row + 1):
            try:
                # Verificar se a linha existe
                if row_num > ws.max_row:
                    break
                    
                row_data = [cell.value for cell in ws[row_num]]
                
                # Pular linhas vazias
                if not any(row_data):
                    continue
                
                # Criar dicionário de dados baseado nos cabeçalhos
                dados = {}
                for i, valor in enumerate(row_data):
                    if i < len(normalized_headers):
                        dados[normalized_headers[i]] = valor
                
                # Verificar se há pelo menos um campo essencial preenchido
                campos_essenciais = ['TIPO', 'DESCRICAO', 'VALOR', 'CATEGORIA', 'DATA_PREVISTA']
                tem_dados = any(str(dados.get(campo, '')).strip() for campo in campos_essenciais)
                
                if not tem_dados:
                    continue
                
                # Processar dados com valores padrão apenas quando necessário
                tipo_raw = str(dados.get('TIPO', '')).strip()
                descricao_raw = str(dados.get('DESCRICAO', '')).strip()
                valor_raw = dados.get('VALOR')
                quantidade_raw = dados.get('QUANTIDADE')
                valor_unitario_raw = dados.get('VALOR_UNITARIO')
                desconto_raw = dados.get('DESCONTO')
                categoria_raw = str(dados.get('CATEGORIA', '')).strip()
                data_prevista_raw = dados.get('DATA_PREVISTA')
                
                # Aplicar valores padrão apenas se campos estiverem vazios mas linha tiver dados
                tipo = normalizar_tipo(tipo_raw) if tipo_raw else 'entrada'
                descricao = descricao_raw if descricao_raw else f"Lançamento importado - Linha {row_num}"
                
                # Processar quantidade
                quantidade = processar_valor(quantidade_raw) if quantidade_raw else 1
                if quantidade is not None and quantidade <= 0:
                    erros.append(f"Linha {row_num}: Quantidade inválida ({quantidade_raw}). Convertido para 1.")
                    quantidade = 1  # Padrão mínimo
                elif quantidade is None:
                    quantidade = 1

                # Processar valor unitário
                valor_unitario = processar_valor(valor_unitario_raw) if valor_unitario_raw else None
                if valor_unitario is not None and valor_unitario <= 0:
                    erros.append(f"Linha {row_num}: Valor unitário inválido ({valor_unitario_raw}). Ignorado.")
                    valor_unitario = None

                # Processar desconto
                desconto = processar_valor(desconto_raw) if desconto_raw else 0
                if desconto is not None and desconto < 0:
                    erros.append(f"Linha {row_num}: Desconto negativo ({desconto_raw}). Convertido para 0.")
                    desconto = 0
                elif desconto is None:
                    desconto = 0

                # CALCULAR VALOR AUTOMATICAMENTE baseado em Quantidade, Valor Unitário e Desconto
                # Fórmula: valor = (quantidade × valor_unitario) - desconto
                if valor_unitario is not None and quantidade is not None:
                    # Calcular automaticamente (prioridade)
                    valor = (quantidade * valor_unitario) - desconto
                    if valor < 0:
                        valor = 0  # Não permitir valor negativo
                    valor_total = valor  # Para vendas/compras
                elif valor_raw:
                    # Fallback: usar valor da coluna VALOR se fornecido (compatibilidade com planilhas antigas)
                    valor = processar_valor(valor_raw)
                    if valor is None or valor <= 0:
                        valor = 0.01  # Valor padrão mínimo
                    valor_total = None  # Usar valor da parcela se não houver cálculo
                else:
                    # Sem dados suficientes para calcular
                    erros.append(f"Linha {row_num}: Dados insuficientes para calcular valor. Forneça Quantidade e Valor Unitário, ou Valor direto.")
                    valor = 0.01  # Valor padrão mínimo para não bloquear
                    valor_total = None
                
                # Buscar ou criar categoria no plano de contas automaticamente
                if categoria_raw:
                    categoria_id = buscar_ou_criar_categoria_plano_contas(categoria_raw, tipo, usuario.id, usuario.empresa_id)
                    if categoria_id:
                        categoria = categoria_raw
                    else:
                        categoria = 'Importado'  # Fallback se não conseguir criar
                else:
                    categoria = 'Importado'
                
                # Processar datas com múltiplos formatos
                data_prevista = processar_data(data_prevista_raw)
                if not data_prevista:
                    erros.append(f"Linha {row_num}: Data prevista ausente ou inválida ({data_prevista_raw}). Usando data atual.")
                    data_prevista = datetime.now().date()  # Data atual como padrão

                data_realizada = processar_data(dados.get('DATA_REALIZADA')) if dados.get('DATA_REALIZADA') else None
                
                # Debug: verificar se a data foi processada corretamente
                if dados.get('DATA_REALIZADA'):
                    app.logger.info(f"Linha {row_num}: Data Realizada original: '{dados.get('DATA_REALIZADA')}' -> Processada: {data_realizada}")
                
                # Processar campos de parcelamento
                avista_parcelado_str = str(dados.get('PARCELADO', '')).strip().lower()
                quantidade_parcelas_str = str(dados.get('QUANTIDADE_PARCELAS', '')).strip()
                
                # Determinar se é parcelado
                is_parcelado = avista_parcelado_str in ['sim', 's', 'yes', 'y', '1', 'true']
                quantidade_parcelas = 1
                
                if is_parcelado and quantidade_parcelas_str:
                    try:
                        quantidade_parcelas = int(quantidade_parcelas_str)
                        if quantidade_parcelas < 1:
                            quantidade_parcelas = 1
                    except (ValueError, TypeError):
                        quantidade_parcelas = 1
                
                # Processar campos opcionais
                observacoes = str(dados.get('OBSERVACOES', '')).strip() if dados.get('OBSERVACOES') else None
                tipo_produto_servico = str(dados.get('TIPO_PRODUTO_SERVICO', '')).strip() if dados.get('TIPO_PRODUTO_SERVICO') else None
                nome_produto_servico = str(dados.get('NOME_PRODUTO_SERVICO', '')).strip() if dados.get('NOME_PRODUTO_SERVICO') else None
                produto_servico = str(dados.get('PRODUTO_SERVICO', '')).strip() if dados.get('PRODUTO_SERVICO') else None
                
                # Importação: se data_realizada hoje ou passada => realizado; futura => agendado
                hoje = datetime.now().date()
                if data_realizada and isinstance(data_realizada, datetime):
                    data_realizada = data_realizada.date()
                realizado = bool(data_realizada and data_realizada <= hoje)
                
                # Buscar ou criar conta caixa automaticamente
                conta_caixa_input = dados.get('CONTA_CAIXA')
                conta_caixa_id = buscar_ou_criar_conta_caixa(conta_caixa_input, contas_caixa_cache, usuario.id, usuario.empresa_id)
                conta_caixa_nome = None
                
                # Debug: mostrar busca de conta caixa
                app.logger.info(f"Linha {row_num} - Busca conta caixa:")
                app.logger.info(f"  Input: '{conta_caixa_input}'")
                app.logger.info(f"  Cache disponível: {list(contas_caixa_cache.keys())}")
                app.logger.info(f"  ID encontrado: {conta_caixa_id}")
                
                if conta_caixa_id:
                    # Encontrar o nome da conta caixa pelo ID
                    for nome_cached, conta in contas_caixa_cache.items():
                        if conta.id == conta_caixa_id:
                            conta_caixa_nome = conta.nome
                            break
                    app.logger.info(f"  Nome da conta: '{conta_caixa_nome}'")
                else:
                    if conta_caixa_input:
                        app.logger.warning(f"  Conta caixa '{conta_caixa_input}' não pôde ser criada")
                    else:
                        app.logger.info(f"  Campo conta caixa vazio")
                
                # Debug: mostrar dados processados
                app.logger.info(f"Linha {row_num} - Dados processados:")
                app.logger.info(f"  TIPO_CLIENTE_FORNECEDOR: '{dados.get('TIPO_CLIENTE_FORNECEDOR')}'")
                app.logger.info(f"  CLIENTE_FORNECEDOR: '{dados.get('CLIENTE_FORNECEDOR')}'")
                app.logger.info(f"  TIPO_PRODUTO_SERVICO: '{dados.get('TIPO_PRODUTO_SERVICO')}'")
                app.logger.info(f"  NOME_PRODUTO_SERVICO: '{dados.get('NOME_PRODUTO_SERVICO')}'")
                
                # Processar cliente/fornecedor baseado no tipo
                tipo_cliente_fornecedor = str(dados.get('TIPO_CLIENTE_FORNECEDOR', '')).strip().lower()
                nome_cliente_fornecedor = str(dados.get('CLIENTE_FORNECEDOR', '')).strip()
                
                cliente_id = None
                fornecedor_id = None
                cliente_nome = None
                fornecedor_nome = None
                
                if tipo_cliente_fornecedor and nome_cliente_fornecedor:
                    if tipo_cliente_fornecedor == 'cliente':
                        cliente_id = buscar_cliente(nome_cliente_fornecedor, clientes_cache)
                        if cliente_id:
                            cliente_nome = clientes_cache[list(clientes_cache.keys())[list(clientes_cache.values()).index(next(c for c in clientes_cache.values() if c.id == cliente_id))]].nome
                    elif tipo_cliente_fornecedor == 'fornecedor':
                        fornecedor_id = buscar_fornecedor(nome_cliente_fornecedor, fornecedores_cache)
                        if fornecedor_id:
                            fornecedor_nome = fornecedores_cache[list(fornecedores_cache.keys())[list(fornecedores_cache.values()).index(next(f for f in fornecedores_cache.values() if f.id == fornecedor_id))]].nome
                
                # Determinar se criará venda/compra
                criara_venda = bool(cliente_id)
                criara_compra = bool(fornecedor_id)
                
                # VALIDAÇÃO: Verificar se o Valor Total está correto
                valor_total_calculado = None
                valor_total_divergente = False
                erro_validacao = None
                
                if valor_unitario is not None and quantidade is not None:
                    valor_total_calculado = (quantidade * valor_unitario) - desconto
                    if valor_total is not None:
                        # Verificar se há divergência (tolerância de 0.01 para arredondamentos)
                        diferenca = abs(valor_total - valor_total_calculado)
                        if diferenca > 0.01:
                            valor_total_divergente = True
                            erro_validacao = f"Valor Total divergente! Esperado: R$ {valor_total_calculado:.2f}, Informado: R$ {valor_total:.2f}"
                            app.logger.warning(f"Linha {row_num}: {erro_validacao}")
                
                # Adicionar à prévia
                preview_data.append({
                    'linha': row_num,
                    'tipo': tipo,
                    'descricao': descricao,
                    'valor': valor,
                    'quantidade': quantidade,
                    'valor_unitario': valor_unitario,
                    'desconto': desconto,
                    'valor_total': valor_total,
                    'valor_total_calculado': valor_total_calculado,
                    'valor_total_divergente': valor_total_divergente,
                    'erro_validacao': erro_validacao,
                    'categoria': categoria,
                    'data_prevista': data_prevista.strftime('%d/%m/%Y'),
                    'data_realizada': data_realizada.strftime('%d/%m/%Y') if data_realizada else None,
                    'realizado': realizado,
                    'nota_fiscal': str(dados.get('NOTA_FISCAL', '')).strip() if dados.get('NOTA_FISCAL') else None,
                    'conta_caixa': conta_caixa_nome,
                    'tipo_cliente_fornecedor': tipo_cliente_fornecedor,
                    'cliente_fornecedor': nome_cliente_fornecedor,
                    'observacoes': observacoes,
                    'tipo_produto_servico': tipo_produto_servico,
                    'nome_produto_servico': nome_produto_servico,
                    'produto_servico': produto_servico,
                    'parcelado': 'SIM' if is_parcelado and quantidade_parcelas > 1 else 'NÃO',
                    'quantidade_parcelas': quantidade_parcelas if is_parcelado and quantidade_parcelas > 1 else '',
                    'valor_parcela': f"R$ {(valor_total if valor_total is not None else valor)/quantidade_parcelas:.2f}" if is_parcelado and quantidade_parcelas > 1 else '',
                    'criara_venda': criara_venda,
                    'criara_compra': criara_compra,
                    'status': 'Realizado' if realizado else ('Agendado' if data_realizada and (data_realizada.date() if isinstance(data_realizada, datetime) else data_realizada) > datetime.now().date() else 'Pendente')
                })
                
            except Exception as e:
                erro_msg = f"Linha {row_num}: {str(e)}"
                erros.append(erro_msg)
                app.logger.error(f"Erro na linha {row_num}: {str(e)}")
                # Continuar processamento mesmo com erro
                continue
        
        # Contar divergências de Valor Total
        divergencias_valor_total = [p for p in preview_data if p['valor_total_divergente']]

        # Verificar categorias faltantes no plano de contas
        categorias_importadas = set()
        for item in preview_data:
            if item.get('categoria'):
                categorias_importadas.add(item['categoria'])

        # Buscar categorias existentes por nome ou código
        empresa_id = obter_empresa_id_sessao(session, usuario)
        categorias_existentes = PlanoConta.query.filter(
            PlanoConta.empresa_id == empresa_id,
            PlanoConta.ativo == True,
            db.or_(
                PlanoConta.nome.in_(categorias_importadas),
                PlanoConta.codigo.in_(categorias_importadas)
            )
        ).all()

        nomes_existentes = {c.nome for c in categorias_existentes}
        codigos_existentes = {c.codigo for c in categorias_existentes}

        # Identificar categorias faltantes
        categorias_faltantes = []
        for cat in categorias_importadas:
            if cat not in nomes_existentes and cat not in codigos_existentes:
                categorias_faltantes.append(cat)

        return jsonify({
            'success': True,
            'preview_data': preview_data,
            'erros': erros,
            'total_linhas': len(preview_data),
            'total_erros': len(erros),
            'divergencias_valor_total': len(divergencias_valor_total),
            'tem_divergencias': len(divergencias_valor_total) > 0,
            'categorias_faltantes': sorted(categorias_faltantes),
            'tem_categorias_faltantes': len(categorias_faltantes) > 0,
            'resumo': {
                'lancamentos': len(preview_data),
                'vendas': len([p for p in preview_data if p['criara_venda']]),
                'compras': len([p for p in preview_data if p['criara_compra']]),
                'realizados': len([p for p in preview_data if p['realizado']]),
                'agendados': len([p for p in preview_data if p['status'] == 'Agendado']),
                'pendentes': len([p for p in preview_data if p['status'] == 'Pendente']),
                'divergencias_valor_total': len(divergencias_valor_total)
            }
        })
        
    except Exception as e:
        app.logger.error(f"Erro ao gerar prévia: {str(e)}")
        import traceback
        app.logger.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({'error': f'Erro ao processar arquivo: {str(e)}'}), 500

@app.route('/api/importacao/importar', methods=['POST'])
def api_importar_dados():
    """Importa dados de planilha Excel com validação inteligente e integração completa"""
    if 'usuario_id' not in session:
        return jsonify({'error': 'Usuário não autenticado'}), 401
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario or usuario.tipo == 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    
    try:
        if 'arquivo' not in request.files:
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400
        
        arquivo = request.files['arquivo']
        if arquivo.filename == '':
            return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
        
        if not arquivo.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Formato de arquivo não suportado. Use .xlsx ou .xls'}), 400
        
        # Ler arquivo Excel
        # data_only=True garante que, se houver fórmulas, seja lido o valor calculado
        from openpyxl import load_workbook
        wb = load_workbook(arquivo, data_only=True)
        
        # Detectar automaticamente a aba com dados
        ws = None
        
        # 1. Tentar encontrar aba específica de dados (planilha modelo)
        if "📊 DADOS" in wb.sheetnames:
            ws = wb["📊 DADOS"]
        elif "DADOS" in wb.sheetnames:
            ws = wb["DADOS"]
        else:
            # 2. Procurar por aba que contenha cabeçalhos de dados
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                if sheet.max_row > 0:
                    # Verificar se a primeira linha contém cabeçalhos de dados
                    first_row = [str(cell.value).strip() if cell.value else '' for cell in sheet[1]]
                    if (len(first_row) >= 2 and 
                        first_row[0].lower() in ['tipo'] and 
                        first_row[1].lower() in ['descrição', 'descricao', 'description']):
                        ws = sheet
                        break
            
            # 3. Se não encontrou, usar aba ativa
            if not ws:
                ws = wb.active
        
        # Encontrar automaticamente a linha com os cabeçalhos
        header_row = 1
        headers = []
        
        # Detectar cabeçalhos automaticamente
        headers = []
        header_row = 1
        
        # Procurar pela linha com cabeçalhos de dados
        for row_num in range(1, min(10, ws.max_row + 1)):  # Verificar até 10 linhas
            row_headers = [str(cell.value).strip() if cell.value else '' for cell in ws[row_num]]
            
            # Verificar se esta linha contém cabeçalhos de dados
            if (len(row_headers) >= 2 and 
                row_headers[0].lower() in ['tipo'] and 
                row_headers[1].lower() in ['descrição', 'descricao', 'description']):
                headers = row_headers
                header_row = row_num
                break
        
        # Se não encontrou, usar a primeira linha como fallback
        if not headers:
            headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
        
        app.logger.info(f"Aba selecionada: {ws.title}")
        app.logger.info(f"Total de linhas na aba: {ws.max_row}")
        app.logger.info(f"Cabeçalhos encontrados na linha {header_row}: {headers}")
        
        # Mapear títulos para nomes internos
        header_mapping = {
            'Tipo': 'TIPO',
            'Descrição': 'DESCRICAO', 
            'Valor': 'VALOR',
            'Quantidade': 'QUANTIDADE',
            'Qtd': 'QUANTIDADE',
            'Qtd.': 'QUANTIDADE',
            'Valor Unitário': 'VALOR_UNITARIO',
            'Valor Unitario': 'VALOR_UNITARIO',
            'Preço Unitário': 'VALOR_UNITARIO',
            'Preco Unitario': 'VALOR_UNITARIO',
            'Preço': 'VALOR_UNITARIO',
            'Preco': 'VALOR_UNITARIO',
            'Desconto': 'DESCONTO',
            'Desc': 'DESCONTO',
            'Categoria': 'CATEGORIA',
            'Data Prevista': 'DATA_PREVISTA',
            'Data Realizada': 'DATA_REALIZADA',
            'Conta Caixa': 'CONTA_CAIXA',
            'Conta': 'CONTA_CAIXA',
            'Caixa': 'CONTA_CAIXA',
            'Conta/Caixa': 'CONTA_CAIXA',
            'Conta - Caixa': 'CONTA_CAIXA',
            'Conta_Caixa': 'CONTA_CAIXA',
            'CONTA_CAIXA': 'CONTA_CAIXA',
            'conta_caixa': 'CONTA_CAIXA',
            'conta caixa': 'CONTA_CAIXA',
            'conta': 'CONTA_CAIXA',
            'caixa': 'CONTA_CAIXA',
            'Tipo Cliente/Fornecedor': 'TIPO_CLIENTE_FORNECEDOR',
            'Cliente/Fornecedor': 'CLIENTE_FORNECEDOR',
            'Observações': 'OBSERVACOES',
            'Tipo Produto/Serviço': 'TIPO_PRODUTO_SERVICO',
            'Nome Produto/Serviço': 'NOME_PRODUTO_SERVICO',
            'Tipo Produto/Servico': 'TIPO_PRODUTO_SERVICO',
            'Nome Produto/Servico': 'NOME_PRODUTO_SERVICO',
            'Parcelado': 'PARCELADO',
            'Quantidade de Parcelas': 'QUANTIDADE_PARCELAS',
            # Manter compatibilidade com títulos antigos
            'TIPO': 'TIPO',
            'DESCRICAO': 'DESCRICAO',
            'VALOR': 'VALOR',
            'CATEGORIA': 'CATEGORIA',
            'DATA_PREVISTA': 'DATA_PREVISTA',
            'DATA_REALIZADA': 'DATA_REALIZADA',
            'CONTA_CAIXA': 'CONTA_CAIXA',
            'CLIENTE': 'CLIENTE',
            'FORNECEDOR': 'FORNECEDOR',
            'OBSERVACOES': 'OBSERVACOES',
            'TAGS': 'TAGS',
            'CENTRO_CUSTO': 'CENTRO_CUSTO',
            'PROJETO': 'PROJETO',
            'DOCUMENTO': 'DOCUMENTO',
            # Variações adicionais para garantir compatibilidade
            'Tipo Produto/Serviço': 'TIPO_PRODUTO_SERVICO',
            'Nome Produto/Serviço': 'NOME_PRODUTO_SERVICO',
            'Parcelado?': 'PARCELADO',
            'TIPO_PRODUTO_SERVICO': 'TIPO_PRODUTO_SERVICO',
            'NOME_PRODUTO_SERVICO': 'NOME_PRODUTO_SERVICO',
            'PARCELADO': 'PARCELADO',
            'QUANTIDADE_PARCELAS': 'QUANTIDADE_PARCELAS'
        }
        
        # Normalizar cabeçalhos usando mapeamento direto
        normalized_headers = []
        
        # Garantir que headers não está vazio
        if not headers:
            headers = ['Tipo', 'Descrição', 'Quantidade', 'Valor Unitário', 'Desconto', 'Categoria', 'Data Prevista']
        
        for header in headers:
            # Usar o mapeamento existente ou converter para maiúsculas
            normalized_header = header_mapping.get(header, header.upper())
            normalized_headers.append(normalized_header)
        
        app.logger.info(f"Cabeçalhos normalizados: {normalized_headers}")
        
        # Não validar cabeçalhos obrigatórios - aceitar qualquer planilha
        # Apenas verificar se há pelo menos uma linha de dados
        if ws.max_row < 2:
            return jsonify({
                'error': 'Planilha não contém dados para importar'
            }), 400
        
        # Verificar se há dados após os cabeçalhos
        if header_row >= ws.max_row:
            return jsonify({
                'error': 'Planilha não contém dados após os cabeçalhos'
            }), 400
        
        # Buscar todos os usuários da empresa para validações
        empresa_id = obter_empresa_id_sessao(session, usuario)

        usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
        usuarios_ids = [u.id for u in usuarios_empresa]
        
        # Cache de dados para otimização
        contas_caixa_cache = {c.nome.lower(): c for c in ContaCaixa.query.filter(
            ContaCaixa.usuario_id.in_(usuarios_ids), ContaCaixa.ativo==True
        ).all()}
        
        clientes_cache = {c.nome.lower(): c for c in Cliente.query.filter(
            Cliente.empresa_id == empresa_id
        ).all()}
        
        fornecedores_cache = {f.nome.lower(): f for f in Fornecedor.query.filter(
            Fornecedor.empresa_id == empresa_id
        ).all()}
        
        # Processar linhas
        sucessos = 0
        erros = []
        lancamentos_criados = []
        vendas_criadas = []
        compras_criadas = []
        
        for row_num in range(header_row + 1, ws.max_row + 1):
            try:
                # Verificar se a linha existe
                if row_num > ws.max_row:
                    break
                    
                row_data = [cell.value for cell in ws[row_num]]
                
                # Pular linhas vazias
                if not any(row_data):
                    continue
                
                # Criar dicionário de dados baseado nos cabeçalhos
                dados = {}
                for i, valor in enumerate(row_data):
                    if i < len(normalized_headers):
                        dados[normalized_headers[i]] = valor
                
                # Verificar se há pelo menos um campo essencial preenchido
                campos_essenciais = ['TIPO', 'DESCRICAO', 'VALOR', 'CATEGORIA', 'DATA_PREVISTA']
                tem_dados = any(str(dados.get(campo, '')).strip() for campo in campos_essenciais)
                
                if not tem_dados:
                    continue
                
                # Processar dados com valores padrão apenas quando necessário
                tipo_raw = str(dados.get('TIPO', '')).strip()
                descricao_raw = str(dados.get('DESCRICAO', '')).strip()
                valor_raw = dados.get('VALOR')
                quantidade_raw = dados.get('QUANTIDADE')
                valor_unitario_raw = dados.get('VALOR_UNITARIO')
                desconto_raw = dados.get('DESCONTO')
                categoria_raw = str(dados.get('CATEGORIA', '')).strip()
                data_prevista_raw = dados.get('DATA_PREVISTA')
                
                # Aplicar valores padrão apenas se campos estiverem vazios mas linha tiver dados
                tipo = normalizar_tipo(tipo_raw) if tipo_raw else 'entrada'
                descricao = descricao_raw if descricao_raw else f"Lançamento importado - Linha {row_num}"
                
                # Processar quantidade
                quantidade = processar_valor(quantidade_raw) if quantidade_raw else 1
                if quantidade is not None and quantidade <= 0:
                    erros.append(f"Linha {row_num}: Quantidade inválida ({quantidade_raw}). Convertido para 1.")
                    quantidade = 1  # Padrão mínimo
                elif quantidade is None:
                    quantidade = 1

                # Processar valor unitário
                valor_unitario = processar_valor(valor_unitario_raw) if valor_unitario_raw else None
                if valor_unitario is not None and valor_unitario <= 0:
                    erros.append(f"Linha {row_num}: Valor unitário inválido ({valor_unitario_raw}). Ignorado.")
                    valor_unitario = None

                # Processar desconto
                desconto = processar_valor(desconto_raw) if desconto_raw else 0
                if desconto is not None and desconto < 0:
                    erros.append(f"Linha {row_num}: Desconto negativo ({desconto_raw}). Convertido para 0.")
                    desconto = 0
                elif desconto is None:
                    desconto = 0

                # CALCULAR VALOR AUTOMATICAMENTE baseado em Quantidade, Valor Unitário e Desconto
                # Fórmula: valor = (quantidade × valor_unitario) - desconto
                if valor_unitario is not None and quantidade is not None:
                    # Calcular automaticamente (prioridade)
                    valor = (quantidade * valor_unitario) - desconto
                    if valor < 0:
                        valor = 0  # Não permitir valor negativo
                    valor_total = valor  # Para vendas/compras
                elif valor_raw:
                    # Fallback: usar valor da coluna VALOR se fornecido (compatibilidade com planilhas antigas)
                    valor = processar_valor(valor_raw)
                    if valor is None or valor <= 0:
                        valor = 0.01  # Valor padrão mínimo
                    valor_total = None  # Usar valor da parcela se não houver cálculo
                else:
                    # Sem dados suficientes para calcular
                    erros.append(f"Linha {row_num}: Dados insuficientes para calcular valor. Forneça Quantidade e Valor Unitário, ou Valor direto.")
                    valor = 0.01  # Valor padrão mínimo para não bloquear
                    valor_total = None
                
                # Buscar ou criar categoria no plano de contas automaticamente
                if categoria_raw:
                    categoria_id = buscar_ou_criar_categoria_plano_contas(categoria_raw, tipo, usuario.id, usuario.empresa_id)
                    if categoria_id:
                        categoria = categoria_raw
                    else:
                        categoria = 'Importado'  # Fallback se não conseguir criar
                else:
                    categoria = 'Importado'
                
                # Processar datas com múltiplos formatos
                data_prevista = processar_data(data_prevista_raw)
                if not data_prevista:
                    erros.append(f"Linha {row_num}: Data prevista ausente ou inválida ({data_prevista_raw}). Usando data atual.")
                    data_prevista = datetime.now().date()  # Data atual como padrão

                data_realizada = processar_data(dados.get('DATA_REALIZADA')) if dados.get('DATA_REALIZADA') else None
                
                # Debug: verificar se a data foi processada corretamente
                if dados.get('DATA_REALIZADA'):
                    app.logger.info(f"Linha {row_num}: Data Realizada original: '{dados.get('DATA_REALIZADA')}' -> Processada: {data_realizada}")
                
                # Processar campos de parcelamento
                avista_parcelado_str = str(dados.get('PARCELADO', '')).strip().lower()
                quantidade_parcelas_str = str(dados.get('QUANTIDADE_PARCELAS', '')).strip()
                
                # Determinar se é parcelado
                is_parcelado = avista_parcelado_str in ['sim', 's', 'yes', 'y', '1', 'true']
                quantidade_parcelas = 1
                
                if is_parcelado and quantidade_parcelas_str:
                    try:
                        quantidade_parcelas = int(quantidade_parcelas_str)
                        if quantidade_parcelas < 1:
                            quantidade_parcelas = 1
                    except (ValueError, TypeError):
                        quantidade_parcelas = 1
                
                # Processar campos opcionais
                observacoes = str(dados.get('OBSERVACOES', '')).strip() if dados.get('OBSERVACOES') else None
                tipo_produto_servico = str(dados.get('TIPO_PRODUTO_SERVICO', '')).strip().lower() if dados.get('TIPO_PRODUTO_SERVICO') else None
                nome_produto_servico = str(dados.get('NOME_PRODUTO_SERVICO', '')).strip() if dados.get('NOME_PRODUTO_SERVICO') else None
                nota_fiscal = str(dados.get('NOTA_FISCAL', '')).strip() if dados.get('NOTA_FISCAL') else None
                
                # Determinar se é produto ou serviço para controle de estoque
                eh_produto = tipo_produto_servico in ['produto', 'Produto']
                eh_servico = tipo_produto_servico in ['serviço', 'servico', 'Serviço', 'Servico']
                
                # Usar nome_produto_servico se disponível, senão usar produto_servico (compatibilidade)
                produto_servico_final = nome_produto_servico if nome_produto_servico else str(dados.get('PRODUTO_SERVICO', '')).strip() if dados.get('PRODUTO_SERVICO') else None
                
                # NOVA LÓGICA: Processar vinculação automática baseada no tipo cliente/fornecedor
                tipo_cliente_fornecedor = str(dados.get('TIPO_CLIENTE_FORNECEDOR', '')).strip().lower()
                nome_cliente_fornecedor = str(dados.get('CLIENTE_FORNECEDOR', '')).strip()
                
                cliente_id = None
                fornecedor_id = None
                
                # Vinculação inteligente baseada no tipo cliente/fornecedor da planilha
                if tipo_cliente_fornecedor and nome_cliente_fornecedor:
                    if tipo_cliente_fornecedor in ['cliente', 'client']:
                        # Buscar ou criar cliente automaticamente
                        cliente_id = buscar_cliente(nome_cliente_fornecedor, clientes_cache)
                        if not cliente_id:
                            # Criar cliente automaticamente se não existir
                            novo_cliente = Cliente(
                                nome=nome_cliente_fornecedor,
                                usuario_id=usuario.id,
                                empresa_id=usuario.empresa_id
                            )
                            db.session.add(novo_cliente)
                            db.session.flush()
                            cliente_id = novo_cliente.id
                            clientes_cache[nome_cliente_fornecedor.lower()] = novo_cliente
                            app.logger.info(f"Cliente criado automaticamente: {nome_cliente_fornecedor} (ID: {cliente_id})")
                    
                    elif tipo_cliente_fornecedor in ['fornecedor', 'supplier']:
                        # Buscar ou criar fornecedor automaticamente
                        fornecedor_id = buscar_fornecedor(nome_cliente_fornecedor, fornecedores_cache)
                        if not fornecedor_id:
                            # Criar fornecedor automaticamente se não existir
                            novo_fornecedor = Fornecedor(
                                nome=nome_cliente_fornecedor,
                                usuario_id=usuario.id,
                                empresa_id=usuario.empresa_id
                            )
                            db.session.add(novo_fornecedor)
                            db.session.flush()
                            fornecedor_id = novo_fornecedor.id
                            fornecedores_cache[nome_cliente_fornecedor.lower()] = novo_fornecedor
                            app.logger.info(f"Fornecedor criado automaticamente: {nome_cliente_fornecedor} (ID: {fornecedor_id})")
                else:
                    # Fallback para lógica antiga se não houver tipo específico
                    # Buscar ou criar conta caixa automaticamente
                    conta_caixa_id = buscar_ou_criar_conta_caixa(row_data[6] if len(row_data) > 6 else None, contas_caixa_cache, usuario.id, usuario.empresa_id)
                    
                    # Buscar cliente com busca inteligente
                    cliente_id = buscar_cliente(row_data[7] if len(row_data) > 7 else None, clientes_cache)
                    
                    # Buscar fornecedor com busca inteligente
                    fornecedor_id = buscar_fornecedor(row_data[8] if len(row_data) > 8 else None, fornecedores_cache)
                
                # Buscar conta caixa (se não foi processada no fallback)
                if 'conta_caixa_id' not in locals():
                    conta_caixa_id = buscar_ou_criar_conta_caixa(dados.get('CONTA_CAIXA'), contas_caixa_cache, usuario.id, usuario.empresa_id)
                
                # Calcular status automaticamente baseado na data de realização
                hoje = datetime.now().date()
                if data_realizada:
                    # Garantir que data_realizada seja um objeto date
                    if isinstance(data_realizada, datetime):
                        data_realizada = data_realizada.date()
                    
                    if data_realizada <= hoje:
                        realizado = True  # Data passada ou presente = realizado
                    else:
                        realizado = False  # Data futura = agendado
                else:
                    realizado = False  # Sem data = pendente
                
                # Criar lançamento financeiro
                if is_parcelado and quantidade_parcelas > 1:
                    # Criar lançamentos parcelados
                    # Usar valor_total da operação (incluindo descontos) para calcular valor da parcela
                    valor_total_operacao = valor_total if valor_total is not None else valor
                    valor_parcela = valor_total_operacao / quantidade_parcelas

                    for i in range(quantidade_parcelas):
                        # Calcular data da parcela (mês a mês) - adiciona meses corretamente
                        if i == 0:
                            data_parcela = data_prevista
                        else:
                            # Adicionar meses considerando variação de dias (28-31)
                            mes = data_prevista.month + i
                            ano = data_prevista.year
                            while mes > 12:
                                mes -= 12
                                ano += 1
                            # Garantir que o dia existe no mês destino
                            import calendar
                            ultimo_dia_mes = calendar.monthrange(ano, mes)[1]
                            dia = min(data_prevista.day, ultimo_dia_mes)
                            data_parcela = date(ano, mes, dia)
                        
                        # Criar lançamento da parcela
                        novo_lancamento = Lancamento(
                            descricao=f"{descricao} - Parcela {i+1}/{quantidade_parcelas}",
                            valor=valor_parcela,
                            tipo=tipo,
                            categoria=categoria,
                            data_prevista=data_parcela,
                            data_realizada=data_realizada if i == 0 and data_realizada else None,
                            realizado=realizado if i == 0 else False,
                            usuario_id=usuario.id,
                            empresa_id=empresa_id,
                            conta_caixa_id=conta_caixa_id,
                            cliente_id=cliente_id,
                            fornecedor_id=fornecedor_id,
                            observacoes=f"{observacoes or ''} - Parcela {i+1}/{quantidade_parcelas}".strip(),
                            produto_servico=produto_servico_final,
                            tipo_produto_servico=tipo_produto_servico,
                            nota_fiscal=nota_fiscal,
                            usuario_criacao_id=usuario.id  # Registrar quem criou
                        )
                        
                        db.session.add(novo_lancamento)
                        db.session.flush()  # Para obter o ID
                        
                        lancamentos_criados.append(novo_lancamento)
                        
                        # LÓGICA MELHORADA: Vinculação automática aos módulos baseada no tipo cliente/fornecedor
                        if i == 0:  # Apenas na primeira parcela
                            deve_criar_venda = False
                            deve_criar_compra = False
                            
                            # Lógica baseada no tipo cliente/fornecedor da planilha
                            if tipo_cliente_fornecedor == 'cliente':
                                deve_criar_venda = True
                                app.logger.info(f"Vinculação automática: Cliente identificado → Criando VENDA (ID Cliente: {cliente_id})")
                                
                            elif tipo_cliente_fornecedor == 'fornecedor':
                                deve_criar_compra = True
                                app.logger.info(f"Vinculação automática: Fornecedor identificado → Criando COMPRA (ID Fornecedor: {fornecedor_id})")
                                
                                # Se fornecedor + mercadoria, vincular também ao estoque
                                if eh_produto:
                                    app.logger.info(f"Vinculação automática: Fornecedor + Produto → Vinculando aos módulos COMPRAS + ESTOQUE")
                            
                            # Fallback para lógica antiga se não houver tipo específico
                            elif not tipo_cliente_fornecedor:
                                # Se tem cliente, sempre criar venda
                                if cliente_id:
                                    deve_criar_venda = True
                                    app.logger.info(f"Fallback: Cliente encontrado → Criando venda (Cliente ID: {cliente_id})")
                                
                                # Se tem fornecedor, sempre criar compra
                                if fornecedor_id:
                                    deve_criar_compra = True
                                    app.logger.info(f"Fallback: Fornecedor encontrado → Criando compra (Fornecedor ID: {fornecedor_id})")
                                
                                # Se não tem cliente/fornecedor específico, usar lógica baseada no tipo e produto/serviço
                                if not cliente_id and not fornecedor_id and produto_servico_final:
                                    if tipo == 'entrada':
                                        deve_criar_venda = True
                                        app.logger.info(f"Fallback: Entrada com produto/serviço → Criando venda automática: {produto_servico_final}")
                                    elif tipo == 'saida':
                                        deve_criar_compra = True
                                        app.logger.info(f"Fallback: Saída com produto/serviço → Criando compra automática: {produto_servico_final}")
                            
                            # Criar venda se necessário
                            if deve_criar_venda:
                                venda = criar_venda_automatica(novo_lancamento, usuario.id, valor_total, quantidade)
                                if venda:
                                    vendas_criadas.append(venda)
                                    # Criar vínculo
                                    criar_vinculo('lancamento', novo_lancamento.id, 'venda', venda.id, usuario.id)
                                    app.logger.info(f"Venda criada com sucesso: ID {venda.id}")
                                    
                                    # Atualizar estoque se for produto
                                    if eh_produto and produto_servico_final:
                                        sucesso_estoque, mensagem_estoque = atualizar_estoque_venda(venda, usuario.id)
                                        if sucesso_estoque:
                                            app.logger.info(f"Estoque atualizado para venda {venda.id}: {mensagem_estoque}")
                                        else:
                                            app.logger.warning(f"Aviso no estoque para venda {venda.id}: {mensagem_estoque}")
                                    elif eh_servico:
                                        app.logger.info(f"Venda de serviço {venda.id} - não afeta estoque")
                                else:
                                    app.logger.error("Falha ao criar venda")
                            
                            # Criar compra se necessário
                            if deve_criar_compra:
                                compra = criar_compra_automatica(novo_lancamento, usuario.id, valor_total, quantidade)
                                if compra:
                                    compras_criadas.append(compra)
                                    # Criar vínculo
                                    criar_vinculo('lancamento', novo_lancamento.id, 'compra', compra.id, usuario.id)
                                    app.logger.info(f"Compra criada com sucesso: ID {compra.id}")
                                    
                                    # Atualizar estoque se for produto
                                    if eh_produto and produto_servico_final:
                                        sucesso_estoque, mensagem_estoque = atualizar_estoque_compra(compra, usuario.id)
                                        if sucesso_estoque:
                                            app.logger.info(f"Estoque atualizado para compra {compra.id}: {mensagem_estoque}")
                                        else:
                                            app.logger.warning(f"Aviso no estoque para compra {compra.id}: {mensagem_estoque}")
                                    elif eh_servico:
                                        app.logger.info(f"Compra de serviço {compra.id} - não afeta estoque")
                                else:
                                    app.logger.error("Falha ao criar compra")
                else:
                    # Criar lançamento único
                    novo_lancamento = Lancamento(
                        descricao=descricao,
                        valor=valor,
                        tipo=tipo,
                        categoria=categoria,
                        data_prevista=data_prevista,
                        data_realizada=data_realizada,
                        realizado=realizado,
                        usuario_id=usuario.id,
                        empresa_id=empresa_id,
                        conta_caixa_id=conta_caixa_id,
                        cliente_id=cliente_id,
                        fornecedor_id=fornecedor_id,
                        observacoes=observacoes,
                        produto_servico=produto_servico_final,
                        tipo_produto_servico=tipo_produto_servico,
                        nota_fiscal=nota_fiscal,
                        usuario_criacao_id=usuario.id  # Registrar quem criou
                    )
                    
                    db.session.add(novo_lancamento)
                    db.session.flush()  # Para obter o ID
                    
                    lancamentos_criados.append(novo_lancamento)
                    
                    # LÓGICA MELHORADA: Vinculação automática aos módulos baseada no tipo cliente/fornecedor
                    deve_criar_venda = False
                    deve_criar_compra = False
                    
                    # Lógica baseada no tipo cliente/fornecedor da planilha
                    if tipo_cliente_fornecedor == 'cliente':
                        deve_criar_venda = True
                        app.logger.info(f"Vinculação automática: Cliente identificado → Criando VENDA (ID Cliente: {cliente_id})")
                        
                    elif tipo_cliente_fornecedor == 'fornecedor':
                        deve_criar_compra = True
                        app.logger.info(f"Vinculação automática: Fornecedor identificado → Criando COMPRA (ID Fornecedor: {fornecedor_id})")
                        
                        # Se fornecedor + mercadoria, vincular também ao estoque
                        if eh_produto:
                            app.logger.info(f"Vinculação automática: Fornecedor + Produto → Vinculando aos módulos COMPRAS + ESTOQUE")
                    
                    # Fallback para lógica antiga se não houver tipo específico
                    elif not tipo_cliente_fornecedor:
                        # Se tem cliente, sempre criar venda
                        if cliente_id:
                            deve_criar_venda = True
                            app.logger.info(f"Fallback: Cliente encontrado → Criando venda (Cliente ID: {cliente_id})")
                        
                        # Se tem fornecedor, sempre criar compra
                        if fornecedor_id:
                            deve_criar_compra = True
                            app.logger.info(f"Fallback: Fornecedor encontrado → Criando compra (Fornecedor ID: {fornecedor_id})")
                        
                        # Se não tem cliente/fornecedor específico, usar lógica baseada no tipo e produto/serviço
                        if not cliente_id and not fornecedor_id and produto_servico_final:
                            if tipo == 'entrada':
                                deve_criar_venda = True
                                app.logger.info(f"Fallback: Entrada com produto/serviço → Criando venda automática: {produto_servico_final}")
                            elif tipo == 'saida':
                                deve_criar_compra = True
                                app.logger.info(f"Fallback: Saída com produto/serviço → Criando compra automática: {produto_servico_final}")
                    
                    # Criar venda se necessário
                    if deve_criar_venda:
                        venda = criar_venda_automatica(novo_lancamento, usuario.id, valor_total, quantidade)
                        if venda:
                            vendas_criadas.append(venda)
                            # Criar vínculo
                            criar_vinculo('lancamento', novo_lancamento.id, 'venda', venda.id, usuario.id)
                            app.logger.info(f"Venda criada com sucesso: ID {venda.id}")
                            
                            # Atualizar estoque se for produto
                            if eh_produto and produto_servico_final:
                                sucesso_estoque, mensagem_estoque = atualizar_estoque_venda(venda, usuario.id)
                                if sucesso_estoque:
                                    app.logger.info(f"Estoque atualizado para venda {venda.id}: {mensagem_estoque}")
                                else:
                                    app.logger.warning(f"Aviso no estoque para venda {venda.id}: {mensagem_estoque}")
                            elif eh_servico:
                                app.logger.info(f"Venda de serviço {venda.id} - não afeta estoque")
                        else:
                            app.logger.error("Falha ao criar venda")
                    
                    # Criar compra se necessário
                    if deve_criar_compra:
                        compra = criar_compra_automatica(novo_lancamento, usuario.id, valor_total, quantidade)
                        if compra:
                            compras_criadas.append(compra)
                            # Criar vínculo
                            criar_vinculo('lancamento', novo_lancamento.id, 'compra', compra.id, usuario.id)
                            app.logger.info(f"Compra criada com sucesso: ID {compra.id}")
                            
                            # Atualizar estoque se for produto
                            if eh_produto and produto_servico_final:
                                sucesso_estoque, mensagem_estoque = atualizar_estoque_compra(compra, usuario.id)
                                if sucesso_estoque:
                                    app.logger.info(f"Estoque atualizado para compra {compra.id}: {mensagem_estoque}")
                                else:
                                    app.logger.warning(f"Aviso no estoque para compra {compra.id}: {mensagem_estoque}")
                            elif eh_servico:
                                app.logger.info(f"Compra de serviço {compra.id} - não afeta estoque")
                        else:
                            app.logger.error("Falha ao criar compra")

                # Nota: Saldo da conta caixa é calculado dinamicamente, não precisa atualizar aqui

                sucessos += 1
                
            except Exception as e:
                erro_msg = f"Linha {row_num}: {str(e)}"
                erros.append(erro_msg)
                app.logger.error(f"Erro na linha {row_num}: {str(e)}")
                # Continuar processamento mesmo com erro
                continue
        
        # Commit das alterações
        db.session.commit()
        
        # Log da importação
        app.logger.info(f"Importação concluída: {sucessos} sucessos, {len(erros)} erros")
        
        # Salvar registro da importação
        total_receitas_import = sum([l.valor for l in lancamentos_criados if l.tipo == 'entrada'])
        total_despesas_import = sum([l.valor for l in lancamentos_criados if l.tipo == 'saida'])
        
        # Salvar IDs dos registros criados para poder desfazer depois
        import json
        lancamentos_ids = [str(l.id) for l in lancamentos_criados]
        vendas_ids = [str(v.id) for v in vendas_criadas]
        compras_ids = [str(c.id) for c in compras_criadas]
        
        registro_importacao = Importacao(
            usuario_id=usuario.id,
            nome_arquivo=arquivo.filename,
            total_lancamentos=len(lancamentos_criados),
            total_entradas=total_receitas_import,
            total_saidas=total_despesas_import,
            total_vendas=len(vendas_criadas),
            total_compras=len(compras_criadas),
            sucessos=sucessos,
            erros=len(erros),
            status='concluida',
            lancamentos_ids=json.dumps(lancamentos_ids),
            vendas_ids=json.dumps(vendas_ids),
            compras_ids=json.dumps(compras_ids)
        )
        db.session.add(registro_importacao)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'sucessos': sucessos,
            'erros': erros,
            'total_processados': sucessos + len(erros),
            'lancamentos_criados': len(lancamentos_criados),
            'vendas_criadas': len(vendas_criadas),
            'compras_criadas': len(compras_criadas),
            'importacao_id': registro_importacao.id,
            'detalhes': {
                'lancamentos': [{'id': l.id, 'descricao': l.descricao} for l in lancamentos_criados],
                'vendas': [{'id': v.id, 'cliente': v.cliente.nome if v.cliente else 'N/A'} for v in vendas_criadas],
                'compras': [{'id': c.id, 'fornecedor': c.fornecedor.nome if c.fornecedor else 'N/A'} for c in compras_criadas]
            }
        })
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao importar dados: {str(e)}")
        return jsonify({'error': f'Erro ao processar arquivo: {str(e)}'}), 500

@app.route('/api/importacao/listar')
def api_listar_importacoes():
    """Lista todas as importações do usuário"""
    if 'usuario_id' not in session:
        return jsonify({'error': 'Usuário não autenticado'}), 401
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario or usuario.tipo == 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    
    # Buscar importações do usuário
    importacoes = Importacao.query.filter_by(usuario_id=usuario.id).order_by(Importacao.data_importacao.desc()).all()
    
    dados = []
    for imp in importacoes:
        dados.append({
            'id': imp.id,
            'nome_arquivo': imp.nome_arquivo,
            'data_importacao': imp.data_importacao.strftime('%d/%m/%Y %H:%M'),
            'total_lancamentos': imp.total_lancamentos,
            'total_entradas': imp.total_entradas,
            'total_saidas': imp.total_saidas,
            'total_vendas': imp.total_vendas,
            'total_compras': imp.total_compras,
            'sucessos': imp.sucessos,
            'erros': imp.erros,
            'status': imp.status
        })
    
    return jsonify({'importacoes': dados})

@app.route('/api/importacao/<int:importacao_id>/desfazer', methods=['POST'])
def api_desfazer_importacao(importacao_id):
    """Desfaz uma importação específica"""
    if 'usuario_id' not in session:
        return jsonify({'error': 'Usuário não autenticado'}), 401
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario or usuario.tipo == 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    
    try:
        # Buscar a importação
        importacao = Importacao.query.filter_by(id=importacao_id, usuario_id=usuario.id).first()
        if not importacao:
            return jsonify({'error': 'Importação não encontrada'}), 404
        
        if importacao.status == 'desfeita':
            return jsonify({'error': 'Importação já foi desfeita'}), 400
        
        # Buscar lançamentos, vendas e compras pelos IDs salvos
        import json
        lancamentos_ids = json.loads(importacao.lancamentos_ids) if importacao.lancamentos_ids else []
        vendas_ids = json.loads(importacao.vendas_ids) if importacao.vendas_ids else []
        compras_ids = json.loads(importacao.compras_ids) if importacao.compras_ids else []
        
        # Buscar registros pelos IDs
        lancamentos = Lancamento.query.filter(Lancamento.id.in_(lancamentos_ids)).all()
        vendas = Venda.query.filter(Venda.id.in_(vendas_ids)).all()
        compras = Compra.query.filter(Compra.id.in_(compras_ids)).all()
        
        # Se não há IDs salvos, usar o método por data (fallback)
        if not lancamentos_ids and not vendas_ids and not compras_ids:
            data_inicio = importacao.data_importacao - timedelta(hours=1)
            data_fim = importacao.data_importacao + timedelta(hours=1)
            
            lancamentos = Lancamento.query.filter(
                Lancamento.usuario_id == usuario.id,
                Lancamento.data_criacao >= data_inicio,
                Lancamento.data_criacao <= data_fim
            ).all()
            
            vendas = Venda.query.filter(
                Venda.usuario_id == usuario.id,
                Venda.data_criacao >= data_inicio,
                Venda.data_criacao <= data_fim
            ).all()
            
            compras = Compra.query.filter(
                Compra.usuario_id == usuario.id,
                Compra.data_criacao >= data_inicio,
                Compra.data_criacao <= data_fim
            ).all()
        
        # Deletar lançamentos e suas parcelas
        for lancamento in lancamentos:
            # Deletar parcelas
            for parcela in lancamento.parcelas:
                db.session.delete(parcela)
            db.session.delete(lancamento)
        
        # Deletar vendas e suas parcelas
        for venda in vendas:
            for parcela in venda.parcelas:
                db.session.delete(parcela)
            db.session.delete(venda)
        
        # Deletar compras e suas parcelas
        for compra in compras:
            for parcela in compra.parcelas:
                db.session.delete(parcela)
            db.session.delete(compra)
        
        # Deletar completamente o registro de importação
        db.session.delete(importacao)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Importação desfeita com sucesso. {len(lancamentos)} lançamentos, {len(vendas)} vendas e {len(compras)} compras removidos.'
        })
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao desfazer importação {importacao_id}: {str(e)}")
        return jsonify({'error': f'Erro ao desfazer importação: {str(e)}'}), 500

def normalizar_tipo(tipo_str):
    """Normaliza tipo de lançamento com múltiplas variações"""
    tipo_str = tipo_str.lower().strip()
    
    tipos_entrada = ['entrada', 'entradas', 'entrada', 'entradas', 'income', 'r']
    tipos_saida = ['saida', 'saidas', 'saída', 'saidas', 'saida', 'expense', 'd']
    
    if tipo_str in tipos_entrada:
        return 'entrada'
    elif tipo_str in tipos_saida:
        return 'saida'
    
    # Retorna 'entrada' como padrão se não conseguir normalizar
    return 'entrada'

def processar_valor(valor_input):
    """Processa valor com múltiplos formatos"""
    if valor_input is None:
        return None
    
    try:
        # Se já é um número
        if isinstance(valor_input, (int, float)):
            return float(valor_input)
        
        # Converter string
        valor_str = str(valor_input).strip()
        
        # Remover caracteres não numéricos exceto vírgula e ponto
        valor_limpo = ''.join(c for c in valor_str if c.isdigit() or c in '.,-')
        
        # Tratar vírgula como separador decimal
        if ',' in valor_limpo and '.' in valor_limpo:
            # Formato brasileiro: 1.000,50
            valor_limpo = valor_limpo.replace('.', '').replace(',', '.')
        elif ',' in valor_limpo:
            # Formato brasileiro: 1000,50
            valor_limpo = valor_limpo.replace(',', '.')
        
        return float(valor_limpo)
        
    except (ValueError, TypeError):
        return None

def processar_data(data_input):
    """Processa data com múltiplos formatos"""
    if not data_input:
        return None
    
    try:
        # Se já é um objeto date, retornar diretamente
        if isinstance(data_input, date) and not isinstance(data_input, datetime):
            return data_input
        
        # Se é um objeto datetime, converter para date
        if isinstance(data_input, datetime):
            return data_input.date()
        
        # Se tem método date (como pandas Timestamp), usar
        if hasattr(data_input, 'date') and callable(getattr(data_input, 'date')):
            return data_input.date()
        
        data_str = str(data_input).strip()
        
        # Formatos suportados
        formatos = [
            '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y',
            '%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d',
            '%d/%m/%y', '%d-%m-%y', '%d.%m.%y',
            '%m/%d/%Y', '%m-%d-%Y', '%m.%d.%Y'
        ]
        
        for formato in formatos:
            try:
                return datetime.strptime(data_str, formato).date()
            except ValueError:
                continue
        
        return None
        
    except (ValueError, TypeError):
        return None

def processar_status_realizado(status_input):
    """Processa status realizado com múltiplas variações"""
    if not status_input:
        return False
    
    status_str = str(status_input).lower().strip()
    
    valores_true = ['sim', 's', 'yes', 'y', 'true', '1', 'verdadeiro', 'realizado', 'ok']
    valores_false = ['não', 'nao', 'n', 'no', 'false', '0', 'falso', 'pendente', 'nok']
    
    if status_str in valores_true:
        return True
    elif status_str in valores_false:
        return False
    
    return False

def buscar_conta_caixa(nome_input, cache):
    """Busca conta caixa com busca inteligente"""
    if not nome_input:
        return None
    
    nome = str(nome_input).strip().lower()
    
    # Busca exata
    if nome in cache:
        return cache[nome].id
    
    # Busca parcial
    for nome_cached, conta in cache.items():
        if nome in nome_cached or nome_cached in nome:
            return conta.id
    
    return None

def buscar_ou_criar_conta_caixa(nome_input, cache, usuario_id, empresa_id):
    """Busca conta caixa e cria automaticamente se não existir"""
    if not nome_input:
        return None
    
    nome_original = str(nome_input).strip()
    nome = nome_original.lower()
    
    # Primeiro, tentar buscar normalmente
    conta_id = buscar_conta_caixa(nome_input, cache)
    if conta_id:
        return conta_id
    
    # Se não encontrou, criar automaticamente
    try:
        # Determinar tipo baseado no nome
        tipo_conta = 'Dinheiro'  # Padrão
        if any(palavra in nome for palavra in ['banco', 'conta', 'corrente', 'poupança']):
            tipo_conta = 'Conta Corrente'
        elif any(palavra in nome for palavra in ['cartão', 'cartao', 'credito', 'crédito']):
            tipo_conta = 'Cartão'
        elif any(palavra in nome for palavra in ['nubank', 'inter', 'itau', 'bradesco', 'santander']):
            tipo_conta = 'Conta Digital'
        
        # Criar nova conta caixa
        nova_conta = ContaCaixa(
            nome=nome_original,
            tipo=tipo_conta,
            saldo_inicial=0.0,
            saldo_atual=0.0,
            ativo=True,
            descricao=f'Conta criada automaticamente durante importação - {nome_original}',
            usuario_id=usuario_id,
            data_criacao=datetime.utcnow()
        )
        
        db.session.add(nova_conta)
        db.session.commit()
        
        # Atualizar cache
        cache[nome] = nova_conta
        
        app.logger.info(f"✅ Conta caixa '{nome_original}' criada automaticamente (ID: {nova_conta.id})")
        
        return nova_conta.id
        
    except Exception as e:
        app.logger.error(f"❌ Erro ao criar conta caixa '{nome_original}': {str(e)}")
        return None

def buscar_ou_criar_categoria_plano_contas(nome_categoria, tipo_lancamento, usuario_id, empresa_id):
    """Busca categoria no plano de contas e cria automaticamente se não existir"""
    if not nome_categoria:
        return None
    
    nome_original = str(nome_categoria).strip()
    
    try:
        # Buscar todos os usuários da empresa
        usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
        usuarios_ids = [u.id for u in usuarios_empresa]
        
        # Verificar se a categoria já existe
        categoria_existe = PlanoConta.query.filter(
            PlanoConta.empresa_id == empresa_id,
            PlanoConta.tipo == tipo_lancamento,
            PlanoConta.nome == nome_original,
            PlanoConta.ativo == True
        ).first()
        
        if categoria_existe:
            app.logger.info(f"✅ Categoria '{nome_original}' já existe no plano de contas (ID: {categoria_existe.id})")
            return categoria_existe.id
        
        # Se não encontrou, criar automaticamente
        # Gerar código único para evitar duplicatas
        codigo_base = f"99.{nome_original[:5].upper()}"
        codigo = codigo_base
        contador = 1
        while PlanoConta.query.filter_by(empresa_id=empresa_id, codigo=codigo, ativo=True).first():
            codigo = f"{codigo_base}.{contador}"
            contador += 1
            if contador > 999:  # Limite de segurança
                codigo = f"99.IMP{usuario_id}{int(datetime.utcnow().timestamp())}"
                break

        nova_categoria = PlanoConta(
            nome=nome_original,
            tipo=tipo_lancamento,
            natureza='analitica',
            nivel=1,
            codigo=codigo,  # Código único garantido
            descricao=f"Categoria criada automaticamente durante importação - {nome_original}",
            usuario_id=usuario_id,
            empresa_id=empresa_id,
            data_criacao=datetime.utcnow()
        )
        
        db.session.add(nova_categoria)
        db.session.commit()
        
        app.logger.info(f"✅ Categoria '{nome_original}' criada automaticamente no plano de contas (ID: {nova_categoria.id})")
        
        return nova_categoria.id
        
    except Exception as e:
        app.logger.error(f"❌ Erro ao criar categoria '{nome_original}': {str(e)}")
        return None

def criar_venda_automatica(lancamento, usuario_id, valor_total=None, quantidade=1):
    """Cria uma venda automaticamente baseada no lançamento financeiro"""
    try:
        
        # Usar valor_total se fornecido, senão usar valor do lançamento
        valor_venda = valor_total if valor_total is not None else lancamento.valor
        
        # Determinar tipo de venda baseado no produto/serviço
        if lancamento.tipo_produto_servico and lancamento.tipo_produto_servico.lower() in ['produto', 'produtos']:
            tipo_venda = 'produto'
        elif lancamento.tipo_produto_servico and lancamento.tipo_produto_servico.lower() in ['serviço', 'servico', 'serviços', 'servicos']:
            tipo_venda = 'servico'
        else:
            tipo_venda = 'servico'  # Padrão
        
        # Calcular valor da parcela baseado no valor total da operação
        numero_parcelas = getattr(lancamento, 'numero_parcelas', 1) or 1
        valor_parcela = valor_venda / numero_parcelas if numero_parcelas > 1 else valor_venda
        
        # Criar venda
        venda = Venda(
            produto=lancamento.produto_servico or 'Produto/Serviço Importado',
            quantidade=quantidade,  # Usar quantidade da importação
            valor=valor_venda,
            tipo_venda=tipo_venda,
            data_prevista=lancamento.data_prevista,
            data_realizada=lancamento.data_realizada,
            realizado=lancamento.realizado,
            observacoes=f'Venda criada automaticamente durante importação - Lançamento ID: {lancamento.id}',
            usuario_id=usuario_id,
            cliente_id=lancamento.cliente_id,
            numero_parcelas=numero_parcelas,
            valor_parcela=valor_parcela,
            nota_fiscal=lancamento.nota_fiscal
        )
        
        db.session.add(venda)
        db.session.flush()  # Para obter o ID
        
        app.logger.info(f"✅ Venda criada automaticamente: ID {venda.id}, Produto: {venda.produto}, Valor: R$ {venda.valor}")
        
        return venda
        
    except Exception as e:
        app.logger.error(f"❌ Erro ao criar venda automática: {str(e)}")
        return None

def criar_compra_automatica(lancamento, usuario_id, valor_total=None, quantidade=1):
    """Cria uma compra automaticamente baseada no lançamento financeiro"""
    try:
        
        # Usar valor_total se fornecido, senão usar valor do lançamento
        valor_compra = valor_total if valor_total is not None else lancamento.valor
        
        # Determinar tipo de compra baseado no produto/serviço
        if lancamento.tipo_produto_servico and lancamento.tipo_produto_servico.lower() in ['produto', 'produtos']:
            # Normalizar para 'mercadoria' quando lançamento indica produto
            tipo_compra = 'mercadoria'
        elif lancamento.tipo_produto_servico and lancamento.tipo_produto_servico.lower() in ['serviço', 'servico', 'serviços', 'servicos']:
            tipo_compra = 'servico'
        else:
            tipo_compra = 'servico'  # Padrão
        
        # Calcular valor da parcela baseado no valor total da operação
        numero_parcelas = getattr(lancamento, 'numero_parcelas', 1) or 1
        valor_parcela = valor_compra / numero_parcelas if numero_parcelas > 1 else valor_compra
        
        # Criar compra
        compra = Compra(
            produto=lancamento.produto_servico or 'Produto/Serviço Importado',
            quantidade=quantidade,  # Usar quantidade da importação
            preco_custo=valor_compra,
            valor=valor_compra,
            tipo_compra=tipo_compra,
            data_prevista=lancamento.data_prevista,
            data_realizada=lancamento.data_realizada,
            realizado=lancamento.realizado,
            observacoes=f'Compra criada automaticamente durante importação - Lançamento ID: {lancamento.id}',
            usuario_id=usuario_id,
            fornecedor_id=lancamento.fornecedor_id,
            numero_parcelas=numero_parcelas,
            valor_parcela=valor_parcela,
            nota_fiscal=lancamento.nota_fiscal
        )
        
        db.session.add(compra)
        db.session.flush()  # Para obter o ID
        
        app.logger.info(f"✅ Compra criada automaticamente: ID {compra.id}, Produto: {compra.produto}, Valor: R$ {compra.valor}")
        
        return compra
        
    except Exception as e:
        app.logger.error(f"❌ Erro ao criar compra automática: {str(e)}")
        return None

def atualizar_estoque_venda(venda, usuario_id):
    """Atualiza o estoque após uma venda (independente do status)"""
    try:
        
        if not hasattr(venda, 'tipo_venda') or venda.tipo_venda != 'produto':
            return True, "Venda não é de produto - estoque não será atualizado"
        
        if not hasattr(venda, 'produto') or not venda.produto:
            return True, "Produto não especificado - estoque não será atualizado"
        
        # Normalizar nome do produto
        nome_produto_normalizado = normalizar_nome_produto(venda.produto)
        venda.produto = nome_produto_normalizado
        
        # Buscar ou criar produto no estoque
        produto_estoque = Produto.query.filter_by(
            nome=nome_produto_normalizado,
            usuario_id=usuario_id
        ).first()
        
        if not produto_estoque:
            # Criar produto no estoque
            produto_estoque = Produto(
                nome=nome_produto_normalizado,
                descricao=f'Produto criado automaticamente via venda - {venda.produto}',
                preco_custo=0.0,
                preco_venda=venda.valor,
                estoque=0,  # Será calculado
                usuario_id=usuario_id
            )
            db.session.add(produto_estoque)
            db.session.flush()
            app.logger.info(f"✅ Produto '{venda.produto}' criado no estoque (ID: {produto_estoque.id})")
        
        # Calcular estoque real baseado em todas as compras e vendas
        estoque_real = calcular_estoque_produto(venda.produto, usuario_id)
        
        if estoque_real >= 0:
            produto_estoque.estoque = estoque_real
            db.session.commit()
            return True, f"Estoque atualizado: {estoque_real} unidades"
        else:
            return False, f"Estoque insuficiente: {produto_estoque.estoque} disponíveis, {venda.quantidade} solicitadas"
        
    except Exception as e:
        app.logger.error(f"❌ Erro ao atualizar estoque da venda: {str(e)}")
        return False, f"Erro ao atualizar estoque: {str(e)}"

def atualizar_estoque_compra(compra, usuario_id):
    """Atualiza o estoque após uma compra (independente do status) - Suporta múltiplos produtos"""
    try:
        # Verificar se a compra é de mercadoria
        if not hasattr(compra, 'tipo_compra') or compra.tipo_compra == 'servico':
            return True, "Compra não é de produto - estoque não será atualizado"

        if not hasattr(compra, 'produto') or not compra.produto:
            return True, "Produto não especificado - estoque não será atualizado"

        # Buscar lançamento vinculado para pegar itens_carrinho
        lancamento = Lancamento.query.filter_by(compra_id=compra.id).first()
        itens_para_processar = []

        if lancamento and lancamento.itens_carrinho:
            # Compra com múltiplos produtos (carrinho)
            try:
                import json
                itens_carrinho = json.loads(lancamento.itens_carrinho)

                for item in itens_carrinho:
                    # Processar apenas mercadorias
                    if item.get('tipo') == 'mercadoria':
                        itens_para_processar.append({
                            'nome': item['nome'],
                            'quantidade': item['qtd'],
                            'preco_custo': item['preco']
                        })

                app.logger.info(f"🛒 Processando {len(itens_para_processar)} produtos do carrinho")
            except Exception as e:
                app.logger.warning(f"⚠️ Erro ao parsear itens_carrinho: {e}, processando como produto único")
                itens_para_processar = []

        # Se não tem itens do carrinho, processar como produto único (comportamento antigo)
        if not itens_para_processar:
            itens_para_processar = [{
                'nome': compra.produto,
                'quantidade': compra.quantidade if hasattr(compra, 'quantidade') else 1,
                'preco_custo': compra.preco_custo
            }]

        # Processar cada item
        mensagens = []
        for item in itens_para_processar:
            nome_produto_normalizado = normalizar_nome_produto(item['nome'])

            # Buscar ou criar produto no estoque
            produto_estoque = Produto.query.filter_by(
                nome=nome_produto_normalizado,
                usuario_id=usuario_id
            ).first()

            if not produto_estoque:
                # Criar produto no estoque
                produto_estoque = Produto(
                    nome=nome_produto_normalizado,
                    descricao=f'Produto criado automaticamente via compra - {item["nome"]}',
                    preco_custo=item['preco_custo'],
                    preco_venda=item['preco_custo'] * 1.3,  # Margem de 30%
                    estoque=0,  # Será calculado
                    usuario_id=usuario_id
                )
                db.session.add(produto_estoque)
                db.session.flush()
                app.logger.info(f"✅ Produto '{item['nome']}' criado no estoque (ID: {produto_estoque.id})")

            # Calcular estoque e preço médio reais baseado em todas as compras e vendas
            estoque_real = calcular_estoque_produto(nome_produto_normalizado, usuario_id)
            preco_medio_real = calcular_preco_medio_produto(nome_produto_normalizado, usuario_id)

            produto_estoque.estoque = estoque_real
            produto_estoque.preco_custo = preco_medio_real

            mensagens.append(f"{item['nome']}: {estoque_real} un, R$ {preco_medio_real:.2f}")

        db.session.commit()

        if len(mensagens) > 1:
            return True, f"Estoque atualizado: {len(mensagens)} produtos processados"
        else:
            return True, f"Estoque atualizado: {mensagens[0]}"

    except Exception as e:
        app.logger.error(f"❌ Erro ao atualizar estoque da compra: {str(e)}")
        return False, f"Erro ao atualizar estoque: {str(e)}"

def buscar_cliente(nome_input, cache):
    """Busca cliente com busca inteligente"""
    if not nome_input:
        return None
    
    nome = str(nome_input).strip().lower()
    
    # Busca exata
    if nome in cache:
        return cache[nome].id
    
    # Busca parcial
    for nome_cached, cliente in cache.items():
        if nome in nome_cached or nome_cached in nome:
            return cliente.id
    
    return None

def buscar_fornecedor(nome_input, cache):
    """Busca fornecedor com busca inteligente"""
    if not nome_input:
        return None
    
    nome = str(nome_input).strip().lower()
    
    # Busca exata
    if nome in cache:
        return cache[nome].id
    
    # Busca parcial
    for nome_cached, fornecedor in cache.items():
        if nome in nome_cached or nome_cached in nome:
            return fornecedor.id
    
    return None



# REMOVIDO: Função desnecessária - saldo é calculado dinamicamente
# def atualizar_saldo_conta_caixa(conta_id, valor, tipo):

@app.route('/api/lancamentos/agendados', methods=['GET'])
def api_lancamentos_agendados():
    """API para buscar lançamentos agendados que precisam de confirmação"""
    # Verificar se está logado (pode ser usuario_id ou sub_usuario_id)
    if 'usuario_id' not in session and 'sub_usuario_id' not in session:
        return jsonify({'error': 'Usuário não autenticado'}), 401
    
    # Se for sub-usuário, retornar lista vazia (não há lançamentos agendados para contadores)
    if 'sub_usuario_id' in session:
        return jsonify({
            'success': True,
            'lancamentos': [],
            'count': 0
        })
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario or usuario.tipo == 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    
    try:
        # Buscar todos os usuários da mesma empresa
        empresa_id = obter_empresa_id_sessao(session, usuario)

        usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
        usuarios_ids = [u.id for u in usuarios_empresa]
        
        # Data atual
        hoje = datetime.now().date()
        
        # Buscar lançamentos agendados para hoje que ainda não foram confirmados como realizados
        lancamentos_agendados = Lancamento.query.filter(
            Lancamento.empresa_id == empresa_id,
            Lancamento.data_realizada == hoje,
            Lancamento.realizado == False  # Apenas os que ainda não foram confirmados
        ).order_by(Lancamento.data_realizada.asc()).all()
        
        lancamentos_data = []
        for lancamento in lancamentos_agendados:
            lancamento_data = {
                'id': lancamento.id,
                'descricao': lancamento.descricao,
                'valor': lancamento.valor,
                'tipo': lancamento.tipo,
                'categoria': lancamento.categoria,
                'data_prevista': lancamento.data_prevista.strftime('%d/%m/%Y'),
                'conta_caixa': lancamento.conta_caixa.nome if lancamento.conta_caixa else None
            }
            lancamentos_data.append(lancamento_data)
        
        return jsonify({
            'success': True,
            'lancamentos': lancamentos_data,
            'count': len(lancamentos_data)
        })
        
    except Exception as e:
        app.logger.error(f"Erro ao buscar lançamentos agendados: {str(e)}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/lancamentos/excluir-lote', methods=['POST'])
def api_excluir_lote():
    """API para excluir lançamentos em lote"""
    if 'usuario_id' not in session:
        return jsonify({'error': 'Usuário não autenticado'}), 401

    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario or usuario.tipo == 'admin':
        return jsonify({'error': 'Acesso negado'}), 403

    try:
        data = request.get_json()
        lancamento_ids = data.get('lancamento_ids', [])

        if not lancamento_ids:
            return jsonify({'error': 'Nenhum lançamento selecionado'}), 400

        # Obter empresa_id correta da sessão
        empresa_id = obter_empresa_id_sessao(session, usuario)

        # Buscar lançamentos do usuário E da empresa
        lancamentos = Lancamento.query.filter(
            Lancamento.id.in_(lancamento_ids),
            Lancamento.usuario_id == usuario.id,
            Lancamento.empresa_id == empresa_id
        ).all()
        
        if not lancamentos:
            return jsonify({'error': 'Nenhum lançamento válido encontrado'}), 404
        
        # CORREÇÃO: Verificar vínculos tanto diretos quanto através da tabela Vinculo
        lancamentos_vinculados = []
        for lancamento in lancamentos:
            # Verificar vínculos diretos (lógica antiga)
            if hasattr(lancamento, 'compra') and lancamento.compra:
                lancamentos_vinculados.append(lancamento.id)
                continue
            if hasattr(lancamento, 'venda') and lancamento.venda:
                lancamentos_vinculados.append(lancamento.id)
                continue
                
            # Verificar vínculos através da tabela Vinculo (nova lógica)
            vinculo_existe = Vinculo.query.filter(
                db.or_(
                    db.and_(Vinculo.lado_a_tipo == 'lancamento', Vinculo.lado_a_id == lancamento.id),
                    db.and_(Vinculo.lado_b_tipo == 'lancamento', Vinculo.lado_b_id == lancamento.id)
                )
            ).first()
            
            if vinculo_existe:
                lancamentos_vinculados.append(lancamento.id)
        
        # Se houver lançamentos vinculados, oferecer opção de exclusão em cascata
        if lancamentos_vinculados:
            # Verificar se foi solicitada exclusão forçada (em cascata)
            forcar_exclusao = data.get('forcar_exclusao', False)
            
            if not forcar_exclusao:
                return jsonify({
                    'error': f'Lançamentos {lancamentos_vinculados} estão vinculados a transações',
                    'vinculados': lancamentos_vinculados,
                    'pode_forcar': True,
                    'message': 'Deseja excluir os lançamentos e suas transações vinculadas?'
                }), 400
            else:
                # Exclusão em cascata: excluir também as transações vinculadas
                app.logger.info(f"Iniciando exclusão em cascata para lançamentos: {lancamento_ids}")
                
                for lancamento in lancamentos:
                    # Excluir vínculos e transações relacionadas
                    vinculos = Vinculo.query.filter(
                        db.or_(
                            db.and_(Vinculo.lado_a_tipo == 'lancamento', Vinculo.lado_a_id == lancamento.id),
                            db.and_(Vinculo.lado_b_tipo == 'lancamento', Vinculo.lado_b_id == lancamento.id)
                        )
                    ).all()
                    
                    for vinculo in vinculos:
                        # Identificar e excluir a transação vinculada
                        if vinculo.lado_a_tipo == 'venda' and vinculo.lado_a_id:
                            venda = db.session.get(Venda, vinculo.lado_a_id)
                            if venda:
                                db.session.delete(venda)
                                app.logger.info(f"Venda {venda.id} excluída em cascata")
                        elif vinculo.lado_b_tipo == 'venda' and vinculo.lado_b_id:
                            venda = db.session.get(Venda, vinculo.lado_b_id)
                            if venda:
                                db.session.delete(venda)
                                app.logger.info(f"Venda {venda.id} excluída em cascata")
                        elif vinculo.lado_a_tipo == 'compra' and vinculo.lado_a_id:
                            compra = db.session.get(Compra, vinculo.lado_a_id)
                            if compra:
                                db.session.delete(compra)
                                app.logger.info(f"Compra {compra.id} excluída em cascata")
                        elif vinculo.lado_b_tipo == 'compra' and vinculo.lado_b_id:
                            compra = db.session.get(Compra, vinculo.lado_b_id)
                            if compra:
                                db.session.delete(compra)
                                app.logger.info(f"Compra {compra.id} excluída em cascata")
                        
                        # Excluir o vínculo
                        db.session.delete(vinculo)
                    
                    # Excluir parcelas se existirem
                    if hasattr(lancamento, 'parcelas') and lancamento.parcelas:
                        for parcela in lancamento.parcelas:
                            db.session.delete(parcela)
        
        # Excluir lançamentos (incluindo parcelas para lançamentos não vinculados)
        count = 0
        for lancamento in lancamentos:
            # Excluir parcelas se existirem (para lançamentos não vinculados)
            if lancamento.id not in lancamentos_vinculados:
                if hasattr(lancamento, 'parcelas') and lancamento.parcelas:
                    for parcela in lancamento.parcelas:
                        db.session.delete(parcela)
                        app.logger.info(f"Parcela {parcela.id} excluída para lançamento {lancamento.id}")
            
            db.session.delete(lancamento)
            count += 1
            app.logger.info(f"Lançamento {lancamento.id} excluído com sucesso")
        
        db.session.commit()
        
        # Mensagem de sucesso diferenciada
        if lancamentos_vinculados and data.get('forcar_exclusao', False):
            message = f'{count} lançamento(s) e suas transações vinculadas excluídos com sucesso'
        else:
            message = f'{count} lançamento(s) excluído(s) com sucesso'
        
        return jsonify({
            'success': True,
            'count': count,
            'message': message
        })
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao excluir lançamentos em lote: {str(e)}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/clientes/excluir-lote', methods=['POST'])
def api_excluir_clientes_lote():
    """API para excluir clientes em lote"""
    if 'usuario_id' not in session:
        return jsonify({'error': 'Usuário não autenticado'}), 401

    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario or usuario.tipo == 'admin':
        return jsonify({'error': 'Acesso negado'}), 403

    try:
        data = request.get_json()
        cliente_ids = data.get('cliente_ids', [])

        if not cliente_ids:
            return jsonify({'error': 'Nenhum cliente selecionado'}), 400

        # Obter empresa_id correta da sessão
        empresa_id = obter_empresa_id_sessao(session, usuario)

        # Buscar clientes do usuário E da empresa
        clientes = Cliente.query.filter(
            Cliente.id.in_(cliente_ids),
            Cliente.usuario_id == usuario.id,
            Cliente.empresa_id == empresa_id
        ).all()
        
        if not clientes:
            return jsonify({'error': 'Nenhum cliente válido encontrado'}), 404
        
        # Verificar se algum cliente tem vendas associadas
        clientes_com_vendas = []
        for cliente in clientes:
            vendas = Venda.query.filter_by(cliente_id=cliente.id).count()
            if vendas > 0:
                clientes_com_vendas.append(cliente.id)
        
        if clientes_com_vendas:
            return jsonify({
                'error': f'Clientes {clientes_com_vendas} não podem ser excluídos pois possuem vendas associadas'
            }), 400
        
        # Excluir clientes
        count = 0
        for cliente in clientes:
            # Desvincular lançamentos órfãos primeiro
            Lancamento.query.filter_by(cliente_id=cliente.id).update({Lancamento.cliente_id: None})
            
            db.session.delete(cliente)
            count += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'count': count,
            'message': f'{count} cliente(s) excluído(s) com sucesso'
        })
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao excluir clientes em lote: {str(e)}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/fornecedores/excluir-lote', methods=['POST'])
def api_excluir_fornecedores_lote():
    """API para excluir fornecedores em lote"""
    if 'usuario_id' not in session:
        return jsonify({'error': 'Usuário não autenticado'}), 401

    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario or usuario.tipo == 'admin':
        return jsonify({'error': 'Acesso negado'}), 403

    try:
        data = request.get_json()
        fornecedor_ids = data.get('fornecedor_ids', [])

        if not fornecedor_ids:
            return jsonify({'error': 'Nenhum fornecedor selecionado'}), 400

        # Obter empresa_id correta da sessão
        empresa_id = obter_empresa_id_sessao(session, usuario)

        # Buscar fornecedores do usuário E da empresa
        fornecedores = Fornecedor.query.filter(
            Fornecedor.id.in_(fornecedor_ids),
            Fornecedor.usuario_id == usuario.id,
            Fornecedor.empresa_id == empresa_id
        ).all()
        
        if not fornecedores:
            return jsonify({'error': 'Nenhum fornecedor válido encontrado'}), 404
        
        # Verificar se algum fornecedor tem compras associadas
        fornecedores_com_compras = []
        for fornecedor in fornecedores:
            compras = Compra.query.filter_by(fornecedor_id=fornecedor.id).count()
            if compras > 0:
                fornecedores_com_compras.append(fornecedor.id)
        
        if fornecedores_com_compras:
            return jsonify({
                'error': f'Fornecedores {fornecedores_com_compras} não podem ser excluídos pois possuem compras associadas'
            }), 400
        
        # Excluir fornecedores
        count = 0
        for fornecedor in fornecedores:
            db.session.delete(fornecedor)
            count += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'count': count,
            'message': f'{count} fornecedor(es) excluído(s) com sucesso'
        })
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao excluir fornecedores em lote: {str(e)}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/criar-cliente-fornecedor', methods=['POST'])
def api_criar_cliente_fornecedor():
    """API para criar novo cliente ou fornecedor"""
    if 'usuario_id' not in session:
        return jsonify({'error': 'Usuário não autenticado'}), 401
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario or usuario.tipo == 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    
    try:
        data = request.get_json()
        nome = data.get('nome', '').strip()
        tipo = data.get('tipo', '').strip().lower()
        
        if not nome:
            return jsonify({'error': 'Nome é obrigatório'}), 400
        
        if tipo not in ['cliente', 'fornecedor']:
            return jsonify({'error': 'Tipo deve ser cliente ou fornecedor'}), 400
        
        # Verificar se já existe com mesmo nome
        if tipo == 'cliente':
            existente = Cliente.query.filter(
                Cliente.usuario_id == usuario.id,
                Cliente.nome.ilike(nome)
            ).first()
        else:
            existente = Fornecedor.query.filter(
                Fornecedor.usuario_id == usuario.id,
                Fornecedor.nome.ilike(nome)
            ).first()
        
        if existente:
            return jsonify({'error': f'{tipo.capitalize()} com este nome já existe'}), 400
        
        # Obter empresa_id correto (considerando contexto de contador)
        empresa_id = obter_empresa_id_sessao(session, usuario)

        # Criar novo registro
        if tipo == 'cliente':
            novo_registro = Cliente(
                nome=nome,
                usuario_id=usuario.id,
                empresa_id=empresa_id
            )
        else:
            novo_registro = Fornecedor(
                nome=nome,
                usuario_id=usuario.id,
                empresa_id=empresa_id
            )
        
        db.session.add(novo_registro)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'id': novo_registro.id,
            'nome': novo_registro.nome,
            'tipo': tipo,
            'message': f'{tipo.capitalize()} criado com sucesso'
        })
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao criar {tipo}: {str(e)}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/vendas/marcar-status-lote', methods=['POST'])
def api_marcar_vendas_status_lote():
    """API para marcar status de vendas em lote"""
    if 'usuario_id' not in session:
        return jsonify({'error': 'Usuário não autenticado'}), 401
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario or usuario.tipo == 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    
    try:
        data = request.get_json()
        venda_ids = data.get('venda_ids', [])
        realizado = data.get('realizado', True)
        
        if not venda_ids:
            return jsonify({'error': 'Nenhuma venda selecionada'}), 400
        
        # Buscar vendas do usuário
        vendas = Venda.query.filter(
            Venda.id.in_(venda_ids),
            Venda.usuario_id == usuario.id
        ).all()
        
        if not vendas:
            return jsonify({'error': 'Nenhuma venda válida encontrada'}), 404
        
        # Atualizar status
        count = 0
        for venda in vendas:
            venda.realizado = realizado
            count += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'count': count,
            'message': f'{count} venda(s) atualizada(s) com sucesso'
        })
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao marcar status de vendas em lote: {str(e)}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/vendas/excluir-lote', methods=['POST'])
def api_excluir_vendas_lote():
    """API para excluir vendas em lote"""
    if 'usuario_id' not in session:
        return jsonify({'error': 'Usuário não autenticado'}), 401

    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario or usuario.tipo == 'admin':
        return jsonify({'error': 'Acesso negado'}), 403

    try:
        data = request.get_json()
        venda_ids = data.get('venda_ids', [])

        if not venda_ids:
            return jsonify({'error': 'Nenhuma venda selecionada'}), 400

        # Obter empresa_id correta da sessão
        empresa_id = obter_empresa_id_sessao(session, usuario)

        # Buscar vendas do usuário E da empresa
        vendas = Venda.query.filter(
            Venda.id.in_(venda_ids),
            Venda.usuario_id == usuario.id,
            Venda.empresa_id == empresa_id
        ).all()
        
        if not vendas:
            return jsonify({'error': 'Nenhuma venda válida encontrada'}), 404
        
        # Excluir vendas e seus lançamentos vinculados
        count = 0
        for venda in vendas:
            # Excluir TODOS os lançamentos financeiros vinculados à venda
            # 1. Lançamentos com venda_id direto
            lancamentos_diretos = Lancamento.query.filter_by(venda_id=venda.id).all()
            for lancamento in lancamentos_diretos:
                db.session.delete(lancamento)
                app.logger.info(f"✅ Lançamento direto {lancamento.id} excluído para venda {venda.id}")

            # 2. Lançamentos vinculados através da tabela Vinculo
            vinculos = Vinculo.query.filter(
                db.or_(
                    db.and_(Vinculo.lado_a_tipo == 'venda', Vinculo.lado_a_id == venda.id),
                    db.and_(Vinculo.lado_b_tipo == 'venda', Vinculo.lado_b_id == venda.id)
                )
            ).all()

            for vinculo in vinculos:
                # Buscar lançamento vinculado
                if vinculo.lado_a_tipo == 'lancamento':
                    lancamento_vinculado = db.session.get(Lancamento, vinculo.lado_a_id)
                elif vinculo.lado_b_tipo == 'lancamento':
                    lancamento_vinculado = db.session.get(Lancamento, vinculo.lado_b_id)
                else:
                    lancamento_vinculado = None

                # Excluir lançamento vinculado se existir
                if lancamento_vinculado:
                    db.session.delete(lancamento_vinculado)
                    app.logger.info(f"✅ Lançamento vinculado {lancamento_vinculado.id} excluído para venda {venda.id}")

                # Excluir o vínculo
                db.session.delete(vinculo)

            # Excluir a venda
            db.session.delete(venda)
            count += 1

        db.session.commit()
        
        return jsonify({
            'success': True,
            'count': count,
            'message': f'{count} venda(s) excluída(s) com sucesso'
        })
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao excluir vendas em lote: {str(e)}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

 
@app.route('/api/compras/marcar-status-lote', methods=['POST'])
def api_marcar_compras_status_lote():
    """API para marcar status de compras em lote"""
    if 'usuario_id' not in session:
        return jsonify({'error': 'Usuário não autenticado'}), 401

    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario or usuario.tipo == 'admin':
        return jsonify({'error': 'Acesso negado'}), 403

    try:
        data = request.get_json()
        compra_ids = data.get('compra_ids', [])
        realizado = data.get('realizado', True)

        if not compra_ids:
            return jsonify({'error': 'Nenhuma compra selecionada'}), 400

        # Buscar compras do usuário
        compras = Compra.query.filter(
            Compra.id.in_(compra_ids),
            Compra.usuario_id == usuario.id
        ).all()

        if not compras:
            return jsonify({'error': 'Nenhuma compra válida encontrada'}), 404

        # Atualizar status
        count = 0
        for compra in compras:
            compra.realizado = realizado
            count += 1

        db.session.commit()

        return jsonify({
            'success': True,
            'count': count,
            'message': f'{count} compra(s) atualizada(s) com sucesso'
        })

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao marcar status de compras em lote: {str(e)}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/compras/excluir-lote', methods=['POST'])
def api_excluir_compras_lote():
    """API para excluir compras em lote"""
    if 'usuario_id' not in session:
        return jsonify({'error': 'Usuário não autenticado'}), 401

    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario or usuario.tipo == 'admin':
        return jsonify({'error': 'Acesso negado'}), 403

    try:
        data = request.get_json()
        compra_ids = data.get('compra_ids', [])

        if not compra_ids:
            return jsonify({'error': 'Nenhuma compra selecionada'}), 400

        # Obter empresa_id correta da sessão
        empresa_id = obter_empresa_id_sessao(session, usuario)

        # Buscar compras do usuário E da empresa
        compras = Compra.query.filter(
            Compra.id.in_(compra_ids),
            Compra.usuario_id == usuario.id,
            Compra.empresa_id == empresa_id
        ).all()

        if not compras:
            return jsonify({'error': 'Nenhuma compra válida encontrada'}), 404

        # Excluir compras e seus lançamentos vinculados
        count = 0
        for compra in compras:
            # Excluir TODOS os lançamentos financeiros vinculados à compra
            # 1. Lançamentos com compra_id direto
            lancamentos_diretos = Lancamento.query.filter_by(compra_id=compra.id).all()
            for lancamento in lancamentos_diretos:
                db.session.delete(lancamento)
                app.logger.info(f"✅ Lançamento direto {lancamento.id} excluído para compra {compra.id}")

            # 2. Lançamentos vinculados através da tabela Vinculo
            vinculos = Vinculo.query.filter(
                db.or_(
                    db.and_(Vinculo.lado_a_tipo == 'compra', Vinculo.lado_a_id == compra.id),
                    db.and_(Vinculo.lado_b_tipo == 'compra', Vinculo.lado_b_id == compra.id)
                )
            ).all()

            for vinculo in vinculos:
                # Buscar lançamento vinculado
                if vinculo.lado_a_tipo == 'lancamento':
                    lancamento_vinculado = db.session.get(Lancamento, vinculo.lado_a_id)
                elif vinculo.lado_b_tipo == 'lancamento':
                    lancamento_vinculado = db.session.get(Lancamento, vinculo.lado_b_id)
                else:
                    lancamento_vinculado = None

                # Excluir lançamento vinculado se existir
                if lancamento_vinculado:
                    db.session.delete(lancamento_vinculado)
                    app.logger.info(f"✅ Lançamento vinculado {lancamento_vinculado.id} excluído para compra {compra.id}")

                # Excluir o vínculo
                db.session.delete(vinculo)

            # Excluir a compra
            db.session.delete(compra)
            count += 1

        db.session.commit()

        return jsonify({
            'success': True,
            'count': count,
            'message': f'{count} compra(s) excluída(s) com sucesso'
        })

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao excluir compras em lote: {str(e)}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

def buscar_plano_conta_automatico(usuario_id, tipo_lancamento, categoria_nome):
    """
    Busca o PlanoConta analítico mais adequado para um lançamento automático.
    """
    # 1. Busca exata por nome
    pc = PlanoConta.query.filter(
        PlanoConta.usuario_id == usuario_id,
        PlanoConta.nome == categoria_nome,
        PlanoConta.tipo == tipo_lancamento,
        PlanoConta.natureza == 'analitica',
        PlanoConta.ativo == True
    ).first()
    
    if pc:
        return pc
        
    # 2. Busca por termos similares
    termo = ""
    if categoria_nome == 'Vendas':
        termo = 'Venda'
    elif categoria_nome == 'Serviços':
        termo = 'Serviço'
    elif categoria_nome == 'Compras':
        termo = 'Compra'
        
    if termo:
        pc = PlanoConta.query.filter(
            PlanoConta.usuario_id == usuario_id,
            PlanoConta.nome.like(f'%{termo}%'),
            PlanoConta.tipo == tipo_lancamento,
            PlanoConta.natureza == 'analitica',
            PlanoConta.ativo == True
        ).first()
        
    return pc

def criar_lancamento_financeiro_automatico(venda_ou_compra, tipo, usuario_id):
    """
    Cria automaticamente um lançamento financeiro para uma venda ou compra
    """
    try:
        print(f"🔍 Iniciando criação de lançamento financeiro para {tipo} ID: {venda_ou_compra.id}")
        
        # Verificar se já existe um lançamento financeiro
        if tipo == 'venda':
            lancamento_existente = Lancamento.query.filter_by(venda_id=venda_ou_compra.id).first()
        else:
            lancamento_existente = Lancamento.query.filter_by(compra_id=venda_ou_compra.id).first()
        
        if lancamento_existente:
            print(f"⚠️ Lançamento financeiro já existe para {tipo} {venda_ou_compra.id}")
            return lancamento_existente
        
        # Determinar se é venda ou compra
        if tipo == 'venda':
            cliente = db.session.get(Cliente, venda_ou_compra.cliente_id)
            cliente_nome = cliente.nome if cliente else 'Cliente não encontrado'
            descricao = f"Venda - {venda_ou_compra.produto} - {cliente_nome}"
            valor = venda_ou_compra.valor_final if hasattr(venda_ou_compra, 'valor_final') and venda_ou_compra.valor_final else venda_ou_compra.valor
            tipo_lancamento = 'entrada'
            categoria = 'Vendas' if venda_ou_compra.tipo_venda == 'produto' else 'Serviços'
            venda_id = venda_ou_compra.id
            compra_id = None
            
            plano_conta_id = None
            pc = buscar_plano_conta_automatico(usuario_id, tipo_lancamento, categoria)
            if pc:
                plano_conta_id = pc.id
        else:  # compra
            fornecedor = db.session.get(Fornecedor, venda_ou_compra.fornecedor_id)
            fornecedor_nome = fornecedor.nome if fornecedor else 'Fornecedor não encontrado'
            descricao = f"Compra - {venda_ou_compra.produto} - {fornecedor_nome}"
            valor = venda_ou_compra.valor
            tipo_lancamento = 'saida'
            categoria = 'Compras' if venda_ou_compra.tipo_compra == 'mercadoria' else 'Serviços'
            venda_id = None
            compra_id = venda_ou_compra.id
            
            plano_conta_id = None
            pc = buscar_plano_conta_automatico(usuario_id, tipo_lancamento, categoria)
            if pc:
                plano_conta_id = pc.id
        
        # Buscar o usuário para obter o empresa_id
        usuario = db.session.get(Usuario, usuario_id)
        if not usuario:
            print(f"❌ Usuário {usuario_id} não encontrado")
            return None
        
        # Obter empresa_id correta da sessão (considera acesso_contador)
        from flask import session as flask_session
        empresa_id_correta = obter_empresa_id_sessao(flask_session, usuario)
        if not empresa_id_correta:
            print(f"❌ Não foi possível obter empresa_id para usuário {usuario_id}")
            return None
        
        # Criar o lançamento financeiro
        novo_lancamento = Lancamento(
            descricao=descricao,
            valor=valor,
            tipo=tipo_lancamento,
            categoria=categoria,
            plano_conta_id=plano_conta_id,
            data_prevista=venda_ou_compra.data_prevista,
            data_realizada=venda_ou_compra.data_realizada if venda_ou_compra.realizado else None,
            realizado=venda_ou_compra.realizado,
            usuario_id=usuario_id,
            empresa_id=empresa_id_correta,  # Usar empresa_id correta da sessão
            venda_id=venda_id,
            compra_id=compra_id,
            cliente_id=venda_ou_compra.cliente_id if tipo == 'venda' else None,
            fornecedor_id=venda_ou_compra.fornecedor_id if tipo == 'compra' else None,
            usuario_criacao_id=usuario_id  # Registrar quem criou
        )
        
        db.session.add(novo_lancamento)
        db.session.commit()
        
        print(f"✅ Lançamento financeiro criado: {descricao} - R$ {valor:.2f}")
        return novo_lancamento
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Erro ao criar lançamento financeiro: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

@app.route('/api/criar_lancamento_financeiro/<tipo>/<int:id>', methods=['POST'])
def api_criar_lancamento_financeiro(tipo, id):
    """API para criar lançamento financeiro para venda ou compra"""
    if 'usuario_id' not in session:
        return jsonify({'success': False, 'error': 'Usuário não autenticado'}), 401
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return jsonify({'success': False, 'error': 'Acesso negado'}), 403
    
    try:
        if tipo == 'venda':
            venda_ou_compra = db.session.get(Venda, id)
            if not venda_ou_compra or venda_ou_compra.usuario_id != usuario.id:
                return jsonify({'success': False, 'error': 'Venda não encontrada'}), 404
        elif tipo == 'compra':
            venda_ou_compra = db.session.get(Compra, id)
            if not venda_ou_compra or venda_ou_compra.usuario_id != usuario.id:
                return jsonify({'success': False, 'error': 'Compra não encontrada'}), 404
        else:
            return jsonify({'success': False, 'error': 'Tipo inválido'}), 400
        
        # Verificar se já existe um lançamento financeiro
        if venda_ou_compra.lancamento_financeiro:
            return jsonify({'success': False, 'error': 'Lançamento financeiro já existe'}), 400
        
        # Criar o lançamento financeiro
        lancamento = criar_lancamento_financeiro_automatico(venda_ou_compra, tipo, usuario.id)
        
        if lancamento:
            return jsonify({'success': True, 'message': 'Lançamento financeiro criado com sucesso'})
        else:
            return jsonify({'success': False, 'error': 'Erro ao criar lançamento financeiro'}), 500
            
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/criar_lancamentos_existentes')
def criar_lancamentos_existentes():
    """Cria lançamentos financeiros para vendas e compras existentes que não têm lançamentos"""
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if usuario.tipo == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    try:
        # Buscar todos os usuários da mesma empresa
        empresa_id = obter_empresa_id_sessao(session, usuario)

        usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
        usuarios_ids = [u.id for u in usuarios_empresa]
        
        # Buscar vendas sem lançamentos financeiros (limitando para evitar loops)
        vendas_sem_lancamento = Venda.query.filter(
            Venda.empresa_id == empresa_id,
            ~Venda.id.in_(db.session.query(Lancamento.venda_id).filter(Lancamento.venda_id.isnot(None)))
        ).limit(100).all()  # Limite de 100 para evitar loops
        
        # Buscar compras sem lançamentos financeiros (limitando para evitar loops)
        compras_sem_lancamento = Compra.query.filter(
            Compra.empresa_id == empresa_id,
            ~Compra.id.in_(db.session.query(Lancamento.compra_id).filter(Lancamento.compra_id.isnot(None)))
        ).limit(100).all()  # Limite de 100 para evitar loops
        
        lancamentos_criados = 0
        
        # Criar lançamentos para vendas
        for venda in vendas_sem_lancamento:
            try:
                # Verificar se já existe um lançamento para esta venda
                lancamento_existente = Lancamento.query.filter_by(venda_id=venda.id).first()
                if not lancamento_existente:
                    lancamento = criar_lancamento_financeiro_automatico(venda, 'venda', venda.usuario_id)
                    if lancamento:
                        lancamentos_criados += 1
            except Exception as e:
                db.session.rollback()
                print(f"Erro ao criar lançamento para venda {venda.id}: {e}")
                continue
        
        # Criar lançamentos para compras
        for compra in compras_sem_lancamento:
            try:
                # Verificar se já existe um lançamento para esta compra
                lancamento_existente = Lancamento.query.filter_by(compra_id=compra.id).first()
                if not lancamento_existente:
                    lancamento = criar_lancamento_financeiro_automatico(compra, 'compra', compra.usuario_id)
                    if lancamento:
                        lancamentos_criados += 1
            except Exception as e:
                print(f"Erro ao criar lançamento para compra {compra.id}: {e}")
                continue
        
        flash(f'✅ {lancamentos_criados} lançamentos financeiros criados automaticamente para vendas e compras existentes!', 'success')
        
    except Exception as e:
        print(f"❌ Erro ao criar lançamentos existentes: {str(e)}")
        flash('❌ Erro ao criar lançamentos financeiros para vendas e compras existentes.', 'danger')
    
    return redirect(url_for('lancamentos'))

def exportar_relatorio_fornecedores_excel(fornecedores_dados):
    """Exporta relatório de fornecedores para Excel"""
    try:
        # Criar um novo workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Fornecedores"
        
        # Cabeçalhos
        headers = ['Nome', 'Email', 'Telefone', 'CPF/CNPJ', 'Endereço', 'Data de Criação']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = Font(bold=True)
        
        # Dados
        for row, fornecedor in enumerate(fornecedores_dados, 2):
            ws.cell(row=row, column=1, value=fornecedor['nome'])
            ws.cell(row=row, column=2, value=fornecedor['email'] or '')
            ws.cell(row=row, column=3, value=fornecedor['telefone'] or '')
            ws.cell(row=row, column=4, value=fornecedor['cpf_cnpj'] or '')
            ws.cell(row=row, column=5, value=fornecedor['endereco'] or '')
            ws.cell(row=row, column=6, value=fornecedor['data_criacao'].strftime('%d/%m/%Y') if fornecedor['data_criacao'] else '')
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar arquivo
        filename = f"relatorio_fornecedores_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        wb.save(filepath)
        
        return send_file(filepath, as_attachment=True, download_name=filename)
        
    except Exception as e:
        print(f"Erro ao exportar relatório de fornecedores para Excel: {str(e)}")
        return None

def exportar_relatorio_produtos_pdf(produtos_dados):
    """Exporta relatório de produtos para PDF"""
    try:
        # Criar PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=getSampleStyleSheet()['Title'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Centralizado
        )
        title = Paragraph("Relatório de Produtos", title_style)
        elements.append(title)
        
        # Data do relatório
        data_style = ParagraphStyle(
            'CustomData',
            parent=getSampleStyleSheet()['Normal'],
            fontSize=10,
            spaceAfter=20,
            alignment=1  # Centralizado
        )
        data_relatorio = Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}", data_style)
        elements.append(data_relatorio)
        
        # Tabela de dados
        if produtos_dados:
            # Cabeçalhos da tabela
            headers = ['Produto', 'Quantidade Vendida', 'Valor Total', 'Número de Vendas']
            data = [headers]
            
            # Dados dos produtos
            for produto in produtos_dados:
                data.append([
                    produto['nome'],
                    str(produto['quantidade']),
                    f"R$ {produto['valor_total']:.2f}".replace('.', ','),
                    str(produto['num_vendas'])
                ])
            
            # Criar tabela
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),  # Valores numéricos à direita
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),     # Nome do produto à esquerda
            ]))
            
            elements.append(table)
        else:
            # Mensagem quando não há dados
            no_data_style = ParagraphStyle(
                'NoData',
                parent=getSampleStyleSheet()['Normal'],
                fontSize=12,
                spaceAfter=20,
                alignment=1  # Centralizado
            )
            no_data = Paragraph("Nenhum produto encontrado para o período selecionado.", no_data_style)
            elements.append(no_data)
        
        # Construir PDF
        doc.build(elements)
        buffer.seek(0)
        
        # Salvar arquivo
        filename = f"relatorio_produtos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        with open(filepath, 'wb') as f:
            f.write(buffer.getvalue())
        
        return send_file(filepath, as_attachment=True, download_name=filename)
        
    except Exception as e:
        print(f"Erro ao exportar relatório de produtos para PDF: {str(e)}")
        return None

def exportar_relatorio_produtos_excel(produtos_dados):
    """Exporta relatório de produtos para Excel"""
    try:
        # Criar um novo workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Produtos"
        
        # Cabeçalhos
        headers = ['Produto', 'Quantidade Vendida', 'Valor Total', 'Número de Vendas']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = Font(bold=True)
        
        # Dados
        for row, produto in enumerate(produtos_dados, 2):
            ws.cell(row=row, column=1, value=produto['nome'])
            ws.cell(row=row, column=2, value=produto['quantidade'])
            ws.cell(row=row, column=3, value=produto['valor_total'])
            ws.cell(row=row, column=4, value=produto['num_vendas'])
        
        # Formatação para valores monetários
        for row in range(2, len(produtos_dados) + 2):
            ws.cell(row=row, column=3).number_format = 'R$ #,##0.00'
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar arquivo
        filename = f"relatorio_produtos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        wb.save(filepath)
        
        return send_file(filepath, as_attachment=True, download_name=filename)
        
    except Exception as e:
        print(f"Erro ao exportar relatório de produtos para Excel: {str(e)}")
        return None

# ========================================
# SERVIÇOS DE DOMÍNIO - ESTOQUE E FINANCEIRO
# ========================================

def criar_hash_evento(tipo_evento, origem_tipo, origem_id):
    """Cria hash único para idempotência de eventos"""
    import hashlib
    texto = f"{tipo_evento}:{origem_tipo}:{origem_id}"
    return hashlib.sha256(texto.encode()).hexdigest()

def verificar_evento_existente(tipo_evento, origem_tipo, origem_id):
    """Verifica se um evento já foi processado"""
    hash_evento = criar_hash_evento(tipo_evento, origem_tipo, origem_id)
    return EventLog.query.filter_by(hash_evento=hash_evento).first()

def registrar_evento(tipo_evento, origem_tipo, origem_id, dados_evento, usuario_id):
    """Registra um evento no log para idempotência"""
    hash_evento = criar_hash_evento(tipo_evento, origem_tipo, origem_id)
    evento = EventLog(
        tipo_evento=tipo_evento,
        origem_tipo=origem_tipo,
        origem_id=origem_id,
        hash_evento=hash_evento,
        dados_evento=dados_evento,
        usuario_id=usuario_id
    )
    db.session.add(evento)
    return evento

def aplicar_movimento_estoque(produto_id, quantidade, tipo_movimento, usuario_id):
    """Aplica movimento de estoque (incremento/decremento)"""
    produto = db.session.get(Produto, produto_id)
    
    if not produto:
        return None
    
    if tipo_movimento == 'incremento':
        produto.estoque += quantidade
    elif tipo_movimento == 'decremento':
        produto.estoque = max(0, produto.estoque - quantidade)
    
    return produto

def criar_lancamento_financeiro_venda(venda, usuario_id):
    """Cria lançamento financeiro para uma venda"""
    app.logger.info(f"🔧 Iniciando criação de lançamento financeiro para venda {venda.id}")
    
    try:
        # Buscar o usuário para obter a empresa_id
        usuario = db.session.get(Usuario, usuario_id)
        if not usuario:
            app.logger.error(f"Usuário {usuario_id} não encontrado")
            raise ValueError(f"Usuário {usuario_id} não encontrado")
        
        app.logger.info(f"✅ Usuário encontrado: {usuario.nome} (ID: {usuario.id})")
        
        # Obter empresa_id correta da sessão (considera acesso_contador)
        from flask import session as flask_session
        empresa_id_correta = obter_empresa_id_sessao(flask_session, usuario)
        if not empresa_id_correta:
            app.logger.error(f"Usuário {usuario.id} não tem empresa_id definido ou acesso_contador inválido")
            raise ValueError(f"Usuário {usuario.id} não tem empresa associada")
        
        app.logger.info(f"✅ Empresa ID: {empresa_id_correta}")
        
        # Verificar se a venda tem cliente
        if not venda.cliente_id:
            app.logger.error(f"Venda {venda.id} não tem cliente associado")
            raise ValueError(f"Venda {venda.id} não tem cliente associado")
        
        app.logger.info(f"✅ Cliente ID: {venda.cliente_id}")
        
        # Verificar se já existe um lançamento para esta venda
        lancamento_existente = Lancamento.query.filter_by(venda_id=venda.id).first()
        if lancamento_existente:
            app.logger.info(f"✅ Lançamento financeiro já existe para venda {venda.id}")
            return lancamento_existente
        
        # Calcular valor total da venda (usando campos existentes)
        valor_total = venda.valor_final if venda.valor_final else venda.valor
        app.logger.info(f"💰 Valor total calculado: R$ {valor_total:.2f}")
        
        # Validar dados antes de criar o lançamento
        if not venda.data_prevista:
            app.logger.error(f"Venda {venda.id} não tem data_prevista definida")
            raise ValueError(f"Venda {venda.id} não tem data prevista definida")
        
        # Criar lançamento a receber
        app.logger.info(f"📝 Criando objeto Lancamento...")
        lancamento = Lancamento(
            descricao=f"Venda #{venda.id} - Cliente ID {venda.cliente_id}"[:100],
            valor=valor_total,
            tipo='entrada',
            categoria='Vendas',  # Categoria padrão para vendas
            data_prevista=venda.data_prevista,
            realizado=False,
            usuario_id=usuario_id,
            empresa_id=empresa_id_correta,
            venda_id=venda.id
        )
        
        app.logger.info(f"✅ Objeto Lancamento criado com sucesso")
        
        # Adicionar à sessão
        db.session.add(lancamento)
        app.logger.info(f"✅ Lançamento adicionado à sessão")
        
        # Fazer flush para obter o ID
        db.session.flush()
        app.logger.info(f"✅ Flush executado, ID do lançamento: {lancamento.id}")
        
        app.logger.info(f"🎉 Lançamento financeiro criado com sucesso para venda {venda.id}: R$ {valor_total:.2f}")
        return lancamento
        
    except Exception as e:
        app.logger.error(f"❌ Erro ao criar lançamento financeiro para venda {venda.id}: {str(e)}")
        raise

def criar_lancamento_financeiro_compra(compra, usuario_id):
    """Cria lançamento financeiro para uma compra"""
    app.logger.info(f"🔧 Iniciando criação de lançamento financeiro para compra {compra.id}")
    
    try:
        # Buscar o usuário para obter a empresa_id
        usuario = db.session.get(Usuario, usuario_id)
        if not usuario:
            app.logger.error(f"Usuário {usuario_id} não encontrado")
            raise ValueError(f"Usuário {usuario_id} não encontrado")
        
        app.logger.info(f"✅ Usuário encontrado: {usuario.nome} (ID: {usuario.id})")
        
        # Obter empresa_id correta da sessão (considera acesso_contador)
        from flask import session as flask_session
        empresa_id_correta = obter_empresa_id_sessao(flask_session, usuario)
        if not empresa_id_correta:
            app.logger.error(f"Usuário {usuario.id} não tem empresa_id definido ou acesso_contador inválido")
            raise ValueError(f"Usuário {usuario.id} não tem empresa associada")
        
        app.logger.info(f"✅ Empresa ID: {empresa_id_correta}")
        
        # Verificar se a compra tem fornecedor
        if not compra.fornecedor_id:
            app.logger.error(f"Compra {compra.id} não tem fornecedor associado")
            raise ValueError(f"Compra {compra.id} não tem fornecedor associado")
        
        app.logger.info(f"✅ Fornecedor ID: {compra.fornecedor_id}")
        
        # Verificar se já existe um lançamento para esta compra
        lancamento_existente = Lancamento.query.filter_by(compra_id=compra.id).first()
        if lancamento_existente:
            app.logger.info(f"✅ Lançamento financeiro já existe para compra {compra.id}")
            return lancamento_existente
        
        # Calcular valor total da compra (quantidade × valor)
        valor_total = compra.quantidade * compra.valor
        app.logger.info(f"💰 Valor total calculado: {compra.quantidade} x R$ {compra.valor} = R$ {valor_total:.2f}")
        
        # Validar dados antes de criar o lançamento
        if not compra.data_prevista:
            app.logger.error(f"Compra {compra.id} não tem data_prevista definida")
            raise ValueError(f"Compra {compra.id} não tem data prevista definida")
        
        # Criar lançamento a pagar
        app.logger.info(f"📝 Criando objeto Lancamento...")
        lancamento = Lancamento(
            descricao=f"Compra #{compra.id} - Fornecedor ID {compra.fornecedor_id}"[:100],
            valor=valor_total,
            tipo='saida',
            categoria='Compras',  # Categoria padrão para compras
            data_prevista=compra.data_prevista,
            realizado=False,
            usuario_id=usuario_id,
            empresa_id=empresa_id_correta,
            compra_id=compra.id
        )
        
        app.logger.info(f"✅ Objeto Lancamento criado com sucesso")
        
        # Adicionar à sessão
        db.session.add(lancamento)
        app.logger.info(f"✅ Lançamento adicionado à sessão")
        
        # Fazer flush para obter o ID
        db.session.flush()
        app.logger.info(f"✅ Flush executado, ID do lançamento: {lancamento.id}")
        
        app.logger.info(f"🎉 Lançamento financeiro criado com sucesso para compra {compra.id}: R$ {valor_total:.2f}")
        return lancamento
        
    except Exception as e:
        app.logger.error(f"❌ Erro ao criar lançamento financeiro para compra {compra.id}: {str(e)}")
        raise

def criar_vinculo(lado_a_tipo, lado_a_id, lado_b_tipo, lado_b_id, usuario_id):
    """Cria vínculo entre entidades do sistema"""
    # Verificar se o usuário existe
    usuario = db.session.get(Usuario, usuario_id)
    if not usuario:
        raise ValueError(f"Usuário {usuario_id} não encontrado para criar vínculo")
    
    # Verificar se já existe um vínculo similar
    vinculo_existente = Vinculo.query.filter_by(
        lado_a_tipo=lado_a_tipo,
        lado_a_id=lado_a_id,
        lado_b_tipo=lado_b_tipo,
        lado_b_id=lado_b_id,
        usuario_id=usuario_id
    ).first()
    
    if vinculo_existente:
        app.logger.info(f"Vínculo já existe entre {lado_a_tipo} {lado_a_id} e {lado_b_tipo} {lado_b_id}")
        return vinculo_existente
    
    vinculo = Vinculo(
        lado_a_tipo=lado_a_tipo,
        lado_a_id=lado_a_id,
        lado_b_tipo=lado_b_tipo,
        lado_b_id=lado_b_id,
        usuario_id=usuario_id
    )
    db.session.add(vinculo)
    app.logger.info(f"Vínculo criado: {lado_a_tipo} {lado_a_id} <-> {lado_b_tipo} {lado_b_id}")
    return vinculo

def processar_venda_criada(venda_id, usuario_id):
    """Processa uma venda recém-criada com atualização automática de estoque (independente do status realizado)"""
    try:
        # Verificar se já foi processada (idempotência) - TEMPORARIAMENTE DESABILITADO
        # if verificar_evento_existente('venda_criada', 'venda', venda_id):
        #     app.logger.info(f"Venda {venda_id} já foi processada anteriormente")
        #     return True, "Venda já processada anteriormente"
        
        # Registrar início do processamento - TEMPORARIAMENTE DESABILITADO
        # evento = registrar_evento('venda_criada', 'venda', venda_id, 
        #                         f"Iniciando processamento da venda {venda_id}", usuario_id)
        
        # Buscar a venda
        venda = db.session.get(Venda, venda_id)
        if not venda:
            app.logger.error(f"Venda {venda_id} não encontrada")
            # evento.status = 'erro' - TEMPORARIAMENTE DESABILITADO
            # evento.dados_evento = f"Venda {venda_id} não encontrada" - TEMPORARIAMENTE DESABILITADO
            # Removido db.session.commit() - será gerenciado pela transação principal
            raise ValueError(f"Venda {venda_id} não encontrada")
        
        # Processar estoque se for produto (movimento imediato - independente do status realizado)
        if venda.tipo_venda == 'produto':
            # Buscar produto pelo nome na empresa (não apenas no usuário)
            usuario = db.session.get(Usuario, usuario_id)
            if not usuario:
                raise ValueError(f"Usuário {usuario_id} não encontrado")
            
            # Buscar produto por nome em toda a empresa
            empresa_id = obter_empresa_id_sessao(session, usuario)

            usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
            usuarios_ids = [u.id for u in usuarios_empresa]
            
            produto = Produto.query.filter(
                Produto.nome == venda.produto,
                Produto.usuario_id.in_(usuarios_ids)
            ).first()
            
            if not produto:
                app.logger.error(f"Produto {venda.produto} não encontrado no estoque da empresa {usuario.empresa_id}")
                # evento.status = 'erro' - TEMPORARIAMENTE DESABILITADO
                # evento.dados_evento = f"Produto {venda.produto} não encontrado no estoque da empresa {usuario.empresa_id}" - TEMPORARIAMENTE DESABILITADO
                # Removido db.session.commit() - será gerenciado pela transação principal
                raise ValueError(f"Produto {venda.produto} não encontrado no estoque da empresa {usuario.empresa_id}")
            
            # ATUALIZAÇÃO: Sempre atualizar o estoque, independente do status "realizado"
            # O estoque será calculado baseado em todas as compras e vendas da empresa
            estoque_real = calcular_estoque_produto(venda.produto, usuario_id)
            produto.estoque = estoque_real
            app.logger.info(f"Estoque atualizado para produto {venda.produto}: {estoque_real} (venda processada independente do status)")
        
        # Criar lançamento financeiro
        lancamento = criar_lancamento_financeiro_venda(venda, usuario_id)
        
        # Fazer flush para obter o ID do lançamento
        db.session.flush()
        
        # Criar vínculo entre venda e lançamento
        vinculo = criar_vinculo('lancamento', lancamento.id, 'venda', venda_id, usuario_id)
        
        app.logger.info(f"Vínculo criado entre lançamento {lancamento.id} e venda {venda_id}")
        
        # Marcar como processada - TEMPORARIAMENTE DESABILITADO
        # evento.status = 'processado'
        # evento.dados_evento = f"Venda processada com sucesso - Estoque: {'atualizado' if venda.tipo_venda == 'produto' else 'não aplicável'}, Financeiro: criado, Vínculo: criado"
        
        app.logger.info(f"Venda {venda_id} processada com sucesso - Estoque: {'atualizado' if venda.tipo_venda == 'produto' else 'não aplicável'}, Financeiro: criado, Vínculo: criado")
        
        return True, "Venda processada com sucesso"
        
    except ValueError as e:
        app.logger.error(f"Erro de validação ao processar venda {venda_id}: {str(e)}")
        # if 'evento' in locals(): - TEMPORARIAMENTE DESABILITADO
        #     evento.status = 'erro' - TEMPORARIAMENTE DESABILITADO
        #     evento.dados_evento = f"Erro de validação: {str(e)}" - TEMPORARIAMENTE DESABILITADO
        #     # Removido db.session.commit() - será gerenciado pela transação principal
        raise
    except Exception as e:
        app.logger.error(f"Erro inesperado ao processar venda {venda_id}: {str(e)}")
        # if 'evento' in locals(): - TEMPORARIAMENTE DESABILITADO
        #     evento.status = 'erro' - TEMPORARIAMENTE DESABILITADO
        #     evento.dados_evento = f"Erro inesperado: {str(e)}" - TEMPORARIAMENTE DESABILITADO
        #     # Removido db.session.commit() - será gerenciado pela transação principal
        raise

def processar_compra_criada(compra_id, usuario_id):
    """Processa uma compra recém-criada com atualização automática de estoque (independente do status realizado)"""
    app.logger.info(f"Iniciando processamento da compra {compra_id} para usuário {usuario_id}")
    
    try:
        # Buscar a compra
        app.logger.info(f"Buscando compra {compra_id} no banco de dados")
        compra = db.session.get(Compra, compra_id)
        if not compra:
            app.logger.error(f"Compra {compra_id} não encontrada no banco de dados")
            raise ValueError(f"Compra {compra_id} não encontrada")
        
        app.logger.info(f"Compra {compra_id} encontrada: {compra.produto} x{compra.quantidade}")
        
        # Processar estoque se for mercadoria (movimento imediato - independente do status realizado)
        if compra.tipo_compra == 'mercadoria':
            app.logger.info(f"Processando estoque para mercadoria: {compra.produto}")
            
            # Buscar produto pelo nome na empresa (não apenas no usuário)
            usuario = db.session.get(Usuario, usuario_id)
            if not usuario:
                app.logger.error(f"Usuário {usuario_id} não encontrado ao processar estoque")
                raise ValueError(f"Usuário {usuario_id} não encontrado")
            
            app.logger.info(f"Usuário encontrado: {usuario.nome} (Empresa ID: {usuario.empresa_id})")
            
            # Buscar produto por nome em toda a empresa
            empresa_id = obter_empresa_id_sessao(session, usuario)

            usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
            usuarios_ids = [u.id for u in usuarios_empresa]
            app.logger.info(f"Usuários ativos da empresa {usuario.empresa_id}: {usuarios_ids}")
            
            produto = Produto.query.filter(
                Produto.nome == compra.produto,
                Produto.usuario_id.in_(usuarios_ids)
            ).first()
            
            if not produto:
                app.logger.info(f"Produto {compra.produto} não encontrado, criando automaticamente...")
                
                # ATUALIZAÇÃO: Criar produto automaticamente na empresa se não existir
                novo_produto = Produto(
                    nome=compra.produto,
                    descricao=f'Produto adicionado via compra - {compra.observacoes}' if compra.observacoes else f'Produto adicionado via compra',
                    preco_custo=compra.preco_custo,
                    preco_venda=compra.preco_custo * 1.3,  # Margem de 30% por padrão
                    estoque=0,  # Será calculado abaixo
                    ativo=False,  # Inativo até ter estoque
                    usuario_id=usuario_id  # Criar no usuário que fez a compra
                )
                db.session.add(novo_produto)
                db.session.flush()  # Para obter o ID
                produto = novo_produto
                app.logger.info(f"Produto {compra.produto} criado automaticamente no estoque da empresa {usuario.empresa_id} (ID: {produto.id})")
            else:
                app.logger.info(f"Produto {compra.produto} já existe no estoque (ID: {produto.id})")
            
            # ATUALIZAÇÃO: Sempre atualizar o estoque, independente do status "realizado"
            # O estoque será calculado baseado em todas as compras e vendas da empresa
            app.logger.info(f"Calculando estoque para produto {compra.produto}")
            estoque_real = calcular_estoque_produto(compra.produto, usuario_id)
            produto.estoque = estoque_real
            app.logger.info(f"Estoque calculado para {compra.produto}: {estoque_real}")
            
            # Atualizar preço médio também
            app.logger.info(f"Calculando preço médio para produto {compra.produto}")
            preco_medio_real = calcular_preco_medio_produto(compra.produto, usuario_id)
            produto.preco_custo = preco_medio_real
            app.logger.info(f"Preço médio calculado para {compra.produto}: R$ {preco_medio_real:.2f}")
            
            app.logger.info(f"Estoque atualizado para produto {compra.produto}: {estoque_real} (compra processada independente do status)")
        
        # Criar lançamento financeiro
        app.logger.info(f"Criando lançamento financeiro para compra {compra_id}")
        lancamento = criar_lancamento_financeiro_compra(compra, usuario_id)
        app.logger.info(f"Lançamento financeiro criado com ID: {lancamento.id}")
        
        # Fazer flush para obter o ID do lançamento
        db.session.flush()
        
        # Criar vínculo entre compra e lançamento
        app.logger.info(f"Criando vínculo entre lançamento {lancamento.id} e compra {compra_id}")
        vinculo = criar_vinculo('lancamento', lancamento.id, 'compra', compra_id, usuario_id)
        app.logger.info(f"Vínculo criado entre lançamento {lancamento.id} e compra {compra_id}")
        
        # Marcar como processada - TEMPORARIAMENTE DESABILITADO
        # evento.status = 'processado'
        # evento.dados_evento = f"Compra processada com sucesso - Estoque: {'atualizado' if compra.tipo_compra == 'mercadoria' else 'não aplicável'}, Financeiro: criado, Vínculo: criado"
        
        app.logger.info(f"Compra {compra_id} processada com sucesso - Estoque: {'atualizado' if compra.tipo_compra == 'mercadoria' else 'não aplicável'}, Financeiro: criado, Vínculo: criado")
        
        return True, "Compra processada com sucesso"
        
    except ValueError as e:
        app.logger.error(f"Erro de validação ao processar compra {compra_id}: {str(e)}")
        # if 'evento' in locals(): - TEMPORARIAMENTE DESABILITADO
        #     evento.status = 'erro' - TEMPORARIAMENTE DESABILITADO
        #     evento.dados_evento = f"Erro de validação: {str(e)}" - TEMPORARIAMENTE DESABILITADO
        #     # Removido db.session.commit() - será gerenciado pela transação principal
        raise
    except Exception as e:
        app.logger.error(f"Erro inesperado ao processar compra {compra_id}: {str(e)}")
        # if 'evento' in locals(): - TEMPORARIAMENTE DESABILITADO
        #     evento.status = 'erro' - TEMPORARIAMENTE DESABILITADO
        #     evento.dados_evento = f"Erro inesperado: {str(e)}" - TEMPORARIAMENTE DESABILITADO
        #     # Removido db.session.commit() - será gerenciado pela transação principal
        raise

def reverter_movimento_estoque_venda(venda_id, usuario_id):
    """Reverte movimento de estoque de uma venda cancelada"""
    try:
        venda = db.session.get(Venda, venda_id)
        if not venda:
            return False, "Venda não encontrada"
        
        # Aplicar movimentos reversos de estoque (usando campos existentes)
        if venda.tipo_venda == 'produto':
            # Buscar produto pelo nome
            produto = Produto.query.filter_by(
                nome=venda.produto,
                usuario_id=usuario_id
            ).first()
            
            if produto:
                aplicar_movimento_estoque(produto.id, venda.quantidade, 'incremento', usuario_id)
        
        # Cancelar lançamentos financeiros vinculados
        vinculos = Vinculo.query.filter_by(lado_b_tipo='venda', lado_b_id=venda_id).all()
        for vinculo in vinculos:
            if vinculo.lado_a_tipo == 'lancamento':
                lancamento = db.session.get(Lancamento, vinculo.lado_a_id)
                if lancamento:
                    lancamento.realizado = False  # Marcar como não realizado
        
        # Removido db.session.commit() - será gerenciado pela transação principal
        return True, "Movimento de estoque revertido com sucesso"
        
    except Exception as e:
        # Removido db.session.rollback() - será gerenciado pela transação principal
        return False, f"Erro ao reverter movimento: {str(e)}"

def reverter_movimento_estoque_compra(compra_id, usuario_id):
    """Reverte movimento de estoque de uma compra cancelada"""
    try:
        compra = db.session.get(Compra, compra_id)
        if not compra:
            return False, "Compra não encontrada"
        
        # Aplicar movimentos reversos de estoque (usando campos existentes)
        if compra.tipo_compra == 'mercadoria':
            # Buscar produto pelo nome
            produto = Produto.query.filter_by(
                nome=compra.produto,
                usuario_id=usuario_id
            ).first()
            
            if produto:
                aplicar_movimento_estoque(produto.id, compra.quantidade, 'decremento', usuario_id)
        
        # Cancelar lançamentos financeiros vinculados
        vinculos = Vinculo.query.filter_by(lado_b_tipo='compra', lado_b_id=compra_id).all()
        for vinculo in vinculos:
            if vinculo.lado_a_tipo == 'lancamento':
                lancamento = db.session.get(Lancamento, vinculo.lado_a_id)
                if lancamento:
                    lancamento.realizado = False  # Marcar como não realizado
        
        # Removido db.session.commit() - será gerenciado pela transação principal
        return True, "Movimento de estoque revertido com sucesso"
        
    except Exception as e:
        # Removido db.session.rollback() - será gerenciado pela transação principal
        return False, f"Erro ao reverter movimento: {str(e)}"

def processar_data(valor_data):
    """
    Processa e converte diferentes formatos de data para objeto date do Python.
    Lida com fórmulas do Excel, strings de data e objetos datetime.
    
    Args:
        valor_data: Valor da data (string, datetime, date, ou fórmula Excel)
    
    Returns:
        date: Objeto date do Python ou None se não conseguir processar
    """
    if not valor_data:
        return None
    
    try:
        # Se já é um objeto date, retornar diretamente
        if isinstance(valor_data, date):
            return valor_data
        
        # Se é um objeto datetime, converter para date
        if isinstance(valor_data, datetime):
            return valor_data.date()
        
        # Converter para string para processamento
        valor_str = str(valor_data).strip()
        
        # Se está vazio após strip, retornar None
        if not valor_str:
            return None
        
        # Verificar se é uma fórmula do Excel (começa com =)
        if valor_str.startswith('='):
            app.logger.warning(f"Fórmula Excel detectada: {valor_str} - não é possível processar")
            return None
        
        # Tentar diferentes formatos de data
        formatos_data = [
            '%d/%m/%Y',      # DD/MM/AAAA
            '%d-%m-%Y',      # DD-MM-AAAA
            '%Y-%m-%d',      # AAAA-MM-DD
            '%d/%m/%y',      # DD/MM/AA
            '%d-%m-%y',      # DD-MM-AA
            '%Y/%m/%d',      # AAAA/MM/DD
        ]
        
        for formato in formatos_data:
            try:
                return datetime.strptime(valor_str, formato).date()
            except ValueError:
                continue
        
        # Se nenhum formato funcionou, tentar interpretar como número serial do Excel
        try:
            # Excel usa 1 de janeiro de 1900 como dia 1 (mas há um bug no Excel que conta 1900 como bissexto)
            valor_numerico = float(valor_str)
            if valor_numerico > 0:
                # Ajuste para o bug do Excel
                if valor_numerico > 59:  # Depois de 28/02/1900
                    valor_numerico -= 1
                
                # Calcular a data
                data_base = datetime(1899, 12, 30)  # 30 de dezembro de 1899 é o dia 0 do Excel
                data_calculada = data_base + timedelta(days=valor_numerico)
                return data_calculada.date()
        except (ValueError, OverflowError):
            pass
        
        app.logger.warning(f"Não foi possível processar a data: {valor_str}")
        return None
        
    except Exception as e:
        app.logger.error(f"Erro ao processar data '{valor_data}': {str(e)}")
        return None

# ===== FUNÇÕES DE EXPORTAÇÃO =====

def exportar_relatorio_excel(dados, nome_arquivo, titulo, usuario=None, filtros=None):
    """Exporta relatório para Excel"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório"
        
        linha_atual = 1
        
        # Cabeçalho da empresa
        if usuario and usuario.empresa:
            empresa_nome = usuario.empresa.razao_social or usuario.empresa.nome_fantasia or 'Empresa'
            empresa_cnpj = usuario.empresa.cnpj or ''
            
            ws[f'A{linha_atual}'] = empresa_nome
            ws[f'A{linha_atual}'].font = Font(bold=True, size=16)
            linha_atual += 1
            
            if empresa_cnpj:
                ws[f'A{linha_atual}'] = f"CNPJ: {empresa_cnpj}"
                linha_atual += 1
        
        # Título do relatório
        ws[f'A{linha_atual}'] = titulo
        ws[f'A{linha_atual}'].font = Font(bold=True, size=14)
        linha_atual += 1
        
        # Informações do período e filtros
        if filtros:
            if filtros.get('data_inicio') and filtros.get('data_fim'):
                ws[f'A{linha_atual}'] = f"Período: {filtros['data_inicio']} a {filtros['data_fim']}"
            elif filtros.get('data_inicio'):
                ws[f'A{linha_atual}'] = f"A partir de: {filtros['data_inicio']}"
            elif filtros.get('data_fim'):
                ws[f'A{linha_atual}'] = f"Até: {filtros['data_fim']}"
            else:
                ws[f'A{linha_atual}'] = "Período: Todos os registros"
            linha_atual += 1
            
            if filtros.get('tipo'):
                ws[f'A{linha_atual}'] = f"Tipo: {filtros['tipo'].title()}"
                linha_atual += 1
            if filtros.get('categoria'):
                ws[f'A{linha_atual}'] = f"Categoria: {filtros['categoria']}"
                linha_atual += 1
            if filtros.get('status'):
                ws[f'A{linha_atual}'] = f"Status: {filtros['status'].title()}"
                linha_atual += 1
        
        # Data de geração
        ws[f'A{linha_atual}'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')} por {usuario.nome if usuario else 'Sistema'}"
        linha_atual += 2  # Espaço extra antes da tabela
        
        # Cabeçalhos
        if 'categorias_receitas' in dados and 'categorias_despesas' in dados:
            # Relatório de saldos
            ws[f'A{linha_atual}'] = "Categoria"
            ws[f'B{linha_atual}'] = "Entradas Realizadas"
            ws[f'C{linha_atual}'] = "Entradas Pendentes"
            ws[f'D{linha_atual}'] = "Total Entradas"
            ws[f'E{linha_atual}'] = "Saídas Realizadas"
            ws[f'F{linha_atual}'] = "Saídas Pendentes"
            ws[f'G{linha_atual}'] = "Total Saídas"
            ws[f'H{linha_atual}'] = "Saldo da Categoria"
            
            # Formatar cabeçalhos
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                ws[f'{col}{linha_atual}'].font = Font(bold=True)
            
            linha_atual += 1
            
            # Dados das categorias
            todas_categorias = set()
            if dados['categorias_receitas']:
                todas_categorias.update(dados['categorias_receitas'].keys())
            if dados['categorias_despesas']:
                todas_categorias.update(dados['categorias_despesas'].keys())
            
            for categoria in sorted(todas_categorias):
                entrada = dados['categorias_receitas'].get(categoria, {'realizado': 0, 'pendente': 0})
                saida = dados['categorias_despesas'].get(categoria, {'realizado': 0, 'pendente': 0})
                
                ws[f'A{linha_atual}'] = categoria
                ws[f'B{linha_atual}'] = entrada['realizado']
                ws[f'C{linha_atual}'] = entrada['pendente']
                ws[f'D{linha_atual}'] = entrada['realizado'] + entrada['pendente']
                ws[f'E{linha_atual}'] = saida['realizado']
                ws[f'F{linha_atual}'] = saida['pendente']
                ws[f'G{linha_atual}'] = saida['realizado'] + saida['pendente']
                ws[f'H{linha_atual}'] = (entrada['realizado'] + entrada['pendente']) - (saida['realizado'] + saida['pendente'])
                linha_atual += 1
            
            # Total geral
            linha_atual += 1
            ws[f'A{linha_atual}'] = "TOTAL GERAL"
            ws[f'A{linha_atual}'].font = Font(bold=True)
            ws[f'B{linha_atual}'] = dados.get('total_entradas', 0)
            ws[f'C{linha_atual}'] = dados.get('total_receitas_pendentes', 0)
            ws[f'D{linha_atual}'] = dados.get('total_entradas', 0) + dados.get('total_receitas_pendentes', 0)
            ws[f'E{linha_atual}'] = dados.get('total_saidas', 0)
            ws[f'F{linha_atual}'] = dados.get('total_despesas_pendentes', 0)
            ws[f'G{linha_atual}'] = dados.get('total_saidas', 0) + dados.get('total_despesas_pendentes', 0)
            ws[f'H{linha_atual}'] = dados.get('saldo_geral', 0)
            
            # Formatar total geral
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                ws[f'{col}{linha_atual}'].font = Font(bold=True)
                
        elif 'lancamentos' in dados:
            # Relatório de lançamentos
            ws[f'A{linha_atual}'] = "Data Vencimento"
            ws[f'B{linha_atual}'] = "Data Realizado"
            ws[f'C{linha_atual}'] = "Descrição"
            ws[f'D{linha_atual}'] = "Categoria"
            ws[f'E{linha_atual}'] = "Tipo"
            ws[f'F{linha_atual}'] = "Cliente/Fornecedor"
            ws[f'G{linha_atual}'] = "Valor"
            ws[f'H{linha_atual}'] = "NF"
            ws[f'I{linha_atual}'] = "Status"
            
            # Formatar cabeçalhos
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                ws[f'{col}{linha_atual}'].font = Font(bold=True)
            
            linha_atual += 1
            
            # Dados dos lançamentos
            for lancamento in dados['lancamentos']:
                # Determinar status baseado na lógica atual
                hoje = datetime.now().date()
                if lancamento.data_realizada and lancamento.data_realizada <= hoje:
                    status = "Realizado"
                elif lancamento.data_realizada and lancamento.data_realizada > hoje:
                    status = "Agendado"
                elif not lancamento.data_realizada and lancamento.data_prevista and lancamento.data_prevista < hoje:
                    status = "Vencido"
                else:
                    status = "Pendente"
                
                # Determinar cliente/fornecedor
                cliente_fornecedor = ""
                if lancamento.cliente:
                    cliente_fornecedor = lancamento.cliente.nome
                elif lancamento.fornecedor:
                    cliente_fornecedor = lancamento.fornecedor.nome
                elif lancamento.venda and lancamento.venda.cliente:
                    cliente_fornecedor = lancamento.venda.cliente.nome
                elif lancamento.compra and lancamento.compra.fornecedor:
                    cliente_fornecedor = lancamento.compra.fornecedor.nome
                
                ws[f'A{linha_atual}'] = lancamento.data_prevista.strftime('%d/%m/%Y') if lancamento.data_prevista else ''
                ws[f'B{linha_atual}'] = lancamento.data_realizada.strftime('%d/%m/%Y') if lancamento.data_realizada else ''
                ws[f'C{linha_atual}'] = lancamento.descricao
                ws[f'D{linha_atual}'] = lancamento.categoria or ''
                ws[f'E{linha_atual}'] = lancamento.tipo.title()
                ws[f'F{linha_atual}'] = cliente_fornecedor
                ws[f'G{linha_atual}'] = lancamento.valor
                ws[f'H{linha_atual}'] = lancamento.nota_fiscal or ''
                ws[f'I{linha_atual}'] = status
                linha_atual += 1
                
        elif 'clientes_dados' in dados:
            # Relatório de clientes
            ws[f'A{linha_atual}'] = "Cliente"
            ws[f'B{linha_atual}'] = "Email"
            ws[f'C{linha_atual}'] = "Telefone"
            ws[f'D{linha_atual}'] = "Total Entradas"
            ws[f'E{linha_atual}'] = "Total Saídas"
            ws[f'F{linha_atual}'] = "Total Geral"
            ws[f'G{linha_atual}'] = "Saldo Aberto"
            ws[f'H{linha_atual}'] = "Qtd. Lançamentos"
            
            # Formatar cabeçalhos
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                ws[f'{col}{linha_atual}'].font = Font(bold=True)
            
            linha_atual += 1
            
            # Dados dos clientes
            for cliente_dados in dados['clientes_dados']:
                cliente = cliente_dados['cliente']
                ws[f'A{linha_atual}'] = cliente.nome
                ws[f'B{linha_atual}'] = cliente.email or ''
                ws[f'C{linha_atual}'] = cliente.telefone or ''
                ws[f'D{linha_atual}'] = cliente_dados['total_entradas']
                ws[f'E{linha_atual}'] = cliente_dados['total_saidas']
                ws[f'F{linha_atual}'] = cliente_dados['total_geral']
                ws[f'G{linha_atual}'] = cliente_dados['saldo_aberto']
                ws[f'H{linha_atual}'] = cliente_dados['quantidade_lancamentos']
                linha_atual += 1
                
        elif 'produtos' in dados:
            # Relatório de produtos/estoque
            ws[f'A{linha_atual}'] = "ID"
            ws[f'B{linha_atual}'] = "Nome"
            ws[f'C{linha_atual}'] = "Descrição"
            ws[f'D{linha_atual}'] = "Preço Custo"
            ws[f'E{linha_atual}'] = "Preço Venda"
            ws[f'F{linha_atual}'] = "Estoque"
            ws[f'G{linha_atual}'] = "Valor Estoque"
            ws[f'H{linha_atual}'] = "Usuário Criador"
            
            # Formatar cabeçalhos
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                ws[f'{col}{linha_atual}'].font = Font(bold=True)
            
            linha_atual += 1
            
            # Dados dos produtos
            for produto in dados['produtos']:
                ws[f'A{linha_atual}'] = produto.id
                ws[f'B{linha_atual}'] = produto.nome
                ws[f'C{linha_atual}'] = produto.descricao or ''
                ws[f'D{linha_atual}'] = produto.preco_custo or 0
                ws[f'E{linha_atual}'] = produto.preco_venda or 0
                ws[f'F{linha_atual}'] = produto.estoque or 0
                ws[f'G{linha_atual}'] = getattr(produto, 'valor_estoque', 0)
                ws[f'H{linha_atual}'] = getattr(produto, 'usuario_criador', {}).get('nome', '') if hasattr(produto, 'usuario_criador') else ''
                linha_atual += 1
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar arquivo
        caminho_arquivo = os.path.join('uploads', f"{nome_arquivo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb.save(caminho_arquivo)
        return caminho_arquivo
        
    except Exception as e:
        app.logger.error(f"Erro ao exportar Excel: {str(e)}")
        return None

def exportar_relatorio_pdf(dados, nome_arquivo, titulo, usuario=None, filtros=None):
    """Exporta relatório para PDF"""
    try:
        caminho_arquivo = os.path.join('uploads', f"{nome_arquivo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
        doc = SimpleDocTemplate(caminho_arquivo, pagesize=A4)
        story = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=20,
            alignment=1,  # Centralizado
            textColor=colors.darkblue
        )
        
        header_style = ParagraphStyle(
            'Header',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=10,
            alignment=1,
            textColor=colors.black
        )
        
        info_style = ParagraphStyle(
            'Info',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=5,
            textColor=colors.darkgrey
        )
        
        # Cabeçalho da empresa
        if usuario and usuario.empresa:
            empresa_nome = usuario.empresa.razao_social or usuario.empresa.nome_fantasia or 'Empresa'
            empresa_cnpj = usuario.empresa.cnpj or ''
            empresa_endereco = usuario.empresa.endereco or ''
            
            story.append(Paragraph(f"<b>{empresa_nome}</b>", title_style))
            if empresa_cnpj:
                story.append(Paragraph(f"CNPJ: {empresa_cnpj}", header_style))
            if empresa_endereco.strip():
                story.append(Paragraph(empresa_endereco.strip(), header_style))
            story.append(Spacer(1, 10))
        
        # Título do relatório
        story.append(Paragraph(f"<b>{titulo}</b>", title_style))
        
        # Informações do período e filtros
        periodo_info = []
        if filtros:
            if filtros.get('data_inicio') and filtros.get('data_fim'):
                periodo_info.append(f"Período: {filtros['data_inicio']} a {filtros['data_fim']}")
            elif filtros.get('data_inicio'):
                periodo_info.append(f"A partir de: {filtros['data_inicio']}")
            elif filtros.get('data_fim'):
                periodo_info.append(f"Até: {filtros['data_fim']}")
            
            if filtros.get('tipo'):
                periodo_info.append(f"Tipo: {filtros['tipo'].title()}")
            if filtros.get('categoria'):
                periodo_info.append(f"Categoria: {filtros['categoria']}")
            if filtros.get('status'):
                periodo_info.append(f"Status: {filtros['status'].title()}")
        
        if not periodo_info:
            periodo_info.append("Período: Todos os registros")
        
        for info in periodo_info:
            story.append(Paragraph(info, info_style))
        
        story.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')} por {usuario.nome if usuario else 'Sistema'}", info_style))
        story.append(Spacer(1, 20))
        
        if 'categorias_receitas' in dados and 'categorias_despesas' in dados:
            # Relatório de saldos
            data_table = [['Categoria', 'Rec. Real.', 'Rec. Pend.', 'Total Rec.', 'Desp. Real.', 'Desp. Pend.', 'Total Desp.', 'Saldo']]
            
            todas_categorias = set()
            if dados['categorias_receitas']:
                todas_categorias.update(dados['categorias_receitas'].keys())
            if dados['categorias_despesas']:
                todas_categorias.update(dados['categorias_despesas'].keys())
            
            for categoria in sorted(todas_categorias):
                entrada = dados['categorias_receitas'].get(categoria, {'realizado': 0, 'pendente': 0})
                saida = dados['categorias_despesas'].get(categoria, {'realizado': 0, 'pendente': 0})
                
                linha = [
                    categoria,
                    f"R$ {entrada['realizado']:.2f}",
                    f"R$ {entrada['pendente']:.2f}",
                    f"R$ {entrada['realizado'] + entrada['pendente']:.2f}",
                    f"R$ {saida['realizado']:.2f}",
                    f"R$ {saida['pendente']:.2f}",
                    f"R$ {saida['realizado'] + saida['pendente']:.2f}",
                    f"R$ {(entrada['realizado'] + entrada['pendente']) - (saida['realizado'] + saida['pendente']):.2f}"
                ]
                data_table.append(linha)
            
            # Total geral
            data_table.append([
                "TOTAL GERAL",
                f"R$ {dados.get('total_entradas', 0):.2f}",
                f"R$ {dados.get('total_receitas_pendentes', 0):.2f}",
                f"R$ {dados.get('total_entradas', 0) + dados.get('total_receitas_pendentes', 0):.2f}",
                f"R$ {dados.get('total_saidas', 0):.2f}",
                f"R$ {dados.get('total_despesas_pendentes', 0):.2f}",
                f"R$ {dados.get('total_saidas', 0) + dados.get('total_despesas_pendentes', 0):.2f}",
                f"R$ {dados.get('saldo_geral', 0):.2f}"
            ])
            
        elif 'lancamentos' in dados:
            # Relatório de lançamentos
            data_table = [['Data Vencimento', 'Data Realizado', 'Descrição', 'Categoria', 'Tipo', 'Cliente/Fornecedor', 'Valor', 'Status']]
            
            for lancamento in dados['lancamentos']:
                # Determinar status baseado na lógica atual
                hoje = datetime.now().date()
                if lancamento.data_realizada and lancamento.data_realizada <= hoje:
                    status = "Realizado"
                elif lancamento.data_realizada and lancamento.data_realizada > hoje:
                    status = "Agendado"
                elif not lancamento.data_realizada and lancamento.data_prevista and lancamento.data_prevista < hoje:
                    status = "Vencido"
                else:
                    status = "Pendente"
                
                # Determinar cliente/fornecedor
                cliente_fornecedor = ""
                if lancamento.cliente:
                    cliente_fornecedor = lancamento.cliente.nome
                elif lancamento.fornecedor:
                    cliente_fornecedor = lancamento.fornecedor.nome
                elif lancamento.venda and lancamento.venda.cliente:
                    cliente_fornecedor = lancamento.venda.cliente.nome
                elif lancamento.compra and lancamento.compra.fornecedor:
                    cliente_fornecedor = lancamento.compra.fornecedor.nome
                
                linha = [
                    lancamento.data_prevista.strftime('%d/%m/%Y') if lancamento.data_prevista else '',
                    lancamento.data_realizada.strftime('%d/%m/%Y') if lancamento.data_realizada else '',
                    lancamento.descricao[:30] + '...' if len(lancamento.descricao) > 30 else lancamento.descricao,
                    lancamento.categoria or '',
                    lancamento.tipo.title(),
                    cliente_fornecedor[:25] + '...' if len(cliente_fornecedor) > 25 else cliente_fornecedor,
                    f"R$ {lancamento.valor:.2f}",
                    status
                ]
                data_table.append(linha)
                
        elif 'clientes_dados' in dados:
            # Relatório de clientes
            data_table = [['Cliente', 'Email', 'Telefone', 'Total Rec.', 'Total Desp.', 'Total Geral', 'Saldo Aberto', 'Qtd. Lanc.']]
            
            for cliente_dados in dados['clientes_dados']:
                cliente = cliente_dados['cliente']
                linha = [
                    cliente.nome[:20] + '...' if len(cliente.nome) > 20 else cliente.nome,
                    cliente.email[:20] + '...' if cliente.email and len(cliente.email) > 20 else (cliente.email or ''),
                    cliente.telefone or '',
                    f"R$ {cliente_dados['total_entradas']:.2f}",
                    f"R$ {cliente_dados['total_saidas']:.2f}",
                    f"R$ {cliente_dados['total_geral']:.2f}",
                    f"R$ {cliente_dados['saldo_aberto']:.2f}",
                    str(cliente_dados['quantidade_lancamentos'])
                ]
                data_table.append(linha)
                
        elif 'produtos' in dados:
            # Relatório de produtos/estoque
            data_table = [['ID', 'Nome', 'Descrição', 'Preço Custo', 'Preço Venda', 'Estoque', 'Valor Estoque', 'Usuário Criador']]
            
            for produto in dados['produtos']:
                linha = [
                    str(produto.id),
                    produto.nome[:25] + '...' if len(produto.nome) > 25 else produto.nome,
                    produto.descricao[:20] + '...' if produto.descricao and len(produto.descricao) > 20 else (produto.descricao or ''),
                    f"R$ {produto.preco_custo:.2f}" if produto.preco_custo else "R$ 0,00",
                    f"R$ {produto.preco_venda:.2f}" if produto.preco_venda else "R$ 0,00",
                    str(produto.estoque or 0),
                    f"R$ {getattr(produto, 'valor_estoque', 0):.2f}",
                    getattr(produto, 'usuario_criador', {}).get('nome', '') if hasattr(produto, 'usuario_criador') else ''
                ]
                data_table.append(linha)
        
        # Criar tabela
        table = Table(data_table)
        
        # Ajustar largura das colunas baseado no tipo de relatório
        if 'lancamentos' in dados:
            # Para relatório de lançamentos com 8 colunas
            col_widths = [1.2, 1.2, 2.5, 1.5, 0.8, 2.0, 1.0, 1.0]  # Larguras proporcionais
        elif 'produtos' in dados:
            # Para relatório de produtos com 8 colunas
            col_widths = [0.8, 2.0, 2.0, 1.2, 1.2, 1.0, 1.2, 1.5]  # Larguras proporcionais
        else:
            # Para outros relatórios, usar larguras automáticas
            col_widths = None
        
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white]),
        ]))
        
        # Aplicar larguras das colunas se definidas
        if col_widths:
            col_width_styles = []
            for i, width in enumerate(col_widths):
                col_width_styles.append(('COLWIDTH', (i, 0), (i, -1), width * 72))  # Converter para pontos
            
            # Aplicar todas as larguras de uma vez
            table.setStyle(TableStyle(col_width_styles))
        
        story.append(table)
        doc.build(story)
        return caminho_arquivo
        
    except Exception as e:
        app.logger.error(f"Erro ao exportar PDF: {str(e)}")
        return None

@app.route('/api/lancamentos/totais', methods=['GET'])
def api_lancamentos_totais():
    """API para retornar totais atualizados dos lançamentos baseado nos filtros"""
    if 'usuario_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario:
        return jsonify({'success': False, 'message': 'Usuário não encontrado'}), 401
    
    # Obter empresa_id correta da sessão
    empresa_id_correta = obter_empresa_id_sessao(session, usuario)
    if not empresa_id_correta:
        return jsonify({'success': False, 'message': 'Erro ao obter empresa associada'}), 400
    
    # Buscar todos os usuários da empresa correta
    usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id_correta, ativo=True).all()
    usuarios_ids = [u.id for u in usuarios_empresa]
    
    # Aplicar os mesmos filtros da rota /lancamentos
    query = Lancamento.query.options(
        db.joinedload(Lancamento.cliente),
        db.joinedload(Lancamento.fornecedor),
        db.joinedload(Lancamento.usuario)
    ).filter(Lancamento.empresa_id == empresa_id_correta)
    
    # Filtro por tipo
    tipo = request.args.get('tipo')
    if tipo:
        query = query.filter(Lancamento.tipo == tipo)
    
    # Filtro por categoria
    categoria = request.args.get('categoria')
    if categoria:
        query = query.filter(Lancamento.categoria == categoria)
    
    # Filtro por data
    def _parse_data(valor: str):
        if not valor:
            return None
        valor = valor.strip()
        try:
            if '/' in valor:
                return datetime.strptime(valor, '%d/%m/%Y').date()
            return datetime.strptime(valor, '%Y-%m-%d').date()
        except ValueError:
            return None

    data_inicio_raw = request.args.get('data_inicio', '')
    data_fim_raw = request.args.get('data_fim', '')
    data_inicio_obj = _parse_data(data_inicio_raw)
    data_fim_obj = _parse_data(data_fim_raw)

    if data_inicio_obj and data_fim_obj and data_inicio_obj > data_fim_obj:
        data_inicio_obj, data_fim_obj = data_fim_obj, data_inicio_obj

    if data_inicio_obj and data_fim_obj:
        query = query.filter(Lancamento.data_prevista.between(data_inicio_obj, data_fim_obj))
    elif data_inicio_obj:
        query = query.filter(Lancamento.data_prevista >= data_inicio_obj)
    elif data_fim_obj:
        query = query.filter(Lancamento.data_prevista <= data_fim_obj)
    
    # Filtro por status
    status = request.args.get('status')
    hoje = datetime.now().date()
    if status == 'realizado':
        query = query.filter(Lancamento.data_realizada.isnot(None), Lancamento.data_realizada <= hoje)
    elif status == 'pendente':
        query = query.filter(
            db.and_(
                Lancamento.data_realizada.is_(None),
                Lancamento.data_prevista >= hoje
            )
        )
    elif status == 'agendado':
        query = query.filter(
            db.and_(
                Lancamento.data_realizada.isnot(None),
                Lancamento.data_realizada > hoje
            )
        )
    elif status == 'vencido':
        query = query.filter(
            db.and_(
                Lancamento.data_realizada.is_(None),
                Lancamento.data_prevista < hoje
            )
        )
    
    # Filtro por busca
    busca = request.args.get('busca')
    if busca and busca.strip():
        busca_clean = busca.strip()
        query = query.join(Cliente, Lancamento.cliente_id == Cliente.id, isouter=True).join(
            Fornecedor, Lancamento.fornecedor_id == Fornecedor.id, isouter=True)
        
        try:
            valor_busca = float(busca_clean.replace(',', '.'))
            query = query.filter(
                db.or_(
                    Lancamento.descricao.ilike(f'%{busca_clean}%'),
                    Cliente.nome.ilike(f'%{busca_clean}%'),
                    Fornecedor.nome.ilike(f'%{busca_clean}%'),
                    db.func.abs(Lancamento.valor - valor_busca) < 0.01
                )
            )
        except ValueError:
            query = query.filter(
                db.or_(
                    Lancamento.descricao.ilike(f'%{busca_clean}%'),
                    Cliente.nome.ilike(f'%{busca_clean}%'),
                    Fornecedor.nome.ilike(f'%{busca_clean}%')
                )
            )
    
    # Executar query
    lancamentos = query.order_by(Lancamento.data_prevista.desc()).all()
    
    # Calcular totais
    receitas_totais = sum([l.valor for l in lancamentos if l.tipo == 'entrada'])
    despesas_totais = sum([l.valor for l in lancamentos if l.tipo == 'saida'])
    saldo_atual = receitas_totais - despesas_totais
    
    # Totais por status
    entradas_realizadas = sum([l.valor for l in lancamentos if l.tipo == 'entrada' and ((l.data_realizada and l.data_realizada <= hoje) or (l.realizado and not l.data_realizada))])
    saidas_realizadas = sum([l.valor for l in lancamentos if l.tipo == 'saida' and ((l.data_realizada and l.data_realizada <= hoje) or (l.realizado and not l.data_realizada))])
    entradas_agendadas = sum([l.valor for l in lancamentos if l.tipo == 'entrada' and l.data_realizada and l.data_realizada > hoje])
    saidas_agendadas = sum([l.valor for l in lancamentos if l.tipo == 'saida' and l.data_realizada and l.data_realizada > hoje])
    receitas_pendentes = sum([l.valor for l in lancamentos if l.tipo == 'entrada' and not l.realizado and not l.data_realizada])
    despesas_pendentes = sum([l.valor for l in lancamentos if l.tipo == 'saida' and not l.realizado and not l.data_realizada])
    
    # Saldos por status
    saldo_realizado = entradas_realizadas - saidas_realizadas
    saldo_pendente = receitas_pendentes - despesas_pendentes
    saldo_agendado = entradas_agendadas - saidas_agendadas
    
    return jsonify({
        'success': True,
        'total_lancamentos': len(lancamentos),
        'total_entradas': float(receitas_totais),
        'total_saidas': float(despesas_totais),
        'saldo_atual': float(saldo_atual),
        'saldo_realizado': float(saldo_realizado),
        'saldo_pendente': float(saldo_pendente),
        'saldo_agendado': float(saldo_agendado)
    })

@app.route('/lancamentos/<int:lancamento_id>/toggle-status', methods=['POST'])
def api_toggle_lancamento_status(lancamento_id):
    """Rota para alternar status do lançamento via AJAX"""
    # Verificar se está logado (pode ser usuario_id ou sub_usuario_id)
    if 'usuario_id' not in session and 'sub_usuario_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    # Buscar lançamento primeiro (sem filtrar por empresa ainda)
    lancamento = Lancamento.query.filter(Lancamento.id == lancamento_id).first()
    
    if not lancamento:
        return jsonify({'success': False, 'message': 'Lançamento não encontrado'}), 404
    
    # Verificar permissão de acesso
    tipo_conta = session.get('tipo_conta')
    tipo_usuario = session.get('usuario_tipo')
    
    tem_permissao = False
    
    # Caso 1: Usuário normal acessando seus próprios lançamentos
    if 'usuario_id' in session:
        usuario = db.session.get(Usuario, session['usuario_id'])
        if usuario:
            empresa_id_correta = obter_empresa_id_sessao(session, usuario)
            if lancamento.empresa_id == empresa_id_correta:
                tem_permissao = True
    
    # Caso 2: Contador/BPO acessando lançamentos de empresa vinculada
    if not tem_permissao and (tipo_conta == 'contador_bpo' or tipo_usuario == 'sub_contador'):
        if tipo_usuario == 'sub_contador':
            # Sub-usuário: verificar se tem permissão para esta empresa
            sub_usuario_id = session.get('sub_usuario_id')
            contador_id = session.get('contador_id')
            
            permissao = PermissaoSubUsuario.query.filter_by(
                sub_usuario_id=sub_usuario_id,
                empresa_id=lancamento.empresa_id
            ).first()
            
            if permissao:
                # Verificar se o vínculo está autorizado
                vinculo = VinculoContador.query.filter_by(
                    contador_id=contador_id,
                    empresa_id=lancamento.empresa_id,
                    status='autorizado'
                ).first()
                if vinculo:
                    tem_permissao = True
        else:
            # Usuário principal do contador: verificar vínculo autorizado
            contador_id = session.get('empresa_id')
            vinculo = VinculoContador.query.filter_by(
                contador_id=contador_id,
                empresa_id=lancamento.empresa_id,
                status='autorizado'
            ).first()
            if vinculo:
                tem_permissao = True
    
    if not tem_permissao:
        return jsonify({'success': False, 'message': 'Sem permissão para acessar este lançamento'}), 403
    
    try:
        # Verificar se foi passado um novo status específico
        data = request.get_json() or {}
        novo_status = data.get('novo_status')
        
        # Salvar status anterior para cálculo de saldo
        realizado_anterior = lancamento.realizado
        
        # Definir novo status
        if novo_status is not None:
            lancamento.realizado = bool(novo_status)
        else:
            lancamento.realizado = not lancamento.realizado
        
        # Atualizar data_realizada
        if lancamento.realizado and not lancamento.data_realizada:
            lancamento.data_realizada = datetime.now().date()
        elif not lancamento.realizado:
            lancamento.data_realizada = None
        
        # REMOVIDO: Atualização manual de saldo - agora é calculado dinamicamente
        # O saldo_atual é calculado automaticamente pelo método calcular_saldo_atual()
        
        # Vincular com a compra correspondente (se existir)
        if lancamento.compra:
            lancamento.compra.realizado = lancamento.realizado
            lancamento.compra.data_realizada = lancamento.data_realizada
            
            # Atualizar estoque sempre que houver compra vinculada (independente do status)
            sucesso, mensagem = atualizar_estoque_compra(lancamento.compra, usuario.id)
            if not sucesso:
                db.session.rollback()
                return jsonify({'success': False, 'message': mensagem}), 400
        
        # Vincular com a venda correspondente (se existir)
        if lancamento.venda:
            lancamento.venda.realizado = lancamento.realizado
            lancamento.venda.data_realizada = lancamento.data_realizada
            
            # Atualizar estoque sempre que houver venda vinculada (independente do status)
            sucesso, mensagem = atualizar_estoque_venda(lancamento.venda, usuario.id)
            if not sucesso:
                db.session.rollback()
                return jsonify({'success': False, 'message': mensagem}), 400

        # Sincronizar transferência pareada (se existir)
        if lancamento.eh_transferencia and lancamento.transferencia_id:
            lancamento_pareado = Lancamento.query.get(lancamento.transferencia_id)
            if lancamento_pareado:
                lancamento_pareado.realizado = lancamento.realizado
                lancamento_pareado.data_realizada = lancamento.data_realizada

        db.session.commit()
        
        status_texto = 'realizado' if lancamento.realizado else 'pendente'
        
        # Calcular status detalhado para exibição
        hoje = date.today()
        status_detalhado = 'realizado'
        if not lancamento.realizado:
            if lancamento.data_realizada and lancamento.data_realizada > hoje:
                status_detalhado = 'agendado'
            elif lancamento.data_prevista and lancamento.data_prevista < hoje:
                status_detalhado = 'vencido'
            else:
                status_detalhado = 'pendente'
        
        # Formatar data realizada para retorno
        data_realizada_formatada = None
        if lancamento.data_realizada:
            data_realizada_formatada = lancamento.data_realizada.strftime('%d/%m/%y')

        # Recalcular totais para retornar atualizado
        # Buscar todos os lançamentos da empresa para recalcular totais
        todos_lancamentos = Lancamento.query.filter(
            Lancamento.empresa_id == empresa_id_correta
        ).all()

        # Calcular totais
        receitas_totais = sum([l.valor for l in todos_lancamentos if l.tipo == 'entrada'])
        despesas_totais = sum([l.valor for l in todos_lancamentos if l.tipo == 'saida'])

        # Totais realizados
        entradas_realizadas = sum([l.valor for l in todos_lancamentos if l.tipo == 'entrada' and ((l.data_realizada and l.data_realizada <= hoje) or (l.realizado and not l.data_realizada))])
        saidas_realizadas = sum([l.valor for l in todos_lancamentos if l.tipo == 'saida' and ((l.data_realizada and l.data_realizada <= hoje) or (l.realizado and not l.data_realizada))])

        # Totais a vencer
        entradas_a_vencer = sum([l.valor for l in todos_lancamentos if l.tipo == 'entrada' and not l.realizado and not l.data_realizada and l.data_prevista >= hoje])
        saidas_a_vencer = sum([l.valor for l in todos_lancamentos if l.tipo == 'saida' and not l.realizado and not l.data_realizada and l.data_prevista >= hoje])

        # Totais vencidos
        entradas_vencidas = sum([l.valor for l in todos_lancamentos if l.tipo == 'entrada' and not l.realizado and not l.data_realizada and l.data_prevista < hoje])
        saidas_vencidas = sum([l.valor for l in todos_lancamentos if l.tipo == 'saida' and not l.realizado and not l.data_realizada and l.data_prevista < hoje])

        # Totais agendados
        entradas_agendadas = sum([l.valor for l in todos_lancamentos if l.tipo == 'entrada' and l.data_realizada and l.data_realizada > hoje])
        saidas_agendadas = sum([l.valor for l in todos_lancamentos if l.tipo == 'saida' and l.data_realizada and l.data_realizada > hoje])

        # Saldos
        saldo_atual = entradas_realizadas - saidas_realizadas
        saldo_realizado = entradas_realizadas - saidas_realizadas
        saldo_a_vencer = abs(entradas_a_vencer - saidas_a_vencer)  # Valor absoluto
        saldo_vencido = abs(entradas_vencidas - saidas_vencidas)  # Valor absoluto
        saldo_agendado = abs(entradas_agendadas - saidas_agendadas)  # Valor absoluto

        return jsonify({
            'success': True,
            'message': f'Lançamento marcado como {status_texto} com sucesso!',
            'novo_status': status_texto,
            'status_detalhado': status_detalhado,
            'data_realizada': data_realizada_formatada,
            'realizado': lancamento.realizado,
            'totais': {
                # Totais gerais
                'total_entradas': float(receitas_totais),
                'total_saidas': float(despesas_totais),
                'saldo_atual': float(saldo_atual),
                'saldo_realizado': float(saldo_realizado),
                'saldo_a_vencer': float(saldo_a_vencer),
                'saldo_vencido': float(saldo_vencido),
                'saldo_agendado': float(saldo_agendado),
                # Entradas separadas
                'entrada_total': float(receitas_totais),
                'entrada_realizada': float(entradas_realizadas),
                'entrada_a_vencer': float(entradas_a_vencer),
                'entrada_vencida': float(entradas_vencidas),
                'entrada_agendada': float(entradas_agendadas),
                # Saídas separadas
                'saida_total': float(despesas_totais),
                'saida_realizada': float(saidas_realizadas),
                'saida_a_vencer': float(saidas_a_vencer),
                'saida_vencida': float(saidas_vencidas),
                'saida_agendada': float(saidas_agendadas)
            }
        })
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Erro ao alternar status do lançamento: {str(e)}')
        return jsonify({'success': False, 'message': f'Erro ao atualizar status: {str(e)}'}), 500

@app.route('/api/lancamentos/marcar-realizado-lote', methods=['POST'])
def api_marcar_lancamentos_realizado_lote():
    """Marca ou desmarca múltiplos lançamentos como realizados"""
    if 'usuario_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401

    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario:
        return jsonify({'success': False, 'message': 'Usuário não encontrado'}), 401

    # Obter empresa_id correta da sessão (considera acesso_contador)
    empresa_id_correta = obter_empresa_id_sessao(session, usuario)
    if not empresa_id_correta:
        return jsonify({'success': False, 'message': 'Erro ao obter empresa associada'}), 400

    try:
        data = request.get_json() or {}
        lancamento_ids = data.get('lancamento_ids', [])
        realizado = data.get('realizado', True)

        if not lancamento_ids:
            return jsonify({'success': False, 'message': 'Nenhum lançamento selecionado'}), 400

        # Buscar lançamentos pela empresa_id correta
        lancamentos = Lancamento.query.filter(
            Lancamento.id.in_(lancamento_ids),
            Lancamento.empresa_id == empresa_id_correta
        ).all()

        if not lancamentos:
            return jsonify({'success': False, 'message': 'Nenhum lançamento válido encontrado'}), 404

        # Atualizar cada lançamento
        atualizados = 0
        for lancamento in lancamentos:
            # Salvar status anterior para cálculo de saldo
            realizado_anterior = lancamento.realizado

            # Definir novo status
            lancamento.realizado = bool(realizado)

            # Atualizar data_realizada
            if lancamento.realizado and not lancamento.data_realizada:
                lancamento.data_realizada = datetime.now().date()
            elif not lancamento.realizado:
                lancamento.data_realizada = None

            # Atualizar saldo da conta caixa se houver
            if lancamento.conta_caixa_id:
                conta_caixa = db.session.get(ContaCaixa, lancamento.conta_caixa_id)
                if conta_caixa:
                    if lancamento.realizado and not realizado_anterior:
                        # Lançamento foi marcado como realizado
                        if lancamento.tipo == 'entrada':
                            conta_caixa.saldo_atual += lancamento.valor
                        else:  # saida
                            conta_caixa.saldo_atual -= lancamento.valor
                    elif not lancamento.realizado and realizado_anterior:
                        # Lançamento foi desmarcado como realizado
                        if lancamento.tipo == 'entrada':
                            conta_caixa.saldo_atual -= lancamento.valor
                        else:  # saida
                            conta_caixa.saldo_atual += lancamento.valor

            # Vincular com a compra correspondente (se existir)
            if lancamento.compra:
                lancamento.compra.realizado = lancamento.realizado
                lancamento.compra.data_realizada = lancamento.data_realizada
                atualizar_estoque_compra(lancamento.compra, usuario.id)

            # Vincular com a venda correspondente (se existir)
            if lancamento.venda:
                lancamento.venda.realizado = lancamento.realizado
                lancamento.venda.data_realizada = lancamento.data_realizada
                atualizar_estoque_venda(lancamento.venda, usuario.id)

            atualizados += 1

        db.session.commit()

        acao = 'marcados' if realizado else 'desmarcados'
        return jsonify({
            'success': True,
            'message': f'{atualizados} lançamento(s) {acao} como realizado com sucesso!'
        })

    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Erro ao marcar lançamentos em lote: {str(e)}')
        return jsonify({'success': False, 'message': f'Erro ao atualizar: {str(e)}'}), 500

@app.route('/backup/geral')
def exportar_backup_geral():
    """Exporta backup geral com todos os relatórios em abas separadas"""
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    usuario = db.session.get(Usuario, session['usuario_id'])
    if not usuario:
        flash('Usuário não encontrado', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        # Criar workbook
        wb = Workbook()
        
        # Remover planilha padrão
        wb.remove(wb.active)
        
        # Obter todos os usuários da mesma empresa
        empresa_id = obter_empresa_id_sessao(session, usuario)

        usuarios_empresa = Usuario.query.filter_by(empresa_id=empresa_id, ativo=True).all()
        usuarios_ids = [u.id for u in usuarios_empresa]
        
        # 1. ABA: LANÇAMENTOS
        ws_lancamentos = wb.create_sheet("Lançamentos")
        ws_lancamentos.append(["Data Vencimento", "Data Realizado", "Descrição", "Categoria", "Tipo", "Cliente/Fornecedor", "Valor", "Status"])
        
        # Buscar todos os lançamentos
        lancamentos = Lancamento.query.options(
            joinedload(Lancamento.cliente), 
            joinedload(Lancamento.fornecedor),
            joinedload(Lancamento.venda),
            joinedload(Lancamento.compra)
        ).filter(Lancamento.empresa_id == empresa_id).order_by(Lancamento.data_prevista).all()
        
        for lancamento in lancamentos:
            # Determinar status
            hoje = datetime.now().date()
            if lancamento.data_realizada and lancamento.data_realizada <= hoje:
                status = "Realizado"
            elif lancamento.data_realizada and lancamento.data_realizada > hoje:
                status = "Agendado"
            elif not lancamento.data_realizada and lancamento.data_prevista and lancamento.data_prevista < hoje:
                status = "Vencido"
            else:
                status = "Pendente"
            
            # Determinar cliente/fornecedor
            cliente_fornecedor = ""
            if lancamento.cliente:
                cliente_fornecedor = lancamento.cliente.nome
            elif lancamento.fornecedor:
                cliente_fornecedor = lancamento.fornecedor.nome
            elif lancamento.venda and lancamento.venda.cliente:
                cliente_fornecedor = lancamento.venda.cliente.nome
            elif lancamento.compra and lancamento.compra.fornecedor:
                cliente_fornecedor = lancamento.compra.fornecedor.nome
            
            ws_lancamentos.append([
                lancamento.data_prevista.strftime('%d/%m/%Y') if lancamento.data_prevista else '',
                lancamento.data_realizada.strftime('%d/%m/%Y') if lancamento.data_realizada else '',
                lancamento.descricao,
                lancamento.categoria or '',
                lancamento.tipo.title(),
                cliente_fornecedor,
                lancamento.valor,
                status
            ])
        
        # 2. ABA: CLIENTES
        ws_clientes = wb.create_sheet("Clientes")
        ws_clientes.append(["Nome", "Email", "Telefone", "CPF/CNPJ", "Endereço", "Data Criação"])
        
        clientes = Cliente.query.filter(Cliente.empresa_id == empresa_id, Cliente.ativo == True).all()
        for cliente in clientes:
            ws_clientes.append([
                cliente.nome,
                cliente.email or '',
                cliente.telefone or '',
                cliente.cpf_cnpj or '',
                cliente.endereco or '',
                cliente.data_criacao.strftime('%d/%m/%Y') if cliente.data_criacao else ''
            ])
        
        # 3. ABA: FORNECEDORES
        ws_fornecedores = wb.create_sheet("Fornecedores")
        ws_fornecedores.append(["Nome", "Email", "Telefone", "CPF/CNPJ", "Endereço", "Data Criação"])
        
        fornecedores = Fornecedor.query.filter(Fornecedor.empresa_id == empresa_id, Fornecedor.ativo == True).all()
        for fornecedor in fornecedores:
            ws_fornecedores.append([
                fornecedor.nome,
                fornecedor.email or '',
                fornecedor.telefone or '',
                fornecedor.cpf_cnpj or '',
                fornecedor.endereco or '',
                fornecedor.data_criacao.strftime('%d/%m/%Y') if fornecedor.data_criacao else ''
            ])
        
        # 4. ABA: PRODUTOS
        ws_produtos = wb.create_sheet("Produtos")
        ws_produtos.append(["ID", "Nome", "Descrição", "Preço Custo", "Preço Venda", "Estoque", "Valor Estoque", "Usuário Criador"])
        
        produtos = buscar_produtos_empresa(usuario.empresa_id)
        for produto in produtos:
            ws_produtos.append([
                produto.id,
                produto.nome,
                produto.descricao or '',
                produto.preco_custo or 0,
                produto.preco_venda or 0,
                produto.estoque or 0,
                getattr(produto, 'valor_estoque', 0),
                getattr(produto, 'usuario_criador', {}).get('nome', '') if hasattr(produto, 'usuario_criador') else ''
            ])
        
        # 5. ABA: VENDAS
        ws_vendas = wb.create_sheet("Vendas")
        ws_vendas.append(["ID", "Cliente", "Produto", "Quantidade", "Valor", "Data", "Observações", "Status"])
        
        vendas = Venda.query.options(joinedload(Venda.cliente)).filter(Venda.empresa_id == empresa_id).all()
        for venda in vendas:
            ws_vendas.append([
                venda.id,
                venda.cliente.nome if venda.cliente else '',
                venda.produto,
                venda.quantidade,
                venda.valor,
                venda.data.strftime('%d/%m/%Y') if venda.data else '',
                venda.observacoes or '',
                "Realizado" if venda.realizado else "Pendente"
            ])
        
        # 6. ABA: COMPRAS
        ws_compras = wb.create_sheet("Compras")
        ws_compras.append(["ID", "Fornecedor", "Produto", "Quantidade", "Valor", "Data", "Observações", "Status"])
        
        compras = Compra.query.options(joinedload(Compra.fornecedor)).filter(Compra.empresa_id == empresa_id).all()
        for compra in compras:
            ws_compras.append([
                compra.id,
                compra.fornecedor.nome if compra.fornecedor else '',
                compra.produto,
                compra.quantidade,
                compra.valor,
                compra.data.strftime('%d/%m/%Y') if compra.data else '',
                compra.observacoes or '',
                "Realizado" if compra.realizado else "Pendente"
            ])
        
        # Formatar cabeçalhos de todas as abas
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
        
        # Salvar arquivo
        caminho_arquivo = os.path.join('uploads', f"backup_geral_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb.save(caminho_arquivo)
        
        return send_file(caminho_arquivo, as_attachment=True, download_name=f"backup_geral_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        
    except Exception as e:
        app.logger.error(f"Erro ao gerar backup geral: {str(e)}")
        flash('Erro ao gerar backup geral.', 'error')
        return redirect(url_for('dashboard'))

# ===== ROTAS DE API PARA SEGURANÇA =====

@app.route('/api/session-data', methods=['GET'])
def api_session_data():
    """API para fornecer dados da sessão de forma segura"""
    try:
        # Verificar se está logado (pode ser usuario_id ou sub_usuario_id)
        if 'usuario_id' not in session and 'sub_usuario_id' not in session:
            return jsonify({'success': False, 'message': 'Não autenticado'}), 401
        
        # Verificar se é sub-usuário
        if 'sub_usuario_id' in session:
            sub_usuario = db.session.get(SubUsuarioContador, session['sub_usuario_id'])
            if not sub_usuario:
                return jsonify({'success': False, 'message': 'Sub-usuário não encontrado'}), 404
            
            empresa = db.session.get(Empresa, sub_usuario.contador_id) if sub_usuario.contador_id else None
            if not empresa:
                return jsonify({'success': False, 'message': 'Empresa contador não encontrada'}), 404
            
            data = {
                'usuario_nome': sub_usuario.nome,
                'usuario_tipo': 'sub_contador',
                'empresa_nome': empresa.razao_social,
                'empresa_tipo_pessoa': empresa.tipo_pessoa if hasattr(empresa, 'tipo_pessoa') else None,
                'empresa_tipo': empresa.tipo_empresa if hasattr(empresa, 'tipo_empresa') else None,
                'empresa_documento': empresa.cpf if (hasattr(empresa, 'tipo_pessoa') and empresa.tipo_pessoa == 'PF') else (empresa.cnpj if hasattr(empresa, 'cnpj') else None)
            }
        else:
            usuario = db.session.get(Usuario, session['usuario_id'])
            if not usuario:
                return jsonify({'success': False, 'message': 'Usuário não encontrado'}), 404
            
            empresa = db.session.get(Empresa, usuario.empresa_id) if usuario.empresa_id else None
            if not empresa:
                return jsonify({'success': False, 'message': 'Empresa não encontrada'}), 404
            
            # Retornar apenas dados necessários para o frontend
            data = {
                'usuario_nome': usuario.nome,
                'usuario_tipo': usuario.tipo,
                'empresa_nome': empresa.razao_social,
                'empresa_tipo_pessoa': empresa.tipo_pessoa,
                'empresa_tipo': empresa.tipo_empresa,
                'empresa_documento': empresa.cpf if empresa.tipo_pessoa == 'PF' else empresa.cnpj
            }
        
        return jsonify({'success': True, 'data': data})
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro na API session-data: {str(e)}")
        return jsonify({'success': False, 'message': 'Erro interno do servidor'}), 500


# Rota catch-all para servir o React App (SPA)
# Esta deve ser a ÚLTIMA rota definida
# DESABILITADA TEMPORARIAMENTE - usando templates HTML
# @app.route('/<path:path>')
# def catch_all(path):
#     """
#     Rota catch-all para servir o aplicativo React.






# ===== FUNÇÕES AUXILIARES =====

def atualizar_dias_assinatura(empresa):
    """Atualiza os dias restantes de assinatura baseado na data de início"""
    if not empresa or empresa.tipo_conta == 'admin':
        return

    # Se não tem data_inicio_assinatura, definir como hoje e não fazer nada mais
    if not empresa.data_inicio_assinatura:
        empresa.data_inicio_assinatura = datetime.utcnow()
        db.session.commit()
        return

    # Calcular dias decorridos desde a última atualização (data_inicio_assinatura)
    hoje = datetime.utcnow()
    dias_decorridos = (hoje.date() - empresa.data_inicio_assinatura.date()).days

    # Se passou pelo menos 1 dia, atualizar
    if dias_decorridos > 0:
        # Calcular dias restantes
        dias_restantes = empresa.dias_assinatura - dias_decorridos

        # Se os dias restantes são negativos ou zero, definir como 0
        if dias_restantes <= 0:
            empresa.dias_assinatura = 0
            # Suspender conta automaticamente quando os dias acabarem
            if empresa.ativo:
                app.logger.info(f"⚠️ Conta {empresa.razao_social} suspensa automaticamente - assinatura expirada")
        else:
            empresa.dias_assinatura = dias_restantes

        # Atualizar a data de início para hoje (reset do contador diário)
        empresa.data_inicio_assinatura = hoje
        db.session.commit()

def atualizar_todas_assinaturas():
    """Job agendado para atualizar dias de assinatura de todas as empresas"""
    with app.app_context():
        try:
            app.logger.info("🔄 Iniciando atualização automática de assinaturas...")
            empresas = Empresa.query.filter(Empresa.tipo_conta != 'admin').all()

            contas_atualizadas = 0
            contas_suspensas = 0

            for empresa in empresas:
                dias_antes = empresa.dias_assinatura
                atualizar_dias_assinatura(empresa)

                if dias_antes != empresa.dias_assinatura:
                    contas_atualizadas += 1

                if empresa.dias_assinatura == 0 and dias_antes > 0:
                    contas_suspensas += 1

            app.logger.info(f"✅ Atualização concluída: {contas_atualizadas} contas atualizadas, {contas_suspensas} contas suspensas")

        except Exception as e:
            app.logger.error(f"❌ Erro ao atualizar assinaturas: {str(e)}")
            db.session.rollback()

# ===== ROTAS DO PAINEL ADMIN =====

@app.route('/admin/painel-completo')
def admin_painel_completo():
    """Redireciona para admin_usuarios, pois todas as funcionalidades foram movidas para lá"""
    if 'usuario_id' not in session or session.get('usuario_tipo') != 'admin':
        flash('Acesso negado. Apenas administradores.', 'error')
        return redirect(url_for('login'))

    # Redirecionar para admin_usuarios, pois todas as funcionalidades foram movidas para lá
    return redirect(url_for('admin_usuarios'))

@app.route('/admin/vouchers-page')
def admin_vouchers_page():
    """Página de gerenciamento de vouchers"""
    if 'usuario_id' not in session or session.get('usuario_tipo') != 'admin':
        flash('Acesso negado. Apenas administradores.', 'error')
        return redirect(url_for('login'))

    return render_template('admin_vouchers.html')

@app.route('/admin/editar-dias', methods=['POST'])
def admin_editar_dias():
    """Editar dias de assinatura de uma conta"""
    if 'usuario_id' not in session or session.get('usuario_tipo') != 'admin':
        return jsonify({'success': False, 'message': 'Acesso negado'}), 403
    
    try:
        conta_id = request.form.get('conta_id')
        dias_assinatura = request.form.get('dias_assinatura', type=int)
        
        if not conta_id or dias_assinatura is None:
            flash('Dados inválidos.', 'error')
            return redirect(url_for('admin_usuarios'))
        
        conta = Empresa.query.get(conta_id)
        if not conta:
            flash('Conta não encontrada.', 'error')
            return redirect(url_for('admin_usuarios'))
        
        # Não permitir editar dias da conta admin
        if conta.tipo_conta == 'admin':
            flash('Não é possível editar os dias de assinatura da conta administrativa.', 'error')
            return redirect(url_for('admin_usuarios'))
        
        # Atualizar dias de assinatura e resetar a data de início
        conta.dias_assinatura = dias_assinatura
        conta.data_inicio_assinatura = datetime.utcnow()  # Resetar contador

        # Se adicionou dias (maior que 0), reativar a conta automaticamente
        if dias_assinatura > 0 and not conta.ativo:
            conta.ativo = True
            flash(f'Dias de assinatura atualizados para {dias_assinatura} dias. Conta reativada com sucesso!', 'success')
            app.logger.info(f"✅ Conta {conta.razao_social} reativada pelo admin - {dias_assinatura} dias de assinatura")
        else:
            flash(f'Dias de assinatura atualizados para {dias_assinatura} dias.', 'success')

        db.session.commit()
        # Redirecionar para admin_usuarios
        referer = request.headers.get('Referer', '')
        if 'admin/usuarios' in referer:
            return redirect(url_for('admin_usuarios'))
        return redirect(url_for('admin_usuarios'))
    
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao atualizar dias de assinatura: {str(e)}', 'error')
        return redirect(url_for('admin_usuarios'))

@app.route('/admin/toggle-status/<int:conta_id>', methods=['POST'])
def admin_toggle_status(conta_id):
    """Ativar/Desativar conta"""
    if 'usuario_id' not in session or session.get('usuario_tipo') != 'admin':
        return jsonify({'success': False, 'message': 'Acesso negado'}), 403
    
    try:
        data = request.get_json()
        ativo = data.get('ativo', False)
        
        conta = Empresa.query.get(conta_id)
        if not conta:
            return jsonify({'success': False, 'message': 'Conta não encontrada'})
        
        # Não permitir desativar a conta admin
        if conta.tipo_conta == 'admin':
            return jsonify({'success': False, 'message': 'Não é possível desativar a conta administrativa'})
        
        conta.ativo = ativo
        db.session.commit()
        
        return jsonify({'success': True, 'message': f'Conta {"ativada" if ativo else "desativada"} com sucesso'})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/admin/excluir-conta/<int:conta_id>', methods=['POST'])
def admin_excluir_conta(conta_id):
    """Excluir conta"""
    if 'usuario_id' not in session or session.get('usuario_tipo') != 'admin':
        return jsonify({'success': False, 'message': 'Acesso negado'}), 403
    
    try:
        conta = Empresa.query.get(conta_id)
        if not conta:
            return jsonify({'success': False, 'message': 'Conta não encontrada'})
        
        # Não permitir excluir a conta admin
        if conta.tipo_conta == 'admin':
            return jsonify({'success': False, 'message': 'Não é possível excluir a conta administrativa'})
        
        # Pegar todos os usuários da conta para limpeza de dependências
        usuarios = Usuario.query.filter_by(empresa_id=conta_id).all()
        usuarios_ids = [u.id for u in usuarios]
        
        if usuarios_ids:
            # 1. Limpar dependências financeiras (Parcelas dependem de Lancamento/Venda/Compra)
            Parcela.query.filter(Parcela.usuario_id.in_(usuarios_ids)).delete(synchronize_session=False)
            
            # 2. Limpar Lançamentos, Vendas e Compras
            Lancamento.query.filter_by(empresa_id=conta_id).delete(synchronize_session=False)
            Venda.query.filter(Venda.empresa_id == conta_id).delete(synchronize_session=False)
            Compra.query.filter(Compra.empresa_id == conta_id).delete(synchronize_session=False)

            # 3. Limpar Cadastros (Cliente, Fornecedor, Produto, Servico)
            Cliente.query.filter(Cliente.empresa_id == conta_id).delete(synchronize_session=False)
            Fornecedor.query.filter(Fornecedor.empresa_id == conta_id).delete(synchronize_session=False)
            Produto.query.filter(Produto.usuario_id.in_(usuarios_ids)).delete(synchronize_session=False)
            Servico.query.filter(Servico.usuario_id.in_(usuarios_ids)).delete(synchronize_session=False)

            # 4. Limpar Infraestrutura e Logs
            PlanoConta.query.filter(PlanoConta.empresa_id == conta_id).delete(synchronize_session=False)
            ContaCaixa.query.filter(ContaCaixa.usuario_id.in_(usuarios_ids)).delete(synchronize_session=False)
            Importacao.query.filter(Importacao.usuario_id.in_(usuarios_ids)).delete(synchronize_session=False)
            Permissao.query.filter(Permissao.usuario_id.in_(usuarios_ids)).delete(synchronize_session=False)
            EventLog.query.filter(EventLog.usuario_id.in_(usuarios_ids)).delete(synchronize_session=False)
            Vinculo.query.filter(Vinculo.usuario_id.in_(usuarios_ids)).delete(synchronize_session=False)

        # 5. Limpar Categorias e Vínculos da Empresa
        CategoriaUsuario.query.filter_by(empresa_id=conta_id).delete(synchronize_session=False)
        VinculoContador.query.filter(
            or_(
                VinculoContador.contador_id == conta_id,
                VinculoContador.empresa_id == conta_id
            )
        ).delete(synchronize_session=False)
        
        # 6. Limpar Registros de Pagamento e Vouchers
        Pagamento.query.filter_by(empresa_id=conta_id).delete(synchronize_session=False)
        VoucherUso.query.filter_by(empresa_id=conta_id).delete(synchronize_session=False)
        
        # 7. Excluir sub-usuários se for contador
        sub_usuarios = SubUsuarioContador.query.filter_by(contador_id=conta_id).all()
        for sub in sub_usuarios:
            PermissaoSubUsuario.query.filter_by(sub_usuario_id=sub.id).delete(synchronize_session=False)
            db.session.delete(sub)
        
        # 8. Finalmente excluir usuários e a conta
        Usuario.query.filter_by(empresa_id=conta_id).delete(synchronize_session=False)
        db.session.delete(conta)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Conta excluída com sucesso'})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

# ============================================================================
# ROTAS DE VOUCHERS (ADMIN)
# ============================================================================

@app.route('/admin/vouchers', methods=['GET'])
def listar_vouchers():
    """Lista todos os vouchers em formato JSON para o modal"""
    if 'usuario_id' not in session:
        return jsonify({'erro': 'Não autenticado'}), 401

    usuario = Usuario.query.get(session['usuario_id'])
    if not usuario or session.get('usuario_tipo') != 'admin':
        return jsonify({'erro': 'Acesso negado'}), 403

    try:
        vouchers = Voucher.query.order_by(Voucher.data_criacao.desc()).all()

        dados = []
        for v in vouchers:
            dados.append({
                'id': v.id,
                'codigo': v.codigo,
                'dias_assinatura': v.dias_assinatura,
                'validade': v.validade.strftime('%d/%m/%Y %H:%M') if v.validade else '',
                'ativo': v.ativo,
                'pode_usar': v.pode_usar(),
                'usos': len(v.usos),
                'data_criacao': v.data_criacao.strftime('%d/%m/%Y %H:%M') if v.data_criacao else ''
            })

        return jsonify({'sucesso': True, 'vouchers': dados}), 200
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao listar vouchers: {str(e)}")
        return jsonify({'erro': str(e)}), 500

@app.route('/admin/vouchers', methods=['POST'])
def criar_voucher():
    """Cria um novo voucher"""
    if 'usuario_id' not in session:
        return jsonify({'erro': 'Não autenticado'}), 401

    usuario = Usuario.query.get(session['usuario_id'])
    if not usuario or session.get('usuario_tipo') != 'admin':
        return jsonify({'erro': 'Acesso negado'}), 403

    try:
        dados = request.get_json()

        # Validações
        if not dados.get('codigo'):
            return jsonify({'erro': 'Código é obrigatório'}), 400

        if not dados.get('dias_assinatura') or int(dados.get('dias_assinatura', 0)) <= 0:
            return jsonify({'erro': 'Dias de assinatura deve ser maior que 0'}), 400

        if not dados.get('validade'):
            return jsonify({'erro': 'Validade é obrigatória'}), 400

        # Verificar se código já existe
        if Voucher.query.filter_by(codigo=dados['codigo'].upper().strip()).first():
            return jsonify({'erro': 'Código de voucher já existe'}), 400

        # Converter data
        try:
            validade = datetime.strptime(dados['validade'], '%Y-%m-%d')
        except ValueError:
            return jsonify({'erro': 'Formato de data inválido (use YYYY-MM-DD)'}), 400

        # Criar voucher
        novo_voucher = Voucher(
            codigo=dados['codigo'].upper().strip(),
            dias_assinatura=int(dados['dias_assinatura']),
            validade=validade,
            ativo=True,
            criado_por=usuario.id
        )

        db.session.add(novo_voucher)
        db.session.commit()

        return jsonify({
            'sucesso': True,
            'mensagem': f'Voucher {novo_voucher.codigo} criado com sucesso',
            'voucher': {
                'id': novo_voucher.id,
                'codigo': novo_voucher.codigo,
                'dias_assinatura': novo_voucher.dias_assinatura,
                'validade': novo_voucher.validade.strftime('%d/%m/%Y %H:%M'),
                'ativo': novo_voucher.ativo
            }
        }), 201

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao criar voucher: {str(e)}")
        return jsonify({'erro': str(e)}), 500

@app.route('/admin/vouchers/<int:voucher_id>/toggle', methods=['PATCH'])
def toggle_voucher(voucher_id):
    """Ativa ou desativa um voucher"""
    if 'usuario_id' not in session:
        return jsonify({'erro': 'Não autenticado'}), 401

    usuario = Usuario.query.get(session['usuario_id'])
    if not usuario or session.get('usuario_tipo') != 'admin':
        return jsonify({'erro': 'Acesso negado'}), 403

    try:
        voucher = Voucher.query.get(voucher_id)
        if not voucher:
            return jsonify({'erro': 'Voucher não encontrado'}), 404

        voucher.ativo = not voucher.ativo
        db.session.commit()

        status = 'ativado' if voucher.ativo else 'desativado'
        return jsonify({
            'sucesso': True,
            'mensagem': f'Voucher {status} com sucesso',
            'ativo': voucher.ativo
        }), 200

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao toggle voucher: {str(e)}")
        return jsonify({'erro': str(e)}), 500

@app.route('/admin/vouchers/aplicar', methods=['POST'])
def aplicar_voucher():
    """Aplica um voucher a uma empresa"""
    if 'usuario_id' not in session:
        return jsonify({'erro': 'Não autenticado'}), 401

    usuario = Usuario.query.get(session['usuario_id'])
    if not usuario or session.get('usuario_tipo') != 'admin':
        return jsonify({'erro': 'Acesso negado'}), 403

    try:
        dados = request.get_json()

        if not dados.get('codigo'):
            return jsonify({'erro': 'Código do voucher é obrigatório'}), 400

        if not dados.get('empresa_id'):
            return jsonify({'erro': 'ID da empresa é obrigatório'}), 400

        # Buscar voucher
        voucher = Voucher.query.filter_by(codigo=dados['codigo'].upper().strip()).first()
        if not voucher:
            return jsonify({'erro': 'Voucher não encontrado'}), 404

        # Validar voucher
        if not voucher.ativo:
            return jsonify({'erro': 'Voucher está inativo'}), 400

        if datetime.utcnow() > voucher.validade:
            return jsonify({'erro': 'Voucher expirado'}), 400

        if not voucher.pode_usar():
            return jsonify({'erro': 'Voucher já foi utilizado'}), 400

        # Buscar empresa
        empresa = Empresa.query.get(dados['empresa_id'])
        if not empresa:
            return jsonify({'erro': 'Empresa não encontrada'}), 404

        # Não permitir aplicar em admin
        if empresa.tipo_conta == 'admin':
            return jsonify({'erro': 'Não é permitido aplicar voucher em conta admin'}), 400

        # Verificar se já foi usado para esta empresa
        uso_existente = VoucherUso.query.filter_by(
            voucher_id=voucher.id,
            empresa_id=empresa.id
        ).first()
        if uso_existente:
            return jsonify({'erro': 'Este voucher já foi aplicado a esta empresa'}), 400

        # Aplicar voucher
        dias_anteriores = empresa.dias_assinatura or 0

        if dias_anteriores == 0:
            # Se expirado, setar data de início
            empresa.dias_assinatura = voucher.dias_assinatura
            empresa.data_inicio_assinatura = datetime.utcnow()
        else:
            # Se ativo, somar dias
            empresa.dias_assinatura = dias_anteriores + voucher.dias_assinatura

        # Registrar uso
        uso = VoucherUso(
            voucher_id=voucher.id,
            empresa_id=empresa.id,
            dias_creditados=voucher.dias_assinatura,
            usuario_admin_id=usuario.id,
            observacoes=f'Voucher {voucher.codigo} aplicado automaticamente'
        )

        db.session.add(uso)
        db.session.commit()

        return jsonify({
            'sucesso': True,
            'mensagem': f'Voucher aplicado com sucesso! {voucher.dias_assinatura} dias creditados.',
            'empresa': {
                'id': empresa.id,
                'razao_social': empresa.razao_social,
                'dias_anteriores': dias_anteriores,
                'dias_creditados': voucher.dias_assinatura,
                'dias_atuais': empresa.dias_assinatura
            }
        }), 200

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao aplicar voucher: {str(e)}")
        return jsonify({'erro': str(e)}), 500

@app.route('/admin/vouchers/<int:voucher_id>', methods=['DELETE'])
def deletar_voucher(voucher_id):
    """Deleta um voucher (apenas se não foi usado)"""
    if 'usuario_id' not in session:
        return jsonify({'erro': 'Não autenticado'}), 401

    usuario = Usuario.query.get(session['usuario_id'])
    if not usuario or session.get('usuario_tipo') != 'admin':
        return jsonify({'erro': 'Acesso negado'}), 403

    try:
        voucher = Voucher.query.get(voucher_id)
        if not voucher:
            return jsonify({'erro': 'Voucher não encontrado'}), 404

        # Não permitir deletar se foi usado
        if len(voucher.usos) > 0:
            return jsonify({'erro': 'Não é possível deletar um voucher que já foi utilizado'}), 400

        db.session.delete(voucher)
        db.session.commit()

        return jsonify({
            'sucesso': True,
            'mensagem': 'Voucher deletado com sucesso'
        }), 200

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao deletar voucher: {str(e)}")
        return jsonify({'erro': str(e)}), 500

@app.route('/admin/vouchers/usos', methods=['GET'])
def listar_usos_vouchers():
    """Lista histórico de uso de vouchers"""
    if 'usuario_id' not in session:
        return jsonify({'erro': 'Não autenticado'}), 401

    usuario = Usuario.query.get(session['usuario_id'])
    if not usuario or session.get('usuario_tipo') != 'admin':
        return jsonify({'erro': 'Acesso negado'}), 403

    try:
        usos = VoucherUso.query.order_by(VoucherUso.data_uso.desc()).all()

        dados = []
        for uso in usos:
            dados.append({
                'id': uso.id,
                'codigo_voucher': uso.voucher.codigo,
                'empresa': uso.empresa.razao_social,
                'dias_creditados': uso.dias_creditados,
                'data_uso': uso.data_uso.strftime('%d/%m/%Y %H:%M'),
                'admin': uso.usuario_admin.nome if uso.usuario_admin else 'Sistema'
            })

        return jsonify({'sucesso': True, 'usos': dados}), 200
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao listar usos de vouchers: {str(e)}")
        return jsonify({'erro': str(e)}), 500

# ============================================================================
# ROTAS DE PAGAMENTO - MERCADO PAGO
# ============================================================================

@app.route('/checkout/<plano_codigo>')
def checkout(plano_codigo):
    """Página de checkout para escolher plano"""
    plano = Plano.query.filter_by(codigo=plano_codigo, ativo=True).first()
    if not plano:
        flash('Plano não encontrado ou inativo.', 'error')
        return redirect(url_for('precos'))

    return render_template('checkout.html', plano=plano)

@app.route('/criar-preferencia', methods=['POST'])
def criar_preferencia():
    """Cria preferência de pagamento no Mercado Pago e redireciona"""
    try:
        # Importar configurações
        from mercadopago_config import (
            MERCADOPAGO_ACCESS_TOKEN,
            SUCCESS_URL,
            FAILURE_URL,
            PENDING_URL,
            WEBHOOK_URL,
            STATEMENT_DESCRIPTOR,
            validar_configuracao
        )

        # Validar configuração
        valido, mensagem = validar_configuracao()
        if not valido:
            flash(mensagem, 'error')
            return redirect(url_for('precos'))

        import mercadopago
        import uuid

        # Obter dados do formulário
        plano_id = request.form.get('plano_id')
        nome = request.form.get('nome')
        email = request.form.get('email')
        documento = request.form.get('documento', '').replace('.', '').replace('-', '').replace('/', '')

        # Buscar plano
        plano = Plano.query.get_or_404(plano_id)

        # Buscar ou criar empresa temporária
        # Por enquanto, vamos criar um pagamento sem empresa (será vinculado depois)
        empresa_id = None
        if 'usuario_id' in session:
            usuario = Usuario.query.get(session['usuario_id'])
            if usuario:
                empresa_id = usuario.empresa_id

        # Gerar referência única
        external_reference = f"PAG-{uuid.uuid4().hex[:12].upper()}"

        # Criar registro de pagamento
        pagamento = Pagamento(
            empresa_id=empresa_id,
            plano_id=plano.id,
            external_reference=external_reference,
            valor=plano.valor,
            dias_assinatura=plano.dias_assinatura,
            status='pending',
            observacoes=f'Nome: {nome}, Email: {email}, Doc: {documento}'
        )
        db.session.add(pagamento)
        db.session.commit()

        # Inicializar SDK do Mercado Pago
        sdk = mercadopago.SDK(MERCADOPAGO_ACCESS_TOKEN)

        # Criar preferência de pagamento
        preference_data = {
            "items": [
                {
                    "title": plano.nome,
                    "description": plano.descricao or f"Assinatura {plano.nome}",
                    "quantity": 1,
                    "unit_price": float(plano.valor),
                    "currency_id": "BRL"
                }
            ],
            "payer": {
                "name": nome,
                "email": email
            },
            "back_urls": {
                "success": SUCCESS_URL,
                "failure": FAILURE_URL,
                "pending": PENDING_URL
            },
            "auto_return": "all",
            "external_reference": external_reference,
            "notification_url": WEBHOOK_URL,
            "statement_descriptor": STATEMENT_DESCRIPTOR
        }

        # Criar preferência
        preference_response = sdk.preference().create(preference_data)

        # Log da resposta para debug
        app.logger.info(f"Resposta do MP: {preference_response}")

        # A resposta pode vir diretamente ou dentro de 'response'
        if 'response' in preference_response:
            preference = preference_response["response"]
        else:
            preference = preference_response

        # Verificar se tem os campos necessários
        if 'id' not in preference:
            app.logger.error(f"Resposta sem ID: {preference}")
            flash('Erro ao criar preferência de pagamento. Verifique as credenciais.', 'error')
            return redirect(url_for('precos'))

        # Salvar preference_id
        pagamento.preference_id = preference["id"]
        db.session.commit()

        # Redirecionar para Mercado Pago
        init_point = preference.get("init_point") or preference.get("sandbox_init_point")
        if not init_point:
            app.logger.error(f"Sem init_point na resposta: {preference}")
            flash('Erro ao obter link de pagamento.', 'error')
            return redirect(url_for('precos'))

        return redirect(init_point)

    except ImportError:
        flash('SDK do Mercado Pago não instalado. Execute: pip install mercadopago', 'error')
        return redirect(url_for('precos'))
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao criar preferência: {str(e)}")
        app.logger.error(traceback.format_exc())
        flash(f'Erro ao processar pagamento: {str(e)}', 'error')
        return redirect(url_for('precos'))

@app.route('/webhook/mercadopago', methods=['POST', 'GET'])
def webhook_mercadopago():
    """Recebe notificações do Mercado Pago sobre pagamentos"""
    try:
        from mercadopago_config import MERCADOPAGO_ACCESS_TOKEN
        import mercadopago

        # Aceitar GET para teste do Mercado Pago
        if request.method == 'GET':
            app.logger.info("Webhook GET recebido (teste do Mercado Pago)")
            return jsonify({'success': True, 'message': 'Webhook ativo'}), 200

        data = request.get_json() if request.is_json else request.form.to_dict()

        app.logger.info(f"Webhook POST recebido: {data}")

        # Verificar se é notificação de pagamento
        topic = data.get('topic') or data.get('type')
        action = data.get('action')

        # Aceitar notificação de teste
        if not data.get('live_mode', True):
            app.logger.info(f"Notificação de TESTE recebida - Topic: {topic}, Action: {action}")
            return jsonify({'success': True, 'message': 'Teste aceito'}), 200

        if topic == 'payment' or action == 'payment.updated' or action == 'payment.created':
            # Extrair payment_id de diferentes formatos
            payment_id = None
            if isinstance(data.get('data'), dict):
                payment_id = data['data'].get('id')
            elif data.get('id'):
                payment_id = data.get('id')

            # Para query params (algumas notificações antigas)
            if not payment_id and request.args.get('id'):
                payment_id = request.args.get('id')

            if not payment_id:
                app.logger.warning(f"ID de pagamento não encontrado na notificação: {data}")
                return jsonify({'success': True, 'message': 'ID não encontrado, notificação ignorada'}), 200

            # Buscar detalhes do pagamento no Mercado Pago
            sdk = mercadopago.SDK(MERCADOPAGO_ACCESS_TOKEN)

            try:
                payment_info = sdk.payment().get(payment_id)

                # Verificar se a resposta contém erro
                if payment_info.get('status') == 404 or payment_info.get('status') >= 400:
                    app.logger.error(f"Erro ao buscar pagamento {payment_id} no MP: {payment_info}")
                    return jsonify({'success': True, 'message': 'Pagamento não encontrado no MP'}), 200

                payment = payment_info.get("response")
                if not payment:
                    app.logger.error(f"Resposta sem dados do pagamento: {payment_info}")
                    return jsonify({'success': True, 'message': 'Dados do pagamento não disponíveis'}), 200

            except Exception as e:
                app.logger.error(f"Erro ao consultar pagamento {payment_id}: {str(e)}")
                return jsonify({'success': True, 'message': 'Erro ao consultar MP, será processado depois'}), 200

            # Buscar pagamento no banco
            external_reference = payment.get('external_reference')
            if not external_reference:
                app.logger.warning(f"Pagamento {payment_id} sem referência externa")
                return jsonify({'success': True, 'message': 'Sem referência externa'}), 200

            pagamento = Pagamento.query.filter_by(external_reference=external_reference).first()

            if not pagamento:
                app.logger.warning(f"Pagamento não encontrado no banco: {external_reference}")
                return jsonify({'success': True, 'message': 'Pagamento não encontrado no sistema'}), 200

            # Atualizar status do pagamento
            pagamento.payment_id = str(payment_id)
            pagamento.status = payment['status']
            pagamento.status_detail = payment.get('status_detail')
            pagamento.payment_type = payment.get('payment_type_id')
            pagamento.payment_method = payment.get('payment_method_id')
            pagamento.dados_mp = json.dumps(payment)

            # Se aprovado, creditar dias
            if payment['status'] == 'approved' and not pagamento.data_aprovacao:
                pagamento.data_aprovacao = datetime.utcnow()

                # Creditar dias na empresa (se existe)
                if pagamento.empresa_id:
                    empresa = pagamento.empresa

                    if empresa.dias_assinatura and empresa.dias_assinatura > 0:
                        # Somar dias aos existentes
                        empresa.dias_assinatura += pagamento.dias_assinatura
                    else:
                        # Iniciar nova assinatura
                        empresa.dias_assinatura = pagamento.dias_assinatura
                        empresa.data_inicio_assinatura = datetime.utcnow()

                    pagamento.observacoes += f"\n{pagamento.dias_assinatura} dias creditados em {datetime.utcnow()}"
                    app.logger.info(f"✅ Dias creditados: {pagamento.dias_assinatura} para empresa {empresa.id}")

            db.session.commit()
            app.logger.info(f"✅ Pagamento atualizado: {external_reference} - Status: {pagamento.status}")

            return jsonify({'success': True}), 200

        # Outros tipos de notificação
        app.logger.info(f"Notificação não processada - Topic: {topic}, Action: {action}")
        return jsonify({'success': True, 'message': 'Tipo de notificação não processada'}), 200

    except Exception as e:
        app.logger.error(f"❌ Erro no webhook: {str(e)}")
        app.logger.error(traceback.format_exc())
        # Retornar 200 mesmo com erro para evitar retentativas desnecessárias
        return jsonify({'success': False, 'error': 'Erro interno, será processado manualmente'}), 200

@app.route('/pagamento/sucesso')
def pagamento_sucesso():
    """Página de sucesso após pagamento"""
    payment_id = request.args.get('payment_id')
    external_reference = request.args.get('external_reference')

    pagamento = None
    if external_reference:
        pagamento = Pagamento.query.filter_by(external_reference=external_reference).first()

    return render_template('pagamento_sucesso.html', pagamento=pagamento)

@app.route('/pagamento/falha')
def pagamento_falha():
    """Página de falha no pagamento"""
    return render_template('pagamento_falha.html')

@app.route('/pagamento/pendente')
def pagamento_pendente():
    """Página de pagamento pendente"""
    return render_template('pagamento_pendente.html')

@app.route('/admin/editar-conta/<int:conta_id>')
def admin_editar_conta(conta_id):
    """Página para editar dados da conta"""
    if 'usuario_id' not in session or session.get('usuario_tipo') != 'admin':
        flash('Acesso negado. Apenas administradores.', 'error')
        return redirect(url_for('login'))

    conta = Empresa.query.get(conta_id)
    if not conta:
        flash('Conta não encontrada.', 'error')
        return redirect(url_for('admin_usuarios'))
    
    return render_template('admin_editar_conta.html', conta=conta)

@app.route('/admin/salvar-conta/<int:conta_id>', methods=['POST'])
def admin_salvar_conta(conta_id):
    """Salvar alterações na conta"""
    if 'usuario_id' not in session or session.get('usuario_tipo') != 'admin':
        flash('Acesso negado. Apenas administradores.', 'error')
        return redirect(url_for('login'))
    
    try:
        conta = Empresa.query.get(conta_id)
        if not conta:
            flash('Conta não encontrada.', 'error')
            return redirect(url_for('admin_usuarios'))
        
        # Atualizar dados
        conta.razao_social = request.form.get('razao_social', conta.razao_social)
        conta.nome_fantasia = request.form.get('nome_fantasia', conta.nome_fantasia)
        conta.email = request.form.get('email', conta.email)
        conta.telefone = request.form.get('telefone', conta.telefone)
        
        db.session.commit()
        flash('Conta atualizada com sucesso!', 'success')
        # Redirecionar para a página de origem
        referer = request.headers.get('Referer', '')
        if 'admin/usuarios' in referer:
            return redirect(url_for('admin_usuarios'))
        return redirect(url_for('admin_usuarios'))
    
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao atualizar conta: {str(e)}', 'error')
        return redirect(url_for('admin_usuarios'))

# ===== ROTAS DO PAINEL CONTADOR/BPO =====

@app.route('/contador/dashboard')
def dashboard_contador():
    """Dashboard do contador/BPO"""
    # Verificar se está logado (pode ser usuario_id ou sub_usuario_id)
    if 'usuario_id' not in session and 'sub_usuario_id' not in session:
        flash('Faça login para acessar.', 'error')
        return redirect(url_for('login'))
    
    # Verificar se é contador ou sub-usuário
    tipo_conta = session.get('tipo_conta')
    tipo_usuario = session.get('usuario_tipo')
    
    # Verificar se é contador/BPO ou sub-usuário de contador
    if tipo_conta != 'contador_bpo' and tipo_usuario != 'sub_contador':
        flash('Acesso negado. Apenas contadores/BPO.', 'error')
        return redirect(url_for('dashboard'))
    
    if tipo_usuario == 'sub_contador':
        # Sub-usuário
        sub_usuario_id = session.get('sub_usuario_id')
        contador_id = session.get('contador_id')
        
        # Buscar empresas autorizadas para este sub-usuário
        permissoes = PermissaoSubUsuario.query.filter_by(sub_usuario_id=sub_usuario_id).all()
        empresas_ids = [p.empresa_id for p in permissoes] if permissoes else []

        # ✅ OTIMIZAÇÃO: Buscar vínculos autorizados COM joinedload
        from sqlalchemy.orm import joinedload
        if empresas_ids:
            vinculos = VinculoContador.query.options(
                joinedload(VinculoContador.empresa)
            ).filter(
                VinculoContador.contador_id == contador_id,
                VinculoContador.status == 'autorizado',
                VinculoContador.empresa_id.in_(empresas_ids)
            ).all()
        else:
            # Se não há permissões, retornar lista vazia
            vinculos = []

        sub_usuarios = []
    else:
        # Usuário principal do contador
        empresa_id = session.get('empresa_id')

        # ✅ OTIMIZAÇÃO: Usar joinedload para carregar empresas junto com vínculos (evita N queries)
        from sqlalchemy.orm import joinedload
        vinculos = VinculoContador.query.options(
            joinedload(VinculoContador.empresa)
        ).filter_by(
            contador_id=empresa_id
        ).order_by(VinculoContador.data_solicitacao.desc()).all()

        # Buscar sub-usuários
        sub_usuarios = SubUsuarioContador.query.filter_by(contador_id=empresa_id).all()

    # Estatísticas
    stats = {
        'empresas_autorizadas': len([v for v in vinculos if v.status == 'autorizado']),
        'vinculos_pendentes': VinculoContador.query.filter_by(
            contador_id=session.get('empresa_id'),
            status='pendente'
        ).count() if tipo_usuario != 'sub_contador' else 0,
        'sub_usuarios': len(sub_usuarios)
    }

    # ✅ OTIMIZAÇÃO: Buscar lançamentos de hoje de TODAS as empresas de uma vez
    from datetime import date
    from collections import defaultdict
    hoje = date.today()

    # Coletar IDs de empresas autorizadas
    empresas_ids_autorizadas = [v.empresa_id for v in vinculos if v.status == 'autorizado']

    # Buscar TODOS os lançamentos do dia de TODAS as empresas autorizadas em UMA query
    lancamentos_hoje = []
    if empresas_ids_autorizadas:
        lancamentos_hoje = Lancamento.query.filter(
            Lancamento.empresa_id.in_(empresas_ids_autorizadas),
            Lancamento.data_prevista == hoje
        ).all()

    # Agrupar lançamentos por empresa_id em memória
    lancamentos_por_empresa = defaultdict(lambda: {'entradas': [], 'saidas': []})
    for lancamento in lancamentos_hoje:
        if lancamento.tipo == 'entrada':
            lancamentos_por_empresa[lancamento.empresa_id]['entradas'].append(lancamento)
        elif lancamento.tipo == 'saida':
            lancamentos_por_empresa[lancamento.empresa_id]['saidas'].append(lancamento)

    # Montar lista de empresas com lançamentos
    empresas_com_lancamentos = []
    for vinculo in vinculos:
        if vinculo.status == 'autorizado':
            empresa_id_atual = vinculo.empresa_id
            if empresa_id_atual in lancamentos_por_empresa:
                entradas = lancamentos_por_empresa[empresa_id_atual]['entradas']
                saidas = lancamentos_por_empresa[empresa_id_atual]['saidas']

                empresas_com_lancamentos.append({
                    'empresa': vinculo.empresa,
                    'entradas': entradas,
                    'saidas': saidas,
                    'total_entradas': sum(l.valor for l in entradas),
                    'total_saidas': sum(l.valor for l in saidas)
                })
    
    return render_template('contador_dashboard.html',
                         vinculos=vinculos,
                         sub_usuarios=sub_usuarios,
                         stats=stats,
                         empresas_com_lancamentos=empresas_com_lancamentos)

@app.route('/contador/vincular-empresa', methods=['POST'])
def contador_vincular_empresa():
    """Solicitar vínculo com empresa/pessoa física"""
    # Verificar se está logado (pode ser usuario_id ou sub_usuario_id)
    if 'usuario_id' not in session and 'sub_usuario_id' not in session:
        flash('Faça login para acessar.', 'error')
        return redirect(url_for('login'))
    
    # Apenas usuário principal pode vincular empresas
    tipo_usuario = session.get('usuario_tipo')
    if tipo_usuario == 'sub_contador':
        flash('Apenas o usuário principal pode vincular empresas.', 'error')
        return redirect(url_for('dashboard_contador'))
    
    tipo_documento = request.form.get('tipo_documento')
    cpf = request.form.get('cpf', '').strip()
    cnpj = request.form.get('cnpj', '').strip()

    contador_id = session.get('empresa_id')

    # Buscar empresa com busca flexível (com ou sem formatação)
    empresa = None
    if tipo_documento == 'cpf':
        # Tentar buscar por CPF exato
        empresa = Empresa.query.filter_by(cpf=cpf).first()
        if not empresa:
            # Tentar buscar removendo formatação
            cpf_limpo = cpf.replace('.', '').replace('-', '').replace('/', '')
            todas_empresas = Empresa.query.filter(Empresa.cpf.isnot(None)).all()
            for emp in todas_empresas:
                emp_cpf_limpo = emp.cpf.replace('.', '').replace('-', '').replace('/', '') if emp.cpf else ''
                if emp_cpf_limpo == cpf_limpo:
                    empresa = emp
                    break
    elif tipo_documento == 'cnpj':
        # Tentar buscar por CNPJ exato
        empresa = Empresa.query.filter_by(cnpj=cnpj).first()
        if not empresa:
            # Tentar buscar removendo formatação
            cnpj_limpo = cnpj.replace('.', '').replace('-', '').replace('/', '')
            todas_empresas = Empresa.query.filter(Empresa.cnpj.isnot(None)).all()
            for emp in todas_empresas:
                emp_cnpj_limpo = emp.cnpj.replace('.', '').replace('-', '').replace('/', '') if emp.cnpj else ''
                if emp_cnpj_limpo == cnpj_limpo:
                    empresa = emp
                    break

    if not empresa:
        flash('Empresa/Pessoa Física não encontrada.', 'error')
        return redirect(url_for('dashboard_contador'))
    
    # Verificar se já existe vínculo
    vinculo_existente = VinculoContador.query.filter_by(
        contador_id=contador_id,
        empresa_id=empresa.id
    ).first()
    
    if vinculo_existente:
        flash('Já existe um vínculo com esta empresa.', 'info')
        return redirect(url_for('dashboard_contador'))
    
    # Criar vínculo pendente
    novo_vinculo = VinculoContador(
        contador_id=contador_id,
        empresa_id=empresa.id,
        status='pendente'
    )
    
    db.session.add(novo_vinculo)
    db.session.commit()
    
    flash('Solicitação de vínculo enviada! Aguarde autorização do usuário master.', 'success')
    return redirect(url_for('dashboard_contador'))

@app.route('/api/buscar-empresa-por-documento')
def api_buscar_empresa_por_documento():
    """API para buscar empresa por CPF ou CNPJ"""
    tipo_documento = request.args.get('tipo')
    documento = request.args.get('documento', '').strip()

    if not tipo_documento or not documento:
        return jsonify({'success': False, 'message': 'Parâmetros inválidos'})

    # Remover formatação do documento para busca flexível
    documento_limpo = documento.replace('.', '').replace('-', '').replace('/', '')

    empresa = None
    if tipo_documento == 'cpf':
        # Buscar por CPF exato ou removendo formatação
        empresa = Empresa.query.filter_by(cpf=documento).first()
        if not empresa:
            # Tentar buscar comparando sem formatação
            todas_empresas = Empresa.query.filter(Empresa.cpf.isnot(None)).all()
            for emp in todas_empresas:
                cpf_limpo = emp.cpf.replace('.', '').replace('-', '').replace('/', '') if emp.cpf else ''
                if cpf_limpo == documento_limpo:
                    empresa = emp
                    break
    elif tipo_documento == 'cnpj':
        # Buscar por CNPJ exato ou removendo formatação
        empresa = Empresa.query.filter_by(cnpj=documento).first()
        if not empresa:
            # Tentar buscar comparando sem formatação
            todas_empresas = Empresa.query.filter(Empresa.cnpj.isnot(None)).all()
            for emp in todas_empresas:
                cnpj_limpo = emp.cnpj.replace('.', '').replace('-', '').replace('/', '') if emp.cnpj else ''
                if cnpj_limpo == documento_limpo:
                    empresa = emp
                    break

    if empresa:
        return jsonify({
            'success': True,
            'empresa': {
                'id': empresa.id,
                'razao_social': empresa.razao_social,
                'nome_fantasia': empresa.nome_fantasia,
                'tipo_pessoa': empresa.tipo_pessoa,
                'cpf': empresa.cpf,
                'cnpj': empresa.cnpj
            }
        })
    else:
        return jsonify({'success': False, 'message': 'Empresa/Pessoa Física não encontrada'})

@app.route('/contador/vincular/<int:vinculo_id>/excluir', methods=['POST'])
def contador_excluir_vinculo(vinculo_id):
    """Excluir vínculo com empresa"""
    # Verificar se está logado (pode ser usuario_id ou sub_usuario_id)
    if 'usuario_id' not in session and 'sub_usuario_id' not in session:
        return jsonify({'success': False, 'message': 'Acesso negado'}), 403
    
    # Apenas usuário principal pode excluir vínculos
    tipo_usuario = session.get('usuario_tipo')
    if tipo_usuario == 'sub_contador':
        return jsonify({'success': False, 'message': 'Apenas o usuário principal pode excluir vínculos'}), 403
    
    try:
        contador_id = session.get('empresa_id')
        
        vinculo = VinculoContador.query.filter_by(
            id=vinculo_id,
            contador_id=contador_id
        ).first()
        
        if not vinculo:
            return jsonify({'success': False, 'message': 'Vínculo não encontrado'})
        
        # Excluir permissões de sub-usuários relacionadas a este vínculo
        if vinculo.status == 'autorizado':
            # Buscar sub-usuários que têm permissão para esta empresa
            sub_usuarios = SubUsuarioContador.query.filter_by(contador_id=contador_id).all()
            for sub in sub_usuarios:
                PermissaoSubUsuario.query.filter_by(
                    sub_usuario_id=sub.id,
                    empresa_id=vinculo.empresa_id
                ).delete()
        
        # Excluir vínculo
        db.session.delete(vinculo)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Vínculo excluído com sucesso'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/contador/criar-sub-usuario', methods=['POST'])
def contador_criar_sub_usuario():
    """Criar novo sub-usuário"""
    # Verificar se está logado (pode ser usuario_id ou sub_usuario_id)
    if 'usuario_id' not in session and 'sub_usuario_id' not in session:
        flash('Faça login para acessar.', 'error')
        return redirect(url_for('login'))
    
    # Apenas usuário principal pode criar sub-usuários
    tipo_usuario = session.get('usuario_tipo')
    if tipo_usuario == 'sub_contador':
        flash('Apenas o usuário principal pode criar sub-usuários.', 'error')
        return redirect(url_for('dashboard_contador'))
    
    nome = request.form.get('nome', '').strip()
    usuario = request.form.get('usuario', '').strip()
    email = request.form.get('email', '').strip()  # Email opcional
    senha = request.form.get('senha')
    
    # Validações
    if not nome:
        flash('Nome completo é obrigatório.', 'error')
        return redirect(url_for('dashboard_contador'))
    
    if not usuario:
        flash('Nome de usuário é obrigatório.', 'error')
        return redirect(url_for('dashboard_contador'))
    
    if not senha:
        flash('Senha é obrigatória.', 'error')
        return redirect(url_for('dashboard_contador'))
    
    contador_id = session.get('empresa_id')
    
    # Verificar se usuário já existe (globalmente, não apenas no contador)
    sub_existente = SubUsuarioContador.query.filter_by(usuario=usuario).first()
    if sub_existente:
        flash('Nome de usuário já cadastrado. Escolha outro.', 'error')
        return redirect(url_for('dashboard_contador'))
    
    # Removida validação de email único para sub-usuários
    # Criar sub-usuário
    
    # Criar sub-usuário
    senha_hash = generate_password_hash(senha)
    novo_sub = SubUsuarioContador(
        contador_id=contador_id,
        nome=nome,
        usuario=usuario,
        email=email if email else None,
        senha=senha_hash
    )
    
    db.session.add(novo_sub)
    db.session.commit()
    
    flash(f'Sub-usuário {nome} criado com sucesso!', 'success')
    return redirect(url_for('dashboard_contador'))

@app.route('/contador/sub-usuario/<int:sub_usuario_id>/permissoes')
def contador_gerenciar_permissoes(sub_usuario_id):
    """Gerenciar permissões de um sub-usuário"""
    # Verificar se está logado (pode ser usuario_id ou sub_usuario_id)
    if 'usuario_id' not in session and 'sub_usuario_id' not in session:
        flash('Faça login para acessar.', 'error')
        return redirect(url_for('login'))
    
    # Apenas usuário principal pode gerenciar permissões
    tipo_usuario = session.get('usuario_tipo')
    if tipo_usuario == 'sub_contador':
        flash('Apenas o usuário principal pode gerenciar permissões.', 'error')
        return redirect(url_for('dashboard_contador'))
    
    contador_id = session.get('empresa_id')
    
    # Buscar sub-usuário
    sub_usuario = SubUsuarioContador.query.filter_by(
        id=sub_usuario_id,
        contador_id=contador_id
    ).first()
    
    if not sub_usuario:
        flash('Sub-usuário não encontrado.', 'error')
        return redirect(url_for('dashboard_contador'))
    
    # Buscar empresas vinculadas autorizadas
    empresas_disponiveis = VinculoContador.query.filter_by(
        contador_id=contador_id,
        status='autorizado'
    ).all()
    
    # Buscar empresas já autorizadas para este sub-usuário
    permissoes = PermissaoSubUsuario.query.filter_by(sub_usuario_id=sub_usuario_id).all()
    empresas_autorizadas = [p.empresa_id for p in permissoes]
    
    return render_template('contador_permissoes_sub_usuario.html',
                         sub_usuario=sub_usuario,
                         empresas_disponiveis=empresas_disponiveis,
                         empresas_autorizadas=empresas_autorizadas)

@app.route('/contador/sub-usuario/<int:sub_usuario_id>/salvar-permissoes', methods=['POST'])
def contador_salvar_permissoes(sub_usuario_id):
    """Salvar permissões de um sub-usuário"""
    # Verificar se está logado (pode ser usuario_id ou sub_usuario_id)
    if 'usuario_id' not in session and 'sub_usuario_id' not in session:
        flash('Faça login para acessar.', 'error')
        return redirect(url_for('login'))
    
    # Apenas usuário principal pode salvar permissões
    tipo_usuario = session.get('usuario_tipo')
    if tipo_usuario == 'sub_contador':
        flash('Apenas o usuário principal pode salvar permissões.', 'error')
        return redirect(url_for('dashboard_contador'))
    
    contador_id = session.get('empresa_id')
    
    # Buscar sub-usuário
    sub_usuario = SubUsuarioContador.query.filter_by(
        id=sub_usuario_id,
        contador_id=contador_id
    ).first()
    
    if not sub_usuario:
        flash('Sub-usuário não encontrado.', 'error')
        return redirect(url_for('dashboard_contador'))
    
    # Remover permissões antigas
    PermissaoSubUsuario.query.filter_by(sub_usuario_id=sub_usuario_id).delete()
    
    # Adicionar novas permissões
    empresas_autorizadas = request.form.getlist('empresas_autorizadas')
    
    for empresa_id in empresas_autorizadas:
        nova_permissao = PermissaoSubUsuario(
            sub_usuario_id=sub_usuario_id,
            empresa_id=int(empresa_id)
        )
        db.session.add(nova_permissao)
    
    db.session.commit()
    
    flash('Permissões atualizadas com sucesso!', 'success')
    return redirect(url_for('dashboard_contador'))

@app.route('/contador/sub-usuario/<int:sub_usuario_id>/excluir', methods=['POST'])
def contador_excluir_sub_usuario(sub_usuario_id):
    """Excluir sub-usuário"""
    # Verificar se está logado (pode ser usuario_id ou sub_usuario_id)
    if 'usuario_id' not in session and 'sub_usuario_id' not in session:
        return jsonify({'success': False, 'message': 'Acesso negado'}), 403
    
    # Apenas usuário principal pode excluir sub-usuários
    tipo_usuario = session.get('usuario_tipo')
    if tipo_usuario == 'sub_contador':
        return jsonify({'success': False, 'message': 'Apenas o usuário principal pode excluir sub-usuários'}), 403
    
    try:
        contador_id = session.get('empresa_id')
        
        sub_usuario = SubUsuarioContador.query.filter_by(
            id=sub_usuario_id,
            contador_id=contador_id
        ).first()
        
        if sub_usuario:
            # Excluir permissões
            PermissaoSubUsuario.query.filter_by(sub_usuario_id=sub_usuario_id).delete()
            
            # Excluir sub-usuário
            db.session.delete(sub_usuario)
            db.session.commit()
            
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'message': 'Sub-usuário não encontrado'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/contador/acessar-empresa', methods=['POST'])
def contador_acessar_empresa():
    """Permite ao contador acessar uma empresa vinculada"""
    # Verificar se está logado (pode ser usuario_id ou sub_usuario_id)
    if 'usuario_id' not in session and 'sub_usuario_id' not in session:
        flash('Faça login para acessar.', 'error')
        return redirect(url_for('login'))
    
    tipo_conta = session.get('tipo_conta')
    tipo_usuario = session.get('usuario_tipo')
    empresa_id = int(request.form.get('empresa_id'))
    
    # Verificar se é contador/BPO ou sub-usuário
    if tipo_conta != 'contador_bpo' and tipo_usuario != 'sub_contador':
        flash('Acesso negado. Apenas contadores/BPO.', 'error')
        return redirect(url_for('dashboard'))
    
    # Verificar se o vínculo existe e está autorizado
    if tipo_usuario == 'sub_contador':
        sub_usuario_id = session.get('sub_usuario_id')
        contador_id = session.get('contador_id')
        
        # Verificar se o sub-usuário tem permissão para esta empresa
        permissao = PermissaoSubUsuario.query.filter_by(
            sub_usuario_id=sub_usuario_id,
            empresa_id=empresa_id
        ).first()
        
        if not permissao:
            flash('Você não tem permissão para acessar esta empresa.', 'error')
            return redirect(url_for('dashboard_contador'))
        
        vinculo = VinculoContador.query.filter_by(
            contador_id=contador_id,
            empresa_id=empresa_id,
            status='autorizado'
        ).first()
    else:
        contador_id = session.get('empresa_id')
        vinculo = VinculoContador.query.filter_by(
            contador_id=contador_id,
            empresa_id=empresa_id,
            status='autorizado'
        ).first()
    
    if not vinculo:
        flash('Vínculo não encontrado ou não autorizado.', 'error')
        return redirect(url_for('dashboard_contador'))
    
    # Buscar a empresa
    empresa = db.session.get(Empresa, empresa_id)
    if not empresa:
        flash('Empresa não encontrada.', 'error')
        return redirect(url_for('dashboard_contador'))
    
    # Salvar informações originais do contador antes de atualizar
    session['empresa_id_original'] = session.get('empresa_id')
    session['empresa_nome_original'] = session.get('empresa_nome')
    session['empresa_tipo_pessoa_original'] = session.get('empresa_tipo_pessoa')
    session['empresa_nome_fantasia_original'] = session.get('empresa_nome_fantasia')
    session['empresa_cpf_original'] = session.get('empresa_cpf')
    session['empresa_cnpj_original'] = session.get('empresa_cnpj')
    session['sub_usuario_id_original'] = session.get('sub_usuario_id')  # Salvar sub_usuario_id original
    session['usuario_id_original'] = session.get('usuario_id')  # Salvar usuario_id original se existir
    
    # Se for sub-usuário, buscar um usuário da empresa vinculada para usar como referência
    if tipo_usuario == 'sub_contador':
        # Buscar o primeiro usuário ativo da empresa vinculada (ou criar um temporário)
        usuario_empresa = Usuario.query.filter_by(empresa_id=empresa.id, ativo=True).first()
        if usuario_empresa:
            # Usar um usuário existente da empresa como referência
            session['usuario_id'] = usuario_empresa.id
            session['usuario_id_temporario'] = True  # Flag para indicar que é temporário
        else:
            # Se não houver usuário, criar um temporário (não salvar no banco)
            # Usar o sub_usuario_id como base para criar um ID temporário
            session['usuario_id'] = None  # Será tratado na rota dashboard
            session['usuario_id_temporario'] = True
    
    # Atualizar a sessão para acessar como a empresa vinculada
    session['empresa_id'] = empresa.id
    session['empresa_nome'] = empresa.razao_social
    session['empresa_tipo_pessoa'] = empresa.tipo_pessoa
    session['empresa_nome_fantasia'] = empresa.nome_fantasia
    session['empresa_cpf'] = empresa.cpf
    session['empresa_cnpj'] = empresa.cnpj
    session['acesso_contador'] = True  # Flag para indicar que está acessando como contador
    
    flash(f'Acessando como: {empresa.razao_social}', 'success')
    return redirect(url_for('dashboard'))

@app.route('/contador/voltar-perfil', methods=['POST'])
def contador_voltar_perfil():
    """Volta ao perfil original do contador"""
    # Verificar se está logado (pode ser usuario_id ou sub_usuario_id)
    if 'usuario_id' not in session and 'sub_usuario_id' not in session:
        flash('Faça login para acessar.', 'error')
        return redirect(url_for('login'))
    
    if not session.get('acesso_contador') or not session.get('empresa_id_original'):
        flash('Você já está no perfil do contador.', 'info')
        return redirect(url_for('dashboard_contador'))
    
    # Restaurar dados originais do contador
    empresa_original = db.session.get(Empresa, session.get('empresa_id_original'))
    if empresa_original:
        session['empresa_id'] = empresa_original.id
        session['empresa_nome'] = empresa_original.razao_social
        session['empresa_tipo_pessoa'] = empresa_original.tipo_pessoa
        session['empresa_nome_fantasia'] = empresa_original.nome_fantasia
        session['empresa_cpf'] = empresa_original.cpf
        session['empresa_cnpj'] = empresa_original.cnpj
    
    # Restaurar usuario_id original (se for sub-usuário, remover o temporário)
    if session.get('usuario_id_temporario'):
        session.pop('usuario_id', None)
        session.pop('usuario_id_temporario', None)
    
    if session.get('usuario_id_original'):
        session['usuario_id'] = session.get('usuario_id_original')
        session.pop('usuario_id_original', None)
    
    # Restaurar sub_usuario_id se existir
    if session.get('sub_usuario_id_original'):
        session['sub_usuario_id'] = session.get('sub_usuario_id_original')
        session.pop('sub_usuario_id_original', None)
    
    session.pop('acesso_contador', None)
    session.pop('empresa_id_original', None)
    session.pop('empresa_nome_original', None)
    session.pop('empresa_tipo_pessoa_original', None)
    session.pop('empresa_nome_fantasia_original', None)
    session.pop('empresa_cpf_original', None)
    session.pop('empresa_cnpj_original', None)
    
    flash('Voltou ao perfil do contador.', 'success')
    return redirect(url_for('dashboard_contador'))

@app.route('/contador/gerenciar-vinculos')
def contador_gerenciar_vinculos():
    """Redireciona para o dashboard do contador na aba de gerenciar vínculos"""
    if 'usuario_id' not in session:
        flash('Faça login para acessar.', 'error')
        return redirect(url_for('login'))
    
    tipo_conta = session.get('tipo_conta')
    tipo_usuario = session.get('usuario_tipo')
    
    # Verificar se é contador/BPO (não sub-usuário)
    if tipo_conta != 'contador_bpo' or tipo_usuario == 'sub_contador':
        flash('Acesso negado. Apenas contadores/BPO principais.', 'error')
        return redirect(url_for('dashboard'))
    
    # Redirecionar para dashboard com parâmetro para abrir a aba de vínculos
    return redirect(url_for('dashboard_contador') + '?aba=empresas-vinculadas')

@app.route('/contador/gerenciar-usuarios')
def contador_gerenciar_usuarios():
    """Redireciona para o dashboard do contador na aba de gerenciar usuários"""
    if 'usuario_id' not in session:
        flash('Faça login para acessar.', 'error')
        return redirect(url_for('login'))
    
    tipo_conta = session.get('tipo_conta')
    tipo_usuario = session.get('usuario_tipo')
    
    # Verificar se é contador/BPO (não sub-usuário)
    if tipo_conta != 'contador_bpo' or tipo_usuario == 'sub_contador':
        flash('Acesso negado. Apenas contadores/BPO principais.', 'error')
        return redirect(url_for('dashboard'))
    
    # Redirecionar para dashboard com parâmetro para abrir a aba de sub-usuários
    return redirect(url_for('dashboard_contador') + '?aba=sub-usuarios')

# ===== ROTA PARA AUTORIZAR VÍNCULOS (EMPRESA/PF) =====

@app.route('/vinculos/pendentes')
def vinculos_pendentes():
    """Listar vínculos pendentes de autorização"""
    if 'usuario_id' not in session:
        flash('Faça login para acessar.', 'error')
        return redirect(url_for('login'))
    
    empresa_id = session.get('empresa_id')
    usuario_id = session.get('usuario_id')
    
    # Verificar se é usuário principal
    usuario = Usuario.query.get(usuario_id)
    if usuario.tipo not in ['admin', 'usuario_principal']:
        flash('Apenas o usuário principal pode autorizar vínculos.', 'error')
        return redirect(url_for('dashboard'))
    
    # Buscar vínculos pendentes
    vinculos = VinculoContador.query.filter_by(
        empresa_id=empresa_id,
        status='pendente'
    ).all()
    
    return render_template('vinculos_pendentes.html', vinculos=vinculos)

@app.route('/vinculos/<int:vinculo_id>/autorizar', methods=['POST'])
def autorizar_vinculo(vinculo_id):
    """Autorizar vínculo com contador"""
    if 'usuario_id' not in session:
        flash('Acesso negado', 'error')
        return redirect(url_for('login'))

    try:
        empresa_id = session.get('empresa_id')
        usuario_id = session.get('usuario_id')

        # Verificar se é usuário principal
        usuario = Usuario.query.get(usuario_id)
        if usuario.tipo not in ['admin', 'usuario_principal']:
            flash('Apenas usuário principal pode autorizar vínculos', 'error')
            return redirect(url_for('vinculos_pendentes'))

        vinculo = VinculoContador.query.filter_by(
            id=vinculo_id,
            empresa_id=empresa_id
        ).first()

        if vinculo:
            vinculo.status = 'autorizado'
            vinculo.data_autorizacao = datetime.utcnow()
            db.session.commit()

            flash('Vínculo autorizado com sucesso!', 'success')
            return redirect(url_for('vinculos_pendentes'))
        else:
            flash('Vínculo não encontrado', 'error')
            return redirect(url_for('vinculos_pendentes'))
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao autorizar vínculo: {str(e)}', 'error')
        return redirect(url_for('vinculos_pendentes'))

@app.route('/vinculos/<int:vinculo_id>/rejeitar', methods=['POST'])
def rejeitar_vinculo(vinculo_id):
    """Rejeitar vínculo com contador"""
    if 'usuario_id' not in session:
        flash('Acesso negado', 'error')
        return redirect(url_for('login'))

    try:
        empresa_id = session.get('empresa_id')

        vinculo = VinculoContador.query.filter_by(
            id=vinculo_id,
            empresa_id=empresa_id
        ).first()

        if vinculo:
            vinculo.status = 'rejeitado'
            db.session.commit()

            flash('Vínculo rejeitado com sucesso!', 'success')
            return redirect(url_for('vinculos_pendentes'))
        else:
            flash('Vínculo não encontrado', 'error')
            return redirect(url_for('vinculos_pendentes'))
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao rejeitar vínculo: {str(e)}', 'error')
        return redirect(url_for('vinculos_pendentes'))

@app.route('/reset-admin-senha')
def reset_admin_senha():
    """Rota temporária para resetar senha do admin"""
    try:
        empresa_admin = Empresa.query.filter_by(cnpj='00.000.000/0000-00').first()
        
        if not empresa_admin:
            # Criar empresa admin se não existir
            empresa_admin = Empresa(
                cnpj='00.000.000/0000-00',
                razao_social='Sistema Administrativo',
                nome_fantasia='Admin',
                tipo_empresa='servicos',
                tipo_conta='admin',
                dias_assinatura=999999,
                ativo=True
            )
            db.session.add(empresa_admin)
            db.session.flush()
        
        # Garantir tipo_conta
        if empresa_admin.tipo_conta != 'admin':
            empresa_admin.tipo_conta = 'admin'
            db.session.commit()
        
        usuario_admin = Usuario.query.filter_by(
            empresa_id=empresa_admin.id,
            usuario='admin'
        ).first()
        
        if not usuario_admin:
            # Criar usuário admin
            usuario_admin = Usuario(
                nome='Administrador',
                usuario='admin',
                email='admin@sistema.com',
                senha=generate_password_hash('admin123'),
                tipo='admin',
                empresa_id=empresa_admin.id,
                ativo=True
            )
            db.session.add(usuario_admin)
            resultado = "Usuário ADMIN criado!"
        else:
            # Resetar senha
            usuario_admin.senha = generate_password_hash('admin123')
            resultado = "Senha do ADMIN resetada!"
        
        db.session.commit()
        
        return f"""
        <html>
        <head><title>Senha Resetada</title></head>
        <body style="font-family: Arial; padding: 50px;">
            <h1>✅ {resultado}</h1>
            <hr>
            <h2>🔐 Credenciais do Administrador:</h2>
            <p><strong>Tipo de Acesso:</strong> Empresa (CNPJ)</p>
            <p><strong>CNPJ:</strong> 00.000.000/0000-00</p>
            <p><strong>Usuário:</strong> admin</p>
            <p><strong>Senha:</strong> admin123</p>
            <hr>
            <p><a href="/login">Clique aqui para fazer login</a></p>
        </body>
        </html>
        """
    except Exception as e:
        return f"❌ Erro: {str(e)}", 500

# ===== CONFIGURAÇÃO DO SCHEDULER PARA ATUALIZAÇÃO AUTOMÁTICA =====

# Inicializar o scheduler
scheduler = BackgroundScheduler()

# Agendar a atualização de assinaturas para rodar todos os dias à meia-noite
scheduler.add_job(
    func=atualizar_todas_assinaturas,
    trigger=CronTrigger(hour=0, minute=0),  # Meia-noite todos os dias
    id='atualizar_assinaturas',
    name='Atualizar dias de assinatura',
    replace_existing=True
)

# Também executar a cada 1 hora para garantir que contas sejam suspensas rapidamente
scheduler.add_job(
    func=atualizar_todas_assinaturas,
    trigger='interval',
    hours=1,
    id='atualizar_assinaturas_hourly',
    name='Atualizar assinaturas a cada hora',
    replace_existing=True
)

# Iniciar o scheduler
scheduler.start()
app.logger.info("✅ Scheduler de atualização de assinaturas iniciado")

# Garantir que o scheduler seja desligado quando o app for finalizado
atexit.register(lambda: scheduler.shutdown())

# ============================================================
# ROTA TEMPORÁRIA: MIGRAÇÃO DE TRANSFERÊNCIAS
# REMOVER APÓS EXECUÇÃO EM PRODUÇÃO
# ============================================================
@app.route('/admin/migrar-transferencias-agora', methods=['GET'])
def migrar_transferencias_agora():
    """
    Rota temporária para executar migração de transferências.
    Acesso: /admin/migrar-transferencias-agora
    """
    try:
        # Verificar se as colunas já existem
        result = db.session.execute(text("""
            SELECT column_name
            FROM information_schema.columns
            WHERE table_name = 'lancamento'
            AND column_name IN ('eh_transferencia', 'transferencia_id')
        """))
        colunas_existentes = [row[0] for row in result.fetchall()]

        if 'eh_transferencia' in colunas_existentes and 'transferencia_id' in colunas_existentes:
            return jsonify({
                'success': True,
                'message': '✅ As colunas já existem. Migração não necessária.',
                'colunas': colunas_existentes
            })

        mensagens = []

        # Adicionar eh_transferencia se não existir
        if 'eh_transferencia' not in colunas_existentes:
            db.session.execute(text("""
                ALTER TABLE lancamento
                ADD COLUMN eh_transferencia BOOLEAN DEFAULT FALSE
            """))
            mensagens.append("✅ Coluna 'eh_transferencia' adicionada")
        else:
            mensagens.append("ℹ️  Coluna 'eh_transferencia' já existe")

        # Adicionar transferencia_id se não existir
        if 'transferencia_id' not in colunas_existentes:
            db.session.execute(text("""
                ALTER TABLE lancamento
                ADD COLUMN transferencia_id VARCHAR(36)
            """))
            mensagens.append("✅ Coluna 'transferencia_id' adicionada")
        else:
            mensagens.append("ℹ️  Coluna 'transferencia_id' já existe")

        # Criar índice
        try:
            db.session.execute(text("""
                CREATE INDEX IF NOT EXISTS idx_lancamento_transferencia_id
                ON lancamento(transferencia_id)
            """))
            mensagens.append("✅ Índice criado")
        except Exception as e:
            mensagens.append(f"⚠️  Índice já existe ou erro: {str(e)}")

        # Commit
        db.session.commit()

        # Verificar resultado
        result = db.session.execute(text("""
            SELECT column_name, data_type, is_nullable
            FROM information_schema.columns
            WHERE table_name = 'lancamento'
            AND column_name IN ('eh_transferencia', 'transferencia_id')
            ORDER BY column_name
        """))

        estrutura = [{'coluna': row[0], 'tipo': row[1], 'nullable': row[2]} for row in result.fetchall()]

        return jsonify({
            'success': True,
            'message': '✅ Migração concluída com sucesso!',
            'mensagens': mensagens,
            'estrutura': estrutura
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

# ============================================================
# ROTA TEMPORÁRIA: ATUALIZAR TRANSFERÊNCIAS ANTIGAS
# REMOVER APÓS EXECUÇÃO EM PRODUÇÃO
# ============================================================
@app.route('/admin/atualizar-transferencias-antigas', methods=['GET'])
def atualizar_transferencias_antigas():
    """
    Rota temporária para atualizar transferências antigas que não têm
    os campos eh_transferencia e transferencia_id preenchidos.
    Acesso: /admin/atualizar-transferencias-antigas
    """
    try:
        # Buscar lançamentos com categoria "Transferência entre contas" que ainda não foram marcados
        transferencias_antigas = Lancamento.query.filter(
            Lancamento.categoria == "Transferência entre contas",
            or_(Lancamento.eh_transferencia == None, Lancamento.eh_transferencia == False)
        ).all()

        if not transferencias_antigas:
            return jsonify({
                'success': True,
                'message': 'Nenhuma transferência antiga encontrada para atualizar',
                'total': 0
            })

        # Agrupar por descrição base (sem "(Origem)" ou "(Destino)")
        transferencias_por_desc = {}
        for lanc in transferencias_antigas:
            # Remover "(Origem)" ou "(Destino)" da descrição
            desc_base = lanc.descricao.replace(' (Origem)', '').replace(' (Destino)', '')

            if desc_base not in transferencias_por_desc:
                transferencias_por_desc[desc_base] = {'origem': None, 'destino': None}

            if '(Origem)' in lanc.descricao:
                transferencias_por_desc[desc_base]['origem'] = lanc
            elif '(Destino)' in lanc.descricao:
                transferencias_por_desc[desc_base]['destino'] = lanc

        # Atualizar os pares encontrados
        pares_atualizados = 0
        pares_incompletos = []

        for desc_base, par in transferencias_por_desc.items():
            origem = par['origem']
            destino = par['destino']

            if origem and destino:
                # Marcar como transferência
                origem.eh_transferencia = True
                destino.eh_transferencia = True

                # Vincular um ao outro
                origem.transferencia_id = destino.id
                destino.transferencia_id = origem.id

                pares_atualizados += 1
            else:
                # Par incompleto
                pares_incompletos.append({
                    'descricao': desc_base,
                    'tem_origem': origem is not None,
                    'tem_destino': destino is not None
                })

        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'✅ {pares_atualizados} pares de transferências atualizados com sucesso!',
            'pares_atualizados': pares_atualizados,
            'pares_incompletos': len(pares_incompletos),
            'detalhes_incompletos': pares_incompletos if len(pares_incompletos) <= 10 else pares_incompletos[:10]
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

if __name__ == '__main__':
    # Configurar logging para suprimir logs verbosos do Werkzeug
    import logging
    log = logging.getLogger('werkzeug')
    log.setLevel(logging.ERROR)  # Mostrar apenas erros do Werkzeug

    # Criar tabelas se não existirem (sem resetar dados)
    with app.app_context():
        db.create_all()
        print("Banco de dados inicializado com todas as tabelas")

    # Verificar se há processo rodando na porta 8002
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(1)  # Timeout de 1 segundo para evitar travamento
        result = sock.connect_ex(('127.0.0.1', 8002))
        sock.close()
        if result == 0:
            print("ATENÇÃO: Porta 8002 já está em uso!")
            print("Dica: Feche outras instâncias do servidor ou use uma porta diferente")
            print("Tentando iniciar mesmo assim...")
    except Exception as e:
        print(f"⚠️  Erro ao verificar porta: {e}")
    
    # Obter endereço IP local para acesso na rede
    try:
        # Conectar a um endereço externo para descobrir o IP local
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.settimeout(2)  # Timeout de 2 segundos
        s.connect(("8.8.8.8", 80))
        local_ip = s.getsockname()[0]
        s.close()
    except Exception as e:
        print(f"⚠️  Erro ao obter IP local: {e}")
        local_ip = "127.0.0.1"
    
    print("Iniciando servidor na porta 8002...")
    print("=" * 60)
    print("ACESSO NA REDE LOCAL:")
    print(f"   📍 Endereço Local: http://127.0.0.1:8002")
    print(f"   🌍 Endereço Rede:  http://{local_ip}:8002")
    print("=" * 60)
    print("Para outros usuários acessarem na mesma rede, use o endereço da rede local")
    print("Certifique-se de que o firewall permite conexões na porta 8002")
    print("=" * 60)
    
    # Configurações otimizadas para evitar loops e melhorar performance
    app.run(
        debug=True,  # Ativar debug para ver erros detalhados
        host='0.0.0.0', 
        port=8002,
        threaded=True,  # Habilitar threading para melhor performance
        use_reloader=False  # Desabilitar reloader para evitar loops
    ) 